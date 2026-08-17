"""routers/export.py — ดึงข้อมูลออกเป็น CSV/Excel + เทมเพลตรายงาน

ย้ายมาจาก main.py แบบยกก้อน ไม่ได้แก้ตรรกะใดๆ
⚠ ห้ามประกาศ session_queues / measure_timeouts / subscribers ซ้ำในไฟล์นี้
  ต้องดึงจาก shared.py เท่านั้น ไม่งั้นจะกลายเป็นคนละ object โดยไม่มี error
"""
from fastapi import APIRouter

from shared import *  # noqa: F401,F403

router = APIRouter()


def _csv_header(key: str) -> str:
    """หัวคอลัมน์ในไฟล์ CSV — ใช้ csv_label ถ้ามี ไม่มีก็ใช้ label ตามปกติ

    มีไว้เพราะบางช่องชื่อที่เหมาะกับ "ชิปในหน้าแก้ผังรายงาน" กับ "หัวคอลัมน์ใน
    ไฟล์ CSV" ไม่ใช่ชื่อเดียวกัน เช่น timestamp: ในรายงานเลือกได้ว่าจะเอาแค่
    วันที่หรือเวลา จึงเรียกว่า "Date" แต่ CSV ออกทั้งวันและเวลาในช่องเดียวเสมอ
    เรียกว่า Date เฉยๆ จะสื่อผิด
    """
    c = EXPORT_COLUMNS[key]
    return c.get("csv_label") or c["label"]

# kind ของเทมเพลตมี 3 ค่า: 'csv' | 'pdf' | 'excel' — แยกลิสต์กันคนละหน้า
# ('report' คือชื่อเก่าสมัยที่ PDF/Excel ใช้ลิสต์ร่วมกัน ยังอ่านของเดิมได้อยู่)
# ส่วน "ชุดคอลัมน์" มีแค่ 2 แบบ: csv ใช้คอลัมน์แยก / รายงานใช้บล็อก Tolerance
def _columns_scope(kind: str) -> str:
    return "csv" if kind == "csv" else "report"

@router.get("/api/export/columns")
async def list_export_columns(kind: str = "csv"):
    """คืน catalog คอลัมน์ที่ใส่ในเทมเพลตได้ (key/label/group)
    ใช้สร้างหน้าเลือกคอลัมน์ — frontend ไม่ต้อง hardcode รายชื่อเอง

    kind='csv'          → ได้ Nominal X / Nominal Y / Upper Tol / Lower Tol แยก 4 ช่อง
    kind='pdf'|'excel'  → ได้ช่องรวม "Tolerance" ช่องเดียวแทนทั้ง 4 ช่องนั้น
    คอลัมน์ที่ไม่ระบุ scope โผล่ทั้งสองแบบ
    """
    out = []
    for k, c in EXPORT_COLUMNS.items():
        if c.get("scope") not in (None, _columns_scope(kind)):
            continue
        item = {"key": k, "label": c["label"], "group": c["group"]}
        if c.get("block"):
            # แนบ values ของ "ลูก" ในบล็อกไปด้วย (value_x/value_y ก็ตั้งหน้าตา
            # แยกตามค่าได้เหมือน Result) — ลูกไม่ได้อยู่ใน catalog เป็นตัวของตัวเอง
            blk = dict(c["block"])
            blk["data"] = [
                {**d, **({"values": EXPORT_COLUMNS[d["key"]]["values"]}
                         if EXPORT_COLUMNS.get(d["key"], {}).get("values") else {})}
                for d in c["block"]["data"]
            ]
            item["block"] = blk             # ผังหลายเซลล์ที่ frontend ต้องกางออก
        if c.get("header"):
            item["header"] = c["header"]    # ข้อความหัวตารางเริ่มต้นในรายงาน
        if c.get("values"):
            item["values"] = c["values"]    # ค่าที่เป็นไปได้ → ตั้งหน้าตาแยกตามค่าได้
        if c.get("formats"):
            item["formats"] = c["formats"]  # รูปแบบที่ติ๊กเลือกได้บนเซลล์
        out.append(item)
    return out

def _parse_columns(raw) -> List[str]:
    """แปลงค่า columns_json จาก DB (str หรือ list) เป็น list ของ key ที่ใช้ได้จริง
    — แตก key แบบเก่าเป็นคอลัมน์ย่อย และตัด key ที่ไม่รู้จักทิ้ง
    """
    cols = raw
    if isinstance(cols, str):
        try:
            cols = json.loads(cols)
        except Exception:
            cols = []
    if not isinstance(cols, list):
        cols = []

    out: List[str] = []
    for c in cols:
        for key in _LEGACY_COLUMN_ALIASES.get(c, [c]):
            if key in EXPORT_COLUMNS and key not in out:
                out.append(key)
    return out

def _get_template(cur, export_template_id: int) -> Dict[str, Any]:
    cur.execute(
        "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
        "FROM export_template WHERE export_template_id = %s",
        (export_template_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบเทมเพลตนี้")
    return row

def _parse_layout(raw):
    """แปลง layout_json จาก DB (str หรือ dict) เป็น dict — None ถ้าไม่มี/พัง"""
    if raw is None:
        return None
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return None
    return raw if isinstance(raw, dict) else None

@router.get("/api/export/templates")
async def list_export_templates(kind: str = "csv"):
    """คืนเทมเพลตของชนิดที่ระบุ — csv / pdf / excel แยกลิสต์กันคนละชนิด

    PDF กับ Excel เคยใช้ kind='report' ร่วมกัน ทำให้เทมเพลตปนกันข้ามรูปแบบ
    ตอนนี้แยกแล้ว แต่ยังดึงของเก่าที่เป็น 'report' มาแสดงด้วย จะได้ไม่หายไปเฉยๆ
    """
    kinds = [kind, "report"] if kind in ("pdf", "excel") else [kind]
    ph = ", ".join(["%s"] * len(kinds))
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                f"FROM export_template WHERE kind IN ({ph}) ORDER BY is_default DESC, name",
                kinds,
            )
            return [
                {
                    "export_template_id": r["export_template_id"],
                    "name": r["name"],
                    "kind": r["kind"],
                    "columns": _parse_columns(r["columns_json"]),
                    "layout": _parse_layout(r["layout_json"]),
                    "is_default": bool(r["is_default"]),
                }
                for r in cur.fetchall()
            ]
    finally:
        db.close()

def _validate_template_body(body: ExportTemplateBody) -> List[str]:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")
    cols = [c for c in (body.columns or []) if c in EXPORT_COLUMNS]
    if not cols:
        raise HTTPException(400, "ต้องเลือกอย่างน้อย 1 คอลัมน์")
    return cols

def _template_payload(body: ExportTemplateBody):
    """ตรวจ body แล้วคืน (kind, columns_json, layout_json) ที่พร้อมเขียนลง DB

    เก็บข้อมูลคนละคอลัมน์กันตามชนิด:
      csv          → columns_json (รายชื่อคอลัมน์ + ลำดับ)
      pdf / excel  → layout_json  (ผังตารางทั้งก้อนจากหน้า editor แบบสเปรดชีต)

    ⚠ ทุกอย่างที่ "ไม่ใช่ csv" ถือเป็นรายงานหมด — ห้ามเช็คแบบ `kind == "report"`
    เพราะตอนแยก PDF กับ Excel ออกจากกัน ค่า kind เปลี่ยนเป็น 'pdf'/'excel' แล้ว
    การเช็คชื่อตายตัวทำให้เทมเพลตรายงานตกไปเข้าเส้นทางของ CSV แล้วเด้ง
    "ต้องเลือกอย่างน้อย 1 คอลัมน์" ทั้งที่หน้าจอนั้นไม่มีให้เลือกคอลัมน์เลย
    """
    if not body.name.strip():
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")

    if body.kind != "csv":
        if not body.layout:
            raise HTTPException(400, "เทมเพลตรายงานต้องมีผังตาราง (layout)")
        kind = body.kind if body.kind in ("pdf", "excel", "report") else "pdf"
        return kind, None, json.dumps(body.layout, ensure_ascii=False)

    cols = _validate_template_body(body)
    return "csv", json.dumps(cols, ensure_ascii=False), None

@router.post("/api/export/templates", status_code=201)
async def create_export_template(body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                    "VALUES (%s, %s, %s, %s, 0)",
                    (body.name.strip(), kind, cols_json, layout_json),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"สร้างเทมเพลตไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"export_template_id": cur.lastrowid}
    finally:
        db.close()

@router.patch("/api/export/templates/{export_template_id}")
async def update_export_template(export_template_id: int, body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            # เทมเพลตตั้งต้นของระบบล็อกไว้ — ให้ Duplicate ไปแก้ตัวใหม่แทน
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นแก้ไขไม่ได้ — กด Duplicate แล้วแก้ตัวสำเนาแทน")
            try:
                cur.execute(
                    "UPDATE export_template SET name = %s, kind = %s, columns_json = %s, layout_json = %s "
                    "WHERE export_template_id = %s",
                    (body.name.strip(), kind, cols_json, layout_json, export_template_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"บันทึกไม่สำเร็จ (ชื่อนี้อาจซ้ำกับเทมเพลตอื่น): {exc}")
        return {"ok": True}
    finally:
        db.close()

@router.delete("/api/export/templates/{export_template_id}")
async def delete_export_template(export_template_id: int):
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นลบไม่ได้")
            cur.execute(
                "DELETE FROM export_template WHERE export_template_id = %s", (export_template_id,)
            )
        return {"ok": True}
    finally:
        db.close()

@router.post("/api/export/templates/{export_template_id}/duplicate", status_code=201)
async def duplicate_export_template(export_template_id: int):
    """ทำสำเนาเทมเพลต (ใช้ได้กับทุกตัวรวมทั้งตัวค่าเริ่มต้น) — ตั้งชื่อใหม่ให้
    อัตโนมัติแบบ "<ชื่อเดิม> (สำเนา)" และเติมเลขต่อท้ายถ้าชื่อนั้นถูกใช้ไปแล้ว
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            cur.execute("SELECT name FROM export_template")
            taken = {r["name"] for r in cur.fetchall()}
            base = f"{row['name']} (สำเนา)"
            name, n = base, 2
            while name in taken:
                name, n = f"{base} {n}", n + 1
            # สำเนาต้องเก็บทั้ง 2 ฟิลด์ตามชนิดของต้นฉบับ (csv ใช้ columns_json,
            # report ใช้ layout_json) — คัดลอกดิบๆ ไปเลยไม่ต้องแปลง
            cur.execute(
                "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                "VALUES (%s, %s, %s, %s, 0)",
                (
                    name,
                    row.get("kind") or "csv",
                    json.dumps(_parse_columns(row["columns_json"]), ensure_ascii=False)
                        if row.get("columns_json") else None,
                    json.dumps(_parse_layout(row["layout_json"]), ensure_ascii=False)
                        if row.get("layout_json") else None,
                ),
            )
            return {"export_template_id": cur.lastrowid, "name": name}
    finally:
        db.close()

# subquery หา measurement_id ของ "การวัดครั้งล่าสุดของแต่ละ ALPL"
# คิดจากข้อมูลทั้งตารางก่อนเสมอ (ไม่ขึ้นกับ filter อื่น) แล้วค่อยเอาไป AND กับ
# filter ที่เหลือ — ให้ความหมายตรงกับ ISLatest ที่ใช้ใน Power BI คือ
# "แถวนี้เป็นค่าล่าสุดของ ALPL นั้นไหม" เป็นคุณสมบัติของแถว ไม่ใช่ผลของการกรอง
# (เช่นกรอง Result=NG + ล่าสุด = ALPL ที่ "สถานะปัจจุบันยังไม่ผ่าน" ไม่ใช่
#  "เคยมี NG ครั้งล่าสุดในบรรดา NG ทั้งหมด" ซึ่งคนละความหมายกันคนละเรื่อง)
def _latest_only_sql(inner_where: str) -> str:
    """เงื่อนไข "เอาเฉพาะการวัดล่าสุดของแต่ละ ALPL"

    ⚠ จุดที่เคยพลาด: ของเดิม subquery เป็น `FROM measurements` เปล่าๆ ไม่มี WHERE
    จึงหา "ล่าสุด" จากทั้งตารางก่อน แล้วค่อยเอาผลไปกรองด้วยเงื่อนไขของผู้ใช้ทีหลัง
    ผลคือถ้า ALPL ตัวไหนถูกวัดซ้ำ "นอกช่วง" ที่ผู้ใช้กรอง แถวล่าสุดของมันจะตก
    ตัวกรอง แล้ว ALPL ตัวนั้นหายไปจากรายงานทั้งตัว ทั้งที่ในช่วงที่เลือกมีการวัดจริง
    เช่น กรอง 1–15 ก.ค. แต่ ALPL 400 ถูกวัดซ้ำวันที่ 20 ก.ค. → 400 หายไปเฉยๆ

    ตอนนี้ subquery ใช้ FROM/JOIN และ WHERE ชุดเดียวกับ query หลัก จึงหมายถึง
    "ล่าสุดภายในชุดข้อมูลที่กรองแล้ว" ตามที่ผู้ใช้คาดหวังจริงๆ
    (พารามิเตอร์ของ inner_where ต้องถูกส่งซ้ำอีกชุด — ดู _export_filters)
    """
    return f"""m.measurement_id IN (
    SELECT measurement_id FROM (
        SELECT m.measurement_id,
               ROW_NUMBER() OVER (
                   PARTITION BY m.number_alpl
                   ORDER BY m.timestamp DESC, m.measurement_id DESC
               ) AS rn
        {_EXPORT_FROM}
        {inner_where}
    ) AS latest WHERE latest.rn = 1
)"""

def _parse_int_ranges(spec: str, field_label: str):
    """แปลงข้อความแบบ "400-407,500,600" เป็น (ช่วง, ตัวเดี่ยว)

    คืนค่าเป็น ([(400, 407)], [500, 600]) — แยกช่วงกับตัวเดี่ยวออกจากกัน
    **ตั้งใจไม่แตกช่วงออกเป็นรายตัว** เพราะถ้าผู้ใช้พิมพ์ "1-999999" การแตกจะได้
    ค่าเกือบล้านตัวไปต่อเป็น SQL ยาว ~6.5 MB ซึ่งชน max_allowed_packet ของ MySQL
    หรือช้าจนใช้งานไม่ได้ — เก็บเป็นช่วงแล้วแปลงเป็น BETWEEN ตอนสร้าง SQL แทน
    ต้นทุนเท่ากับไม่กรองเลย (จำนวนแถวขึ้นกับข้อมูลจริง ไม่ใช่ความกว้างของช่วง)

    เคสที่รองรับ: เว้นวรรค, พิมพ์กลับด้าน (407-400), ค่าซ้ำ, ท่อนว่างจากคอมมาเกิน
    """
    ranges, singles = [], []
    for part in str(spec).split(","):
        if not part.strip():
            continue  # เช่น "400,,500" หรือคอมมาท้ายสุด — ข้ามไปเฉยๆ
        m = _RANGE_PART_RE.match(part)
        if not m:
            raise HTTPException(
                400,
                f'รูปแบบ {field_label} ไม่ถูกต้องตรง "{part.strip()}" — '
                f'ใช้ได้เช่น 400 หรือ 400-407 หรือ 400-407,500,600',
            )
        lo = int(m.group(1))
        if m.group(2) is None:
            singles.append(lo)
            continue
        hi = int(m.group(2))
        if lo > hi:
            lo, hi = hi, lo          # พิมพ์กลับด้าน (407-400) ก็ยังใช้ได้
        if lo == hi:
            singles.append(lo)       # "400-400" ก็คือตัวเดียว
        else:
            ranges.append((lo, hi))
    return sorted(set(ranges)), sorted(set(singles))

def _export_filters(f: Dict[str, Any]):
    """สร้าง WHERE + params ของหน้า Export — ใช้ร่วมกันทั้ง preview และ csv
    เพื่อให้สิ่งที่เห็นในตัวอย่างตรงกับไฟล์ที่ดาวน์โหลดจริงเสมอ
    """
    conditions, params = [], []

    def eq(col, key):
        """เงื่อนไขเท่ากับ — รองรับทั้งค่าเดียวและหลายค่า (multi-select)

        หลายค่าจะกลายเป็น IN (...) ซึ่งหมายถึง "เอาอันไหนก็ได้ในกลุ่มนี้" (OR กัน
        ภายในช่องเดียวกัน) ส่วนต่างช่องกันยังเป็น AND เหมือนเดิม เช่น
        เลือก Vendor = A,B และ Result = NG → (vendor A หรือ B) และ ต้องเป็น NG
        """
        v = f.get(key)
        if v is None or v == "":
            return
        if isinstance(v, (list, tuple, set)):
            vals = [x for x in v if x not in (None, "")]
            if not vals:
                return
            conditions.append(f"{col} IN ({','.join(['%s'] * len(vals))})")
            params.extend(vals)
        else:
            conditions.append(f"{col} = %s"); params.append(v)

    # ALPL รับได้ทั้งตัวเดียว, หลายตัวคั่นคอมมา และช่วง — ประกอบเป็นเงื่อนไข
    # เดียวที่ OR กันเอง เช่น (BETWEEN 400 AND 407 OR BETWEEN 500 AND 507 OR IN (600))
    alpl_spec = f.get("number_alpl")
    if alpl_spec not in (None, ""):
        alpl_ranges, alpl_singles = _parse_int_ranges(alpl_spec, "ALPL")
        or_parts = []
        for lo, hi in alpl_ranges:
            or_parts.append("m.number_alpl BETWEEN %s AND %s")
            params.extend([lo, hi])
        if alpl_singles:
            or_parts.append(f"m.number_alpl IN ({','.join(['%s'] * len(alpl_singles))})")
            params.extend(alpl_singles)
        if or_parts:
            conditions.append("(" + " OR ".join(or_parts) + ")")

    eq("m.result",          "result")
    eq("m.session_id",      "session_id")
    eq("op.operator_name",  "operator")
    eq("m.measure_type",    "measure_type")
    eq("v.vendor_name",     "vendor")
    eq("o.owner_name",      "owner")
    eq("pn.part_number_name", "part_number")
    eq("h.handler_name",    "handler")
    eq("ps.package_size",   "package_size")
    eq("p.po_number",       "po_number")

    # ช่วงวันที่ของ "วันที่วัด" (measurements.timestamp)
    if f.get("date_from"):
        conditions.append("m.timestamp >= %s"); params.append(_day_start(f["date_from"]))
    if f.get("date_to"):
        conditions.append("m.timestamp <= %s"); params.append(_day_end(f["date_to"]))
    # ช่วงวันที่ของ "วันที่รับชิ้นงาน" (parts_specifications.recieve_date)
    if f.get("recv_from"):
        conditions.append("p.recieve_date >= %s"); params.append(_day_start(f["recv_from"]))
    if f.get("recv_to"):
        conditions.append("p.recieve_date <= %s"); params.append(_day_end(f["recv_to"]))

    # Description ค้นแบบมีคำนี้อยู่ข้างใน (ไม่ต้องพิมพ์ตรงเป๊ะ)
    if f.get("description"):
        conditions.append("p.description LIKE %s"); params.append(f"%{f['description']}%")

    # ต้องต่อท้ายสุดเสมอ — เงื่อนไขนี้ห่อ WHERE ของ "ทุกข้อข้างบน" ไว้ข้างใน
    # ลำดับพารามิเตอร์จึงเป็น [ของข้อข้างบน..., ของข้อข้างบนซ้ำอีกรอบ...]
    # (ถ้าย้ายไปไว้กลางๆ ลำดับ %s จะเพี้ยนแล้วได้ข้อมูลผิดแบบเงียบๆ)
    if f.get("latest_only"):
        inner_where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        inner_params = list(params)
        conditions.append(_latest_only_sql(inner_where))
        params.extend(inner_params)

    return (("WHERE " + " AND ".join(conditions)) if conditions else ""), params

def _fetch_export_rows(cols: List[str], where: str, params: list, limit: Optional[int] = None):
    """ดึงข้อมูลแล้วแปลงเป็น list ของ list ตามลำดับคอลัมน์ใน `cols`"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp DESC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            rows = cur.fetchall()
    finally:
        db.close()
    data = [[EXPORT_COLUMNS[c]["get"](r) for c in cols] for r in rows]
    return data, total

def _count_export_rows(where: str, params: list) -> int:
    """นับจำนวนแถวอย่างเดียว — ใช้เช็คเพดานก่อนดึงข้อมูลจริงมาสร้างไฟล์"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            return cur.fetchone()["n"]
    finally:
        db.close()

def _fetch_export_raw(where: str, params: list, limit: Optional[int] = None):
    """เหมือน _fetch_export_rows แต่คืน row ดิบ (dict) ไม่ได้แปลงเป็นคอลัมน์
    ใช้กับรายงาน PDF/Excel ที่ต้องหยิบค่าทีละช่องตามผังตาราง ไม่ใช่เรียงเป็นแถว
    เรียง ASC ตามเวลา เพราะรายงานบนกระดาษอ่านจากเก่าไปใหม่ (ต่างจาก CSV ที่
    เรียง DESC ให้เห็นตัวล่าสุดก่อน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp ASC, m.measurement_id ASC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            return cur.fetchall(), total
    finally:
        db.close()

def _cell_out(cell: Dict[str, Any], text: str) -> Dict[str, Any]:
    """แปลงเซลล์ในผังเป็นเซลล์ผลลัพธ์ — เก็บเฉพาะที่ตัววาดต้องใช้"""
    out = {"v": text, "s": cell.get("s") or {}}
    span = cell.get("span")
    if span:
        out["span"] = {"r": int(span.get("r", 1)), "c": int(span.get("c", 1))}
    if cell.get("hidden"):
        out["hidden"] = True
    # ทำเครื่องหมายไว้ว่าเซลล์ไหนมาจากข้อมูลจริง — ตัววาดใช้ตัดสินใจตีเส้นขอบ
    if cell.get("f") or cell.get("spec"):
        out["data"] = True
    if cell.get("hdr"):
        out["head"] = True
    return out

def _render_report(layout: Dict[str, Any], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """คลี่ผัง + ข้อมูล ออกเป็นตารางพร้อมวาด

    คืน {"nCols": int, "rows": [[cell, ...], ...]} โดย cell = {v, s, span?, hidden?}
    """
    grid = layout.get("grid") or []
    n_rows = int(layout.get("nRows") or len(grid))
    n_cols = int(layout.get("nCols") or (len(grid[0]) if grid else 0))
    data_row = int(layout.get("dataRow", 2))
    if not grid or n_cols == 0:
        return {"nCols": 0, "rows": []}
    data_row = max(0, min(data_row, n_rows - 1))

    def tpl_row(r: int) -> List[Dict[str, Any]]:
        row = grid[r] if r < len(grid) else []
        return [row[c] if c < len(row) else {} for c in range(n_cols)]

    def clamp(cells: List[Dict[str, Any]], r: int, last: int) -> List[Dict[str, Any]]:
        """ตัด rowspan ไม่ให้ทะลุออกนอกบล็อกที่กำลังพิมพ์อยู่

        ผู้ใช้ผสานเซลล์คร่อมแถวข้อมูลได้ในหน้าแก้ไข แต่ตอนคลี่ออกมา ส่วนหัวของ
        กลุ่มถัดไปจะมาต่อท้าย ถ้าปล่อย rowspan เดิมไว้มันจะกินทับแถวของกลุ่มถัดไป
        แล้วช่วงผสานซ้อนกัน → Excel เปิดไฟล์ไม่ขึ้น บอกว่าไฟล์เสียหาย
        """
        for cell in cells:
            sp = cell.get("span")
            if sp and sp["r"] > 1:
                sp["r"] = max(1, min(sp["r"], last - r + 1))
        return cells

    def render_static(r: int, spec_text: str) -> List[Dict[str, Any]]:
        """แถวส่วนหัว/ส่วนท้าย — แทนเฉพาะเซลล์ spec ด้วยค่าของกลุ่มปัจจุบัน
        เซลล์ f (ข้อมูลรายชิ้น) ที่หลุดมาอยู่นอกแถวข้อมูลถือว่าไม่มีความหมาย
        ปล่อยว่างไว้ ดีกว่าเอาค่าของชิ้นแรกมาใส่แบบมั่วๆ
        """
        out = []
        for cell in tpl_row(r):
            if cell.get("spec"):
                out.append(_cell_out(cell, spec_text))
            elif cell.get("f"):
                out.append(_cell_out(cell, ""))
            else:
                out.append(_cell_out(cell, cell.get("v") or ""))
        return out

    def render_data(row: Dict[str, Any], item_no: int) -> List[Dict[str, Any]]:
        """แถวข้อมูล 1 แถวต่อ 1 การวัด — rowspan ถูกบีบเป็น 1 เพราะแถวนี้ถูก
        ทำซ้ำหลายรอบ ถ้าปล่อย rowspan เดิมไว้จะกินทับแถวของการวัดชิ้นถัดไป
        """
        out = []
        for cell in tpl_row(data_row):
            key = cell.get("f")
            col = EXPORT_COLUMNS.get(key) if key else None

            # ช่องลำดับแถว (Item) — ค่าไม่ได้อยู่ใน DB ต้องนับตอนคลี่ผัง
            if col and col.get("row_number"):
                text = str(item_no)
            # รูปแบบที่ผู้ใช้ติ๊กไว้บนเซลล์ (ตอนนี้มีแค่ช่องวันเวลา: date/time/ทั้งคู่)
            elif col and col.get("get_fmt") and cell.get("fmt"):
                text = col["get_fmt"](row, cell["fmt"])
            elif col:
                text = col["get"](row)
            else:
                text = cell.get("v") or ""

            c = _cell_out(cell, str(text))
            if "span" in c:
                c["span"]["r"] = 1

            # หน้าตาแยกตามค่า — เช่น Result OK พื้นเขียว / NG พื้นแดง
            # "สถานะ" ของเซลล์มาจากฟังก์ชัน state ของคอลัมน์ ไม่ใช่ข้อความที่แสดง
            # เพราะบางคอลัมน์สถานะกับค่าที่แสดงเป็นคนละเรื่อง เช่น Value X แสดง
            # "4.089" แต่สถานะคือ NG (เกินสเปก) — ถ้าจับคู่ด้วยข้อความจะไม่เจอเลย
            variants = cell.get("variants") or {}
            if variants:
                state = col["state"](row) if col and col.get("state") else str(text)
                if state in variants:
                    c["s"] = variants[state]
            out.append(c)
        return out

    # จัดกลุ่มตามสเปก โดยรักษาลำดับที่เจอครั้งแรกไว้ (ไม่เรียงใหม่)
    #
    # แบ่งกลุ่มเฉพาะเมื่อผังมีช่อง Tolerance อยู่จริงเท่านั้น — ถ้าผู้ใช้ทำ
    # เทมเพลตที่ไม่ได้ลาก Tolerance ลงไป การแบ่งกลุ่มจะทำให้หัวตารางถูกพิมพ์ซ้ำ
    # หลายรอบโดยไม่มีอะไรบนกระดาษบอกว่าทำไมถึงแยกก้อน ผู้ใช้จะงงว่าเป็นบั๊ก
    has_spec = any(
        (cell or {}).get("spec") or (cell or {}).get("hdr") == REPORT_GROUP_BY
        for row in grid for cell in row
    )
    groups: Dict[str, List[Dict[str, Any]]] = {}
    if has_spec:
        for r in rows:
            groups.setdefault(EXPORT_COLUMNS[REPORT_GROUP_BY]["get"](r), []).append(r)
    elif rows:
        groups[""] = list(rows)

    # แถวส่วนหัวแบ่งเป็น 2 พวก:
    #   - แถวที่มีช่องข้อมูลผูกอยู่ (ชื่อคอลัมน์ / สเปกของกลุ่ม) → ซ้ำทุกกลุ่ม
    #     เพราะแต่ละกลุ่มเป็นตารางของมันเอง ต้องมีหัวตารางกำกับ
    #   - แถวที่เป็นข้อความที่ผู้ใช้พิมพ์เองล้วนๆ (เช่น "ตรวจสอบการวัดวันที่ ...")
    #     → พิมพ์ครั้งเดียวตอนกลุ่มแรก เพราะเป็นหัวเรื่องของทั้งรายงาน ไม่ใช่ของ
    #     ตารางใดตารางหนึ่ง ถ้าซ้ำทุกกลุ่มจะกลายเป็นชื่อเรื่องโผล่กลางหน้า
    # เช็คที่เซลล์ต้นของการผสาน (ข้าม hidden) เพราะหัวเรื่องมักผสานยาวทั้งแถว
    def _repeats_per_group(r: int) -> bool:
        return any(
            (c.get("hdr") or c.get("spec") or c.get("f"))
            for c in tpl_row(r) if not c.get("hidden")
        )

    repeat_head = {r: _repeats_per_group(r) for r in range(0, data_row)}

    out_rows: List[List[Dict[str, Any]]] = []
    # ลำดับแถว (Item) นับต่อเนื่องทั้งรายงาน ไม่รีเซ็ตตอนขึ้นกลุ่มใหม่ —
    # ตรงกับรายงานกระดาษที่ห้อง PM Kit ใช้ ซึ่งเลขลำดับคือ "ชิ้นที่เท่าไรของใบนี้"
    item_no = 0
    for gi, (spec_text, members) in enumerate(groups.items()):
        for r in range(0, data_row):
            if gi and not repeat_head[r]:
                continue
            out_rows.append(clamp(render_static(r, spec_text), r, data_row - 1))
        for m in members:
            item_no += 1
            out_rows.append(render_data(m, item_no))
    # ส่วนท้ายพิมพ์ครั้งเดียวตอนจบ (ปกติเป็นช่องเซ็นชื่อ/ผู้ตรวจ)
    for r in range(data_row + 1, n_rows):
        out_rows.append(clamp(render_static(r, ""), r, n_rows - 1))

    # ── ตัดคอลัมน์/แถวว่างที่ห้อยท้ายทิ้ง ──────────────────────────────
    # ผังในหน้าแก้ไขมี 10 คอลัมน์ 5 แถวเสมอ แต่ผู้ใช้ใช้จริงไม่กี่ช่อง ถ้าปล่อย
    # ช่องว่างติดไปด้วย เวลาพิมพ์ PDF ตารางจะถูกจัดกึ่งกลางโดยนับช่องว่างพวกนั้น
    # เข้าไปด้วย ทำให้ส่วนที่มีเนื้อหาดูเบี้ยวไปทางซ้าย และมีเส้นจางๆ ห้อยอยู่
    # ทางขวาของกระดาษ
    def filled(cell: Dict[str, Any]) -> bool:
        return bool((cell.get("v") or "").strip() or cell.get("data") or cell.get("head"))

    last_col = 0
    for row in out_rows:
        for c, cell in enumerate(row):
            if not cell.get("hidden") and filled(cell):
                last_col = max(last_col, c + (cell.get("span", {}) or {}).get("c", 1))
    last_col = last_col or n_cols

    for row in out_rows:
        del row[last_col:]
        for c, cell in enumerate(row):
            sp = cell.get("span")
            if sp and c + sp["c"] > last_col:
                sp["c"] = max(1, last_col - c)

    while out_rows and not any(filled(x) for x in out_rows[-1]):
        out_rows.pop()

    return {"nCols": last_col, "rows": out_rows, "groups": len(groups)}

def _load_report_layout(cur, export_template_id: int) -> Dict[str, Any]:
    tpl = _get_template(cur, export_template_id)
    layout = _parse_layout(tpl.get("layout_json"))
    if not layout or not layout.get("grid"):
        raise HTTPException(400, f'เทมเพลต "{tpl["name"]}" ยังไม่มีผังตาราง — เปิดแก้ไขแล้วบันทึกใหม่อีกครั้ง')
    return {"name": tpl["name"], "kind": tpl.get("kind") or "pdf", "layout": layout}

@router.get("/api/export/preview")
async def export_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    limit:   int = Query(5, ge=1, le=50),
):
    """คืนหัวคอลัมน์ + ข้อมูลตัวอย่างไม่กี่แถว + จำนวนแถวทั้งหมดที่ตรงกับ filter
    ให้หน้าเว็บโชว์ก่อนกดดาวน์โหลดจริง (จะได้รู้ว่ากรองถูกไหม ไฟล์ใหญ่แค่ไหน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            tpl = _get_template(cur, export_template_id)
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, total = _fetch_export_rows(cols, where, params, limit=limit)
    return {
        "template_name": tpl["name"],
        "columns": [_csv_header(c) for c in cols],
        "column_keys": cols,
        "rows": data,
        "total": total,
    }

@router.get("/api/export/csv")
async def export_csv(
    export_template_id: Optional[int] = None,
    filename: Optional[str] = None,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """Export ประวัติ measurement (พร้อม filter) เป็นไฟล์ CSV ให้ดาวน์โหลด

    export_template_id: เลือกว่าจะเอาคอลัมน์ไหนและเรียงลำดับยังไง — ถ้าไม่ส่งมา
    จะใช้เทมเพลตที่ is_default = 1 ให้อัตโนมัติ (พฤติกรรมเดิมของ endpoint นี้
    คือ dump ทุกคอลัมน์ ซึ่งตรงกับเทมเพลตค่าเริ่มต้นพอดี ของเดิมที่เรียกมาโดย
    ไม่ส่งพารามิเตอร์นี้จึงยังใช้งานได้เหมือนเดิม)

    ใช้ utf-8-sig encoding เพื่อให้เปิดใน Excel ได้ถูกต้องแม้มีตัวอักษรไทยอยู่ใน
    ไฟล์ (BOM ช่วยไม่ให้ตัวอักษรเพี้ยน) และใช้ filter ชุดเดียวกับ /api/export/preview
    เพื่อให้ไฟล์ที่ได้ตรงกับตัวอย่างที่เห็นบนหน้าจอเสมอ
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            if export_template_id is not None:
                tpl = _get_template(cur, export_template_id)
            else:
                cur.execute(
                    "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                    "FROM export_template WHERE kind = 'csv' AND is_default = 1 LIMIT 1"
                )
                tpl = cur.fetchone()
                if not tpl:
                    raise HTTPException(404, "ยังไม่มีเทมเพลตค่าเริ่มต้นในระบบ")
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, _ = _fetch_export_rows(cols, where, params)

    df  = pd.DataFrame(data, columns=[_csv_header(c) for c in cols])
    buf = StringIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)

    # ชื่อไฟล์ใส่วันที่ให้ด้วย — ดาวน์โหลดหลายรอบจะได้ไม่ทับกันใน Downloads
    # ชื่อไฟล์ที่ผู้ใช้กรอกในหน้า Export มาก่อน — ถ้าไม่ส่งมาค่อยตั้งชื่อตามเวลาให้
    # กรองตัวอักษรที่ใช้ในชื่อไฟล์ Windows ไม่ได้ออกอีกชั้น (ฝั่งหน้าเว็บกรองแล้ว
    # แต่ endpoint นี้เรียกตรงจาก URL ได้ จึงต้องกันเองด้วย ไม่เชื่อ input จากข้างนอก)
    safe_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "", (filename or "")).strip(". ")
    fname = (f"{safe_name}.csv" if safe_name
             else f"measurements_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
    return StreamingResponse(
        iter(["﻿" + buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

def _guard_report_size(total: int) -> None:
    if total > REPORT_MAX_ROWS:
        raise HTTPException(
            413,
            f"ข้อมูลที่เลือกมี {total:,} แถว เกินเพดาน {REPORT_MAX_ROWS:,} แถวของรายงาน "
            f"— กรองให้แคบลงก่อน (เช่น จำกัดช่วงวันที่) หรือใช้ Export CSV แทน "
            f"ถ้าต้องการข้อมูลดิบทั้งหมด",
        )

@router.get("/api/export/report-preview")
async def export_report_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    full: int = 0,
):
    """คืนผังที่คลี่แล้วพร้อมข้อมูลจริง ให้หน้าเว็บวาดเป็นตาราง preview

    full=0 → ตัดที่ REPORT_PREVIEW_LIMIT แถว (ดูบนจอเฉยๆ ไม่ต้องครบ)
    full=1 → เอาครบทุกแถว ใช้ตอนสั่งพิมพ์เป็น PDF จริง — ไฟล์ที่ได้ต้องไม่ขาด
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    if full:
        _guard_report_size(total_check := _count_export_rows(where, params))
    rows, total = _fetch_export_raw(where, params, limit=None if full else REPORT_PREVIEW_LIMIT)
    out = _render_report(tpl["layout"], rows)
    out.update({
        "template_name": tpl["name"],
        "total": total,
        "shown": len(rows),
        "truncated": total > len(rows),
    })
    return out

@router.get("/api/export/xlsx")
async def export_xlsx(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """สร้างไฟล์ .xlsx จากผังรายงาน — คงฟอนต์/สี/การผสานเซลล์ตามที่จัดไว้"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            500,
            "ยังไม่ได้ติดตั้ง openpyxl — รัน `pip install openpyxl` ในเครื่องที่รัน Backend ก่อน",
        )

    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    rows, total = _fetch_export_raw(where, params)
    _guard_report_size(total)
    rendered = _render_report(tpl["layout"], rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    thick = Side(style="medium", color="000000")
    box = Border(left=thick, right=thick, top=thick, bottom=thick)

    def rgb(v: Optional[str]) -> Optional[str]:
        """#RRGGBB → RRGGBB (openpyxl ไม่รับเครื่องหมาย #)"""
        if not v or not isinstance(v, str):
            return None
        v = v.strip().lstrip("#")
        return v.upper() if re.fullmatch(r"[0-9A-Fa-f]{6}", v) else None

    # ผังเก็บค่าการจัดวางเป็นคำของ CSS แต่ openpyxl ใช้คำคนละชุด ถ้าส่งคำที่มัน
    # ไม่รู้จักไปจะโยน ValueError ทันที (เคสจริงที่เจอ: valign='middle' ของ CSS
    # ซึ่ง openpyxl เรียกว่า 'center') — แปลงและกรองให้เหลือเฉพาะค่าที่ยอมรับได้
    _H = {"left": "left", "center": "center", "right": "right", "justify": "justify"}
    _V = {"top": "top", "middle": "center", "center": "center", "bottom": "bottom"}

    def align_of(s: Dict[str, Any]) -> "Alignment":
        try:
            indent = max(0, min(250, int(s.get("indent") or 0)))
        except (TypeError, ValueError):
            indent = 0
        return Alignment(
            horizontal=_H.get(str(s.get("align") or "").lower(), "center"),
            vertical=_V.get(str(s.get("valign") or "").lower(), "center"),
            indent=indent,
            wrap_text=False,
        )

    def size_of(s: Dict[str, Any]) -> float:
        try:
            v = float(s.get("size") or 11)
        except (TypeError, ValueError):
            return 11.0
        return v if 1 <= v <= 409 else 11.0

    for r, row in enumerate(rendered["rows"], start=1):
        for c, cell in enumerate(row, start=1):
            if cell.get("hidden"):
                continue
            s = cell.get("s") or {}
            x = ws.cell(row=r, column=c, value=cell.get("v") or "")
            x.font = Font(
                name=str(s.get("font") or "Tahoma"),
                size=size_of(s),
                bold=bool(s.get("bold")),
                italic=bool(s.get("italic")),
                underline="single" if s.get("underline") else None,
                color=rgb(s.get("color")) or "000000",
            )
            fill = rgb(s.get("fill"))
            if fill:
                x.fill = PatternFill("solid", fgColor=fill)
            x.alignment = align_of(s)
            # เส้นขอบหนารอบเซลล์ที่มีเนื้อหา — ตรงกับที่ตกลงไว้ว่าในหน้าแก้ไข
            # ไม่ต้องโชว์เส้น แต่ตอน export ต้องมี
            if cell.get("v") or cell.get("data") or cell.get("head"):
                x.border = box

            span = cell.get("span")
            if span and (span["r"] > 1 or span["c"] > 1):
                r2 = r + max(1, span["r"]) - 1
                c2 = min(c + max(1, span["c"]) - 1, rendered["nCols"])
                # ข้ามถ้าซ้อนกับช่วงที่ผสานไปแล้ว — openpyxl ไม่ห้ามตอนสร้าง แต่
                # ไฟล์ที่ได้จะเปิดไม่ขึ้นใน Excel (บอกว่าไฟล์เสียหาย) ยอมให้ผสาน
                # ไม่ครบดีกว่าได้ไฟล์ที่เปิดไม่ได้เลย
                if not any(
                    m.min_row <= r2 and r <= m.max_row and m.min_col <= c2 and c <= m.max_col
                    for m in ws.merged_cells.ranges
                ):
                    ws.merge_cells(start_row=r, start_column=c, end_row=r2, end_column=c2)

    for c in range(1, (rendered["nCols"] or 1) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        log.exception("สร้างไฟล์ xlsx ไม่สำเร็จ")
        raise HTTPException(500, f"สร้างไฟล์ Excel ไม่สำเร็จ: {type(exc).__name__}: {exc}")
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", tpl["name"]).strip("_") or "report"
    fname = f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
