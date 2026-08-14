
import asyncio
import json
import logging
import os
import re
import secrets
import shutil
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

import httpx
import pandas as pd
import pymysql
import pymysql.cursors
from pymysql.constants import CLIENT
from dotenv import load_dotenv
from fastapi import Body, Depends, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from PIL import Image
from pydantic import BaseModel
from sse_starlette.sse import EventSourceResponse

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

DB_CONFIG = dict(
    host=os.getenv("DB_HOST", "localhost"),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "tmx_db"),
    port=int(os.getenv("DB_PORT", 3306)),
    cursorclass=pymysql.cursors.DictCursor,
    autocommit=True,
    client_flag=CLIENT.FOUND_ROWS,
)

_PROJECT_ROOT = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
)

ALPL_IMAGE_DIR = os.getenv("ALPL_IMAGE_DIR", os.path.join(_PROJECT_ROOT, "image_ALPL"))
if not os.path.isabs(ALPL_IMAGE_DIR):
    ALPL_IMAGE_DIR = os.path.abspath(os.path.join(_PROJECT_ROOT, ALPL_IMAGE_DIR))


_TOL_EPS = 1e-6


def _within_tolerance(value: float, nominal: float, upper_tol: float, lower_tol: float) -> bool:
    """เช็คว่าค่าที่วัดได้ของแกนหนึ่งอยู่ในช่วง nominal -lower_tol .. +upper_tol ไหม
    (นับค่าที่ตกขอบพอดีว่าผ่าน — ดู _TOL_EPS)
    """
    return (nominal - lower_tol - _TOL_EPS) <= value <= (nominal + upper_tol + _TOL_EPS)


def _offset_ok(offset: Optional[float], offset_tol: Optional[float]) -> bool:
    """offset ผ่านเกณฑ์ไหม — เทียบกับ offset_tol ของ part_number

    ต่างจาก value_x/value_y ที่เทียบกับช่วง nominal ± tolerance — offset เป็น
    "ค่าความเยื้อง" ที่ยิ่งน้อยยิ่งดี จึงเทียบแค่ว่าไม่เกินเพดานที่ตั้งไว้
    ใช้ abs() เผื่อ TM-X ส่งค่าติดลบมา (เยื้องคนละทิศก็ถือว่าเยื้องเท่ากัน)

    ถ้ายังไม่ได้ตั้ง offset_tol (NULL) ถือว่า "ไม่ตรวจข้อนี้" → ผ่านเสมอ
    ไม่งั้น Part เก่าที่ยังไม่ได้กรอกค่านี้จะกลายเป็น NG ทั้งหมดทันทีที่ deploy
    """
    if offset is None or offset_tol is None:
        return True
    return abs(offset) <= offset_tol + _TOL_EPS



def _offset_limit(measure_type: Optional[str], row) -> Optional[float]:
    """เพดาน offset ที่ต้องตรวจจริงของการวัดครั้งนี้ — `None` = ไม่นับเป็นเกณฑ์

    อย่าเรียก `_offset_ok(offset, row["offset_tol"])` ตรงๆ อีก ให้ผ่านฟังก์ชันนี้
    เสมอ เพราะ `_offset_ok` ผูกกับ "ตั้งค่าไว้ไหม" (NULL = ไม่ตรวจ) ไม่ได้ผูกกับ
    โหมด — ตัดสินใจเรื่องโหมดต้องเกิดที่นี่ที่เดียว
    """
    if (measure_type or "").upper() == "IPM":
        return None
    return row.get("offset_tol") if row else None


def _load_criteria(cur, number_alpl: int, measure_type: str):
    """ดึง nominal/tolerance ที่จะใช้ตัดสิน ALPL ตัวนี้ ตามโหมดที่กำลังวัด

    คืน dict ที่มี nominal_x/nominal_y/upper_tol/lower_tol/offset_tol เหมือนกัน
    ทั้ง 2 โหมด ผู้เรียกจึงใช้ต่อได้โดยไม่ต้องรู้ว่ามาจากตารางไหน
    (offset_tol ที่คืนมาเป็นค่า "ตามตาราง" เฉยๆ — จะเอามาตัดสินจริงไหมให้ถาม
    `_offset_limit()` อีกที)

    ไม่พบ → raise HTTPException พร้อมบอกว่าขาดตรงไหนและไปแก้ที่หน้าไหน
    """
    if (measure_type or "").upper() == "IPM":
        cur.execute(
            "SELECT ps.nominal_x, ps.nominal_y, ps.upper_tol, ps.lower_tol, ps.offset_tol, "
            "       ps.package_size "
            "FROM parts_specifications p "
            "LEFT JOIN part_number pn  ON p.part_number_id = pn.part_number_id "
            "JOIN package_size ps      ON ps.package_size_id = "
            "                             COALESCE(p.package_size_id, pn.package_size_id) "
            "WHERE p.number_alpl = %s",
            (number_alpl,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(
                404,
                f"ALPL {number_alpl} หาเกณฑ์ตัดสินไม่เจอ — ยังไม่ได้ผูก Package Size "
                f"ให้ ALPL นี้ (แก้ที่หน้า Edit › Parts)",
            )
        return row

    cur.execute(
        "SELECT pn.nominal_x, pn.nominal_y, pn.upper_tol, pn.lower_tol, pn.offset_tol, "
        "       pn.part_number_name "
        "FROM parts_specifications p "
        "JOIN part_number pn ON p.part_number_id = pn.part_number_id "
        "WHERE p.number_alpl = %s",
        (number_alpl,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(
            404,
            f"ALPL {number_alpl} หาเกณฑ์ตัดสินไม่เจอ — ยังไม่ได้ตั้ง Part Number "
            f"ให้ ALPL นี้ (โหมด New/Rework ใช้เกณฑ์จาก Part Number)",
        )
    return row


def _judge(value_x: float, value_y: float, offset: Optional[float],
           crit, measure_type: str) -> dict:
    """ตัดสิน OK/NG ครั้งเดียวจบ — ใช้ร่วมกันทั้งตอนวัดจริงและตอนคำนวณใหม่หลังแก้ไข

    คืน dict ที่เอาไปแนบ SSE ได้เลย เพื่อให้ "สิ่งที่บันทึก" กับ "สิ่งที่หน้าเว็บ
    เห็น" มาจากการคำนวณชุดเดียวกันเสมอ (เคยแยกกันแล้วเพี้ยนกันเงียบๆ)
    """
    ok_x = _within_tolerance(value_x, crit["nominal_x"], crit["upper_tol"], crit["lower_tol"])
    ok_y = _within_tolerance(value_y, crit["nominal_y"], crit["upper_tol"], crit["lower_tol"])

    limit = _offset_limit(measure_type, crit)
    offset_counts = limit is not None
    ok_offset = _offset_ok(offset, limit) if offset_counts else None

    passed = ok_x and ok_y and (ok_offset is not False)
    return {
        "result":        "OK" if passed else "NG",
        "ok_x":          ok_x,
        "ok_y":          ok_y,
        "ok_offset":     ok_offset,
        "offset_counts": offset_counts,
        "offset_tol":    limit,
    }


def _thai_date_str(dt: Optional[datetime] = None) -> str:
    """คืนวันที่รูปแบบ DD-MM-YYYY โดยปีเป็น พ.ศ. (ค.ศ. + 543) เช่น 22-07-2569
    ใช้ตั้งชื่อโฟลเดอร์/ไฟล์รูปภาพแบบแยกตามวันที่ (ดู upload_measurement_image)
    ค่า default (dt=None) ใช้เวลาปัจจุบันของเครื่องที่รัน backend ณ ตอนที่รูป
    ถูกอัปโหลดเข้ามา (ไม่ใช่เวลาที่วัด — สองอย่างนี้ในทางปฏิบัติเป็นเวลาเดียวกัน
    เพราะ Agent อัปโหลดรูปทันทีหลัง trigger แต่ละชิ้น)
    """
    dt = dt or datetime.now()
    return f"{dt.day:02d}-{dt.month:02d}-{dt.year + 543}"


DELETED_DIR = os.getenv("DELETED_DIR", os.path.join(_PROJECT_ROOT, "Deleted"))
if not os.path.isabs(DELETED_DIR):
    DELETED_DIR = os.path.abspath(os.path.join(_PROJECT_ROOT, DELETED_DIR))

DELETED_RETENTION_DAYS = int(os.getenv("DELETED_RETENTION_DAYS", 30))


def _json_safe(value):
    """แปลงค่าจาก DB ให้ json.dumps ได้ — datetime/date/Decimal ไม่ใช่ชนิดมาตรฐาน"""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, timedelta):
        return str(value)
    return value


def _archive_before_delete(
    kind: str,
    table: str,
    pk: Dict[str, Any],
    row: Dict[str, Any],
    related: Optional[Dict[str, List[Dict]]] = None,
    image_path: Optional[str] = None,
) -> Optional[str]:
    """เก็บข้อมูลที่กำลังจะถูกลบไว้ใน Deleted/ — คืนชื่อไฟล์ JSON ที่สร้าง

    ⚠ ต้องเรียก "ก่อน" คำสั่ง DELETE เสมอ เพราะต้องอ่านค่าจากแถวที่ยังอยู่
    ⚠ ห้ามโยน exception ออกไป — ถ้าสำรองไม่สำเร็จก็ยังต้องยอมให้ลบต่อได้
      (ผู้ใช้สั่งลบแล้ว การไปขวางเพราะเขียนไฟล์สำรองไม่ได้จะงงกว่า) แต่ต้อง
      log ไว้ให้เห็นชัดว่ารอบนี้ไม่มีตัวสำรอง

    kind ใช้เป็นทั้งชื่อไฟล์และตัวบอกประเภทตอนกู้คืน:
      measurement | part | operator | owner | vendor | handler |
      template | package_size | part_number
    """
    try:
        day_dir = os.path.join(DELETED_DIR, _thai_date_str())
        os.makedirs(day_dir, exist_ok=True)

        pk_value = "_".join(str(v) for v in pk.values())
        stem = f"{kind}_{pk_value}"

        image_file = None
        if image_path and "://" not in image_path:
            src = os.path.realpath(os.path.join(ALPL_IMAGE_DIR, image_path))
            base = os.path.realpath(ALPL_IMAGE_DIR)
            if (src == base or src.startswith(base + os.sep)) and os.path.isfile(src):
                image_file = f"{stem}_{os.path.basename(image_path)}"
                try:
                    shutil.move(src, os.path.join(day_dir, image_file))
                except OSError as exc:
                    log.warning("ย้ายไฟล์รูปเข้าถังขยะไม่สำเร็จ (%s): %s", image_path, exc)
                    image_file = None

        payload = {
            "deleted_at": datetime.now().isoformat(sep=" ", timespec="seconds"),
            "kind":       kind,
            "table":      table,
            "pk":         {k: _json_safe(v) for k, v in pk.items()},
            "row":        {k: _json_safe(v) for k, v in row.items()},
            "related":    {
                t: [{k: _json_safe(v) for k, v in r.items()} for r in rows]
                for t, rows in (related or {}).items()
            },
            "image_file": image_file,
        }

        json_name = f"{stem}.json"
        n = 2
        while os.path.exists(os.path.join(day_dir, json_name)):
            json_name = f"{stem}({n}).json"
            n += 1

        with open(os.path.join(day_dir, json_name), "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        log.info("เก็บเข้าถังขยะแล้ว: %s/%s", _thai_date_str(), json_name)
        return json_name
    except Exception as exc:
        log.error("สำรองข้อมูลก่อนลบไม่สำเร็จ (%s %s): %s — ลบต่อโดยไม่มีตัวสำรอง", kind, pk, exc)
        return None


def _fetch_one(cur, sql: str, params) -> Optional[Dict[str, Any]]:
    cur.execute(sql, params)
    return cur.fetchone()


def _delete_image_file(image_path: Optional[str]) -> bool:
    """ลบไฟล์รูปที่ image_path (path สัมพัทธ์ต่อ ALPL_IMAGE_DIR) ชี้อยู่

    คืน True เฉพาะตอนลบได้จริง และ **ไม่ throw ไม่ว่ากรณีไหน** เพราะทุกจุดที่
    เรียกใช้ถือว่า "ลบไฟล์ไม่ได้" ไม่ใช่เหตุให้ request ทั้งก้อนพัง (แถวใน DB
    ถูกลบ/อัปเดตไปแล้ว จะ rollback เพราะไฟล์ค้างไม่คุ้ม)

    กันไว้ 2 กรณีที่ห้ามแตะไฟล์เด็ดขาด:
      1. **ค่าที่เป็น URL เต็ม** — ข้อมูลตกค้างจากยุค MinIO ที่เก็บทั้ง URL ลง
         คอลัมน์นี้ (เจอจริงในแถว measurement_id=10:
         "http://172.20.10.4:8080/images/test.png") ไม่ใช่ path บนดิสก์เครื่องนี้
         ถ้าเอาไป join กับ ALPL_IMAGE_DIR จะได้ path มั่วๆ ที่ไม่ควรไปยุ่งด้วย
      2. **path ที่หลุดออกนอก ALPL_IMAGE_DIR** — ถ้าค่าใน DB มี ".." ปนมา
         (พลาดเองหรือถูกยัดมา) ต้องไม่ทำให้ลบไฟล์นอกโฟลเดอร์รูปได้
    """
    if not image_path:
        return False

    if "://" in image_path:
        log.warning(
            "ข้ามการลบไฟล์รูป: image_path เป็น URL ไม่ใช่ path บนดิสก์ (%s)", image_path
        )
        return False

    base = os.path.realpath(ALPL_IMAGE_DIR)
    target = os.path.realpath(os.path.join(base, image_path))
    if target != base and not target.startswith(base + os.sep):
        log.warning(
            "ข้ามการลบไฟล์รูป: path หลุดออกนอก ALPL_IMAGE_DIR (%s)", image_path
        )
        return False

    try:
        os.remove(target)
        return True
    except OSError:
        return False

AGENT_HOST      = os.getenv("AGENT_HOST", "localhost")
AGENT_PORT      = int(os.getenv("AGENT_PORT", 9998))
AGENT_BASE_URL  = f"http://{AGENT_HOST}:{AGENT_PORT}"

HEARTBEAT_INTERVAL = int(os.getenv("HEARTBEAT_INTERVAL", 5))
HEARTBEAT_TIMEOUT  = int(os.getenv("HEARTBEAT_TIMEOUT", 15))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [Server] %(message)s")
log = logging.getLogger(__name__)

CORS_ORIGINS = [
    o.strip() for o in os.getenv("CORS_ORIGINS", "*").split(",") if o.strip()
]

subscribers: List[asyncio.Queue] = []

session_queues: Dict[int, Dict[str, Any]] = {}


async def push_event(event_type: str, data: dict) -> None:
    """กระจาย (broadcast) เหตุการณ์ SSE หนึ่งรายการไปให้ทุก client ที่เปิดหน้า dashboard อยู่

    ทำไมต้องทำแบบนี้: backend คือ Single Source of Truth ของสถานะ session/measurement
    ดังนั้นเมื่อสถานะมีการเปลี่ยนแปลง (เริ่ม session, มี measurement ใหม่, timeout ฯลฯ)
    ทุกแท็บ dashboard ที่เปิดอยู่ต้องรู้ทันที แต่ละ subscriber มี asyncio.Queue ของตัวเอง
    (ดู sse_stream ด้านล่าง) เราแค่ส่ง payload เดียวกันลงไปในทุกคิว SSE ไหลทางเดียวจาก
    server ไป client เท่านั้น (ไม่ใช่ request/response แบบ 2 ทาง)
    """
    payload = json.dumps(data, default=str)
    for q in subscribers:
        await q.put({"event": event_type, "data": payload})
    log.info("SSE ▶ %s: %s", event_type, payload)


def get_db():
    """เปิด MySQL connection ใหม่สำหรับ 1 request

    ทำไมต้องเปิดใหม่ทุกครั้งแทนใช้ connection pool: นี่คือระบบที่ deploy บน PC เดียว
    มี concurrency ต่ำ ความเรียบง่ายของ "connect → ใช้งาน → close" จึงคุ้มกว่าความซับซ้อน
    ของการทำ pool ทุก endpoint ด้านล่างจะเปิด connection นี้ใน try/finally แล้วปิดเมื่อใช้เสร็จ

    หมายเหตุ (เพิ่มเข้ามาทีหลัง): endpoint ทุกตัวเรียก get_db() "ก่อน" เข้า try/finally
    ของตัวเอง ถ้า MySQL server ล่ม/ต่อไม่ติดเลย pymysql.connect() จะ raise
    OperationalError ซึ่งเป็น raw exception ที่ FastAPI ไม่รู้จัก — หลุดออกไปกลายเป็น
    500 ดิบไม่มี CORS header แนบมา (ปัญหาเดียวกับที่ _notify_agent_start เจอกับ Agent)
    จับไว้ตรงนี้ที่เดียวแล้ว raise เป็น HTTPException(503) แทน เพื่อให้ CORS header
    ยังติดมาด้วยเสมอ ไม่ต้องไปแก้ทุก endpoint
    """
    try:
        return pymysql.connect(**DB_CONFIG)
    except pymysql.MySQLError as exc:
        log.error("Database connection failed: %s", exc)
        raise HTTPException(503, f"เชื่อมต่อฐานข้อมูลไม่สำเร็จ: {exc}")


async def _reload_session_queues() -> None:
    """โหลด session_queues กลับเข้า memory จากคอลัมน์ sessions.queue_state

    ทำไมต้องมี: session_queues เดิมอยู่ใน memory ของ backend ล้วนๆ ถ้า backend
    ถูก restart (reload ตอน dev, crash แล้ว auto-restart, deploy ใหม่) ระหว่างที่
    มี session แบบ IPM/New กำลัง running อยู่ คิวจะหายไปจาก memory ทันที —
    create_measurement หลังจากนั้นจะไม่มีทางรู้ว่ากำลังวัด ALPL ตัวไหนอยู่

    เดิมกรณีนั้นจะ fallback ไปใช้ req.number_alpl ที่ Agent ส่งมา ซึ่งเป็น ALPL
    ตัวแรกในคิวเสมอ (อ่านมาจาก sessions.number_alpl ที่ไม่เคยถูก UPDATE) ทำให้ทุก
    measurement ที่เหลือถูกบันทึกผิด ALPL แบบไม่มี error เตือนเลย — ตอนนี้ถอด
    fallback นั้นออกแล้ว เปลี่ยนเป็นตอบ 409 ปฏิเสธไปเลย ฟังก์ชันนี้จึงเป็นด่าน
    เดียวที่กันไม่ให้ session ที่กำลังวัดอยู่ต้องล้มทั้งรอบเวลา backend restart

    จึงต้องรันตรงนี้ (ก่อน yield ให้แอปเริ่มรับ request) — ต้อง await ตรงๆ ไม่ใช่
    fire-and-forget แบบ _init_bucket_bg เพราะต้องมั่นใจว่า session_queues ถูก
    เติมกลับให้ครบก่อนที่ request แรกจาก Agent จะเข้ามาได้
    """
    try:
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT session_id, queue_state FROM sessions "
                    "WHERE state = 'running' AND queue_state IS NOT NULL"
                )
                rows = cur.fetchall()
        finally:
            db.close()

        for row in rows:
            try:
                session_queues[row["session_id"]] = json.loads(row["queue_state"])
                log.info("Restored queue_state for session %s from DB", row["session_id"])
            except Exception as exc:
                log.warning("Failed to parse queue_state for session %s: %s", row["session_id"], exc)
    except Exception as exc:
        log.warning("Reload session_queues failed: %s", exc)


async def heartbeat_checker() -> None:
    """ตรวจเป็นระยะว่า session ที่ 'running' ยังได้ heartbeat จาก Agent ต่อเนื่องไหม

    หมายเหตุ: ก่อนหน้านี้เคยถอดกลไกนี้ออกไปเพราะตอนนั้น Agent/Backend/Web รันอยู่
    บนเครื่องเดียวกันหมด คิดว่าไม่จำเป็น — ตอนนี้เอากลับมาใหม่ตามที่ตกลงกันไว้
    เป็น safety net เผื่อ Agent process ตาย/แฮงค์กลาง session (ไม่ใช่แค่ network
    หลุดข้ามเครื่องเหมือนเหตุผลเดิม) ถ้าเงียบเกิน HEARTBEAT_TIMEOUT วิ จะ mark
    session เป็น 'timeout' อัตโนมัติ แล้วแจ้ง web ผ่าน SSE (`session_timeout` —
    index.html มี handler นี้อยู่แล้ว แค่ก่อนหน้านี้ไม่มีใคร emit ให้)
    """
    while True:
        await asyncio.sleep(HEARTBEAT_INTERVAL)
        try:
            db = get_db()
        except HTTPException as exc:
            log.warning("heartbeat_checker: DB unreachable, skip this round: %s", exc.detail)
            continue

        timed_out: List[int] = []
        try:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT session_id FROM sessions "
                    "WHERE state = 'running' AND last_seen < NOW() - INTERVAL %s SECOND",
                    (HEARTBEAT_TIMEOUT,),
                )
                timed_out = [row["session_id"] for row in cur.fetchall()]
                for sid in timed_out:
                    cur.execute(
                        "UPDATE sessions SET state = 'timeout', ended_at = NOW() "
                        "WHERE session_id = %s",
                        (sid,),
                    )
        except Exception as exc:
            log.warning("heartbeat_checker: check failed: %s", exc)
            timed_out = []
        finally:
            db.close()

        for sid in timed_out:
            session_queues.pop(sid, None)
            measure_timeouts.pop(sid, None)
            log.warning("Session %s: ไม่ได้ heartbeat เกิน %ss — mark เป็น 'timeout'", sid, HEARTBEAT_TIMEOUT)
            await push_event("session_timeout", {"session_id": sid})


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Hook ตอน FastAPI เริ่มทำงาน (startup) และตอนปิด (shutdown)

    ทำไม: ตรงนี้คือจุดที่ background task (heartbeat_checker) ถูกสั่งให้เริ่ม
    ทำงานตอนแอป boot แทนที่จะไปสั่งเริ่มภายใน request handler ส่วน
    _reload_session_queues() ต้อง await ให้เสร็จก่อน yield เพราะต้องกู้คืนคิว
    ให้ครบก่อนรับ request
    """
    asyncio.create_task(heartbeat_checker())
    asyncio.create_task(_deleted_purge_loop())
    await _reload_session_queues()
    yield


async def _deleted_purge_loop():
    """ล้างถังขยะที่เกินอายุ — ทำทันทีตอนเริ่ม แล้ววนซ้ำวันละครั้ง

    ทำตอนเริ่มด้วยเพราะเครื่องหน้างานมักถูกปิดกลางคืน ถ้ารอครบ 24 ชั่วโมงอย่างเดียว
    อาจไม่มีวันได้ทำเลย
    """
    while True:
        try:
            _purge_old_deleted()
        except Exception as exc:
            log.warning("ล้างถังขยะไม่สำเร็จ: %s", exc)
        await asyncio.sleep(24 * 60 * 60)


app = FastAPI(title="TM-X Backend Server", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/stream")
async def sse_stream(request: Request):
    """Endpoint SSE ที่ dashboard เชื่อมต่อเข้ามาเพื่อรับข้อมูล real-time

    ทำไม: แทนที่ frontend จะ poll backend ทุกวินาที มันเปิด connection ค้างไว้ทีเดียว
    ที่นี่ แล้วเรา push event ไปให้ตอนมันเกิดขึ้นจริง (ดู push_event) แต่ละ client
    จะมี queue ของตัวเองที่ลงทะเบียนใน `subscribers` เราจะ yield "ping" ทุก 25 วินาที
    ตอนไม่มีอะไรใหม่ แค่เพื่อ keep connection ไว้ไม่ให้ proxy/browser ตัดการเชื่อมต่อ
    ที่ idle อยู่ SSE เป็นทางเดียว (server → client เท่านั้น) — ฝั่ง frontend ยังใช้
    POST request ปกติในการส่งคำสั่งไปที่ backend
    """
    async def generator():
        queue: asyncio.Queue = asyncio.Queue()
        subscribers.append(queue)
        log.info("SSE client connected  (total=%d)", len(subscribers))
        try:
            while True:
                try:
                    event = await asyncio.wait_for(queue.get(), timeout=25)
                    yield event
                except asyncio.TimeoutError:
                    yield {"event": "ping", "data": ""}
        finally:
            if queue in subscribers:
                subscribers.remove(queue)
            log.info("SSE client disconnected (total=%d)", len(subscribers))

    return EventSourceResponse(generator())


class StopSessionRequest(BaseModel):
    session_id: int
    reason: str | None = None


@app.get("/api/session/state")
async def get_session_state():
    """คืนสถานะปัจจุบันของ session ล่าสุด

    ทำไม: ตอน dashboard โหลดครั้งแรก (หรือ refresh) มันต้องรู้ว่า "มี run การวัด
    กำลังทำงานอยู่ไหม" ก่อนที่ SSE connection จะเปิดเสียอีก นี่คือ snapshot
    แบบครั้งเดียวที่ใช้ sync ตอนเริ่ม หลังจากนั้น SSE event จะคอยอัปเดตให้ real-time
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT session_id, state, target_count, measured_count, "
                "queue_state, last_seen, started_at, ended_at "
                "FROM sessions ORDER BY session_id DESC LIMIT 1"
            )
            row = cur.fetchone()
        return row or {"state": "idle"}
    finally:
        db.close()




def _get_template_name_for_alpl(cur, first_alpl: int) -> str:
    """Query หา template_name (ชื่อโปรแกรมวัดของ TM-X) ของ ALPL ที่จะเริ่มวัด
    — ใช้ทั้งโหมด IPM และ Rework (New ส่ง template_name มาใน payload เอง)

    **template ผูกกับ `package_size` ไม่ใช่ `part_number`** — ตัวกำหนดโปรแกรม
    วัดคือขนาดของ package ไม่ใช่รุ่นของ part ดังนั้นสายที่ถูกต้องคือ

        parts_specifications.package_size_id → package_size.template_id → template

    เดิมโค้ดนี้เดินอ้อมผ่าน `part_number` ทั้งที่ปลายทางคือ `package_size` อยู่ดี
    ผลคือ ALPL ที่ลงทะเบียนแบบ IPM (กรอกแค่ ALPL + Package Size ไม่มี Part
    Number) เริ่มวัดไม่ได้เลยทั้งที่ข้อมูลครบพอจะหา template ได้แล้ว

    ยังเก็บทางอ้อมผ่าน `part_number` ไว้เป็น fallback (COALESCE) เพราะ Part ที่
    ลงทะเบียนไว้ก่อนมีคอลัมน์ `parts_specifications.package_size_id` จะมีแต่
    part_number_id — ถ้าตัดทิ้งเลย ข้อมูลเก่าทั้งหมดจะวัดไม่ได้ทันทีที่ deploy

    ใช้ LEFT JOIN ทุกทอดเพื่อ "วินิจฉัย" ได้ว่าสายขาดตรงข้อไหน ของเดิมใช้ INNER
    JOIN แล้วได้ 0 แถว จึงบอกได้แค่ว่า "ยังไม่ได้ตั้ง part_number" ซึ่งชี้ผิดจุด
    บ่อย เพราะจริงๆ อาจตั้งไว้แล้วแต่ Package Size ยังไม่ได้ผูก Template
    """
    cur.execute(
        "SELECT p.part_number_id, p.package_size_id AS part_pkg_id, "
        "       pn.package_size_id AS pn_pkg_id, "
        "       COALESCE(p.package_size_id, pn.package_size_id) AS pkg_id, "
        "       ps.package_size, ps.template_id, t.template_name "
        "FROM parts_specifications p "
        "LEFT JOIN part_number pn  ON p.part_number_id = pn.part_number_id "
        "LEFT JOIN package_size ps ON ps.package_size_id = "
        "                             COALESCE(p.package_size_id, pn.package_size_id) "
        "LEFT JOIN template t      ON ps.template_id = t.template_id "
        "WHERE p.number_alpl = %s",
        (first_alpl,),
    )
    row = cur.fetchone()

    head = f"เริ่มวัด ALPL {first_alpl} ไม่ได้ —"
    if row is None:
        raise HTTPException(404, f"{head} ยังไม่ได้ลงทะเบียน ALPL นี้ในตาราง Parts")
    if row["pkg_id"] is None:
        raise HTTPException(
            404,
            f"{head} ยังไม่ได้ผูก Package Size ให้ ALPL นี้ "
            f"(แก้ที่หน้า Edit › Parts — จะตั้งผ่าน Part Number ก็ได้)",
        )
    if row["package_size"] is None:
        raise HTTPException(
            404,
            f"{head} Package Size ที่ผูกไว้ถูกลบไปแล้ว (package_size_id = {row['pkg_id']}) "
            f"— เลือกใหม่ที่หน้า Edit › Parts",
        )
    if row["template_id"] is None:
        raise HTTPException(
            404,
            f"{head} Package Size \"{row['package_size']}\" ยังไม่ได้ตั้ง Template ของเครื่อง TM-X "
            f"(แก้ที่หน้า Edit › Lookup Tables › Package Size) — ไม่เกี่ยวกับ Part Number",
        )
    if row["template_name"] is None:
        raise HTTPException(
            404,
            f"{head} Template ที่ Package Size \"{row['package_size']}\" ผูกไว้ถูกลบไปแล้ว "
            f"(template_id = {row['template_id']})",
        )
    return row["template_name"]


def _limits_of(row, measure_type: str) -> Dict[str, Any]:
    """แปลง nominal/tolerance เป็น "ขอบเขตสำเร็จรูป" ที่ Pi เอาไปเทียบตรงๆ

    **ส่งขอบ ไม่ส่ง nominal/tol ดิบ** — จำนวนตัวเลขเท่ากัน (4 ตัว) แต่ Pi ไม่ต้อง
    ลอก `_TOL_EPS` มาไว้ที่ตัวเอง ถ้าลืมเมื่อไหร่จะตัดสินไม่ตรงกับ backend
    **เฉพาะชิ้นที่ตกขอบพอดี** ซึ่งเป็นชิ้นที่สำคัญที่สุดและหาสาเหตุยากที่สุด
    (คอลัมน์เป็น FLOAT — `5.02` อ่านกลับได้ `5.0199999809265137`)

    ห้ามปัดเศษให้สวย ต้องเป็นตัวเลขชุดเดียวกับที่ `_within_tolerance` ใช้เป๊ะ

    `offset_max: null` = ไม่ต้องตรวจข้อนี้ — **ไม่ส่ง `measure_type` ไปด้วย**
    เพราะจะเปิดช่องให้มีคนเขียน `if measure_type == "IPM"` ที่ฝั่ง Pi แล้วกฎ
    เรื่องโหมดจะไปอยู่ 2 ที่ (`_offset_limit` ที่นี่ควรเป็นที่เดียว)
    """
    return {
        "x_lo": row["nominal_x"] - row["lower_tol"] - _TOL_EPS,
        "x_hi": row["nominal_x"] + row["upper_tol"] + _TOL_EPS,
        "y_lo": row["nominal_y"] - row["lower_tol"] - _TOL_EPS,
        "y_hi": row["nominal_y"] + row["upper_tol"] + _TOL_EPS,
        "offset_max": _offset_limit(measure_type, row),
    }


def _criteria_from_config(cur, gi: int, group: Dict[str, Any], entry_mode: str):
    """เกณฑ์ของกลุ่ม — อ่านจาก **config ที่ผู้ใช้กรอก** ไม่ใช่จากแถว Part

    จำเป็นต้องเป็นแบบนี้เพราะตอนกด Start ยังไม่มีแถว Part ให้ query เลยในโหมด
    New (Part ถูกสร้างพร้อม measurement ของชิ้นนั้น — ดู `create_measurement`)
    และโหมด IPM ก็มี ALPL บางตัวที่ยังไม่ลงทะเบียน

    ใช้ตารางเดียวกับ `_load_criteria` ตามโหมด (IPM → `package_size`,
    New/Rework → `part_number`) ค่าที่ได้จึงตรงกับที่ backend จะใช้ตัดสินจริง
    ตอน measurement เข้ามา
    """
    if entry_mode == "IPM":
        cur.execute(
            "SELECT nominal_x, nominal_y, upper_tol, lower_tol, offset_tol "
            "FROM package_size WHERE package_size = %s",
            ((group.get("package_size") or "").strip(),),
        )
    else:
        cur.execute(
            "SELECT nominal_x, nominal_y, upper_tol, lower_tol, offset_tol "
            "FROM part_number WHERE part_number_name = %s",
            ((group.get("part_number") or "").strip(),),
        )
    row = cur.fetchone()
    if not row:
        raise HTTPException(400, f"กลุ่มที่ {gi + 1}: หาเกณฑ์ตัดสินของกลุ่มนี้ไม่เจอ")
    return row


def _build_groups(cur, groups, group_of, queue, templates, entry_mode: str):
    """สร้างฟิลด์ `groups` ที่แนบไปกับ `POST /command` ให้ Pi

    Pi เอาไปทำ 2 อย่าง: รู้ว่าถึง ALPL ตัวไหนต้องสลับ `PW` เป็น template อะไร
    และตัดสิน OK/NG เองจากค่าที่อ่านผ่าน `GM` เพื่อสั่ง MCU ได้ทันทีโดยไม่ต้อง
    ถาม backend กลับ (ดู PLAN_criteria_and_multigroup.md ข้อ F)

    **payload ของ IPM กับ New/Rework หน้าตาเหมือนกันเป๊ะ** ต่างแค่ตัวเลข →
    Pi มี code path เดียว ตรงกับที่เป็นอยู่แล้ววันนี้ (Pi ไม่เคยรู้เรื่องโหมด)

    **ไม่ส่ง `queue` แยก** — เป็น `[a for g in groups for a in g["alpl"]]`
    บรรทัดเดียว ส่งซ้ำมีแต่จะเสี่ยงไม่ตรงกันเอง
    """
    out = []
    for gi, g in enumerate(groups):
        alpl = [queue[i] for i, gg in enumerate(group_of) if gg == gi]
        crit = _criteria_from_config(cur, gi, g, entry_mode)

        if entry_mode == "IPM":
            want = _limits_of(crit, entry_mode)
            for a in alpl:
                cur.execute("SELECT 1 FROM parts_specifications WHERE number_alpl = %s", (a,))
                if not cur.fetchone():
                    continue
                got = _limits_of(_load_criteria(cur, a, entry_mode), entry_mode)
                if got != want:
                    raise HTTPException(
                        400,
                        f"เริ่มวัดไม่ได้ — ALPL {a} ที่ลงทะเบียนไว้ใช้เกณฑ์ไม่ตรงกับ "
                        f"Package Size \"{g.get('package_size')}\" ที่กรอกในกลุ่มที่ {gi + 1} "
                        f"(X {got['x_lo']:.3f}–{got['x_hi']:.3f} vs {want['x_lo']:.3f}–{want['x_hi']:.3f}) "
                        f"— แก้ Package Size ของ ALPL นี้ที่หน้า Edit › Parts หรือแยกไปกรอกคนละกลุ่ม",
                    )

        out.append({
            "template_name": templates[gi],
            "alpl": alpl,
            "limits": _limits_of(crit, entry_mode),
        })
    return out


async def _notify_agent_start(
    session_id: int,
    target_count: int,
    groups: List[Dict[str, Any]],
) -> None:
    """ยิง POST ไปที่ Agent (`send_command.py` บน Pi / `mockup.py`) ให้เริ่มวัด

    ```json
    {"action": "start", "session_id": 42, "target_count": 6,
     "groups": [
       {"template_name": "021", "alpl": [400, 401, 402],
        "limits": {"x_lo": 5.009999, "x_hi": 5.030001,
                   "y_lo": 3.389999, "y_hi": 3.410001, "offset_max": null}},
       {"template_name": "007", "alpl": [501, 502, 503],
        "limits": {"x_lo": 3.199999, "x_hi": 3.240001,
                   "y_lo": 3.199999, "y_hi": 3.240001, "offset_max": 0.03}}
     ]}
    ```

    **ไม่มี `template_name` / `number_alpl` ระดับบนสุดแล้ว** — ทั้งคู่เป็นของ
    "กลุ่มแรก" ซึ่งอยู่ใน `groups[0]` อยู่แล้ว ถ้าส่งซ้ำไว้ข้างบนด้วยจะมีแหล่ง
    ความจริง 2 ที่ แล้ววันหนึ่งจะมีโค้ดฝั่ง Agent ที่อ่านตัวข้างบน (= ของกลุ่มแรก)
    ไปใช้กับทุกกลุ่มโดยไม่มีใครรู้ตัว

    **ไม่ส่ง `queue` แยก** — คือ `[a for g in groups for a in g["alpl"]]`
    บรรทัดเดียว ส่งซ้ำมีแต่จะเสี่ยงไม่ตรงกันเอง (`target_count` ส่งไว้เพราะเป็น
    ตัวที่ Agent ใช้วนลูป และตรวจได้ทันทีว่าตรงกับผลรวมของ `groups` ไหม)

    ╔═══ ถ้าสั่งไม่สำเร็จ ต้องเคลียร์ให้สะอาด (Handle_Pi_Error.md ข้อ 1.3) ═══╗
    ของเดิม `await client.post(...)` เฉยๆ ไม่เก็บผลลัพธ์เลย — Pi ตอบ 400/500
    หรือไม่ได้รันอยู่ ก็เดินหน้าต่อเหมือนกันหมด แล้ว session ค้างที่ `running`
    จนกว่า `heartbeat_checker` จะ mark เป็น `timeout` ใน 15 วิ ซึ่ง**ชี้ผิดสาเหตุ**
    (ผู้ใช้เห็นแค่ "กด Start แล้วไม่มีอะไรเกิดขึ้น")

    ตอนนี้ทุกทางที่พังวิ่งไปที่ `_fail_start()` ซึ่งเคลียร์ทุกอย่างแล้ว raise 502
    ออกไปให้หน้าเว็บขึ้น popup — ต้อง raise **ก่อน** `push_event("session_started")`
    ไม่งั้นหน้าเว็บทุกเครื่องจะเข้าโหมดวัดพร้อมกันทั้งที่ไม่มีอะไรเกิดขึ้น

    **timeout แยก 2 ค่า** — `connect=3` เพราะอยู่วง LAN เดียวกัน (ปกติต่อติดใน
    ~10 ms) ถ้า IP ผิดจะรู้ผลใน 3 วิแทนที่จะกิน 10 วิเต็ม · `read=10` ต้องเผื่อ
    ให้ Pi ประมวลผล  → จับเวลาแล้วเดาสาเหตุได้เลย: ~3 วิ = หา Pi ไม่เจอ ·
    ~10 วิ = Pi ค้าง
    ╚═══════════════════════════════════════════════════════════════════════╝
    """
    payload: Dict[str, Any] = {
        "action": "start",
        "session_id": session_id,
        "target_count": target_count,
        "groups": groups,
    }
    log.info("📤 ส่งไป Agent %s/command:\n%s",
             AGENT_BASE_URL, json.dumps(payload, ensure_ascii=False, indent=2))
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{AGENT_BASE_URL}/command", json=payload,
                timeout=httpx.Timeout(connect=3.0, read=10.0, write=10.0, pool=3.0),
            )
    except httpx.ConnectError:
        _fail_start(session_id,
                    f"ติดต่อโปรแกรมบนเครื่อง Pi ไม่ได้ ({AGENT_BASE_URL}) — "
                    f"ตรวจว่า send_command.py รันอยู่ไหม · สาย LAN · IP ของ Pi เปลี่ยนหรือเปล่า")
    except httpx.ConnectTimeout:
        _fail_start(session_id,
                    f"หาเครื่อง Pi ไม่เจอที่ {AGENT_BASE_URL} — ตรวจ IP หรือสาย LAN")
    except httpx.TimeoutException:
        await _notify_agent_action("stop", session_id)
        _fail_start(session_id,
                    "Pi ไม่ตอบภายใน 10 วินาที — อาจติดคำสั่งเดิมค้างอยู่ "
                    "(สั่งหยุดกลับไปแล้ว) ลองรีสตาร์ท send_command.py")
    except Exception as exc:
        _fail_start(session_id, f"สั่งงาน Pi ไม่สำเร็จ: {exc}")

    if resp.status_code != 200:
        _fail_start(session_id, f"Pi ปฏิเสธคำสั่ง (HTTP {resp.status_code}): {resp.text[:300]}")


def _fail_start(session_id: int, msg: str) -> None:
    """เคลียร์ session ที่เพิ่งสร้างแล้วโยน 502 ออกไป — ใช้ตอนสั่ง Pi ไม่สำเร็จ

    ทำ 3 อย่างที่ลืมง่ายทั้งหมด:

    ① `session_queues.pop()` — เป็น dict ในหน่วยความจำล้วนๆ ไม่มีใครมาเก็บกวาดให้
       กด Start ไม่ติด 20 ครั้งก็ค้าง 20 คิว
    ② ปิด session ใน DB ทันที — อย่าทิ้ง `running` ไว้ให้ `heartbeat_checker`
       มาเก็บใน 15 วิ เพราะมันจะขึ้นเป็น `timeout` ซึ่งชี้ผิดสาเหตุ
    ③ `raise HTTPException(502)` — ต้องเกิด **ก่อน** `push_event("session_started")`
       ใน start_session ไม่งั้นหน้าเว็บทุกเครื่องเข้าโหมดวัดพร้อมกันทั้งที่ไม่มี
       อะไรเกิดขึ้น · การ raise ยังทำให้ fetch ฝั่งเว็บพัง → ขึ้น popup แทนที่จะ
       แสดงหน้าจอวัดงานปลอมๆ

    **ไม่ broadcast SSE `session_stopped`** ต่างจากปุ่ม Stop โดยตั้งใจ — เพราะยัง
    ไม่มีแท็บไหนเคยได้ `session_started` เลย (raise เกิดก่อน) จึงไม่มีใครอยู่ใน
    โหมดวัดให้ต้องพาออกมา ส่วนแท็บอื่นจะเห็น state='stopped' เองในรอบ poll ถัดไป
    """
    session_queues.pop(session_id, None)
    measure_timeouts.pop(session_id, None)
    try:
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET state = 'stopped', ended_at = NOW(), "
                    "last_event = 'START_FAILED', last_event_detail = %s, last_event_at = NOW() "
                    "WHERE session_id = %s",
                    (msg, session_id),
                )
        finally:
            db.close()
    except Exception as exc:
        log.error("_fail_start: ปิด session %s ไม่สำเร็จ: %s", session_id, exc)
    log.warning("Start session %s ล้มเหลว — %s", session_id, msg)
    raise HTTPException(502, msg)


def _parse_entry_groups(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """แปลง payload ของ /api/session/start ให้เป็น "ลิสต์ของกลุ่ม" เสมอ

    ฟอร์ม Part Entry กด +Add เพิ่มกลุ่มได้ แต่ละกลุ่มมี ALPL ได้หลายตัวและมี
    config ของตัวเอง (Package Size / Part Number / PO / Vendor / Owner / …)
    ส่วน Operator อยู่นอกกลุ่ม ใช้ร่วมกันทั้ง session (คนวัดคนเดียวกัน)

        {"Measure_Type": "New", "Operator": "somchai",
         "groups": [{"number_alpl": [400,401], "package_size": "5x5", ...},
                    {"number_alpl": [501],     "package_size": "3x3", ...}]}

    ยังรับ payload แบบเก่า (field อยู่ระดับบนสุด ไม่มี `groups`) ได้ด้วย โดยตี
    ความว่าเป็น "กลุ่มเดียว" — ไม่ใช่เพื่อรองรับหน้าเว็บเก่า (ย้ายพร้อมกันอยู่แล้ว)
    แต่เพื่อให้เครื่องมือทดสอบ/สคริปต์เดิมที่ยิง payload ตรงๆ ไม่พังหมดทีเดียว
    """
    groups = data.get("groups")
    if groups is None:
        return [data]
    if not isinstance(groups, list) or not groups:
        raise HTTPException(400, "groups ต้องเป็น array และมีอย่างน้อย 1 กลุ่ม")
    for i, g in enumerate(groups):
        if not isinstance(g, dict):
            raise HTTPException(400, f"groups[{i}] ต้องเป็น object")
    return groups


def _flatten_groups(groups: List[Dict[str, Any]]) -> tuple:
    """คลี่ ALPL ของทุกกลุ่มออกเป็นคิวเส้นเดียว พร้อม "แผนที่ชิ้น → กลุ่ม"

    คืน (queue, group_of) ที่ยาวเท่ากัน — `group_of[i]` คือลำดับกลุ่มของชิ้นที่ i
    create_measurement ใช้ค่านี้หา config ที่ถูกต้องของชิ้นที่กำลังวัดอยู่
    (ไม่ใช่ config ของกลุ่มแรกเสมอแบบเดิม)

    ⚠ ALPL ห้ามซ้ำ **ข้ามกลุ่ม** ด้วย ไม่ใช่แค่ในกลุ่มเดียวกัน — ถ้าปล่อยให้ซ้ำ
      ชิ้นเดียวกันจะถูกวัด 2 ครั้งด้วย config คนละชุด แล้วอันหลังเขียนทับ Part
      ของอันแรกโดยที่ผู้ใช้ไม่รู้ตัว
    """
    queue: List[int] = []
    group_of: List[int] = []
    seen: Dict[int, int] = {}
    for gi, g in enumerate(groups):
        try:
            alpls = [int(x) for x in g["number_alpl"]]
        except KeyError:
            raise HTTPException(400, f"กลุ่มที่ {gi + 1} ไม่มี field 'number_alpl'")
        except (ValueError, TypeError):
            raise HTTPException(400, f"กลุ่มที่ {gi + 1}: number_alpl ต้องเป็นเลขจำนวนเต็มทั้งหมด")
        if not alpls:
            raise HTTPException(400, f"กลุ่มที่ {gi + 1} ต้องมี ALPL อย่างน้อย 1 ตัว")
        for a in alpls:
            if a in seen:
                where = ("ในกลุ่มเดียวกัน" if seen[a] == gi
                         else f"ซ้ำกับกลุ่มที่ {seen[a] + 1}")
                raise HTTPException(400, f"ALPL {a} ซ้ำ ({where}) — แก้ให้ไม่ซ้ำก่อนเริ่มวัด")
            seen[a] = gi
            queue.append(a)
            group_of.append(gi)
    return queue, group_of


def _group_config_for(qstate: Dict[str, Any], pos: int) -> Optional[Dict[str, Any]]:
    """config ของกลุ่มที่ชิ้นตำแหน่ง `pos` ในคิวสังกัดอยู่

    รองรับ `queue_state` รุ่นเก่าที่ค้างอยู่ใน DB ด้วย (มีแต่ `new_part_config`
    ก้อนเดียว ไม่มี `groups`) — สำคัญตอน backend restart ระหว่างที่ session เก่า
    ยังวัดอยู่ ถ้าอ่านไม่ออกจะไม่มี Part ถูกสร้างให้ ALPL ที่เหลือทั้งคิว
    """
    groups = qstate.get("groups")
    if groups:
        gof = qstate.get("group_of") or []
        gi = gof[pos] if pos < len(gof) else 0
        return groups[gi] if 0 <= gi < len(groups) else groups[0]
    return qstate.get("new_part_config")


def _validate_group(cur, gi: int, group: Dict[str, Any], measure_type: str,
                    alpls: List[int]) -> str:
    """ตรวจกลุ่มหนึ่งให้ครบ **โดยไม่เขียนอะไรลง DB เลย** แล้วคืน template_name

    เจตนา: ให้ผู้ใช้รู้ทุกปัญหา "ตั้งแต่กด Start" ไม่ใช่ไปรู้ตอนวัดชิ้นแรกเสร็จ
    แล้ว Part insert ไม่ผ่าน (ดู PLAN_criteria_and_multigroup.md ข้อ C2)

    เงื่อนไข ALPL ต่อโหมดตรงข้ามกัน (ข้อ D5) — หน้าเว็บเช็คให้แล้วตอนกด Save
    แต่ backend ต้องเช็คซ้ำ เพราะหน้าเว็บไม่ใช่ที่กันข้อมูลเสีย มันแค่ทำให้
    ผู้ใช้รู้เร็วขึ้น
    """
    label = f"กลุ่มที่ {gi + 1}"

    pkg = (group.get("package_size") or "").strip()
    if not pkg:
        raise HTTPException(400, f"{label}: ต้องเลือก Package Size")
    cur.execute("SELECT package_size_id FROM package_size WHERE package_size = %s", (pkg,))
    if not cur.fetchone():
        raise HTTPException(400, f"{label}: ไม่รู้จัก Package Size \"{pkg}\"")

    part_number = (group.get("part_number") or "").strip()
    if measure_type in ("New", "Rework"):
        if not part_number:
            raise HTTPException(400, f"{label}: โหมด {measure_type} ต้องเลือก Part Number")
        cur.execute(
            "SELECT 1 FROM part_number WHERE part_number_name = %s", (part_number,)
        )
        if not cur.fetchone():
            raise HTTPException(400, f"{label}: ไม่รู้จัก Part Number \"{part_number}\"")

    placeholders = ", ".join(["%s"] * len(alpls))
    cur.execute(
        f"SELECT number_alpl FROM parts_specifications WHERE number_alpl IN ({placeholders})",
        alpls,
    )
    found = {r["number_alpl"] for r in cur.fetchall()}
    missing = [a for a in alpls if a not in found]
    existing = [a for a in alpls if a in found]

    if measure_type == "New" and existing:
        raise HTTPException(
            409,
            f"{label}: ALPL {', '.join(map(str, existing))} มีอยู่ในระบบแล้ว — "
            f"ถ้าจะวัดซ้ำใช้โหมด IPM · ถ้าเป็นงานแก้จาก vendor ใช้ Rework",
        )
    if measure_type == "Rework" and missing:
        raise HTTPException(
            404,
            f"{label}: ALPL {', '.join(map(str, missing))} ไม่มีในระบบ — "
            f"Rework ต้องเป็นชิ้นที่เคยวัดมาก่อนเท่านั้น ถ้าเป็นชิ้นใหม่ให้ใช้โหมด New",
        )

    cur.execute(
        "SELECT t.template_name FROM package_size ps "
        "LEFT JOIN template t ON ps.template_id = t.template_id "
        "WHERE ps.package_size = %s",
        (pkg,),
    )
    row = cur.fetchone()
    if not row or not row["template_name"]:
        raise HTTPException(
            400,
            f"{label}: Package Size \"{pkg}\" ยังไม่ได้ตั้ง Template ของเครื่อง TM-X "
            f"(แก้ที่หน้า Edit › Lookup Tables › Package Size)",
        )
    return row["template_name"]


@app.post("/api/session/start")
async def start_session(request: Request):
    """เริ่ม session การวัดใหม่ จาก Part Entry card (โหมด IPM, New หรือ Rework)

    **ฟอร์มส่งมาเป็น "กลุ่ม"** — 1 กลุ่ม = ALPL หลายตัวที่ใช้ config ชุดเดียวกัน
    กด +Add เพิ่มกลุ่มได้ (ดู `_parse_entry_groups`) ทั้ง 3 โหมดใช้โครงเดียวกัน
    ต่างกันแค่ลิสต์ field ในกลุ่มและเงื่อนไขว่า ALPL ต้องมี/ต้องไม่มีใน DB

    **ไม่มีการเขียน Part ลง DB ที่นี่เลยทุกโหมด** — ตรวจอย่างเดียว (`_validate_group`)
    แล้วเก็บ config ไว้ใน `queue_state` ให้ `create_measurement` เอาไปสร้าง/อัปเดต
    Part "พร้อมกับ measurement ของชิ้นนั้น" ทีละชิ้น ผลคือ

        กด Start แล้วกด Stop ทันที   → ไม่มีอะไรเกิดขึ้นใน DB เลย
        วัดชิ้นที่ 1 สำเร็จ            → Part + Measurement เกิดพร้อมกัน
        วัดชิ้นที่ 1 ไม่ติด            → ไม่มีอะไรเกิดขึ้น

    เดิม New insert Part ตัวแรกไว้ก่อนเพราะ `sessions.number_alpl` มี FK — ตอนนี้
    คอลัมน์นั้นถูกถอดออกไปแล้ว จึงไม่มีเหตุผลให้ insert ล่วงหน้าอีก

    ทั้ง 3 กรณี — Agent ไม่ต้องรู้ความต่างเลย ได้รับ payload หน้าตาเดียวกัน
    (action/session_id/template_name/target_count/number_alpl ตัวแรก) ส่วน
    การ map ALPL ตัวต่อๆไปในคิวเข้ากับ measurement ที่จะตามมา เป็นเรื่องที่
    backend จัดการเองทั้งหมดผ่าน session_queues (ดู create_measurement)

    หมายเหตุ (เพิ่มเข้ามาทีหลัง): Race condition ตอนกด Start ซ้ำเร็วๆ — ครอบ
    check+insert ด้วย MySQL GET_LOCK/RELEASE_LOCK กันสอง request แข่งกันผ่าน
    Button Guard พร้อมกันได้ (เดิมเช็คแล้ว insert คนละคำสั่ง ไม่มีอะไรล็อก
    ระหว่างนั้นเลย)
    """
    data = await request.json()
    log.info("📥 ได้รับ payload จาก /api/session/start:\n%s", json.dumps(data, ensure_ascii=False, indent=2))

    measure_type = data.get("Measure_Type")
    if measure_type not in ("New", "IPM", "Rework"):
        raise HTTPException(400, "Measure_Type ต้องเป็น 'New', 'IPM' หรือ 'Rework'")

    groups = _parse_entry_groups(data)
    alpl_queue, group_of = _flatten_groups(groups)
    first_alpl = alpl_queue[0]
    target_count = len(alpl_queue)

    entry_mode = "New" if measure_type in ("New", "Rework") else "IPM"
    entry_note = "Rework" if measure_type == "Rework" else None

    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT GET_LOCK('tmx_start_session', 5) AS got")
            if not cur.fetchone()["got"]:
                raise HTTPException(503, "ระบบกำลังประมวลผลคำสั่ง Start อื่นอยู่ ลองใหม่อีกครั้ง")

        try:
            with db.cursor() as cur:
                cur.execute("SELECT session_id FROM sessions WHERE state = 'running'")
                if cur.fetchone():
                    raise HTTPException(400, "A session is already running")

                templates: List[str] = []
                for gi, g in enumerate(groups):
                    alpls_of_group = [a for a, gg in zip(alpl_queue, group_of) if gg == gi]
                    templates.append(_validate_group(cur, gi, g, measure_type, alpls_of_group))

                distinct = sorted(set(templates))
                if len(distinct) > 1:
                    raise HTTPException(
                        400,
                        "กลุ่มที่กรอกมาใช้ Template ของ TM-X คนละตัวกัน "
                        f"({', '.join(distinct)}) — ตอนนี้ยังสลับโปรแกรมกลางคันไม่ได้ "
                        "กรุณาแยกวัดทีละ Template (กลุ่มที่ Package Size ให้ Template "
                        "เดียวกันรวมรอบเดียวกันได้)",
                    )
                template_name = templates[0]

                agent_groups = _build_groups(
                    cur, groups, group_of, alpl_queue, templates, entry_mode
                )

                cur.execute(
                    "INSERT INTO sessions (state, target_count, measured_count) "
                    "VALUES ('running', %s, 0)",
                    (target_count,),
                )
                session_id = cur.lastrowid
        finally:
            with db.cursor() as cur:
                cur.execute("SELECT RELEASE_LOCK('tmx_start_session')")

        queue_state = {
            "entry_mode": entry_mode,
            "measure_mode": measure_type,
            "queue": alpl_queue,
            "group_of": group_of,
            "groups": groups,
            "group_templates": templates,
            "position": 0,
            "operator": data.get("Operator"),
            "note": entry_note,
        }
        session_queues[session_id] = queue_state
        measure_timeouts.pop(session_id, None)

        with db.cursor() as cur:
            cur.execute(
                "UPDATE sessions SET queue_state = %s WHERE session_id = %s",
                (json.dumps(queue_state), session_id),
            )

        await _notify_agent_start(session_id, target_count, agent_groups)

        await push_event(
            "session_started",
            {
                "session_id": session_id,
                "number_alpl": first_alpl,
                "template_name": template_name,
                "target_count": target_count,
            },
        )
        return {"session_id": session_id, "template_name": template_name, "target_count": target_count}
    finally:
        db.close()


@app.post("/api/session/stop")
async def stop_session(req: StopSessionRequest):
    """หยุด session ที่กำลัง running จากปุ่ม Stop บน dashboard

    ทำไมเรื่องนี้สำคัญ: นี่คือ path "web-initiated stop" — มันอัปเดต DB
    (state='stopped', ended_at=NOW()) แล้วบอก Agent ให้หยุด ซึ่งต่างจากปุ่ม
    Stop ทางกายภาพที่ MCU (ในการ implement ปัจจุบันของ Agent) ที่แค่ flip
    flag ใน memory ฝั่ง Agent โดยไม่แตะ DB เลย — เป็นความไม่สมดุล (asymmetry)
    ที่รู้กันอยู่ระหว่าง stop ทั้ง 2 path นี้

    **นี่คือทางเดียวในระบบที่ปิด session ได้** — ทุกปุ่มและทุกเส้นทางวิ่งมาที่นี่หมด
    (ปุ่ม Stop บนเว็บ · "หยุดการวัด" ใน modal · modal หมดเวลา 60 วิ · Pi ล้มเลิกเอง)
    ตั้งใจไม่ให้มีทางที่สอง เพราะถ้าแยกกันแล้วลืมทำอะไรสักอย่างในเส้นทางไหน
    จะเกิดอาการ "กดหยุดจากตรงนี้แล้วค้าง แต่กดจากตรงนั้นแล้วปกติ" ซึ่งหาสาเหตุยากมาก

    reason: Pi ส่งมาตอนล้มเลิกเอง (ER,PW / T1 retry ครบ / สาย TM-X ขาด) เพื่อให้
            หน้าเว็บบอกผู้ใช้ได้ว่าหยุดเพราะอะไร — หน้าเว็บไม่ต้องส่งมา (None)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT state FROM sessions WHERE session_id = %s", (req.session_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Session not found")
            if req.reason:
                cur.execute(
                    "UPDATE sessions SET state = 'stopped', ended_at = NOW(), "
                    "last_event = 'PI_ERROR', last_event_detail = %s, last_event_at = NOW() "
                    "WHERE session_id = %s",
                    (req.reason, req.session_id),
                )
                log.warning("Session %s: Pi ล้มเลิกเอง — %s", req.session_id, req.reason)
            else:
                cur.execute(
                    "UPDATE sessions SET state = 'stopped', ended_at = NOW() "
                    "WHERE session_id = %s",
                    (req.session_id,),
                )

        agent_err = await _notify_agent_action("stop", req.session_id)

        session_queues.pop(req.session_id, None)
        measure_timeouts.pop(req.session_id, None)

        if agent_err:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET last_event = 'STOP_NOT_DELIVERED', "
                    "last_event_detail = %s, last_event_at = NOW() WHERE session_id = %s",
                    (agent_err, req.session_id),
                )

        await push_event(
            "session_stopped",
            {"session_id": req.session_id, "reason": req.reason, "agent_error": agent_err},
        )
        return {"ok": True, "agent_error": agent_err}
    finally:
        db.close()


measure_timeouts: dict[int, dict] = {}


class MeasureTimeoutRequest(BaseModel):
    session_id: int
    piece: int | None = None
    target: int | None = None


class SessionEventRequest(BaseModel):
    """Recieve_tm-x.py แจ้งสาเหตุที่มัน "ทิ้งค่า" ไปโดยไม่บันทึกลง DB

    ไม่มี session_id — Recieve ไม่รู้ในบางด่าน (ด่าน 1-2 เกิดก่อนด่าน 3 ที่เป็น
    ตัวถาม session) Backend จึงแนบเข้ากับ session ที่ running อยู่ตอนนั้นเอง

    persist=False → ยิง SSE เตือนอย่างเดียว ไม่เขียนลง last_event
                    ใช้กับเรื่องที่ "ค่าลง DB ไปแล้ว" เช่นรูปอัปโหลดไม่สำเร็จ
    """
    event:   str
    detail:  str | None = None
    persist: bool = True


class SessionContinueRequest(BaseModel):
    session_id: int


async def _notify_agent_action(action: str, session_id: int | None = None) -> Optional[str]:
    """ยิงคำสั่งสั้นๆ ไปหา Pi (stop / continue) — คืนข้อความ error ถ้าสั่งไม่สำเร็จ

    ╔═══ ทำไม stop ต้องปฏิบัติต่างจาก start (Handle_Pi_Error.md ข้อ 1.4) ═══╗
    **Start พังแล้วเคลียร์ทิ้งได้** เพราะรู้ว่ายังไม่มีอะไรเกิดขึ้น
    **Stop พังแปลว่าเครื่องอาจยังวัดอยู่จริง** — จะไปเคลียร์ session ทิ้งเฉยๆ
    ไม่ได้ เพราะ DB จะบอกว่าจบแล้วทั้งที่ของยังไหลอยู่บนสายพาน

    ที่นี่จึง **ไม่ raise และไม่แตะสถานะ session เลย** — DB ถูกอัปเดตไปแล้วก่อน
    เรียกฟังก์ชันนี้ (ดู stop_session) การโยน exception ออกไปจะทำให้ผู้ใช้เข้าใจ
    ว่า "กดหยุดไม่สำเร็จ" แล้วกดซ้ำ ทั้งที่ฝั่ง DB หยุดไปเรียบร้อยแล้ว

    แค่ **คืนข้อความกลับไปให้ผู้เรียกตัดสินใจ** ว่าจะเอาไปบอกผู้ใช้ยังไง
    (`stop_session` เอาไปแนบใน SSE `session_stopped` → หน้าเว็บเด้งเตือนว่า
     "หยุดในระบบแล้ว แต่สั่ง Pi ไม่ได้ — ไปกดหยุดที่เครื่องด้วย")
    ╚═══════════════════════════════════════════════════════════════════════╝
    """
    body: dict = {"action": action}
    if session_id is not None:
        body["session_id"] = session_id
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{AGENT_BASE_URL}/command", json=body,
                timeout=httpx.Timeout(connect=3.0, read=10.0, write=10.0, pool=3.0),
            )
    except httpx.ConnectError:
        msg = (f"ติดต่อโปรแกรมบนเครื่อง Pi ไม่ได้ ({AGENT_BASE_URL}) — "
               f"ตรวจว่า send_command.py รันอยู่ไหม")
    except httpx.ConnectTimeout:
        msg = f"หาเครื่อง Pi ไม่เจอที่ {AGENT_BASE_URL} — ตรวจ IP หรือสาย LAN"
    except httpx.TimeoutException:
        msg = "Pi ไม่ตอบภายใน 10 วินาที — อาจติดคำสั่งเดิมค้างอยู่"
    except Exception as exc:
        msg = f"สั่งงาน Pi ไม่สำเร็จ: {exc}"
    else:
        if resp.status_code == 200:
            return None
        msg = f"Pi ปฏิเสธคำสั่ง '{action}' (HTTP {resp.status_code}): {resp.text[:200]}"

    log.warning("Agent %s notify failed: %s", action, msg)
    return msg


@app.post("/api/session/event")
async def session_event(body: SessionEventRequest):
    """รับรายงานสาเหตุจาก Recieve_tm-x.py

    ⚠ last_event มีช่องเดียว ค่าใหม่ทับค่าเก่า — จึงเก็บเฉพาะเรื่องที่ "มีคนรอ
      คำตอบอยู่" (ค่าไม่ลง DB แล้ว Pi กำลังนับถอยหลัง) ส่วนเรื่องที่ค่าลงไปแล้ว
      เช่น IMAGE_UPLOAD_FAILED ต้องส่ง persist=False มา ไม่งั้นจะไปทับสาเหตุที่
      Backend ต้องหยิบไปตอบ Pi ตอน measure-timeout
    """
    if body.persist:
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET last_event = %s, last_event_detail = %s, "
                    "last_event_at = NOW() WHERE state = 'running'",
                    (body.event, body.detail),
                )
        finally:
            db.close()

    log.warning("Station event: %s — %s", body.event, body.detail)
    await push_event("station_event", {"event": body.event, "detail": body.detail})
    return {"ok": True}


LAST_EVENT_FRESH_SEC = 30


@app.post("/api/measure-timeout")
async def report_measure_timeout(req: MeasureTimeoutRequest):
    """Pi แจ้งว่ารอค่าการวัดชิ้นนี้จนหมดเวลาแล้วยังไม่มาถึง

    Pi รู้แค่ "ไม่ได้ค่า" ไม่รู้สาเหตุ — Backend เป็นคนเติมให้ 2 อย่างก่อน broadcast:
      number_alpl  จาก queue[position] ของตัวเอง (ห้ามให้ Pi ส่งมา ไม่งั้นมี 2 แหล่ง
                   ที่บอกว่าชิ้นนี้คือ ALPL อะไร แล้วเหลื่อมกันได้)
      detail       จาก last_event ที่ Recieve เขียนไว้ก่อนหน้า (ถ้ายังสด)

    detail เป็น None ได้ — เกิดเมื่อรูปไม่มาเลย (Recieve ไม่เคยรายงาน) หรือ
    last_event เก่าเกินไป · หน้าเว็บต้องรองรับกรณีนี้ด้วยข้อความกลางๆ
    """
    measure_timeouts[req.session_id] = {"piece": req.piece, "target": req.target}

    detail = None
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT last_event_detail, last_event_at FROM sessions WHERE session_id = %s",
                (req.session_id,),
            )
            row = cur.fetchone()
    finally:
        db.close()

    if row and row["last_event_at"]:
        age = (datetime.now() - row["last_event_at"]).total_seconds()
        if age < LAST_EVENT_FRESH_SEC:
            detail = row["last_event_detail"]

    number_alpl = None
    qstate = session_queues.get(req.session_id)
    if qstate is not None:
        queue = qstate.get("queue") or []
        pos = qstate.get("position", 0)
        if 0 <= pos < len(queue):
            number_alpl = queue[pos]

    await push_event(
        "measure_timeout",
        {
            "session_id":  req.session_id,
            "piece":       req.piece,
            "target":      req.target,
            "number_alpl": number_alpl,
            "detail":      detail,
        },
    )
    log.warning(
        "Measure timeout: session=%s piece=%s alpl=%s — รอผู้ใช้ตัดสินใจ (สาเหตุ: %s)",
        req.session_id, req.piece, number_alpl, detail or "ไม่ทราบ",
    )
    return {"ok": True}


@app.post("/api/session/continue")
async def continue_session(body: SessionContinueRequest):
    """ผู้ใช้กด "วัดชิ้นถัดไป" ใน modal — ข้ามชิ้นที่ไม่ได้ค่าแล้ววัดต่อ

    ╔═══ เปลี่ยนจากการ poll เป็นการ push — 7 ส.ค. 2569 ═══════════════════════╗
    เดิม Pi วน GET /api/measure-timeout/{id} ทุก 0.4 วิรอคำตอบ ตอนนี้ Backend
    ยิง POST /command {"action":"continue"} ไปปลุก Pi แทน — Pi ตื่นทันทีที่คำตอบ
    มาถึง ไม่ต้องรอรอบ poll และคำสั่งจากข้างนอกเข้าทาง /command ทางเดียวหมด
    (start / stop / continue) เหมือนกันทั้งระบบ
    ╚═══════════════════════════════════════════════════════════════════════╝

    ⚠ ลำดับสำคัญ: ต้องขยับ position + sync ลง DB ให้เสร็จ "ก่อน" ปลุก Pi
      ไม่งั้น Pi อาจวัดชิ้นถัดไปเสร็จก่อนที่ UPDATE จะลง แล้วผลถูกแปะ ALPL ผิด

    ส่วน "หยุดการวัด" ใน modal ไม่ผ่านที่นี่ — หน้าเว็บเรียก POST /api/session/stop
    ตรงๆ เพื่อให้เส้นทางการหยุด session มีทางเดียวตลอดทั้งระบบ
    """
    session_id = body.session_id
    if measure_timeouts.get(session_id) is None:
        raise HTTPException(404, "ไม่พบคำถามค้างของ session นี้ (อาจหมดอายุไปแล้ว)")
    measure_timeouts.pop(session_id, None)

    qstate = session_queues.get(session_id)
    if qstate is not None:
        qstate["position"] += 1
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET queue_state = %s WHERE session_id = %s",
                    (json.dumps(qstate), session_id),
                )
        finally:
            db.close()
        log.info(
            "Session %s: ข้ามชิ้นงาน — ขยับตำแหน่งคิวเป็น %d/%d เพื่อไม่ให้ "
            "ชิ้นที่เหลือถูกแปะ ALPL ผิด",
            session_id, qstate["position"], len(qstate.get("queue", [])),
        )

    agent_err = await _notify_agent_action("continue", session_id)
    if agent_err:
        raise HTTPException(502, f"สั่งให้ Pi วัดชิ้นถัดไปไม่สำเร็จ — {agent_err}")
    return {"ok": True}




class HeartbeatRequest(BaseModel):
    session_id: Optional[int] = None


@app.post("/api/heartbeat")
async def heartbeat(req: HeartbeatRequest):
    """รับ heartbeat จาก Agent (ดู agent.py heartbeat_loop — ยิงมาทุก
    HEARTBEAT_INTERVAL วิ ไม่ว่าจะมี session running อยู่หรือไม่)

    ถ้าไม่มี session_id (Agent ยัง idle ไม่มีงานอยู่) แค่ตอบ ok เฉยๆ ไม่ต้องแตะ DB
    ถ้ามี session_id จะอัปเดต sessions.last_seen = NOW() ให้ heartbeat_checker()
    เอาไปเทียบว่า session นี้ยังมี Agent ส่งสัญญาณชีพอยู่ไหม — เงื่อนไข
    `state = 'running'` กันไม่ให้ heartbeat ที่มาช้า/ค้างจาก session เก่าที่จบไป
    แล้วไปอัปเดต last_seen ของ session ผิดตัว
    """
    if req.session_id is None:
        return {"ok": True}
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "UPDATE sessions SET last_seen = NOW() "
                "WHERE session_id = %s AND state = 'running'",
                (req.session_id,),
            )
    finally:
        db.close()
    return {"ok": True}


@app.get("/api/operators")
async def list_operators():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT operator_id, operator_name FROM operator ORDER BY operator_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/owners")
async def list_owners():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT owner_id, owner_name FROM owner ORDER BY owner_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/vendors")
async def list_vendors():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT vendor_id, vendor_name FROM vendor ORDER BY vendor_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/handlers")
async def list_handlers():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT handler_id, handler_name FROM handler ORDER BY handler_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/package-sizes")
async def list_package_sizes():
    """คืนรายการ package_size ทั้งหมด พร้อม nominal/tolerance + template_name
    — ใช้เติม datalist ของช่อง Package Size ใน index.html/edit.html และเป็น
    แหล่งข้อมูลของตาราง Lookup Tables → Package Size

    nominal/tolerance ถูกเก็บไว้ 2 ที่โดยตั้งใจ (ไม่ใช่ข้อมูลซ้ำที่ลืมลบ):
      • package_size — ค่ากลางของ "ขนาด" นั้น ใช้ตอนยังไม่รู้ part_number
      • part_number  — ค่าเฉพาะของ part นั้น (part_number เดียวกันอาจมี
        tolerance ต่างกันได้แม้ package_size เดียวกัน) ดู GET /api/part-numbers
    ทั้ง 5 คอลัมน์เป็น NOT NULL ทั้งคู่ (ดู init.sql)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT ps.package_size_id, ps.package_size, "
                "       ps.nominal_x, ps.nominal_y, ps.upper_tol, ps.lower_tol, ps.offset_tol, "
                "       t.template_name "
                "FROM package_size ps "
                "LEFT JOIN template t ON ps.template_id = t.template_id "
                "ORDER BY ps.package_size"
            )
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/part-numbers")
async def list_part_numbers(package_size: str = Query(..., min_length=1)):
    """คืนรายการ part_number (จากตาราง catalog part_number ตรงๆ ไม่ใช่ derive
    จาก parts_specifications ที่เคยลงทะเบียนแล้ว) ที่ผูกกับ package_size ที่
    ระบุ — ใช้เป็น dropdown ของช่อง Part Number ที่ cascade จาก Package Size
    ในฟอร์ม Part Entry (New/Rework/IPM) เป็น dropdown "ปิด" เหมือน
    operator/vendor/handler/owner/package_size — เลือกได้เฉพาะ part_number ที่
    มีอยู่จริงในตาราง part_number เท่านั้น ไม่มีช่องพิมพ์เพิ่มค่าใหม่ (handler
    ของ part นั้นๆ ผูกมากับ part_number อยู่แล้ว ไม่ต้องให้ผู้ใช้เลือก Handler
    เองอีกที)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT pn.part_number_name FROM part_number pn "
                "JOIN package_size ps ON pn.package_size_id = ps.package_size_id "
                "WHERE ps.package_size = %s "
                "ORDER BY pn.part_number_name",
                (package_size,),
            )
            return [row["part_number_name"] for row in cur.fetchall()]
    finally:
        db.close()


@app.get("/api/templates")
async def list_templates():
    """คืนรายการ template ทั้งหมด — ใช้โดย Database Editor (edit.html) ตอน
    จัดการตาราง template (Add/Rename/Delete) และตอนสร้าง/แก้ package_size
    (เลือก template ที่จะผูกให้)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT template_id, template_name FROM template ORDER BY template_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/part-numbers/all")
async def list_all_part_numbers():
    """คืนรายการ part_number ทั้งหมดพร้อมรายละเอียดครบ (package_size/handler/
    nominal/tolerance) — ใช้โดย Database Editor (edit.html) ต่างจาก
    GET /api/part-numbers (คืนแค่ชื่อ กรองด้วย package_size — ใช้เป็น cascading
    dropdown ของฟอร์ม Part Entry)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT pn.part_number_id, pn.part_number_name, ps.package_size, "
                "h.handler_name AS handler, pn.nominal_x, pn.nominal_y, "
                "pn.upper_tol, pn.lower_tol, pn.offset_tol "
                "FROM part_number pn "
                "JOIN package_size ps ON pn.package_size_id = ps.package_size_id "
                "JOIN handler h       ON pn.handler_id = h.handler_id "
                "ORDER BY pn.part_number_name"
            )
            return cur.fetchall()
    finally:
        db.close()


def _create_lookup(table: str, name_col: str, name: str) -> int:
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(f"INSERT INTO {table} ({name_col}) VALUES (%s)", (name,))
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่มไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return cur.lastrowid
    finally:
        db.close()


def _rename_lookup(table: str, id_col: str, name_col: str, id_value: int, name: str) -> None:
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(f"UPDATE {table} SET {name_col} = %s WHERE {id_col} = %s", (name, id_value))
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไขชื่อไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "ไม่พบข้อมูล")
    finally:
        db.close()


_TABLE_DISPLAY_NAME = {
    "parts_specifications": "Part",
    "measurements":         "Measurement",
    "part_number":          "Part Number",
    "package_size":         "Package Size",
}


def _delete_lookup(table: str, id_col: str, id_value: int, references: List[tuple],
                   kind: Optional[str] = None) -> None:
    """ลบ row ของตาราง lookup — ปฏิเสธถ้ายังมีตารางอื่นอ้างอิง id นี้อยู่จริง
    (`references` คือ list ของ (referencing_table, referencing_col)) เพื่อไม่ให้
    Part/Measurement ที่มีอยู่แล้วกลายเป็นข้อมูลกำพร้า (orphan FK) — แนะนำให้
    "เปลี่ยนชื่อ" (rename) แทนถ้ายังมีข้อมูลผูกอยู่ ไม่ใช่ลบทิ้งแล้วสร้างใหม่
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            for ref_table, ref_col in references:
                cur.execute(f"SELECT 1 FROM {ref_table} WHERE {ref_col} = %s LIMIT 1", (id_value,))
                if cur.fetchone():
                    display_name = _TABLE_DISPLAY_NAME.get(ref_table, ref_table)
                    raise HTTPException(
                        409,
                        f"ลบไม่ได้ — ยังมีข้อมูลใน {display_name} ที่อ้างอิงถึงอยู่ "
                        f"กรุณาเปลี่ยนชื่อแทนถ้าต้องการแก้ไข",
                    )
            row = _fetch_one(cur, f"SELECT * FROM {table} WHERE {id_col} = %s", (id_value,))
            if row is None:
                raise HTTPException(404, "ไม่พบข้อมูล")
            _archive_before_delete(
                kind=kind or table, table=table, pk={id_col: id_value}, row=row
            )

            cur.execute(f"DELETE FROM {table} WHERE {id_col} = %s", (id_value,))
            if cur.rowcount == 0:
                raise HTTPException(404, "ไม่พบข้อมูล")
    finally:
        db.close()


class LookupCreate(BaseModel):
    name: str


class LookupUpdate(BaseModel):
    name: str


@app.post("/api/operators", status_code=201)
async def create_operator(body: LookupCreate):
    return {"operator_id": _create_lookup("operator", "operator_name", body.name)}


@app.patch("/api/operators/{operator_id}")
async def rename_operator(operator_id: int, body: LookupUpdate):
    _rename_lookup("operator", "operator_id", "operator_name", operator_id, body.name)
    return {"ok": True}


@app.delete("/api/operators/{operator_id}")
async def delete_operator(operator_id: int):
    _delete_lookup("operator", "operator_id", operator_id, [("measurements", "operator_id")])
    return {"ok": True}


@app.post("/api/owners", status_code=201)
async def create_owner(body: LookupCreate):
    return {"owner_id": _create_lookup("owner", "owner_name", body.name)}


@app.patch("/api/owners/{owner_id}")
async def rename_owner(owner_id: int, body: LookupUpdate):
    _rename_lookup("owner", "owner_id", "owner_name", owner_id, body.name)
    return {"ok": True}


@app.delete("/api/owners/{owner_id}")
async def delete_owner(owner_id: int):
    _delete_lookup("owner", "owner_id", owner_id, [("parts_specifications", "owner_id")])
    return {"ok": True}


@app.post("/api/vendors", status_code=201)
async def create_vendor(body: LookupCreate):
    return {"vendor_id": _create_lookup("vendor", "vendor_name", body.name)}


@app.patch("/api/vendors/{vendor_id}")
async def rename_vendor(vendor_id: int, body: LookupUpdate):
    _rename_lookup("vendor", "vendor_id", "vendor_name", vendor_id, body.name)
    return {"ok": True}


@app.delete("/api/vendors/{vendor_id}")
async def delete_vendor(vendor_id: int):
    _delete_lookup("vendor", "vendor_id", vendor_id, [("parts_specifications", "vendor_id")])
    return {"ok": True}


@app.post("/api/handlers", status_code=201)
async def create_handler(body: LookupCreate):
    return {"handler_id": _create_lookup("handler", "handler_name", body.name)}


@app.patch("/api/handlers/{handler_id}")
async def rename_handler(handler_id: int, body: LookupUpdate):
    _rename_lookup("handler", "handler_id", "handler_name", handler_id, body.name)
    return {"ok": True}


@app.delete("/api/handlers/{handler_id}")
async def delete_handler(handler_id: int):
    _delete_lookup("handler", "handler_id", handler_id, [("part_number", "handler_id")])
    return {"ok": True}


@app.post("/api/templates", status_code=201)
async def create_template(body: LookupCreate):
    return {"template_id": _create_lookup("template", "template_name", body.name)}


@app.patch("/api/templates/{template_id}")
async def rename_template(template_id: int, body: LookupUpdate):
    _rename_lookup("template", "template_id", "template_name", template_id, body.name)
    return {"ok": True}


@app.delete("/api/templates/{template_id}")
async def delete_template(template_id: int):
    _delete_lookup("template", "template_id", template_id, [("package_size", "template_id")])
    return {"ok": True}


class PackageSizeCreate(BaseModel):
    package_size:  str
    nominal_x:     float
    nominal_y:     float
    upper_tol:     float
    lower_tol:     float
    offset_tol:    float
    template_name: Optional[str] = None


class PackageSizeUpdate(BaseModel):
    package_size:  Optional[str] = None
    nominal_x:     Optional[float] = None
    nominal_y:     Optional[float] = None
    upper_tol:     Optional[float] = None
    lower_tol:     Optional[float] = None
    offset_tol:    Optional[float] = None
    template_name: Optional[str] = None


_PKG_NUM_FIELDS = ("nominal_x", "nominal_y", "upper_tol", "lower_tol", "offset_tol")


@app.post("/api/package-sizes", status_code=201)
async def create_package_size(body: PackageSizeCreate):
    db = get_db()
    try:
        with db.cursor() as cur:
            template_id = _lookup_id(cur, "template", "template_id", "template_name", body.template_name)
            try:
                cur.execute(
                    "INSERT INTO package_size "
                    "(package_size, nominal_x, nominal_y, upper_tol, lower_tol, offset_tol, template_id) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (
                        body.package_size,
                        *(getattr(body, f) for f in _PKG_NUM_FIELDS),
                        template_id,
                    ),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่ม Package Size ไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"package_size_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/package-sizes/{package_size_id}")
async def update_package_size(package_size_id: int, body: PackageSizeUpdate):
    db = get_db()
    try:
        with db.cursor() as cur:
            set_parts, values = [], []
            if body.package_size is not None:
                set_parts.append("package_size = %s")
                values.append(body.package_size)
            for f in _PKG_NUM_FIELDS:
                v = getattr(body, f)
                if v is not None:
                    set_parts.append(f"{f} = %s")
                    values.append(v)
            if body.template_name is not None:
                template_id = _lookup_id(cur, "template", "template_id", "template_name", body.template_name)
                set_parts.append("template_id = %s")
                values.append(template_id)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            try:
                cur.execute(
                    f"UPDATE package_size SET {', '.join(set_parts)} WHERE package_size_id = %s",
                    (*values, package_size_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไข Package Size ไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "Package Size not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/package-sizes/{package_size_id}")
async def delete_package_size(package_size_id: int):
    _delete_lookup("package_size", "package_size_id", package_size_id, [("part_number", "package_size_id")])
    return {"ok": True}


class PartNumberCreate(BaseModel):
    part_number_name: str
    package_size:     str
    handler:          str
    nominal_x:        float
    nominal_y:        float
    upper_tol:        float
    lower_tol:        float
    offset_tol:       float = 0


class PartNumberUpdate(BaseModel):
    part_number_name: Optional[str] = None
    package_size:     Optional[str] = None
    handler:          Optional[str] = None
    nominal_x:        Optional[float] = None
    nominal_y:        Optional[float] = None
    upper_tol:        Optional[float] = None
    lower_tol:        Optional[float] = None
    offset_tol:       Optional[float] = None


@app.post("/api/part-numbers", status_code=201)
async def create_part_number(body: PartNumberCreate):
    db = get_db()
    try:
        with db.cursor() as cur:
            package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", body.package_size)
            handler_id      = _lookup_id(cur, "handler",      "handler_id",      "handler_name",  body.handler)
            if package_size_id is None or handler_id is None:
                raise HTTPException(400, "ต้องระบุ package_size และ handler ที่มีอยู่จริงในระบบ")
            try:
                cur.execute(
                    "INSERT INTO part_number "
                    "(part_number_name, package_size_id, handler_id, nominal_x, nominal_y, "
                    " upper_tol, lower_tol, offset_tol) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    (
                        body.part_number_name, package_size_id, handler_id,
                        body.nominal_x, body.nominal_y, body.upper_tol, body.lower_tol,
                        body.offset_tol,
                    ),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่ม Part Number ไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"part_number_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/part-numbers/{part_number_id}")
async def update_part_number(part_number_id: int, body: PartNumberUpdate):
    db = get_db()
    try:
        with db.cursor() as cur:
            set_parts, values = [], []
            if body.part_number_name is not None:
                set_parts.append("part_number_name = %s")
                values.append(body.part_number_name)
            if body.package_size is not None:
                package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", body.package_size)
                set_parts.append("package_size_id = %s")
                values.append(package_size_id)
            if body.handler is not None:
                handler_id = _lookup_id(cur, "handler", "handler_id", "handler_name", body.handler)
                set_parts.append("handler_id = %s")
                values.append(handler_id)
            for field_name in ("nominal_x", "nominal_y", "upper_tol", "lower_tol", "offset_tol"):
                value = getattr(body, field_name)
                if value is not None:
                    set_parts.append(f"{field_name} = %s")
                    values.append(value)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            try:
                cur.execute(
                    f"UPDATE part_number SET {', '.join(set_parts)} WHERE part_number_id = %s",
                    (*values, part_number_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไข Part Number ไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "Part Number not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/part-numbers/{part_number_id}")
async def delete_part_number(part_number_id: int):
    _delete_lookup("part_number", "part_number_id", part_number_id, [("parts_specifications", "part_number_id")])
    return {"ok": True}


PARTS_SELECT = """
    SELECT p.part_id, p.number_alpl, pn.part_number_name AS part_number,
           p.description, p.po_number,
           p.recieve_date   AS recieve_date,
           h.handler_name   AS handler,
           v.vendor_name    AS vendor,
           o.owner_name     AS owner,
           ps.package_size  AS package_size,
           pn.nominal_x     AS nominal_x,
           pn.nominal_y     AS nominal_y,
           pn.upper_tol     AS upper_tol,
           pn.lower_tol     AS lower_tol,
           pn.offset_tol    AS offset_tol,
           ps.nominal_x     AS nominal_x_pkg,
           ps.nominal_y     AS nominal_y_pkg,
           ps.upper_tol     AS upper_tol_pkg,
           ps.lower_tol     AS lower_tol_pkg,
           ps.offset_tol    AS offset_tol_pkg,
           t.template_name  AS template_name
    FROM parts_specifications p
    LEFT JOIN part_number pn  ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h       ON pn.handler_id = h.handler_id
    LEFT JOIN vendor v        ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o         ON p.owner_id = o.owner_id
    LEFT JOIN package_size ps ON ps.package_size_id =
                                 COALESCE(p.package_size_id, pn.package_size_id)
    LEFT JOIN template t      ON ps.template_id = t.template_id
"""


def _lookup_id(cur, table: str, id_col: str, name_col: str, value: Optional[str]) -> Optional[int]:
    """แปลงชื่อ (เช่น vendor_name ที่ frontend ส่งมาจาก dropdown) เป็น id
    (เช่น vendor_id) จากตาราง lookup ที่เกี่ยวข้อง

    ทำไม: dropdown ฝั่ง frontend (Operator/Owner/Vendor/Handler/
    Package Size) เป็น dropdown "ปิด" — เลือกได้เฉพาะค่าที่มีอยู่แล้วใน DB
    เท่านั้น ไม่มีช่องพิมพ์ค่าใหม่ ดังนั้นค่าที่ส่งเข้ามาควรมีอยู่จริงเสมอ แต่ยัง
    defensive เช็คไว้กันกรณี frontend ค้างข้อมูลเก่า/ผิดพลาด — ถ้าหาไม่เจอ
    ให้ 400 ชัดเจนแทนที่จะปล่อยให้ FK constraint error กลายเป็น 500 ตอน insert
    """
    if value in (None, ""):
        return None
    cur.execute(f"SELECT {id_col} FROM {table} WHERE {name_col} = %s", (value,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(400, f"ไม่พบค่า '{value}' ใน {table} (เลือกจาก dropdown เท่านั้น)")
    return row[id_col]


def _block_if_session_running(cur, action: str) -> None:
    """เช็คว่ามี session ไหนกำลัง running อยู่ไหม — ถ้ามี ปฏิเสธการแก้ไข/ลบ Part
    หรือ Measurement ทันที (ทั้งคู่ ไม่ว่า ALPL ไหน) เพราะข้อมูลกำลังถูกวัดอยู่จริง

    ทำไมต้อง block กว้างขนาดนี้ (ไม่ใช่แค่ ALPL ที่กำลังวัดอยู่): ผู้ใช้เลือกไว้
    ชัดเจนว่าอยากให้ Edit/Delete กดไม่ได้เลยทั้ง Part และ Measurement ตราบใดที่
    ยังมี session running อยู่ — เพื่อความคาดเดาได้ง่าย ไม่ต้องตามว่า ALPL ไหน
    "ปลอดภัย" ไหม (ระบบรัน session พร้อมกันได้แค่ 1 อันเสมออยู่แล้ว — ดู
    Button Guard ใน start_session — เช็คแค่ "มี session running อยู่ไหม" จึง
    เทียบเท่ากับ "session ปัจจุบันกำลังวัดอยู่ไหม")

    ╔═══ เพิ่ม Export เข้ามาด้วย — 7 ส.ค. 2569 ════════════════════════════════╗
    เดิมกันแค่ Edit/Delete ส่วน Export กดได้ตลอดเวลา — แต่ endpoint ของ export
    เป็น `async def` ที่เรียก pymysql (sync ล้วน ไม่คืน control) จึง **บล็อก event
    loop ทั้งเส้นระหว่างทำงาน** และ export มีเพดานถึง REPORT_MAX_ROWS แถว

        มีคนกด Export XLSX ระหว่างวัด ใช้ 20 วิ
        → event loop ค้าง → heartbeat ของ Pi เข้าไม่ได้เลย 4 รอบติด
        → loop คลาย → heartbeat_checker เห็น last_seen เก่า 20 วิ
        → session ที่วัดอยู่ดีๆ กลายเป็น 'timeout' ทั้งที่ Pi ปกติทุกอย่าง

    ล็อก Export ตอนวัดปิดช่องนี้ได้หมด โดยไม่ต้องไล่แก้ endpoint อีก 56 ตัวที่เป็น
    `async def` แต่ไม่มี `await` (ตัวอื่น query เล็กระดับมิลลิวินาที ไม่เป็นปัญหา)

    ⚠ ห้ามใส่ guard นี้ใน endpoint ที่ต้องใช้ได้ตอนยังไม่เริ่มวัด เช่น
      /api/parts/check (ดู PLAN_criteria_and_multigroup.md ข้อ D6)
    ╚═══════════════════════════════════════════════════════════════════════╝
    """
    cur.execute("SELECT 1 FROM sessions WHERE state = 'running' LIMIT 1")
    if cur.fetchone():
        raise HTTPException(
            409,
            f"ไม่สามารถ{action}ข้อมูลได้ขณะนี้ — กำลังมีการวัดอยู่ "
            f"กรุณากด Stop ก่อนแล้วค่อยดำเนินการ",
        )


class PartCreate(BaseModel):
    number_alpl:   int
    part_number:   Optional[str] = None
    description:   Optional[str] = None
    vendor:        Optional[str] = None
    po_number:     Optional[int] = None
    package_size:  Optional[str] = None
    owner:         Optional[str] = None
    recieve_date:  Optional[str] = None


class PartsCheckRequest(BaseModel):
    alpl: List[int] = []
    groups: Optional[List[Dict[str, Any]]] = None
    mode: Optional[str] = None


_GROUP_MATCH_FIELDS = {
    "IPM":    [("package_size", "Package Size")],
    "Rework": [("package_size", "Package Size"), ("part_number", "Part Number")],
    "New":    [],
}


@app.post("/api/parts/check")
async def check_parts(body: PartsCheckRequest):
    """ถามทีเดียวว่า ALPL ชุดนี้ตัวไหน "มีอยู่แล้ว" / "ยังไม่มี" ในตาราง Parts

    ใช้ตอนกด Save ในฟอร์ม Part Entry — ทั้ง 3 โหมดใช้ endpoint เดียวกัน ต่างกัน
    แค่ว่าสนใจฝั่งไหนของคำตอบ (ดู PLAN_criteria_and_multigroup.md ข้อ D5/D6)

        IPM     ดู missing → ถามยืนยันว่าจะลงทะเบียนให้แล้ววัดต่อไหม
        Rework  ดู missing → บล็อก (Rework ต้องเคยวัดมาก่อนเท่านั้น)
        New     ดู exists  → บล็อก (ปล่อยผ่านจะเขียนทับ config เดิมที่มีประวัติ)

    ทำไมต้องมี endpoint นี้: ของเดิมหน้าเว็บโหลด Part **ทั้งตาราง** มาไว้ใน
    หน่วยความจำแล้วเทียบเอง (วนดึงทีละ 1000 จนหมด) — ยิ่ง Part เยอะยิ่งช้า
    ทั้งที่จะเช็คแค่ไม่กี่ตัว

    `detail` มีไว้ให้ข้อความเตือนของโหมด New โชว์ได้ว่า ALPL ที่ชนอยู่นั้น
    ปัจจุบันเป็นของ Part Number/Vendor อะไร — ผู้ใช้จะได้แยกออกว่า "พิมพ์เลขผิด"
    หรือ "เลือกโหมดผิด" ซึ่งเป็น 2 สาเหตุที่พบบ่อยที่สุด

    ╔═══ `conflicts` — ALPL ในกลุ่มเดียวกันต้องเข้าชุดกัน ═════════════════════╗
    ถ้าส่ง `groups` + `mode` มาด้วย จะตรวจเพิ่มว่า ALPL ที่ผู้ใช้จับใส่กลุ่ม
    เดียวกัน มี Package Size (และ Part Number ถ้าเป็น Rework) ตรงกันจริงไหม

    ทำไมสำคัญ: 1 กลุ่ม = ชิ้นงานที่ใช้ **config ชุดเดียวกัน** ถ้าในกลุ่มเดียวกัน
    มี ALPL ที่ผูก Package Size คนละอัน แปลว่าเกณฑ์ตัดสิน OK/NG ของแต่ละชิ้น
    ไม่เหมือนกันตั้งแต่ต้น — Pi จะคัดของด้วยเกณฑ์ของกลุ่ม (ชุดเดียว) แต่ backend
    บันทึกด้วยเกณฑ์รายตัว กลายเป็นตัดสินคนละมาตรฐานโดยไม่มีใครรู้จนกว่าจะไปนับ
    ของจริง (เป็นอาการเดียวกับที่ `_build_groups` กันไว้ตอนกด Start — ที่นี่แค่
    ดักให้เร็วขึ้นตั้งแต่ตอนกด Save จะได้แก้ก่อนที่จะเดินไปไกลกว่านั้น)

    ตรวจเฉพาะ ALPL ที่ **มีอยู่จริงใน DB** — ตัวที่ยังไม่ลงทะเบียนไม่มีอะไรให้
    เทียบ และจะถูกสร้างด้วย config ของกลุ่มอยู่แล้วตอนวัดจริง
    ╚═══════════════════════════════════════════════════════════════════════╝

    ⚠ **ห้ามใส่ `_block_if_session_running()`** — เป็นการอ่านอย่างเดียว และต้อง
      ใช้ได้ตอนที่ยังไม่ได้เริ่มวัด (ซึ่งคือตอนเดียวที่มีคนเรียกจริงๆ)
    """
    groups_in = body.groups or []
    flat: List[int] = list(body.alpl or [])
    for g in groups_in:
        flat.extend(int(a) for a in (g.get("alpl") or g.get("number_alpl") or []))

    seen: set = set()
    wanted: List[int] = []
    for a in flat:
        if a not in seen:
            seen.add(a)
            wanted.append(a)

    if not wanted:
        return {"exists": [], "missing": [], "detail": {}, "conflicts": []}
    if len(wanted) > 500:
        raise HTTPException(400, "เช็คได้สูงสุด 500 ALPL ต่อครั้ง")

    db = get_db()
    try:
        with db.cursor() as cur:
            placeholders = ", ".join(["%s"] * len(wanted))
            cur.execute(
                "SELECT p.number_alpl, pn.part_number_name, ps.package_size, "
                "       v.vendor_name, o.owner_name, p.po_number, p.description "
                "FROM parts_specifications p "
                "LEFT JOIN part_number pn ON p.part_number_id = pn.part_number_id "
                "LEFT JOIN package_size ps "
                "       ON ps.package_size_id = COALESCE(p.package_size_id, pn.package_size_id) "
                "LEFT JOIN vendor v ON p.vendor_id = v.vendor_id "
                "LEFT JOIN owner o  ON p.owner_id  = o.owner_id "
                f"WHERE p.number_alpl IN ({placeholders})",
                wanted,
            )
            rows = cur.fetchall()
    finally:
        db.close()

    detail = {
        str(r["number_alpl"]): {
            "part_number":  r["part_number_name"],
            "package_size": r["package_size"],
            "vendor":       r["vendor_name"],
            "owner":        r["owner_name"],
            "po_number":    r["po_number"],
            "description":  r["description"],
        }
        for r in rows
    }
    found = {r["number_alpl"] for r in rows}

    conflicts: List[Dict[str, Any]] = []
    for field, label in _GROUP_MATCH_FIELDS.get((body.mode or "").strip(), []):
        for gi, g in enumerate(groups_in):
            alpls = [int(a) for a in (g.get("alpl") or g.get("number_alpl") or [])]
            by_value: Dict[str, List[int]] = {}
            for a in alpls:
                if a not in found:
                    continue
                v = detail[str(a)].get(field)
                by_value.setdefault("—" if v is None else str(v), []).append(a)

            if len(by_value) > 1:
                parts = " · ".join(
                    f"{', '.join(map(str, ids))} = \"{v}\"" for v, ids in by_value.items()
                )
                conflicts.append({
                    "group":  gi,
                    "field":  field,
                    "label":  label,
                    "alpl":   alpls,
                    "values": by_value,
                    "message": (
                        f"กลุ่มที่ {gi + 1}: ALPL ในกลุ่มเดียวกันมี {label} ไม่ตรงกัน — "
                        f"{parts} · ชิ้นงานในกลุ่มเดียวกันต้องใช้ค่าเดียวกัน "
                        f"ให้แยกไปคนละกลุ่ม หรือแก้ให้ตรงกันที่หน้า Edit › Parts"
                    ),
                })

    return {
        "exists":    [a for a in wanted if a in found],
        "missing":   [a for a in wanted if a not in found],
        "detail":    detail,
        "conflicts": conflicts,
    }


@app.get("/api/parts")
async def list_parts(
    limit:  int = Query(10, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
):
    """คืน config ของ parts แบบ "แบ่งหน้า" (server-side pagination)

    ทำไมเปลี่ยนจากเดิมที่คืน array ของ parts ทั้งหมดทีเดียว มาเป็น object
    {items, total}: ตาราง parts มีแนวโน้มโตขึ้นเรื่อยๆ ตามการใช้งานจริง การ
    ดึงมาทั้งหมดทุกครั้งจะกินแบนด์วิดท์และหน่วยความจำ frontend โดยเปล่าประโยชน์
    จึงให้ดึงมาทีละหน้า (limit/offset) แทน — frontend ของ edit.html แสดงทีละ 10 แถว

    `total` มาจาก COUNT(*) ที่ใช้ WHERE ชุดเดียวกับ query หลัก (ไม่ใช่นับจาก
    items ของหน้าปัจจุบัน) เพื่อให้ frontend คำนวณจำนวนหน้า/ปิดปุ่ม Next ได้ถูก

    `search` (optional) — กรองด้วย number_alpl หรือ part_number แบบ LIKE
    เหตุผลที่ทำ search ฝั่ง server ไม่ใช่ฝั่ง client: เพื่อให้ค้นหาเจอ "ข้ามทุก
    หน้า" ไม่ใช่แค่ 10 แถวของหน้าที่โหลดมาแสดงอยู่ number_alpl เป็น INT จึงต้อง
    CAST เป็น CHAR ก่อนเทียบ LIKE เพื่อให้ค้นหาบางส่วนของตัวเลขได้ (เช่นพิมพ์
    "10" แล้วเจอทั้ง 1011, 1002 ที่ขึ้นต้นด้วย 10)

    หมายเหตุ: เพิ่ม ORDER BY number_alpl เพื่อให้ลำดับของหน้าคงที่ (stable) —
    ถ้าไม่กำหนด ORDER BY การไล่ LIMIT/OFFSET อาจได้ลำดับไม่แน่นอนข้ามหน้า
    """
    conditions, params = [], []
    if search:
        conditions.append("(CAST(p.number_alpl AS CHAR) LIKE %s OR pn.part_number_name LIKE %s)")
        like = f"%{search}%"
        params.extend([like, like])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) AS total FROM parts_specifications p "
                f"LEFT JOIN part_number pn ON p.part_number_id = pn.part_number_id {where}",
                params,
            )
            total = cur.fetchone()["total"]
            cur.execute(
                f"{PARTS_SELECT} {where} ORDER BY p.number_alpl LIMIT %s OFFSET %s",
                (*params, limit, offset),
            )
            items = cur.fetchall()
        return {"items": items, "total": total}
    finally:
        db.close()


@app.get("/api/parts/{part_id}")
async def get_part(part_id: int):
    """คืน config ของ part 1 ตัวตาม ALPL number (รวมชื่อ handler/vendor/
    owner/package_size ที่ join มาจาก lookup table แล้ว ไม่ใช่แค่ id)"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"{PARTS_SELECT} WHERE p.number_alpl = %s", (part_id,))
            row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Part not found")
        return row
    finally:
        db.close()


def _insert_part_row(cur, number_alpl: int, config: Dict[str, Any]) -> None:
    """Insert 1 row ลง table `parts_specifications` โดยใช้ number_alpl ที่ระบุ
    + field อื่นจาก `config` (dict ของ field part_number/vendor/owner ฯลฯ) —
    part_number/vendor/owner รับมาเป็น "ชื่อ" (ตรงกับค่าที่เลือกจาก dropdown
    ฝั่ง frontend) แล้ว resolve เป็น id ก่อน insert

    หมายเหตุ: handler/package_size ไม่ได้ resolve/insert ตรงนี้แล้ว เพราะ
    schema ใหม่ derive ค่าพวกนี้มาจาก part_number_id ทั้งหมด (part_number
    catalog ผูก package_size_id + handler_id ของตัวเองไว้แล้ว — ดู init.sql)

    ใช้ร่วมกันทั้งจาก endpoint POST /api/parts ปกติ, จาก flow ของ New queue ที่
    ALPL หลายตัวใช้ config เดียวกันซ้ำ, และจากการลงทะเบียน part แบบขั้นต่ำตอน
    IPM เจอ ALPL ที่ยังไม่เคยลงทะเบียน (ดู start_session / create_measurement)
    """
    part_number_id = _lookup_id(cur, "part_number", "part_number_id", "part_number_name", config.get("part_number"))
    vendor_id      = _lookup_id(cur, "vendor",      "vendor_id",      "vendor_name",      config.get("vendor"))
    owner_id       = _lookup_id(cur, "owner",       "owner_id",       "owner_name",       config.get("owner"))
    package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", config.get("package_size"))

    columns = [
        "number_alpl", "part_number_id", "package_size_id", "description",
        "vendor_id", "po_number", "owner_id", "recieve_date",
    ]
    values: List[Any] = [
        number_alpl, part_number_id, package_size_id, config.get("description"),
        vendor_id, config.get("po_number"), owner_id,
        config.get("recieve_date") or None,
    ]

    placeholders = ", ".join(["%s"] * len(values))
    cur.execute(
        f"INSERT INTO parts_specifications ({', '.join(columns)}) VALUES ({placeholders})",
        tuple(values),
    )


def _update_part_row(cur, number_alpl: int, config: Dict[str, Any]) -> None:
    """Update 1 row ที่มีอยู่แล้วใน table `parts_specifications` (ตรงข้ามกับ
    _insert_part_row)

    ใช้เฉพาะกรณี Rework: ALPL นี้เคยผ่าน New มาแล้ว (มี Part row อยู่จริง) แต่
    งานไม่ผ่าน ถูกส่งไป Rework แล้วส่งกลับมาวัดใหม่ — ฟอร์ม Rework auto-fill
    ทุกช่องด้วยข้อมูลเดิมของ ALPL นี้ไว้ให้แล้ว (ดู GET /api/parts/{part_id})
    ยกเว้น recieve_date ที่ผู้ใช้ต้องกรอกวันที่รับกลับมาใหม่เอง — Save จึง
    เขียนทับ row เดิม (WHERE number_alpl = %s) แทนที่จะ insert row ใหม่ซ้อน
    ขึ้นมา ซึ่งจะชนกับ UNIQUE constraint ของ number_alpl ทันที
    """
    part_number_id = _lookup_id(cur, "part_number", "part_number_id", "part_number_name", config.get("part_number"))
    vendor_id      = _lookup_id(cur, "vendor",      "vendor_id",      "vendor_name",      config.get("vendor"))
    owner_id       = _lookup_id(cur, "owner",       "owner_id",       "owner_name",       config.get("owner"))
    package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", config.get("package_size"))
    cur.execute(
        "UPDATE parts_specifications SET part_number_id = %s, package_size_id = %s, "
        "description = %s, vendor_id = %s, po_number = %s, owner_id = %s, "
        "recieve_date = %s "
        "WHERE number_alpl = %s",
        (
            part_number_id, package_size_id, config.get("description"),
            vendor_id, config.get("po_number"), owner_id,
            config.get("recieve_date") or None, number_alpl,
        ),
    )


@app.post("/api/parts", status_code=201)
async def create_part(part: PartCreate):
    """ลงทะเบียน ALPL part ใหม่: nominal X/Y, tolerance แยกแกน, และ template ของ TM-X ที่ใช้

    ทำไมต้องแยก endpoint นี้ออกจาก start_session: parts/templates ถูก config
    ไว้ล่วงหน้า (เป็นขั้นตอน setup) เพื่อให้ start_session แค่ lookup template
    จาก number_alpl ได้เลย ไม่ต้องให้ผู้ปฏิบัติงานพิมพ์เองทุกครั้งที่รัน
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _insert_part_row(cur, part.number_alpl, part.dict())
        return {"number_alpl": part.number_alpl}
    finally:
        db.close()


@app.patch("/api/parts/{part_id}")
async def update_part(part_id: int, data: Dict[str, Any] = Body(...)):
    """อัปเดต config ของ part แบบ partial (เฉพาะ field ที่ส่งมาใน body)

    ทำไมต้องมี whitelist (`allowed`): เพื่อไม่ให้ request body ไปเขียนทับ column
    ที่ไม่ควรแก้ผ่าน endpoint นี้ได้โดยไม่ตั้งใจ (หรือถูกใช้ในทางที่ไม่ดี)

    `number_alpl` แก้ไขได้ (จาก edit.html — Edit Part) แม้จะเป็น "business key"
    หลักที่ sessions/measurements อ้างอิงถึงก็ตาม — ถ้า ALPL ตัวนี้มีประวัติ
    session/measurement ผูกอยู่แล้ว MySQL จะปฏิเสธด้วย FK constraint error
    (เพราะ FOREIGN KEY ไม่มี ON UPDATE CASCADE) เราจับ error นั้นแล้วแปลงเป็น
    409 ที่อ่านง่ายแทนที่จะปล่อยให้เป็น 500 ดิบๆ
    """
    direct_fields = {"number_alpl", "description", "po_number", "recieve_date"}
    lookup_fields = {
        "part_number": ("part_number_id", "part_number", "part_number_id", "part_number_name"),
        "vendor":      ("vendor_id",      "vendor",      "vendor_id",      "vendor_name"),
        "owner":       ("owner_id",       "owner",       "owner_id",       "owner_name"),
    }
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "แก้ไข")
            set_parts, values = [], []
            for k, v in data.items():
                if k in direct_fields:
                    set_parts.append(f"{k} = %s")
                    values.append(v)
                elif k in lookup_fields:
                    col, table, id_col, name_col = lookup_fields[k]
                    set_parts.append(f"{col} = %s")
                    values.append(_lookup_id(cur, table, id_col, name_col, v))
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            set_clause = ", ".join(set_parts)
            try:
                cur.execute(
                    f"UPDATE parts_specifications SET {set_clause} WHERE number_alpl = %s",
                    (*values, part_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(
                    409,
                    f"บันทึกไม่สำเร็จ — ALPL ใหม่อาจซ้ำกับ part อื่น หรือ ALPL เดิมมี "
                    f"session/measurement ผูกอยู่แล้ว (เปลี่ยน ALPL ที่มีประวัติไม่ได้): {exc}",
                )
            if cur.rowcount == 0:
                raise HTTPException(404, "Part not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/parts/{part_id}")
async def delete_part(part_id: int):
    """ลบ Part 1 row ออกจากตาราง `parts_specifications`

    **กฎการลบ (ตามที่ตกลงกันไว้)**: ลบได้ต่อเมื่อ ALPL นี้ "ไม่มีข้อมูลการวัด
    (Measurement) เหลืออยู่เลย" เท่านั้น — ถ้ายังมี Measurement อยู่จะปฏิเสธ
    ทันทีด้วย 409 พร้อมข้อความบอกจำนวนรายการที่ติดอยู่ (frontend เอาไปโชว์เป็น
    popup เตือน) ไม่มีโหมด cascade ที่ลบ Measurement ตามไปด้วยอีกต่อไป —
    ประวัติการวัดถือเป็นข้อมูลจริงที่ห้ามหายไปโดยไม่ตั้งใจ ถ้าผู้ใช้ต้องการลบ
    จริงๆ ต้องไปลบ Measurement ทีละรายการเองที่ตาราง Measurements ก่อน

    ทำทั้งหมดเป็น transaction เดียว (ปิด autocommit ชั่วคราว) กัน DB ค้างครึ่งๆ
    กลางๆ ถ้ามีขั้นไหนพังกลางทาง

    ╔═══ บั๊ก HTTP 500 ที่เพิ่งแก้ (13 ส.ค. 2569) ══════════════════════════╗
    ที่นี่เคยมีอีก 2 คำสั่งที่ยุ่งกับ `sessions`:
        SELECT * FROM sessions WHERE number_alpl = %s     ← สำรองก่อนลบ
        DELETE   FROM sessions WHERE number_alpl = %s     ← เคลียร์ให้ FK ไม่ block
    ทั้งคู่ตกค้างมาจากสมัยที่ `sessions` ยังมีคอลัมน์ `number_alpl` + FK ไป
    `parts_specifications` — คอลัมน์นั้นถูกถอดออกไปแล้ว (ดู init.sql) MySQL จึง
    ตอบ `1054 Unknown column 'number_alpl' in 'where clause'` ซึ่งเป็น
    pymysql.MySQLError ที่ไม่มีใคร catch → หลุดออกไปเป็น **HTTP 500** ทุกครั้ง
    ที่กดลบ Part จากหน้า Edit

    ตอนนี้ `sessions` ไม่ได้อ้างถึง Part ตัวไหนแล้ว (คิว ALPL อยู่ใน
    `queue_state` ซึ่งเป็น JSON ไม่ใช่ FK) จึงไม่มีอะไรต้องเคลียร์ก่อนลบ
    และไม่มี FK ตัวไหนมา block ด้วย — ลบ 2 บรรทัดนั้นทิ้งได้เลย
    ╚═══════════════════════════════════════════════════════════════════════╝
    """
    db = get_db()
    try:
        db.autocommit(False)
        with db.cursor() as cur:
            _block_if_session_running(cur, "ลบ")

            cur.execute("SELECT COUNT(*) AS n FROM measurements WHERE number_alpl = %s", (part_id,))
            measurement_count = cur.fetchone()["n"]
            if measurement_count:
                db.rollback()
                raise HTTPException(
                    409,
                    f"ลบไม่ได้ — ALPL {part_id} ยังมีข้อมูลการวัด (Measurement) อยู่ "
                    f"{measurement_count} รายการ กรุณาลบ Measurement ของ ALPL นี้ให้หมดก่อน "
                    f"แล้วค่อยลบ Part",
                )

            part_row = _fetch_one(
                cur, "SELECT * FROM parts_specifications WHERE number_alpl = %s", (part_id,)
            )
            if part_row is None:
                db.rollback()
                raise HTTPException(404, "Part not found")
            _archive_before_delete(
                kind="part", table="parts_specifications",
                pk={"number_alpl": part_id}, row=part_row,
            )

            try:
                cur.execute("DELETE FROM parts_specifications WHERE number_alpl = %s", (part_id,))
            except pymysql.MySQLError as exc:
                db.rollback()
                raise HTTPException(409, f"ลบ ALPL {part_id} ไม่สำเร็จ: {exc}")
            if cur.rowcount == 0:
                db.rollback()
                raise HTTPException(404, "Part not found")
        db.commit()
        return {"ok": True}
    finally:
        db.autocommit(True)
        db.close()


MEASUREMENTS_SELECT = """
    SELECT m.*, op.operator_name AS operator_name,
           CASE WHEN m.measure_type = 'IPM' THEN ips.nominal_x  ELSE pn.nominal_x  END AS nominal_x,
           CASE WHEN m.measure_type = 'IPM' THEN ips.nominal_y  ELSE pn.nominal_y  END AS nominal_y,
           CASE WHEN m.measure_type = 'IPM' THEN ips.upper_tol  ELSE pn.upper_tol  END AS upper_tol,
           CASE WHEN m.measure_type = 'IPM' THEN ips.lower_tol  ELSE pn.lower_tol  END AS lower_tol,
           CASE WHEN m.measure_type = 'IPM' THEN NULL           ELSE pn.offset_tol END AS offset_tol
    FROM measurements m
    LEFT JOIN operator op ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn ON p.part_number_id = pn.part_number_id
    LEFT JOIN package_size ips
           ON ips.package_size_id = COALESCE(p.package_size_id, pn.package_size_id)
"""


class MeasurementCreate(BaseModel):
    session_id:  Optional[int] = None
    number_alpl: Optional[int] = None
    value_x:     float
    value_y:     float
    offset:      float = 0
    note:        Optional[str] = None
    client_uuid: Optional[str] = None


class ImageUpdate(BaseModel):
    image_path:    Optional[str] = None
    upload_failed: bool = False


@app.get("/api/measurements")
async def list_measurements(
    number_alpl: Optional[int] = None,
    result:      Optional[str] = None,
    date_from:   Optional[str] = None,
    date_to:     Optional[str] = None,
    session_id:  Optional[int] = None,
    limit:  int = Query(100, le=1000),
    offset: int = Query(0, ge=0),
):
    """Query ประวัติ measurement พร้อม filter ที่เลือกได้ — ใช้ทั้งโดยตาราง dashboard
    (renderTable) และเป็นฐานของ /api/export/csv filter ทุกตัวเป็น optional
    และรวมกันด้วย AND

    เปลี่ยน response shape เป็น object {items, total} เหมือน /api/parts เพื่อรองรับ
    server-side pagination (เพิ่ม offset เข้ามาคู่กับ limit ที่มีอยู่เดิม) — measurements
    โตเร็วกว่า parts มาก จึงไม่ควรโหลดทั้งหมดมา slice ฝั่ง client

    `total` ใช้ COUNT(*) บน WHERE ชุดเดียวกับ query หลัก (filter เดียวกัน) เพื่อให้
    frontend รู้จำนวนทั้งหมดที่ตรงกับ filter ปัจจุบัน ไว้คำนวณหน้า/ปิดปุ่ม Next

    หมายเหตุ: /api/export/csv เป็น endpoint แยกที่ "ไม่" reuse ฟังก์ชันนี้ (มันสร้าง
    WHERE ของตัวเองและคืนไฟล์ CSV ไม่ใช่ JSON) — การเปลี่ยน shape ตรงนี้จึงไม่กระทบ export
    """
    conditions, params = [], []
    if number_alpl is not None:
        conditions.append("m.number_alpl = %s"); params.append(number_alpl)
    if result:
        conditions.append("m.result = %s"); params.append(result)
    if date_from:
        conditions.append("m.timestamp >= %s"); params.append(_day_start(date_from))
    if date_to:
        conditions.append("m.timestamp <= %s"); params.append(_day_end(date_to))
    if session_id is not None:
        conditions.append("m.session_id = %s"); params.append(session_id)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS total FROM measurements m {where}", params)
            total = cur.fetchone()["total"]
            cur.execute(
                f"{MEASUREMENTS_SELECT} {where} ORDER BY m.timestamp DESC LIMIT %s OFFSET %s",
                (*params, limit, offset),
            )
            items = cur.fetchall()
        return {"items": items, "total": total}
    finally:
        db.close()


@app.post("/api/measurements")
async def create_measurement(req: MeasurementCreate):
    """บันทึก measurement หนึ่งรายการที่ส่งมาจาก Agent และตัดสิน OK/NG

    Endpoint นี้ถูกเรียกครั้งละ 1 ชิ้นงานที่ TM-X วัดได้ ส่งต่อมาโดย Agent
    พร้อม session_id ที่ได้รับตอนเริ่ม session — Agent ส่ง `number_alpl` มาด้วย
    เหมือนเดิมเสมอ (ไม่ต้องแก้ agent.py) แต่ตอนนี้ backend จะตัดสินเองว่าจะใช้
    ALPL ไหนจริงๆ ตามประเภทของ session:

    **ALPL มาจากคิวของ backend เท่านั้น** — เพิกเฉยค่า `req.number_alpl` ที่ Agent
    ส่งมาเสมอ แล้วใช้ ALPL ตามตำแหน่งปัจจุบันในคิวแทน
    (`session_queues[session_id]["queue"][position]`) เพราะ Agent ไม่รู้
    (และไม่จำเป็นต้องรู้) ว่ากำลังวัดตัวไหนอยู่ในคิว มันรู้แค่ว่า
    "วัดเสร็จแล้ว ได้ value_x/value_y เท่านี้"

    เส้นทางที่ถูกถอดออกไปแล้ว (อย่าเอากลับมาโดยไม่คุยกันก่อน):

      - **Manual add จากหน้าเว็บ** (`session_id` เป็น None) — ปุ่ม
        "+ Add Measurement" ใน edit.html ถูกถอดออกแล้วตามที่ตกลงกันว่า
        **ผลวัดต้องมาจากการวัดจริงเท่านั้น ห้ามพิมพ์เอง** จึงไม่รับ POST ที่ไม่มี
        session_id อีกต่อไป (เดิมจะสร้าง session ปลอมให้ 1 แถวแล้วบันทึกด้วย
        measure_type='Manual')
      - **Fallback ใช้ `req.number_alpl` ตอนคิวหาย** — เดิมถ้า `session_queues`
        ไม่มี entry ของ session นี้ (เช่น backend restart แล้วโหลดคิวกลับไม่สำเร็จ)
        จะตกมาใช้ค่าที่ Agent ส่งมา ซึ่ง **ผิดเสมอตั้งแต่ชิ้นที่ 2 เป็นต้นไป**
        เพราะค่านั้นอ่านมาจาก `sessions.number_alpl` ที่ไม่เคยถูก UPDATE เลย
        = ALPL ตัวแรกของคิวตลอดทั้ง session → ข้อมูลผิด ALPL เข้า DB แบบเงียบๆ
        ตอนนี้เปลี่ยนเป็น "ปฏิเสธไปเลย" ดีกว่าเดา (ดู HTTPException 409 ด้านล่าง)
    """
    if req.session_id is None:
        raise HTTPException(
            400,
            "ต้องมี session_id — ระบบไม่รับการเพิ่มผลวัดเองด้วยมืออีกต่อไป "
            "(ผลวัดต้องมาจากการวัดจริงผ่าน TM-X เท่านั้น)",
        )
    qstate = None
    db = get_db()
    try:
        if req.client_uuid:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT measurement_id, session_id, result FROM measurements "
                    "WHERE client_uuid = %s",
                    (req.client_uuid,),
                )
                dup = cur.fetchone()
            if dup:
                with db.cursor() as cur:
                    cur.execute(
                        "SELECT measured_count, target_count FROM sessions WHERE session_id = %s",
                        (dup["session_id"],),
                    )
                    s = cur.fetchone() or {}
                log.info(
                    "Duplicate measurement POST (client_uuid=%s) — คืนผลเดิม measurement_id=%d",
                    req.client_uuid, dup["measurement_id"],
                )
                return {
                    "measurement_id": dup["measurement_id"],
                    "result":  dup["result"],
                    "status":  "duplicate_ignored",
                    "measured": s.get("measured_count"),
                    "target":   s.get("target_count"),
                }

        with db.cursor() as cur:
            session_id = req.session_id
            cur.execute(
                "SELECT state, target_count, measured_count FROM sessions WHERE session_id = %s",
                (session_id,),
            )
            session = cur.fetchone()
            if not session or session["state"] != "running":
                raise HTTPException(400, "Session is not running")

            qstate = session_queues.get(session_id)
            if qstate is None:
                raise HTTPException(
                    409,
                    f"คิว ALPL ของ session {session_id} หายไป (backend อาจถูก restart) "
                    "— กด Stop ที่หน้าเว็บแล้วเริ่ม session ใหม่",
                )

            queue = qstate["queue"]
            pos = qstate["position"]
            if pos >= len(queue):
                raise HTTPException(400, "Measurement queue หมดแล้วสำหรับ session นี้")
            number_alpl = queue[pos]
            measure_type = qstate["entry_mode"]
            operator_name = qstate.get("operator")
            note = qstate.get("note")

            group_cfg = _group_config_for(qstate, pos)
            if group_cfg is not None:
                cur.execute("SELECT 1 FROM parts_specifications WHERE number_alpl = %s", (number_alpl,))
                exists = cur.fetchone() is not None
                try:
                    if not exists:
                        _insert_part_row(cur, number_alpl, group_cfg)
                    elif qstate.get("measure_mode") == "Rework":
                        _update_part_row(cur, number_alpl, group_cfg)
                except pymysql.MySQLError as exc:
                    raise HTTPException(409, f"บันทึก Part ALPL {number_alpl} ไม่สำเร็จ: {exc}")

            part = _load_criteria(cur, number_alpl, measure_type)

            verdict = _judge(req.value_x, req.value_y, req.offset, part, measure_type)
            result = verdict["result"]

            operator_id = _lookup_id(cur, "operator", "operator_id", "operator_name", operator_name)

            try:
                cur.execute(
                    "INSERT INTO measurements "
                    "(session_id, number_alpl, value_x, value_y, `offset`, result, "
                    " measure_type, operator_id, note, client_uuid) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (session_id, number_alpl, req.value_x, req.value_y, req.offset, result,
                     measure_type, operator_id, note, req.client_uuid),
                )
            except pymysql.IntegrityError:
                raise HTTPException(409, "Measurement นี้ถูกบันทึกไปแล้ว (duplicate client_uuid)")
            measurement_id = cur.lastrowid

            cur.execute(
                "UPDATE sessions SET measured_count = measured_count + 1 "
                "WHERE session_id = %s",
                (session_id,),
            )

            cur.execute(
                "SELECT measured_count, target_count FROM sessions WHERE session_id = %s",
                (session_id,),
            )
            updated = cur.fetchone()
            measured = updated["measured_count"]
            target   = updated["target_count"]

        if qstate is not None:
            qstate["position"] += 1
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET queue_state = %s WHERE session_id = %s",
                    (json.dumps(qstate), session_id),
                )

        status = "continue"
        if measured >= target:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET state = 'stopped', ended_at = NOW() "
                    "WHERE session_id = %s",
                    (session_id,),
                )
            status = "complete"
            session_queues.pop(session_id, None)
            measure_timeouts.pop(session_id, None)
            await push_event(
                "session_complete",
                {"session_id": session_id, "measured": measured, "target": target},
            )

        await push_event(
            "measurement",
            {
                "measurement_id": measurement_id,
                "session_id":     session_id,
                "number_alpl":    number_alpl,
                "value_x":        req.value_x,
                "value_y":        req.value_y,
                "result":         result,
                "offset":         req.offset,
                **{k: verdict[k] for k in ("ok_x", "ok_y", "ok_offset", "offset_counts", "offset_tol")},
                "measure_type":   measure_type,
                "nominal_x":      part["nominal_x"],
                "nominal_y":      part["nominal_y"],
                "upper_tol":      part["upper_tol"],
                "lower_tol":      part["lower_tol"],
                "measured":       measured,
                "target":         target,
            },
        )
        return {
            "measurement_id": measurement_id,
            "result":  result,
            "status":  status,
            "measured": measured,
            "target":  target,
        }
    finally:
        db.close()


@app.patch("/api/measurements/{measurement_id}")
async def update_measurement(measurement_id: int, data: Dict[str, Any] = Body(...)):
    """แก้ไข measurement แบบ partial (เฉพาะ field ที่ส่งมาใน body) — ใช้โดยหน้า
    Database Editor (edit.html) ตอนกด Edit แล้ว Save

    ทำไมต้องมี endpoint นี้แยกจาก create_measurement: create_measurement เป็น flow
    ของ Agent (insert ค่าใหม่ที่ TM-X วัดได้ พร้อมตัดสิน OK/NG จาก tolerance) ส่วนการ
    "แก้" ค่าที่บันทึกไว้แล้วเป็นการ override ด้วยมือจากหน้า editor ซึ่งไม่มีมาก่อน

    ใช้ pattern เดียวกับ update_part(): whitelist เฉพาะ field ที่อนุญาตให้แก้ได้
    (number_alpl, operator) เพื่อกัน body ไปเขียนทับ column ที่ไม่ควรแก้ผ่าน
    endpoint นี้ (เช่น session_id, timestamp, image_path, value_x/value_y)

    `number_alpl` แก้ไขได้ (เพิ่มเข้ามาใหม่) — ใช้กรณี IPM พิมพ์/เลือก ALPL ผิดตัว
    (เลือกชิ้นที่มีอยู่จริงในระบบผิดตัว) วิธีแก้ที่ถูกคือ retarget measurement row
    นี้ไปที่ ALPL ที่ถูกต้อง ไม่ใช่ไปแก้ ALPL ที่ตัว Part (เพราะจะกลายเป็นเปลี่ยนชื่อ
    Part จริงที่มีประวัติของตัวเองอยู่แล้ว — ดู update_part สำหรับกรณีพิมพ์ ALPL
    ผิดตอน "New" ซึ่งเหมาะจะแก้ที่ตัว Part แทน)

    ไม่มี `result` ใน allowed อีกต่อไป (เดิมให้ผู้ใช้เลือก OK/NG เองตรงๆ) — เพราะ
    ตอนนี้ ALPL/value เปลี่ยนได้ ทำให้ tolerance ที่ใช้ตัดสิน OK/NG เปลี่ยนตามไปด้วย
    จึงคำนวณ result ใหม่เสมอหลัง update ทุกครั้ง (ไม่ว่าจะแก้ field ไหนก็ตาม) แทนที่
    จะรับค่าจาก frontend ตรงๆ กัน Result ค้างไม่ตรงกับ ALPL/ค่าที่วัดได้จริง
    """
    allowed = {"number_alpl"}
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "แก้ไข")

            set_parts, values = [], []
            for k, v in data.items():
                if k in allowed:
                    set_parts.append(f"{k} = %s")
                    values.append(v)
            if "operator" in data:
                operator_id = _lookup_id(cur, "operator", "operator_id", "operator_name", data["operator"])
                if operator_id is None:
                    raise HTTPException(400, "ต้องเลือก Operator (เว้นว่างไม่ได้)")
                set_parts.append("operator_id = %s")
                values.append(operator_id)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")

            set_clause = ", ".join(set_parts)
            try:
                cur.execute(
                    f"UPDATE measurements SET {set_clause} WHERE measurement_id = %s",
                    (*values, measurement_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(
                    409,
                    f"บันทึกไม่สำเร็จ — ALPL ใหม่อาจยังไม่ได้ลงทะเบียนใน Parts: {exc}",
                )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")

            cur.execute(
                "SELECT number_alpl, value_x, value_y, `offset`, measure_type "
                "FROM measurements WHERE measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Measurement not found")
            crit = _load_criteria(cur, row["number_alpl"], row["measure_type"])
            new_result = _judge(
                row["value_x"], row["value_y"], row.get("offset"), crit, row["measure_type"]
            )["result"]
            cur.execute(
                "UPDATE measurements SET result = %s WHERE measurement_id = %s",
                (new_result, measurement_id),
            )
        return {"ok": True, "result": new_result}
    finally:
        db.close()


@app.patch("/api/measurements/{measurement_id}/image")
async def update_image(measurement_id: int, req: ImageUpdate):
    """แนบ path ของรูปภาพเข้ากับ measurement หลังจาก Agent จัดเก็บรูปเรียบร้อยแล้ว
    (เดิมคือหลังอัปโหลดขึ้น MinIO — architecture ใหม่จะเป็น path ของไฟล์ใน
    โฟลเดอร์บนเครื่อง PC แทน รอดีไซน์การจัดเก็บ finalize ก่อน)

    ทำไมต้องแยก call นี้ออกจาก create_measurement: Agent จัดเก็บรูป inspection
    *หลังจาก* ค่า measurement ถูกบันทึกไปแล้ว endpoint นี้แค่ patch path ของ
    รูปเข้ากับ measurement ทีหลัง การ broadcast 'image_updated' ทำให้ dashboard
    เปลี่ยนปุ่มรูปภาพในแถวนั้นได้โดยไม่ต้อง refresh ตารางทั้งหมด
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "UPDATE measurements SET image_path = %s, image_upload_failed = %s "
                "WHERE measurement_id = %s",
                (req.image_path, req.upload_failed, measurement_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")
        await push_event(
            "image_updated",
            {
                "measurement_id": measurement_id,
                "image_path": req.image_path,
                "upload_failed": req.upload_failed,
            },
        )
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/measurements/{measurement_id}/image-upload")
async def upload_measurement_image(measurement_id: int, file: UploadFile = File(...)):
    """รับไฟล์รูปจริง (multipart) จาก Agent แล้วบันทึกลงดิสก์ของเครื่อง PC ที่
    รัน backend นี้เอง — แทนที่ MinIO เดิมทั้งหมด (ดูหมายเหตุ ALPL_IMAGE_DIR
    ด้านบน) ต่างจาก update_image (PATCH /image) ตรงที่ endpoint นั้นรับแค่
    "path" ที่ Agent อ้างว่าเก็บไว้แล้ว (ใช้ได้ตอน Agent+Backend อยู่เครื่อง
    เดียวกัน) แต่ตอนนี้ Agent อยู่คนละเครื่อง (Pi) กับ backend (PC) จึงต้องรับ
    "เนื้อไฟล์จริง" มาด้วยเลย แล้ว backend เป็นคนตัดสินใจ path ปลายทางเอง

    path ปลายทาง (เปลี่ยนจากเดิมที่แยกตาม package_size มาเป็นแยกตามวันที่ ตามที่
    ตกลงกันไว้): ALPL_IMAGE_DIR/<DD-MM-YYYY พ.ศ.>/<number_alpl>_<DD-MM-YYYY พ.ศ.>.jpg
    เช่น "22-07-2569/203_22-07-2569.jpg" — ไฟล์ต้นทางจาก TM-X (ปกติเป็น .bmp)
    ถูกแปลงเป็น .jpg เสมอด้วย Pillow ก่อนเซฟ (ไม่เก็บไฟล์ต้นฉบับ .bmp ไว้เลย)
    เก็บเป็น "path สัมพัทธ์" (relative ต่อ ALPL_IMAGE_DIR) ลงคอลัมน์
    measurements.image_path — ไม่เก็บ absolute path เต็มๆ ลง DB เพื่อไม่ให้
    รั่วโครงสร้างไฟล์ระบบจริงออกไป และให้ /api/image-url ต่อ URL ได้ตรงๆ จาก
    ค่านี้ (ดู get_image_url ด้านล่าง กับ static mount /media/alpl ท้ายไฟล์)

    หมายเหตุ (อัปเดต): ชื่อไฟล์เปลี่ยนจาก "number_alpl + วันที่" เป็น
    "number_alpl + measurement_id + ตัวสุ่ม" (เช่น "203_5821_a3f9c8d4e5b6c7d8.jpg")
    — measurement_id ทำให้ชื่อไฟล์ไม่ซ้ำกันเองอยู่แล้วโดยธรรมชาติ (เลขรันของแต่ละ
    การวัด) ส่วนตัวสุ่ม (hex 16 ตัวอักษร จาก secrets.token_hex) กันคนในวง office
    network เดาชื่อไฟล์ ALPL ตัวอื่นถูกโดยไล่เลข measurement_id เอา (ตัว IP
    allowlist ของ office network เป็นชั้นป้องกันหลักที่กันคนนอกอยู่แล้ว ตัวสุ่มนี้
    เป็นชั้นเสริมเท่านั้น) โฟลเดอร์ปลายทางยังคงแยกตามวันที่วัดเหมือนเดิม (ดู
    dest_dir ด้านล่าง) แค่ชื่อไฟล์ข้างในไม่มีวันที่ปนแล้ว
    ผลข้างเคียง: ชื่อไฟล์ไม่ซ้ำกันเองอัตโนมัติอีกต่อไปเหมือนก่อนหน้านี้ (เดิมใช้
    number_alpl+วันที่ ล้วนๆ ทำให้เขียนทับไฟล์เดิมเองโดยไม่ต้องทำอะไรเพิ่ม) จึง
    ต้องลบไฟล์เก่าด้วยมือด้านล่าง (`old_image_path`) หลังเซฟไฟล์ใหม่สำเร็จ
    ไม่งั้นรูปเก่าจะค้างสะสมในโฟลเดอร์ทุกครั้งที่ ALPL เดียวกันถูกวัดซ้ำในวันเดียวกัน
    — ยังคงพฤติกรรมเดิมไว้ว่า "เก็บแค่รูปล่าสุดของ ALPL นั้นในแต่ละวัน"
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT number_alpl, image_path FROM measurements WHERE measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Measurement not found")

            old_image_path = row["image_path"]

            date_str = _thai_date_str()
            dest_dir = os.path.join(ALPL_IMAGE_DIR, date_str)
            os.makedirs(dest_dir, exist_ok=True)

            token = secrets.token_hex(8)
            filename = f"{row['number_alpl']}_{measurement_id}_{token}.jpg"
            dest_path_abs = os.path.join(dest_dir, filename)
            image_path_rel = f"{date_str}/{filename}"

            try:
                image_bytes = await file.read()
                img = Image.open(BytesIO(image_bytes))
                if img.mode != "RGB":
                    img = img.convert("RGB")
                img.save(dest_path_abs, "JPEG", quality=90)
            except Exception as exc:
                raise HTTPException(500, f"บันทึก/แปลงไฟล์รูปเป็น .jpg ไม่สำเร็จ: {exc}")
            finally:
                await file.close()

            cur.execute(
                "UPDATE measurements SET image_path = %s, image_upload_failed = 0 "
                "WHERE measurement_id = %s",
                (image_path_rel, measurement_id),
            )

            if old_image_path and old_image_path != image_path_rel:
                _delete_image_file(old_image_path)
        await push_event(
            "image_updated",
            {
                "measurement_id": measurement_id,
                "image_path": image_path_rel,
                "upload_failed": False,
            },
        )
        return {"ok": True, "image_path": image_path_rel}
    finally:
        db.close()


@app.delete("/api/measurements/{measurement_id}")
async def delete_measurement(measurement_id: int):
    """ลบ measurement 1 row (เช่น ลบค่าที่อ่านผิดพลาด/เป็นการทดสอบ) + ลบไฟล์รูปด้วย

    เดิมลบแค่แถวใน DB ทำให้ไฟล์รูปกลายเป็น "ไฟล์กำพร้า" ค้างสะสมใน
    ALPL_IMAGE_DIR ไปเรื่อยๆ โดยไม่มีอะไรอ้างถึงอีกเลย — และไล่เก็บทีหลังไม่ได้
    ด้วย เพราะข้อมูลที่จะใช้เทียบว่าไฟล์ไหนกำพร้า (แถวใน measurements) ถูกลบไป
    พร้อมกันแล้ว

    ลำดับสำคัญ: **อ่าน image_path ก่อน → ลบแถว → ค่อยลบไฟล์**
    ถ้าลบแถวไม่สำเร็จ (ไม่เจอ row / ติด session guard) ไฟล์ต้องยังอยู่ครบ
    ส่วนถ้าลบแถวสำเร็จแล้วลบไฟล์พลาด ก็ยอมให้ไฟล์ค้าง ดีกว่าทำให้ request พัง
    ทั้งที่ข้อมูลถูกลบไปเรียบร้อยแล้ว (autocommit=True แถวถูก commit ทันที)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "ลบ")

            row = _fetch_one(
                cur, "SELECT * FROM measurements WHERE measurement_id = %s", (measurement_id,)
            )
            if not row:
                raise HTTPException(404, "Measurement not found")
            image_path = row["image_path"]

            archived = _archive_before_delete(
                kind="measurement", table="measurements",
                pk={"measurement_id": measurement_id}, row=row,
                image_path=image_path,
            )

            cur.execute(
                "DELETE FROM measurements WHERE measurement_id = %s", (measurement_id,)
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")

        if archived is None and _delete_image_file(image_path):
            log.info("ลบไฟล์รูปของ measurement %s แล้ว (%s)", measurement_id, image_path)
        return {"ok": True}
    finally:
        db.close()


@app.get("/api/image-url/{measurement_id}")
async def get_image_url(measurement_id: int):
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT image_path, image_upload_failed FROM measurements WHERE measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Measurement not found")
        if not row["image_path"]:
            detail = (
                "Agent อัปโหลดรูปไม่สำเร็จ (ลองครบ 3 ครั้งแล้ว)"
                if row["image_upload_failed"]
                else "ยังไม่มีรูปสำหรับ measurement นี้"
            )
            raise HTTPException(404, detail)
        return {"url": f"/media/alpl/{row['image_path']}"}
    finally:
        db.close()


EXPORT_SELECT = """
    SELECT m.measurement_id, m.session_id, m.number_alpl, m.value_x, m.value_y,
           m.`offset` AS `offset`,
           m.result, m.note, m.measure_type, m.timestamp,
           op.operator_name,
           pn.part_number_name,
           CASE WHEN m.measure_type = 'IPM' THEN ps.nominal_x  ELSE pn.nominal_x  END AS nominal_x,
           CASE WHEN m.measure_type = 'IPM' THEN ps.nominal_y  ELSE pn.nominal_y  END AS nominal_y,
           CASE WHEN m.measure_type = 'IPM' THEN ps.upper_tol  ELSE pn.upper_tol  END AS upper_tol,
           CASE WHEN m.measure_type = 'IPM' THEN ps.lower_tol  ELSE pn.lower_tol  END AS lower_tol,
           CASE WHEN m.measure_type = 'IPM' THEN NULL          ELSE pn.offset_tol END AS offset_tol,
           h.handler_name, ps.package_size, t.template_name,
           v.vendor_name, o.owner_name,
           p.po_number, p.description, p.recieve_date
    FROM measurements m
    LEFT JOIN operator op             ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p  ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn          ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h               ON pn.handler_id = h.handler_id
    LEFT JOIN package_size ps
           ON ps.package_size_id = COALESCE(p.package_size_id, pn.package_size_id)
    LEFT JOIN template t              ON ps.template_id = t.template_id
    LEFT JOIN vendor v                ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o                 ON p.owner_id = o.owner_id
"""


def _fmt_num(v, digits: int = 3):
    return "" if v is None else f"{float(v):.{digits}f}"


def _csv_header(key: str) -> str:
    """หัวคอลัมน์ในไฟล์ CSV — ใช้ csv_label ถ้ามี ไม่มีก็ใช้ label ตามปกติ

    มีไว้เพราะบางช่องชื่อที่เหมาะกับ "ชิปในหน้าแก้ผังรายงาน" กับ "หัวคอลัมน์ใน
    ไฟล์ CSV" ไม่ใช่ชื่อเดียวกัน เช่น timestamp: ในรายงานเลือกได้ว่าจะเอาแค่
    วันที่หรือเวลา จึงเรียกว่า "Date" แต่ CSV ออกทั้งวันและเวลาในช่องเดียวเสมอ
    เรียกว่า Date เฉยๆ จะสื่อผิด
    """
    c = EXPORT_COLUMNS[key]
    return c.get("csv_label") or c["label"]


def _fmt_timestamp(r, fmt: str = "datetime") -> str:
    """วันเวลาของการวัด — เลือกได้ว่าจะเอาอะไรบ้าง (ผู้ใช้ติ๊กในหน้าแก้ผังรายงาน)
    date      → 30/07/2569 แบบ DD/MM/YYYY
    time      → 14:23:11
    datetime  → ทั้งคู่
    """
    ts = r["timestamp"]
    if not ts:
        return ""
    if fmt == "time":
        return ts.strftime("%H:%M:%S")
    if fmt == "date":
        return ts.strftime("%d/%m/%Y")
    return ts.strftime("%d/%m/%Y %H:%M:%S")


_TIME_FORMATS = [
    {"key": "date",     "label": "Date"},
    {"key": "time",     "label": "Time"},
    {"key": "datetime", "label": "Date & Time"},
]


def _axis_state(r, axis: str) -> str:
    """ค่าที่วัดได้ของแกนนี้ อยู่ในสเปกไหม — คืน "OK"/"NG" ("" ถ้าเทียบไม่ได้)

    ต่างจาก Result ที่เป็นค่าดิบใน DB อยู่แล้ว — อันนี้ต้องคำนวณเทียบสเปกทีละแกน
    เพราะ Result เป็นผลรวมของทั้ง X และ Y (X ผ่านแต่ Y ไม่ผ่าน → Result = NG
    ทั้งที่ตัวเลข X เองอยู่ในสเปก) ถ้าเอา Result มาใช้ระบายสีช่อง X จะสีผิด
    """
    val = r[f"value_{axis}"]
    nom = r[f"nominal_{axis}"]
    if val is None or nom is None or r["upper_tol"] is None or r["lower_tol"] is None:
        return ""
    return "OK" if _within_tolerance(val, nom, r["upper_tol"], r["lower_tol"]) else "NG"


def _tolerance_spec(r) -> str:
    """สเปกขนาดชิ้นงานแบบย่อบรรทัดเดียว — ใช้ในรายงาน PDF/Excel เท่านั้น

    รูปแบบ: "<nominal>  +<upper>  -<lower>" เช่น "9.02  +0.03  -0.01"
    ถ้า nominal X กับ Y ไม่เท่ากัน (package ไม่ใช่ทรงจัตุรัส) จะคั่นด้วย "/"
    เป็น "3.28/7.43  +0.02  -0.01" — ที่ยุบเลขเดียวตอน X=Y เพราะรายงาน
    ที่ห้อง PM Kit ใช้อยู่เขียนแบบนั้น (ดูตัวอย่างที่ผู้ใช้ให้มา)

    ทศนิยม 2 ตำแหน่งเท่านั้น (ไม่ใช่ 3 ตำแหน่งแบบ _fmt_num ปกติ) เพราะสเปก
    บนแบบชิ้นงานเขียนแค่ 2 ตำแหน่ง — ส่วนค่าที่วัดได้จริง (value_x/value_y)
    ยังใช้ 3 ตำแหน่งเหมือนเดิม เพราะต้องเห็นความละเอียดของการวัด

    หมายเหตุ: หน้า Export CSV ยังใช้คอลัมน์แยก (nominal_x/nominal_y/
    upper_tol/lower_tol) เหมือนเดิม ไม่ได้ถูกยุบตาม — ตั้งใจให้ CSV แยกช่อง
    เพื่อเอาไปคำนวณต่อได้ ส่วนรายงานเน้นอ่านง่ายบนกระดาษ
    """
    d = 2
    nx, ny = r["nominal_x"], r["nominal_y"]
    if nx is None and ny is None:
        nom = ""
    elif nx is None or ny is None:
        nom = _fmt_num(nx if nx is not None else ny, d)
    elif abs(float(nx) - float(ny)) < _TOL_EPS:
        nom = _fmt_num(nx, d)
    else:
        nom = f"{_fmt_num(nx, d)}/{_fmt_num(ny, d)}"

    parts = [p for p in (nom,
                         f"+{_fmt_num(r['upper_tol'], d)}" if r["upper_tol"] is not None else "",
                         f"-{_fmt_num(r['lower_tol'], d)}" if r["lower_tol"] is not None else "") if p]
    return "  ".join(parts)


EXPORT_COLUMNS: Dict[str, Dict[str, Any]] = {
    "item":          {"label": "Item",          "group": "ข้อมูลการวัด", "scope": "report",
                      "row_number": True, "get": lambda r: ""},
    "number_alpl":   {"label": "ALPL",          "group": "ข้อมูลการวัด",
                      "header": "Number ALPL", "get": lambda r: r["number_alpl"]},
    "value_x":       {"label": "Value X",       "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"], "state": lambda r: _axis_state(r, "x"),
                      "get": lambda r: _fmt_num(r["value_x"])},
    "value_y":       {"label": "Value Y",       "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"], "state": lambda r: _axis_state(r, "y"),
                      "get": lambda r: _fmt_num(r["value_y"])},
    "offset":        {"label": "Offset",        "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"],
                      "state": lambda r: (
                          "" if r.get("offset") is None or r.get("offset_tol") is None
                          else ("OK" if _offset_ok(r.get("offset"), r.get("offset_tol")) else "NG")),
                      "get": lambda r: _fmt_num(r.get("offset"))},
    "tolerance_spec": {
        "label": "Tolerance + Value X/Y", "group": "ข้อมูลการวัด", "scope": "report",
        "get": _tolerance_spec,
        "block": {
            "cols": 2,
            "header": "Tolerance",
            "data": [{"key": "value_x", "label": "Value X"},
                     {"key": "value_y", "label": "Value Y"}],
        },
    },
    "result":        {"label": "Result",        "group": "ข้อมูลการวัด", "values": ["OK", "NG"],
                      "state": lambda r: r["result"] or "",
                      "get": lambda r: r["result"]},
    "note":          {"label": "Note",          "group": "ข้อมูลการวัด", "get": lambda r: r["note"] or ""},
    "operator":      {"label": "Operator",      "group": "ข้อมูลการวัด", "get": lambda r: r["operator_name"] or ""},
    "measure_type":  {"label": "Measure Type",  "group": "ข้อมูลการวัด", "get": lambda r: r["measure_type"] or ""},
    "timestamp":     {"label": "Date",          "csv_label": "Timestamp",
                      "group": "ข้อมูลการวัด",
                      "header": "Date", "formats": _TIME_FORMATS,
                      "get": lambda r: _fmt_timestamp(r, "datetime"),
                      "get_fmt": _fmt_timestamp},
    "session_id":    {"label": "Session",       "group": "ข้อมูลการวัด", "get": lambda r: r["session_id"]},

    "part_number":   {"label": "Part Number",   "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["part_number_name"] or ""},
    "handler":       {"label": "Handler",       "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["handler_name"] or ""},
    "package_size":  {"label": "Package Size",  "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["package_size"] or ""},
    "template_name": {"label": "Template",      "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["template_name"] or ""},
    "nominal_x":     {"label": "Nominal X",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["nominal_x"])},
    "nominal_y":     {"label": "Nominal Y",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["nominal_y"])},
    "upper_tol":     {"label": "Upper Tol",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["upper_tol"])},
    "lower_tol":     {"label": "Lower Tol",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["lower_tol"])},
    "vendor":        {"label": "Vendor",        "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["vendor_name"] or ""},
    "owner":         {"label": "Owner",         "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["owner_name"] or ""},
    "po_number":     {"label": "PO Number",     "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["po_number"] if r["po_number"] is not None else ""},
    "description":   {"label": "Description",   "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["description"] or ""},
    "recieve_date":  {"label": "Receive Date",  "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["recieve_date"].strftime("%d/%m/%Y") if r["recieve_date"] else ""},
}


def _columns_scope(kind: str) -> str:
    return "csv" if kind == "csv" else "report"


@app.get("/api/export/columns")
async def list_export_columns(kind: str = "csv"):
    """คืน catalog คอลัมน์ที่ใส่ในเทมเพลตได้ (key/label/group)
    ใช้สร้างหน้าเลือกคอลัมน์ — frontend ไม่ต้อง hardcode รายชื่อเอง

    kind='csv'          → ได้ Nominal X / Nominal Y / Upper Tol / Lower Tol แยก 4 ช่อง
    kind='pdf'|'excel'  → ได้ช่องรวม "Tolerance" ช่องเดียวแทนทั้ง 4 ช่องนั้น
    คอลัมน์ที่ไม่ระบุ scope โผล่ทั้งสองแบบ
    """
    out = []
    for k, c in EXPORT_COLUMNS.items():
        if c.get("scope") not in (None, _columns_scope(kind)):
            continue
        item = {"key": k, "label": c["label"], "group": c["group"]}
        if c.get("block"):
            blk = dict(c["block"])
            blk["data"] = [
                {**d, **({"values": EXPORT_COLUMNS[d["key"]]["values"]}
                         if EXPORT_COLUMNS.get(d["key"], {}).get("values") else {})}
                for d in c["block"]["data"]
            ]
            item["block"] = blk
        if c.get("header"):
            item["header"] = c["header"]
        if c.get("values"):
            item["values"] = c["values"]
        if c.get("formats"):
            item["formats"] = c["formats"]
        out.append(item)
    return out


_LEGACY_COLUMN_ALIASES = {
    "nominal_xy": ["nominal_x", "nominal_y"],
    "tolerance":  ["upper_tol", "lower_tol"],
}


def _parse_columns(raw) -> List[str]:
    """แปลงค่า columns_json จาก DB (str หรือ list) เป็น list ของ key ที่ใช้ได้จริง
    — แตก key แบบเก่าเป็นคอลัมน์ย่อย และตัด key ที่ไม่รู้จักทิ้ง
    """
    cols = raw
    if isinstance(cols, str):
        try:
            cols = json.loads(cols)
        except Exception:
            cols = []
    if not isinstance(cols, list):
        cols = []

    out: List[str] = []
    for c in cols:
        for key in _LEGACY_COLUMN_ALIASES.get(c, [c]):
            if key in EXPORT_COLUMNS and key not in out:
                out.append(key)
    return out


def _get_template(cur, export_template_id: int) -> Dict[str, Any]:
    cur.execute(
        "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
        "FROM export_template WHERE export_template_id = %s",
        (export_template_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบเทมเพลตนี้")
    return row


def _parse_layout(raw):
    """แปลง layout_json จาก DB (str หรือ dict) เป็น dict — None ถ้าไม่มี/พัง"""
    if raw is None:
        return None
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return None
    return raw if isinstance(raw, dict) else None


@app.get("/api/export/templates")
async def list_export_templates(kind: str = "csv"):
    """คืนเทมเพลตของชนิดที่ระบุ — csv / pdf / excel แยกลิสต์กันคนละชนิด

    PDF กับ Excel เคยใช้ kind='report' ร่วมกัน ทำให้เทมเพลตปนกันข้ามรูปแบบ
    ตอนนี้แยกแล้ว แต่ยังดึงของเก่าที่เป็น 'report' มาแสดงด้วย จะได้ไม่หายไปเฉยๆ
    """
    kinds = [kind, "report"] if kind in ("pdf", "excel") else [kind]
    ph = ", ".join(["%s"] * len(kinds))
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                f"FROM export_template WHERE kind IN ({ph}) ORDER BY is_default DESC, name",
                kinds,
            )
            return [
                {
                    "export_template_id": r["export_template_id"],
                    "name": r["name"],
                    "kind": r["kind"],
                    "columns": _parse_columns(r["columns_json"]),
                    "layout": _parse_layout(r["layout_json"]),
                    "is_default": bool(r["is_default"]),
                }
                for r in cur.fetchall()
            ]
    finally:
        db.close()


class ExportTemplateBody(BaseModel):
    name: str
    columns: Optional[List[str]] = None
    layout:  Optional[Dict[str, Any]] = None
    kind:    str = "csv"


def _validate_template_body(body: ExportTemplateBody) -> List[str]:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")
    cols = [c for c in (body.columns or []) if c in EXPORT_COLUMNS]
    if not cols:
        raise HTTPException(400, "ต้องเลือกอย่างน้อย 1 คอลัมน์")
    return cols


def _template_payload(body: ExportTemplateBody):
    """ตรวจ body แล้วคืน (kind, columns_json, layout_json) ที่พร้อมเขียนลง DB

    เก็บข้อมูลคนละคอลัมน์กันตามชนิด:
      csv          → columns_json (รายชื่อคอลัมน์ + ลำดับ)
      pdf / excel  → layout_json  (ผังตารางทั้งก้อนจากหน้า editor แบบสเปรดชีต)

    ⚠ ทุกอย่างที่ "ไม่ใช่ csv" ถือเป็นรายงานหมด — ห้ามเช็คแบบ `kind == "report"`
    เพราะตอนแยก PDF กับ Excel ออกจากกัน ค่า kind เปลี่ยนเป็น 'pdf'/'excel' แล้ว
    การเช็คชื่อตายตัวทำให้เทมเพลตรายงานตกไปเข้าเส้นทางของ CSV แล้วเด้ง
    "ต้องเลือกอย่างน้อย 1 คอลัมน์" ทั้งที่หน้าจอนั้นไม่มีให้เลือกคอลัมน์เลย
    """
    if not body.name.strip():
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")

    if body.kind != "csv":
        if not body.layout:
            raise HTTPException(400, "เทมเพลตรายงานต้องมีผังตาราง (layout)")
        kind = body.kind if body.kind in ("pdf", "excel", "report") else "pdf"
        return kind, None, json.dumps(body.layout, ensure_ascii=False)

    cols = _validate_template_body(body)
    return "csv", json.dumps(cols, ensure_ascii=False), None


@app.post("/api/export/templates", status_code=201)
async def create_export_template(body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                    "VALUES (%s, %s, %s, %s, 0)",
                    (body.name.strip(), kind, cols_json, layout_json),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"สร้างเทมเพลตไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"export_template_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/export/templates/{export_template_id}")
async def update_export_template(export_template_id: int, body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นแก้ไขไม่ได้ — กด Duplicate แล้วแก้ตัวสำเนาแทน")
            try:
                cur.execute(
                    "UPDATE export_template SET name = %s, kind = %s, columns_json = %s, layout_json = %s "
                    "WHERE export_template_id = %s",
                    (body.name.strip(), kind, cols_json, layout_json, export_template_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"บันทึกไม่สำเร็จ (ชื่อนี้อาจซ้ำกับเทมเพลตอื่น): {exc}")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/export/templates/{export_template_id}")
async def delete_export_template(export_template_id: int):
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นลบไม่ได้")
            cur.execute(
                "DELETE FROM export_template WHERE export_template_id = %s", (export_template_id,)
            )
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/export/templates/{export_template_id}/duplicate", status_code=201)
async def duplicate_export_template(export_template_id: int):
    """ทำสำเนาเทมเพลต (ใช้ได้กับทุกตัวรวมทั้งตัวค่าเริ่มต้น) — ตั้งชื่อใหม่ให้
    อัตโนมัติแบบ "<ชื่อเดิม> (สำเนา)" และเติมเลขต่อท้ายถ้าชื่อนั้นถูกใช้ไปแล้ว
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            cur.execute("SELECT name FROM export_template")
            taken = {r["name"] for r in cur.fetchall()}
            base = f"{row['name']} (สำเนา)"
            name, n = base, 2
            while name in taken:
                name, n = f"{base} {n}", n + 1
            cur.execute(
                "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                "VALUES (%s, %s, %s, %s, 0)",
                (
                    name,
                    row.get("kind") or "csv",
                    json.dumps(_parse_columns(row["columns_json"]), ensure_ascii=False)
                        if row.get("columns_json") else None,
                    json.dumps(_parse_layout(row["layout_json"]), ensure_ascii=False)
                        if row.get("layout_json") else None,
                ),
            )
            return {"export_template_id": cur.lastrowid, "name": name}
    finally:
        db.close()


def _latest_only_sql(inner_where: str) -> str:
    """เงื่อนไข "เอาเฉพาะการวัดล่าสุดของแต่ละ ALPL"

    ⚠ จุดที่เคยพลาด: ของเดิม subquery เป็น `FROM measurements` เปล่าๆ ไม่มี WHERE
    จึงหา "ล่าสุด" จากทั้งตารางก่อน แล้วค่อยเอาผลไปกรองด้วยเงื่อนไขของผู้ใช้ทีหลัง
    ผลคือถ้า ALPL ตัวไหนถูกวัดซ้ำ "นอกช่วง" ที่ผู้ใช้กรอง แถวล่าสุดของมันจะตก
    ตัวกรอง แล้ว ALPL ตัวนั้นหายไปจากรายงานทั้งตัว ทั้งที่ในช่วงที่เลือกมีการวัดจริง
    เช่น กรอง 1–15 ก.ค. แต่ ALPL 400 ถูกวัดซ้ำวันที่ 20 ก.ค. → 400 หายไปเฉยๆ

    ตอนนี้ subquery ใช้ FROM/JOIN และ WHERE ชุดเดียวกับ query หลัก จึงหมายถึง
    "ล่าสุดภายในชุดข้อมูลที่กรองแล้ว" ตามที่ผู้ใช้คาดหวังจริงๆ
    (พารามิเตอร์ของ inner_where ต้องถูกส่งซ้ำอีกชุด — ดู _export_filters)
    """
    return f"""m.measurement_id IN (
    SELECT measurement_id FROM (
        SELECT m.measurement_id,
               ROW_NUMBER() OVER (
                   PARTITION BY m.number_alpl
                   ORDER BY m.timestamp DESC, m.measurement_id DESC
               ) AS rn
        {_EXPORT_FROM}
        {inner_where}
    ) AS latest WHERE latest.rn = 1
)"""


_RANGE_PART_RE = re.compile(r"^\s*(\d+)\s*(?:-\s*(\d+)\s*)?$")


def _parse_int_ranges(spec: str, field_label: str):
    """แปลงข้อความแบบ "400-407,500,600" เป็น (ช่วง, ตัวเดี่ยว)

    คืนค่าเป็น ([(400, 407)], [500, 600]) — แยกช่วงกับตัวเดี่ยวออกจากกัน
    **ตั้งใจไม่แตกช่วงออกเป็นรายตัว** เพราะถ้าผู้ใช้พิมพ์ "1-999999" การแตกจะได้
    ค่าเกือบล้านตัวไปต่อเป็น SQL ยาว ~6.5 MB ซึ่งชน max_allowed_packet ของ MySQL
    หรือช้าจนใช้งานไม่ได้ — เก็บเป็นช่วงแล้วแปลงเป็น BETWEEN ตอนสร้าง SQL แทน
    ต้นทุนเท่ากับไม่กรองเลย (จำนวนแถวขึ้นกับข้อมูลจริง ไม่ใช่ความกว้างของช่วง)

    เคสที่รองรับ: เว้นวรรค, พิมพ์กลับด้าน (407-400), ค่าซ้ำ, ท่อนว่างจากคอมมาเกิน
    """
    ranges, singles = [], []
    for part in str(spec).split(","):
        if not part.strip():
            continue
        m = _RANGE_PART_RE.match(part)
        if not m:
            raise HTTPException(
                400,
                f'รูปแบบ {field_label} ไม่ถูกต้องตรง "{part.strip()}" — '
                f'ใช้ได้เช่น 400 หรือ 400-407 หรือ 400-407,500,600',
            )
        lo = int(m.group(1))
        if m.group(2) is None:
            singles.append(lo)
            continue
        hi = int(m.group(2))
        if lo > hi:
            lo, hi = hi, lo
        if lo == hi:
            singles.append(lo)
        else:
            ranges.append((lo, hi))
    return sorted(set(ranges)), sorted(set(singles))


_DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _day_end(v):
    """ครอบปลายช่วงให้เต็มวัน — "2026-07-30" → "2026-07-30 23:59:59"

    ถ้าไม่ทำ MySQL จะตีค่าเป็น 00:00:00 แล้ว `timestamp <= ...` จะตัดการวัด
    ของวันสุดท้ายที่ผู้ใช้เลือกทิ้งทั้งวัน หน้าเว็บเติมเวลาให้อยู่แล้ว แต่ย้าย
    มาบังคับที่ backend ด้วย เพราะเป็นกติกาที่หน้าใหม่ (หรือคนที่ยิง API ตรงๆ)
    ไม่มีทางรู้ แล้วจะเจอบั๊กเดิมซ้ำ
    """
    return f"{v} 23:59:59" if isinstance(v, str) and _DATE_ONLY_RE.match(v.strip()) else v


def _day_start(v):
    return f"{v} 00:00:00" if isinstance(v, str) and _DATE_ONLY_RE.match(v.strip()) else v


def _export_filters(f: Dict[str, Any]):
    """สร้าง WHERE + params ของหน้า Export — ใช้ร่วมกันทั้ง preview และ csv
    เพื่อให้สิ่งที่เห็นในตัวอย่างตรงกับไฟล์ที่ดาวน์โหลดจริงเสมอ
    """
    conditions, params = [], []

    def eq(col, key):
        """เงื่อนไขเท่ากับ — รองรับทั้งค่าเดียวและหลายค่า (multi-select)

        หลายค่าจะกลายเป็น IN (...) ซึ่งหมายถึง "เอาอันไหนก็ได้ในกลุ่มนี้" (OR กัน
        ภายในช่องเดียวกัน) ส่วนต่างช่องกันยังเป็น AND เหมือนเดิม เช่น
        เลือก Vendor = A,B และ Result = NG → (vendor A หรือ B) และ ต้องเป็น NG
        """
        v = f.get(key)
        if v is None or v == "":
            return
        if isinstance(v, (list, tuple, set)):
            vals = [x for x in v if x not in (None, "")]
            if not vals:
                return
            conditions.append(f"{col} IN ({','.join(['%s'] * len(vals))})")
            params.extend(vals)
        else:
            conditions.append(f"{col} = %s"); params.append(v)

    alpl_spec = f.get("number_alpl")
    if alpl_spec not in (None, ""):
        alpl_ranges, alpl_singles = _parse_int_ranges(alpl_spec, "ALPL")
        or_parts = []
        for lo, hi in alpl_ranges:
            or_parts.append("m.number_alpl BETWEEN %s AND %s")
            params.extend([lo, hi])
        if alpl_singles:
            or_parts.append(f"m.number_alpl IN ({','.join(['%s'] * len(alpl_singles))})")
            params.extend(alpl_singles)
        if or_parts:
            conditions.append("(" + " OR ".join(or_parts) + ")")

    eq("m.result",          "result")
    eq("m.session_id",      "session_id")
    eq("op.operator_name",  "operator")
    eq("m.measure_type",    "measure_type")
    eq("v.vendor_name",     "vendor")
    eq("o.owner_name",      "owner")
    eq("pn.part_number_name", "part_number")
    eq("h.handler_name",    "handler")
    eq("ps.package_size",   "package_size")
    eq("p.po_number",       "po_number")

    if f.get("date_from"):
        conditions.append("m.timestamp >= %s"); params.append(_day_start(f["date_from"]))
    if f.get("date_to"):
        conditions.append("m.timestamp <= %s"); params.append(_day_end(f["date_to"]))
    if f.get("recv_from"):
        conditions.append("p.recieve_date >= %s"); params.append(_day_start(f["recv_from"]))
    if f.get("recv_to"):
        conditions.append("p.recieve_date <= %s"); params.append(_day_end(f["recv_to"]))

    if f.get("description"):
        conditions.append("p.description LIKE %s"); params.append(f"%{f['description']}%")

    if f.get("latest_only"):
        inner_where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        inner_params = list(params)
        conditions.append(_latest_only_sql(inner_where))
        params.extend(inner_params)

    return (("WHERE " + " AND ".join(conditions)) if conditions else ""), params


_EXPORT_FROM = """
    FROM measurements m
    LEFT JOIN operator op             ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p  ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn          ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h               ON pn.handler_id = h.handler_id
    LEFT JOIN package_size ps         ON pn.package_size_id = ps.package_size_id
    LEFT JOIN template t              ON ps.template_id = t.template_id
    LEFT JOIN vendor v                ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o                 ON p.owner_id = o.owner_id
"""


def _fetch_export_rows(cols: List[str], where: str, params: list, limit: Optional[int] = None):
    """ดึงข้อมูลแล้วแปลงเป็น list ของ list ตามลำดับคอลัมน์ใน `cols`"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp DESC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            rows = cur.fetchall()
    finally:
        db.close()
    data = [[EXPORT_COLUMNS[c]["get"](r) for c in cols] for r in rows]
    return data, total


def _count_export_rows(where: str, params: list) -> int:
    """นับจำนวนแถวอย่างเดียว — ใช้เช็คเพดานก่อนดึงข้อมูลจริงมาสร้างไฟล์"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            return cur.fetchone()["n"]
    finally:
        db.close()


def _fetch_export_raw(where: str, params: list, limit: Optional[int] = None):
    """เหมือน _fetch_export_rows แต่คืน row ดิบ (dict) ไม่ได้แปลงเป็นคอลัมน์
    ใช้กับรายงาน PDF/Excel ที่ต้องหยิบค่าทีละช่องตามผังตาราง ไม่ใช่เรียงเป็นแถว
    เรียง ASC ตามเวลา เพราะรายงานบนกระดาษอ่านจากเก่าไปใหม่ (ต่างจาก CSV ที่
    เรียง DESC ให้เห็นตัวล่าสุดก่อน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp ASC, m.measurement_id ASC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            return cur.fetchall(), total
    finally:
        db.close()


REPORT_GROUP_BY = "tolerance_spec"


def _cell_out(cell: Dict[str, Any], text: str) -> Dict[str, Any]:
    """แปลงเซลล์ในผังเป็นเซลล์ผลลัพธ์ — เก็บเฉพาะที่ตัววาดต้องใช้"""
    out = {"v": text, "s": cell.get("s") or {}}
    span = cell.get("span")
    if span:
        out["span"] = {"r": int(span.get("r", 1)), "c": int(span.get("c", 1))}
    if cell.get("hidden"):
        out["hidden"] = True
    if cell.get("f") or cell.get("spec"):
        out["data"] = True
    if cell.get("hdr"):
        out["head"] = True
    return out


def _render_report(layout: Dict[str, Any], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """คลี่ผัง + ข้อมูล ออกเป็นตารางพร้อมวาด

    คืน {"nCols": int, "rows": [[cell, ...], ...]} โดย cell = {v, s, span?, hidden?}
    """
    grid = layout.get("grid") or []
    n_rows = int(layout.get("nRows") or len(grid))
    n_cols = int(layout.get("nCols") or (len(grid[0]) if grid else 0))
    data_row = int(layout.get("dataRow", 2))
    if not grid or n_cols == 0:
        return {"nCols": 0, "rows": []}
    data_row = max(0, min(data_row, n_rows - 1))

    def tpl_row(r: int) -> List[Dict[str, Any]]:
        row = grid[r] if r < len(grid) else []
        return [row[c] if c < len(row) else {} for c in range(n_cols)]

    def clamp(cells: List[Dict[str, Any]], r: int, last: int) -> List[Dict[str, Any]]:
        """ตัด rowspan ไม่ให้ทะลุออกนอกบล็อกที่กำลังพิมพ์อยู่

        ผู้ใช้ผสานเซลล์คร่อมแถวข้อมูลได้ในหน้าแก้ไข แต่ตอนคลี่ออกมา ส่วนหัวของ
        กลุ่มถัดไปจะมาต่อท้าย ถ้าปล่อย rowspan เดิมไว้มันจะกินทับแถวของกลุ่มถัดไป
        แล้วช่วงผสานซ้อนกัน → Excel เปิดไฟล์ไม่ขึ้น บอกว่าไฟล์เสียหาย
        """
        for cell in cells:
            sp = cell.get("span")
            if sp and sp["r"] > 1:
                sp["r"] = max(1, min(sp["r"], last - r + 1))
        return cells

    def render_static(r: int, spec_text: str) -> List[Dict[str, Any]]:
        """แถวส่วนหัว/ส่วนท้าย — แทนเฉพาะเซลล์ spec ด้วยค่าของกลุ่มปัจจุบัน
        เซลล์ f (ข้อมูลรายชิ้น) ที่หลุดมาอยู่นอกแถวข้อมูลถือว่าไม่มีความหมาย
        ปล่อยว่างไว้ ดีกว่าเอาค่าของชิ้นแรกมาใส่แบบมั่วๆ
        """
        out = []
        for cell in tpl_row(r):
            if cell.get("spec"):
                out.append(_cell_out(cell, spec_text))
            elif cell.get("f"):
                out.append(_cell_out(cell, ""))
            else:
                out.append(_cell_out(cell, cell.get("v") or ""))
        return out

    def render_data(row: Dict[str, Any], item_no: int) -> List[Dict[str, Any]]:
        """แถวข้อมูล 1 แถวต่อ 1 การวัด — rowspan ถูกบีบเป็น 1 เพราะแถวนี้ถูก
        ทำซ้ำหลายรอบ ถ้าปล่อย rowspan เดิมไว้จะกินทับแถวของการวัดชิ้นถัดไป
        """
        out = []
        for cell in tpl_row(data_row):
            key = cell.get("f")
            col = EXPORT_COLUMNS.get(key) if key else None

            if col and col.get("row_number"):
                text = str(item_no)
            elif col and col.get("get_fmt") and cell.get("fmt"):
                text = col["get_fmt"](row, cell["fmt"])
            elif col:
                text = col["get"](row)
            else:
                text = cell.get("v") or ""

            c = _cell_out(cell, str(text))
            if "span" in c:
                c["span"]["r"] = 1

            variants = cell.get("variants") or {}
            if variants:
                state = col["state"](row) if col and col.get("state") else str(text)
                if state in variants:
                    c["s"] = variants[state]
            out.append(c)
        return out

    has_spec = any(
        (cell or {}).get("spec") or (cell or {}).get("hdr") == REPORT_GROUP_BY
        for row in grid for cell in row
    )
    groups: Dict[str, List[Dict[str, Any]]] = {}
    if has_spec:
        for r in rows:
            groups.setdefault(EXPORT_COLUMNS[REPORT_GROUP_BY]["get"](r), []).append(r)
    elif rows:
        groups[""] = list(rows)

    def _repeats_per_group(r: int) -> bool:
        return any(
            (c.get("hdr") or c.get("spec") or c.get("f"))
            for c in tpl_row(r) if not c.get("hidden")
        )

    repeat_head = {r: _repeats_per_group(r) for r in range(0, data_row)}

    out_rows: List[List[Dict[str, Any]]] = []
    item_no = 0
    for gi, (spec_text, members) in enumerate(groups.items()):
        for r in range(0, data_row):
            if gi and not repeat_head[r]:
                continue
            out_rows.append(clamp(render_static(r, spec_text), r, data_row - 1))
        for m in members:
            item_no += 1
            out_rows.append(render_data(m, item_no))
    for r in range(data_row + 1, n_rows):
        out_rows.append(clamp(render_static(r, ""), r, n_rows - 1))

    def filled(cell: Dict[str, Any]) -> bool:
        return bool((cell.get("v") or "").strip() or cell.get("data") or cell.get("head"))

    last_col = 0
    for row in out_rows:
        for c, cell in enumerate(row):
            if not cell.get("hidden") and filled(cell):
                last_col = max(last_col, c + (cell.get("span", {}) or {}).get("c", 1))
    last_col = last_col or n_cols

    for row in out_rows:
        del row[last_col:]
        for c, cell in enumerate(row):
            sp = cell.get("span")
            if sp and c + sp["c"] > last_col:
                sp["c"] = max(1, last_col - c)

    while out_rows and not any(filled(x) for x in out_rows[-1]):
        out_rows.pop()

    return {"nCols": last_col, "rows": out_rows, "groups": len(groups)}


def _load_report_layout(cur, export_template_id: int) -> Dict[str, Any]:
    tpl = _get_template(cur, export_template_id)
    layout = _parse_layout(tpl.get("layout_json"))
    if not layout or not layout.get("grid"):
        raise HTTPException(400, f'เทมเพลต "{tpl["name"]}" ยังไม่มีผังตาราง — เปิดแก้ไขแล้วบันทึกใหม่อีกครั้ง')
    return {"name": tpl["name"], "kind": tpl.get("kind") or "pdf", "layout": layout}


def export_filters_dep(
    number_alpl:  Optional[str] = None,
    date_from:    Optional[str] = None,
    date_to:      Optional[str] = None,
    recv_from:    Optional[str] = None,
    recv_to:      Optional[str] = None,
    session_id:   Optional[int] = None,
    po_number:    Optional[int] = None,
    description:  Optional[str] = None,
    result:       Optional[List[str]] = Query(None),
    operator:     Optional[List[str]] = Query(None),
    measure_type: Optional[List[str]] = Query(None),
    vendor:       Optional[List[str]] = Query(None),
    owner:        Optional[List[str]] = Query(None),
    part_number:  Optional[List[str]] = Query(None),
    handler:      Optional[List[str]] = Query(None),
    package_size: Optional[List[str]] = Query(None),
    latest_only:  bool = True,
) -> Dict[str, Any]:
    """พารามิเตอร์ filter ที่ /api/export/preview กับ /api/export/csv รับเหมือนกัน
    ประกาศไว้ที่เดียวแล้วใช้ร่วมกันผ่าน Depends — กันลืมเพิ่มข้างใดข้างหนึ่ง
    จนตัวอย่างบนหน้าจอกับไฟล์ที่ดาวน์โหลดจริงไม่ตรงกัน
    """
    return {
        "number_alpl": number_alpl, "date_from": date_from, "date_to": date_to,
        "recv_from": recv_from, "recv_to": recv_to, "session_id": session_id,
        "po_number": po_number, "description": description,
        "result": result, "operator": operator, "measure_type": measure_type,
        "vendor": vendor, "owner": owner, "part_number": part_number,
        "handler": handler, "package_size": package_size,
        "latest_only": latest_only,
    }


@app.get("/api/export/preview")
async def export_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    limit:   int = Query(5, ge=1, le=50),
):
    """คืนหัวคอลัมน์ + ข้อมูลตัวอย่างไม่กี่แถว + จำนวนแถวทั้งหมดที่ตรงกับ filter
    ให้หน้าเว็บโชว์ก่อนกดดาวน์โหลดจริง (จะได้รู้ว่ากรองถูกไหม ไฟล์ใหญ่แค่ไหน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            tpl = _get_template(cur, export_template_id)
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, total = _fetch_export_rows(cols, where, params, limit=limit)
    return {
        "template_name": tpl["name"],
        "columns": [_csv_header(c) for c in cols],
        "column_keys": cols,
        "rows": data,
        "total": total,
    }


@app.get("/api/export/csv")
async def export_csv(
    export_template_id: Optional[int] = None,
    filename: Optional[str] = None,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """Export ประวัติ measurement (พร้อม filter) เป็นไฟล์ CSV ให้ดาวน์โหลด

    export_template_id: เลือกว่าจะเอาคอลัมน์ไหนและเรียงลำดับยังไง — ถ้าไม่ส่งมา
    จะใช้เทมเพลตที่ is_default = 1 ให้อัตโนมัติ (พฤติกรรมเดิมของ endpoint นี้
    คือ dump ทุกคอลัมน์ ซึ่งตรงกับเทมเพลตค่าเริ่มต้นพอดี ของเดิมที่เรียกมาโดย
    ไม่ส่งพารามิเตอร์นี้จึงยังใช้งานได้เหมือนเดิม)

    ใช้ utf-8-sig encoding เพื่อให้เปิดใน Excel ได้ถูกต้องแม้มีตัวอักษรไทยอยู่ใน
    ไฟล์ (BOM ช่วยไม่ให้ตัวอักษรเพี้ยน) และใช้ filter ชุดเดียวกับ /api/export/preview
    เพื่อให้ไฟล์ที่ได้ตรงกับตัวอย่างที่เห็นบนหน้าจอเสมอ
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            if export_template_id is not None:
                tpl = _get_template(cur, export_template_id)
            else:
                cur.execute(
                    "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                    "FROM export_template WHERE kind = 'csv' AND is_default = 1 LIMIT 1"
                )
                tpl = cur.fetchone()
                if not tpl:
                    raise HTTPException(404, "ยังไม่มีเทมเพลตค่าเริ่มต้นในระบบ")
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, _ = _fetch_export_rows(cols, where, params)

    df  = pd.DataFrame(data, columns=[_csv_header(c) for c in cols])
    buf = StringIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)

    safe_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "", (filename or "")).strip(". ")
    fname = (f"{safe_name}.csv" if safe_name
             else f"measurements_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
    return StreamingResponse(
        iter(["﻿" + buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )



REPORT_PREVIEW_LIMIT = 300

REPORT_MAX_ROWS = int(os.getenv("REPORT_MAX_ROWS", 20000))


def _guard_report_size(total: int) -> None:
    if total > REPORT_MAX_ROWS:
        raise HTTPException(
            413,
            f"ข้อมูลที่เลือกมี {total:,} แถว เกินเพดาน {REPORT_MAX_ROWS:,} แถวของรายงาน "
            f"— กรองให้แคบลงก่อน (เช่น จำกัดช่วงวันที่) หรือใช้ Export CSV แทน "
            f"ถ้าต้องการข้อมูลดิบทั้งหมด",
        )


@app.get("/api/export/report-preview")
async def export_report_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    full: int = 0,
):
    """คืนผังที่คลี่แล้วพร้อมข้อมูลจริง ให้หน้าเว็บวาดเป็นตาราง preview

    full=0 → ตัดที่ REPORT_PREVIEW_LIMIT แถว (ดูบนจอเฉยๆ ไม่ต้องครบ)
    full=1 → เอาครบทุกแถว ใช้ตอนสั่งพิมพ์เป็น PDF จริง — ไฟล์ที่ได้ต้องไม่ขาด
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    if full:
        _guard_report_size(total_check := _count_export_rows(where, params))
    rows, total = _fetch_export_raw(where, params, limit=None if full else REPORT_PREVIEW_LIMIT)
    out = _render_report(tpl["layout"], rows)
    out.update({
        "template_name": tpl["name"],
        "total": total,
        "shown": len(rows),
        "truncated": total > len(rows),
    })
    return out


@app.get("/api/export/xlsx")
async def export_xlsx(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """สร้างไฟล์ .xlsx จากผังรายงาน — คงฟอนต์/สี/การผสานเซลล์ตามที่จัดไว้"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            500,
            "ยังไม่ได้ติดตั้ง openpyxl — รัน `pip install openpyxl` ในเครื่องที่รัน Backend ก่อน",
        )

    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "Export")
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    rows, total = _fetch_export_raw(where, params)
    _guard_report_size(total)
    rendered = _render_report(tpl["layout"], rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    thick = Side(style="medium", color="000000")
    box = Border(left=thick, right=thick, top=thick, bottom=thick)

    def rgb(v: Optional[str]) -> Optional[str]:
        """#RRGGBB → RRGGBB (openpyxl ไม่รับเครื่องหมาย #)"""
        if not v or not isinstance(v, str):
            return None
        v = v.strip().lstrip("#")
        return v.upper() if re.fullmatch(r"[0-9A-Fa-f]{6}", v) else None

    _H = {"left": "left", "center": "center", "right": "right", "justify": "justify"}
    _V = {"top": "top", "middle": "center", "center": "center", "bottom": "bottom"}

    def align_of(s: Dict[str, Any]) -> "Alignment":
        try:
            indent = max(0, min(250, int(s.get("indent") or 0)))
        except (TypeError, ValueError):
            indent = 0
        return Alignment(
            horizontal=_H.get(str(s.get("align") or "").lower(), "center"),
            vertical=_V.get(str(s.get("valign") or "").lower(), "center"),
            indent=indent,
            wrap_text=False,
        )

    def size_of(s: Dict[str, Any]) -> float:
        try:
            v = float(s.get("size") or 11)
        except (TypeError, ValueError):
            return 11.0
        return v if 1 <= v <= 409 else 11.0

    for r, row in enumerate(rendered["rows"], start=1):
        for c, cell in enumerate(row, start=1):
            if cell.get("hidden"):
                continue
            s = cell.get("s") or {}
            x = ws.cell(row=r, column=c, value=cell.get("v") or "")
            x.font = Font(
                name=str(s.get("font") or "Tahoma"),
                size=size_of(s),
                bold=bool(s.get("bold")),
                italic=bool(s.get("italic")),
                underline="single" if s.get("underline") else None,
                color=rgb(s.get("color")) or "000000",
            )
            fill = rgb(s.get("fill"))
            if fill:
                x.fill = PatternFill("solid", fgColor=fill)
            x.alignment = align_of(s)
            if cell.get("v") or cell.get("data") or cell.get("head"):
                x.border = box

            span = cell.get("span")
            if span and (span["r"] > 1 or span["c"] > 1):
                r2 = r + max(1, span["r"]) - 1
                c2 = min(c + max(1, span["c"]) - 1, rendered["nCols"])
                if not any(
                    m.min_row <= r2 and r <= m.max_row and m.min_col <= c2 and c <= m.max_col
                    for m in ws.merged_cells.ranges
                ):
                    ws.merge_cells(start_row=r, start_column=c, end_row=r2, end_column=c2)

    for c in range(1, (rendered["nCols"] or 1) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        log.exception("สร้างไฟล์ xlsx ไม่สำเร็จ")
        raise HTTPException(500, f"สร้างไฟล์ Excel ไม่สำเร็จ: {type(exc).__name__}: {exc}")
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", tpl["name"]).strip("_") or "report"
    fname = f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )



_DELETED_KIND_LABEL = {
    "measurement":  "ผลการวัด",
    "part":         "Part (ALPL)",
    "operator":     "Operator",
    "owner":        "Owner",
    "vendor":       "Vendor",
    "handler":      "Handler",
    "template":     "Template",
    "package_size": "Package Size",
    "part_number":  "Part Number",
}


def _deleted_files():
    """ไล่หาไฟล์ .json ทุกอันในถังขยะ — คืน (โฟลเดอร์วัน, ชื่อไฟล์, path เต็ม)"""
    if not os.path.isdir(DELETED_DIR):
        return []
    out = []
    for day in sorted(os.listdir(DELETED_DIR), reverse=True):
        day_dir = os.path.join(DELETED_DIR, day)
        if not os.path.isdir(day_dir):
            continue
        for name in sorted(os.listdir(day_dir), reverse=True):
            if name.endswith(".json"):
                out.append((day, name, os.path.join(day_dir, name)))
    return out


def _deleted_summary(payload: Dict[str, Any]) -> str:
    """ข้อความสรุปสั้นๆ ให้คนอ่านออกว่าแถวที่ลบไปคืออะไร โดยไม่ต้องเปิดไฟล์ดู"""
    kind, row = payload.get("kind"), payload.get("row", {})
    if kind == "measurement":
        vx, vy = row.get("value_x"), row.get("value_y")
        return (f"ALPL {row.get('number_alpl')} · {vx} / {vy} · {row.get('result')}"
                f" · {str(row.get('timestamp') or '')[:19]}")
    if kind == "part":
        n = len(payload.get("related", {}).get("sessions", []))
        return f"ALPL {row.get('number_alpl')}" + (f" (+ {n} sessions)" if n else "")
    for k, v in row.items():
        if k.endswith("_name") or k == "package_size":
            return str(v)
    return ", ".join(f"{k}={v}" for k, v in list(row.items())[:3])


def _purge_old_deleted() -> int:
    """ลบโฟลเดอร์ในถังขยะที่เก่ากว่า DELETED_RETENTION_DAYS วัน — คืนจำนวนที่ลบ

    อ่านวันจาก "ชื่อโฟลเดอร์" (DD-MM-YYYY พ.ศ.) ไม่ใช่ mtime ของไฟล์ เพราะ mtime
    เปลี่ยนได้ง่ายเวลาก๊อป/ย้ายไฟล์ ส่วนชื่อโฟลเดอร์คือวันที่ลบจริงเสมอ
    """
    if not os.path.isdir(DELETED_DIR):
        return 0
    cutoff = datetime.now().date() - timedelta(days=DELETED_RETENTION_DAYS)
    removed = 0
    for day in os.listdir(DELETED_DIR):
        day_dir = os.path.join(DELETED_DIR, day)
        if not os.path.isdir(day_dir):
            continue
        try:
            d, m, y = day.split("-")
            day_date = date(int(y) - 543, int(m), int(d))
        except (ValueError, TypeError):
            log.warning("ข้ามโฟลเดอร์ในถังขยะที่ชื่อไม่ใช่รูปแบบวันที่: %s", day)
            continue
        if day_date < cutoff:
            try:
                shutil.rmtree(day_dir)
                removed += 1
                log.info("ลบถังขยะของวันที่ %s (เกิน %s วัน)", day, DELETED_RETENTION_DAYS)
            except OSError as exc:
                log.warning("ลบโฟลเดอร์ถังขยะ %s ไม่สำเร็จ: %s", day, exc)
    return removed


@app.get("/api/deleted")
async def list_deleted():
    """รายการทุกอย่างที่อยู่ในถังขยะ เรียงจากลบล่าสุดก่อน"""
    items = []
    for day, name, path in _deleted_files():
        try:
            with open(path, encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as exc:
            log.warning("อ่านไฟล์ถังขยะ %s ไม่สำเร็จ: %s", name, exc)
            continue
        kind = payload.get("kind", "?")
        items.append({
            "id":         f"{day}/{name}",
            "deleted_at": payload.get("deleted_at"),
            "kind":       kind,
            "kind_label": _DELETED_KIND_LABEL.get(kind, kind),
            "summary":    _deleted_summary(payload),
            "has_image":  bool(payload.get("image_file")),
        })
    return {"items": items, "total": len(items), "retention_days": DELETED_RETENTION_DAYS}


def _deleted_path(item_id: str) -> str:
    """แปลง id ("<วัน>/<ไฟล์>.json") เป็น path จริง พร้อมกันหลุดออกนอกถังขยะ"""
    base = os.path.realpath(DELETED_DIR)
    target = os.path.realpath(os.path.join(base, item_id))
    if not target.startswith(base + os.sep) or not target.endswith(".json"):
        raise HTTPException(404, "ไม่พบรายการนี้ในถังขยะ")
    if not os.path.isfile(target):
        raise HTTPException(404, "ไม่พบรายการนี้ในถังขยะ")
    return target


@app.post("/api/deleted/restore")
async def restore_deleted(body: Dict[str, str] = Body(...)):
    """กู้คืน 1 รายการจากถังขยะ — insert แถวกลับ + ย้ายไฟล์รูปกลับที่เดิม

    ทำเป็น transaction เดียว ถ้าขั้นไหนพังจะ rollback ทั้งหมด แล้วไฟล์ JSON
    ยังอยู่ในถังขยะเหมือนเดิม (ลองใหม่ได้)

    ⚠ ลำดับการกู้สำคัญ: ถ้าแถวนี้อ้าง FK ไปหาของที่ถูกลบไปแล้วเหมือนกัน
      (เช่นกู้ measurement ที่อ้าง operator ที่โดนลบไป) MySQL จะปฏิเสธ
      ต้องกู้ตัวที่ถูกอ้างก่อน — ข้อความ error จะบอกให้ผู้ใช้รู้
    """
    item_id = (body or {}).get("id", "")
    path = _deleted_path(item_id)
    with open(path, encoding="utf-8") as f:
        payload = json.load(f)

    table   = payload["table"]
    row     = payload["row"]
    related = payload.get("related") or {}

    db = get_db()
    try:
        db.autocommit(False)
        with db.cursor() as cur:
            _block_if_session_running(cur, "กู้คืน")

            def insert(tbl: str, data: Dict[str, Any]):
                cols = ", ".join(f"`{c}`" for c in data)
                marks = ", ".join(["%s"] * len(data))
                cur.execute(f"INSERT INTO `{tbl}` ({cols}) VALUES ({marks})", list(data.values()))

            try:
                insert(table, row)
                for tbl, rows in related.items():
                    for r in rows:
                        insert(tbl, r)
            except pymysql.err.IntegrityError as exc:
                db.rollback()
                code = exc.args[0] if exc.args else 0
                if code == 1062:
                    raise HTTPException(409, "กู้คืนไม่ได้ — มีข้อมูล id นี้อยู่ในระบบแล้ว")
                if code == 1452:
                    raise HTTPException(
                        409,
                        "กู้คืนไม่ได้ — ข้อมูลนี้อ้างอิงถึงรายการอื่นที่ถูกลบไปแล้ว "
                        "กรุณากู้คืนรายการที่ถูกอ้างถึงก่อน (เช่น Operator / Part Number)",
                    )
                raise HTTPException(409, f"กู้คืนไม่สำเร็จ: {exc}")

        image_file  = payload.get("image_file")
        image_path  = row.get("image_path")
        if image_file and image_path:
            src = os.path.join(os.path.dirname(path), image_file)
            dst = os.path.join(ALPL_IMAGE_DIR, image_path)
            try:
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.move(src, dst)
            except OSError as exc:
                log.warning("กู้ข้อมูลสำเร็จแต่ย้ายรูปกลับไม่ได้ (%s): %s", image_file, exc)

        db.commit()
        os.remove(path)
        log.info("กู้คืนจากถังขยะ: %s", item_id)
        return {"ok": True, "kind": payload.get("kind"), "table": table}
    finally:
        db.autocommit(True)
        db.close()


@app.post("/api/deleted/remove")
async def remove_deleted(body: Dict[str, str] = Body(...)):
    """ลบ 1 รายการออกจากถังขยะ **ถาวร** — กู้คืนไม่ได้อีก

    ลบทั้งไฟล์ .json และไฟล์รูปที่เก็บคู่กัน ไม่แตะฐานข้อมูลเลย (แถวนั้นถูกลบ
    ไปตั้งแต่ตอนกดลบครั้งแรกแล้ว ตรงนี้แค่ทิ้งตัวสำรอง)
    """
    path = _deleted_path((body or {}).get("id", ""))
    try:
        with open(path, encoding="utf-8") as f:
            image_file = json.load(f).get("image_file")
    except Exception:
        image_file = None

    if image_file:
        img = os.path.join(os.path.dirname(path), os.path.basename(image_file))
        try:
            os.remove(img)
        except OSError:
            pass

    os.remove(path)
    log.info("ลบถาวรจากถังขยะ: %s", (body or {}).get("id"))
    return {"ok": True}


@app.delete("/api/deleted")
async def purge_deleted_now():
    """สั่งลบของในถังขยะที่เกินอายุทันที (ปกติทำอัตโนมัติตอน backend เริ่ม)"""
    return {"ok": True, "removed_days": _purge_old_deleted()}


os.makedirs(ALPL_IMAGE_DIR, exist_ok=True)
app.mount("/media/alpl", StaticFiles(directory=ALPL_IMAGE_DIR), name="alpl-images")


_frontend_dir = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Frontend")
)
app.mount(
    "/",
    StaticFiles(directory=_frontend_dir, html=True),
    name="static",
)
