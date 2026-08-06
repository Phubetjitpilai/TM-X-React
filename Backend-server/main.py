# Backend-server/main.py
# How to run:
#   cd Backend-server
#   pip install -r requirements.txt
#   uvicorn main:app --reload --host 0.0.0.0 --port 8000

import asyncio
import json
import logging
import os
import re
import secrets
import shutil
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

import httpx
import pandas as pd
import pymysql
import pymysql.cursors
from pymysql.constants import CLIENT
from dotenv import load_dotenv
from fastapi import Body, Depends, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from PIL import Image
from pydantic import BaseModel
from sse_starlette.sse import EventSourceResponse

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

# ── Config ──────────────────────────────────────────────────────────────────
DB_CONFIG = dict(
    host=os.getenv("DB_HOST", "localhost"),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "tmx_db"),
    # default 3306 = port ปกติของ MySQL ที่ติดตั้งบนเครื่องโดยตรง (เดิม 3307
    # คือ port ที่ map ออกมาจาก Docker container ซึ่งเลิกใช้แล้ว)
    port=int(os.getenv("DB_PORT", 3306)),
    cursorclass=pymysql.cursors.DictCursor,
    autocommit=True,
    # CLIENT.FOUND_ROWS: ค่า default ของ MySQL/pymysql คือ cur.rowcount หลัง UPDATE
    # จะนับเฉพาะ "แถวที่ค่าจริงเปลี่ยน" ไม่ใช่ "แถวที่ WHERE เจอ" — ทำให้กด Save โดย
    # ไม่แก้อะไรเลย (ส่ง payload ค่าเดิมกลับมา) แล้ว rowcount == 0 ทั้งที่แถวมีอยู่จริง
    # โค้ดที่เช็ค `if cur.rowcount == 0: raise 404 not found` (update_part,
    # update_measurement) เลยฟ้อง "not found" หลอกๆ ตั้ง flag นี้เพื่อให้ rowcount
    # นับจากแถวที่ WHERE จับคู่เจอแทน ทำให้เช็ค 404 เดิมถูกต้องอีกครั้ง
    client_flag=CLIENT.FOUND_ROWS,
)

# หมายเหตุ (architecture ใหม่): เลิกใช้ MinIO แล้ว — รูปภาพเก็บเป็นไฟล์จริงใน
# โฟลเดอร์บนเครื่อง PC ที่รัน backend นี้เอง ดีไซน์สรุปแล้ว (ดู
# POST /api/measurements/{id}/image-upload ด้านล่าง):
#   Agent (Pi) รับภาพจาก TM-X ผ่าน FTP ของตัวเองเก็บไว้ที่ Store_image_temporary
#   ก่อน แล้วอัปโหลดไฟล์จริง (multipart) มาที่ endpoint นี้ผ่าน HTTP — backend
#   เป็นคนตัดสินใจ path ปลายทางเอง: ALPL_IMAGE_DIR/<วันที่ DD-MM-YYYY (พ.ศ.)>/
#   <number_alpl>_<วันที่ DD-MM-YYYY (พ.ศ.)>.jpg (เปลี่ยนจากเดิมที่แยกโฟลเดอร์ตาม
#   package_size มาเป็นแยกตามวันที่วัดแทน ตามที่ตกลงกันไว้ — ดู
#   upload_measurement_image ด้านล่าง) ไฟล์ต้นทางจาก TM-X (มักเป็น .bmp) ถูก
#   แปลงเป็น .jpg เสมอก่อนเซฟ (ไม่ใช่ Agent ส่ง path ตรงๆ มาแบบเดิมสมัย MinIO
#   เพราะ Agent อยู่คนละเครื่องกับ backend แล้ว path ฝั่ง Agent ไม่มีความหมายกับ
#   backend เลย)
_PROJECT_ROOT = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
)

ALPL_IMAGE_DIR = os.getenv("ALPL_IMAGE_DIR", os.path.join(_PROJECT_ROOT, "image_ALPL"))
# ⚠ ค่าใน .env เป็น path สัมพัทธ์ ("./image_ALPL") ซึ่งถ้าปล่อยไว้จะอิงกับ
#   "โฟลเดอร์ที่สั่งรัน uvicorn" ไม่ใช่รากโปรเจกต์ — สั่ง `cd Backend-server`
#   ก่อนรันตามคู่มือ รูปจะไปลงที่ Backend-server/image_ALPL/ แทน แล้วหน้าเว็บ
#   หารูปไม่เจอเพราะ static mount ชี้คนละที่ (เคยเกิดกับ TEMP_IMAGE_DIR มาแล้ว)
if not os.path.isabs(ALPL_IMAGE_DIR):
    ALPL_IMAGE_DIR = os.path.abspath(os.path.join(_PROJECT_ROOT, ALPL_IMAGE_DIR))


# ระยะเผื่อสำหรับเทียบทศนิยม — ค่าที่ตกขอบ tolerance พอดี (เช่น nominal 3.02
# lower_tol 0.01 แล้ววัดได้ 3.010 เป๊ะ) ต้องถือว่า "ผ่าน" แต่ถ้าเทียบตรงๆ จะได้ NG
# เพราะ 3.02 - 0.01 ในเลขทศนิยมฐานสองได้ 3.0100000000000002 (มากกว่า 3.010)
# ซ้ำร้ายคอลัมน์พวกนี้เป็น FLOAT (single precision) ค่าที่อ่านกลับจาก DB จึงคลาด
# จากค่าที่ตั้งไว้อีกชั้น (3.02 → 3.0199999809265137) — 1e-6 เล็กกว่าความละเอียด
# ของการวัดจริง (ทศนิยม 3 ตำแหน่ง) มาก จึงไม่ทำให้ชิ้นที่หลุดสเปคจริงกลายเป็น OK
_TOL_EPS = 1e-6


def _within_tolerance(value: float, nominal: float, upper_tol: float, lower_tol: float) -> bool:
    """เช็คว่าค่าที่วัดได้ของแกนหนึ่งอยู่ในช่วง nominal -lower_tol .. +upper_tol ไหม
    (นับค่าที่ตกขอบพอดีว่าผ่าน — ดู _TOL_EPS)
    """
    return (nominal - lower_tol - _TOL_EPS) <= value <= (nominal + upper_tol + _TOL_EPS)


def _offset_ok(offset: Optional[float], offset_tol: Optional[float]) -> bool:
    """offset ผ่านเกณฑ์ไหม — เทียบกับ offset_tol ของ part_number

    ต่างจาก value_x/value_y ที่เทียบกับช่วง nominal ± tolerance — offset เป็น
    "ค่าความเยื้อง" ที่ยิ่งน้อยยิ่งดี จึงเทียบแค่ว่าไม่เกินเพดานที่ตั้งไว้
    ใช้ abs() เผื่อ TM-X ส่งค่าติดลบมา (เยื้องคนละทิศก็ถือว่าเยื้องเท่ากัน)

    ถ้ายังไม่ได้ตั้ง offset_tol (NULL) ถือว่า "ไม่ตรวจข้อนี้" → ผ่านเสมอ
    ไม่งั้น Part เก่าที่ยังไม่ได้กรอกค่านี้จะกลายเป็น NG ทั้งหมดทันทีที่ deploy
    """
    if offset is None or offset_tol is None:
        return True
    return abs(offset) <= offset_tol + _TOL_EPS


def _thai_date_str(dt: Optional[datetime] = None) -> str:
    """คืนวันที่รูปแบบ DD-MM-YYYY โดยปีเป็น พ.ศ. (ค.ศ. + 543) เช่น 22-07-2569
    ใช้ตั้งชื่อโฟลเดอร์/ไฟล์รูปภาพแบบแยกตามวันที่ (ดู upload_measurement_image)
    ค่า default (dt=None) ใช้เวลาปัจจุบันของเครื่องที่รัน backend ณ ตอนที่รูป
    ถูกอัปโหลดเข้ามา (ไม่ใช่เวลาที่วัด — สองอย่างนี้ในทางปฏิบัติเป็นเวลาเดียวกัน
    เพราะ Agent อัปโหลดรูปทันทีหลัง trigger แต่ละชิ้น)
    """
    dt = dt or datetime.now()
    return f"{dt.day:02d}-{dt.month:02d}-{dt.year + 543}"


# ══════════════════════════════════════════════════════════════════════════════
# ถังขยะ (Deleted) — สำรองข้อมูลก่อนลบ เพื่อให้กู้คืนได้
# ══════════════════════════════════════════════════════════════════════════════
# ทำไมเลือกวิธีนี้แทน soft delete (เพิ่มคอลัมน์ deleted_at):
#   มี 12 จุดในไฟล์นี้ที่ query ตาราง measurements — soft delete ต้องเติม
#   "WHERE deleted_at IS NULL" ให้ครบทุกจุด ลืมจุดเดียว = ข้อมูลที่ลบไปแล้ว
#   โผล่กลับมาในรายงาน/export เงียบๆ โดยไม่มีอะไรเตือน (เคยเจอบั๊กแบบนี้มาแล้ว
#   ตอน latest_only ที่ลืม WHERE ใน subquery)
#   วิธีนี้แค่ "แทรกขั้นตอนอ่านข้อมูลไปเก็บก่อนลบ" — ตรรกะเดิมไม่ถูกแตะเลย
#
# โครงสร้างที่เก็บ:
#   Deleted/<DD-MM-YYYY พ.ศ. ของวันที่ลบ>/<kind>_<id>.json
#   + ไฟล์รูป (ถ้ามี) วางคู่กันในโฟลเดอร์เดียวกัน
# แยกตาม "วันที่ลบ" ไม่ใช่วันที่วัด เพื่อให้ตัวลบอัตโนมัติลบทั้งโฟลเดอร์ได้เลย
DELETED_DIR = os.getenv("DELETED_DIR", os.path.join(_PROJECT_ROOT, "Deleted"))
if not os.path.isabs(DELETED_DIR):
    DELETED_DIR = os.path.abspath(os.path.join(_PROJECT_ROOT, DELETED_DIR))

# เก็บของในถังขยะกี่วันก่อนลบทิ้งถาวร
DELETED_RETENTION_DAYS = int(os.getenv("DELETED_RETENTION_DAYS", 30))


def _json_safe(value):
    """แปลงค่าจาก DB ให้ json.dumps ได้ — datetime/date/Decimal ไม่ใช่ชนิดมาตรฐาน"""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, timedelta):
        return str(value)
    return value


def _archive_before_delete(
    kind: str,
    table: str,
    pk: Dict[str, Any],
    row: Dict[str, Any],
    related: Optional[Dict[str, List[Dict]]] = None,
    image_path: Optional[str] = None,
) -> Optional[str]:
    """เก็บข้อมูลที่กำลังจะถูกลบไว้ใน Deleted/ — คืนชื่อไฟล์ JSON ที่สร้าง

    ⚠ ต้องเรียก "ก่อน" คำสั่ง DELETE เสมอ เพราะต้องอ่านค่าจากแถวที่ยังอยู่
    ⚠ ห้ามโยน exception ออกไป — ถ้าสำรองไม่สำเร็จก็ยังต้องยอมให้ลบต่อได้
      (ผู้ใช้สั่งลบแล้ว การไปขวางเพราะเขียนไฟล์สำรองไม่ได้จะงงกว่า) แต่ต้อง
      log ไว้ให้เห็นชัดว่ารอบนี้ไม่มีตัวสำรอง

    kind ใช้เป็นทั้งชื่อไฟล์และตัวบอกประเภทตอนกู้คืน:
      measurement | part | operator | owner | vendor | handler |
      template | package_size | part_number
    """
    try:
        day_dir = os.path.join(DELETED_DIR, _thai_date_str())
        os.makedirs(day_dir, exist_ok=True)

        pk_value = "_".join(str(v) for v in pk.values())
        stem = f"{kind}_{pk_value}"

        # ── ย้ายไฟล์รูปมาเก็บคู่กับ JSON (ไม่ลบทิ้ง) ────────────────────────
        image_file = None
        if image_path and "://" not in image_path:
            src = os.path.realpath(os.path.join(ALPL_IMAGE_DIR, image_path))
            base = os.path.realpath(ALPL_IMAGE_DIR)
            # กัน path ที่หลุดออกนอกโฟลเดอร์รูป (ค่าใน DB อาจถูกแก้มาก่อน)
            if (src == base or src.startswith(base + os.sep)) and os.path.isfile(src):
                image_file = f"{stem}_{os.path.basename(image_path)}"
                try:
                    shutil.move(src, os.path.join(day_dir, image_file))
                except OSError as exc:
                    log.warning("ย้ายไฟล์รูปเข้าถังขยะไม่สำเร็จ (%s): %s", image_path, exc)
                    image_file = None

        payload = {
            "deleted_at": datetime.now().isoformat(sep=" ", timespec="seconds"),
            "kind":       kind,
            "table":      table,
            "pk":         {k: _json_safe(v) for k, v in pk.items()},
            "row":        {k: _json_safe(v) for k, v in row.items()},
            "related":    {
                t: [{k: _json_safe(v) for k, v in r.items()} for r in rows]
                for t, rows in (related or {}).items()
            },
            "image_file": image_file,
        }

        # กันชื่อชนกันถ้าลบ id เดิมซ้ำในวันเดียวกัน (ลบ → กู้ → ลบอีก)
        json_name = f"{stem}.json"
        n = 2
        while os.path.exists(os.path.join(day_dir, json_name)):
            json_name = f"{stem}({n}).json"
            n += 1

        with open(os.path.join(day_dir, json_name), "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        log.info("เก็บเข้าถังขยะแล้ว: %s/%s", _thai_date_str(), json_name)
        return json_name
    except Exception as exc:
        log.error("สำรองข้อมูลก่อนลบไม่สำเร็จ (%s %s): %s — ลบต่อโดยไม่มีตัวสำรอง", kind, pk, exc)
        return None


def _fetch_one(cur, sql: str, params) -> Optional[Dict[str, Any]]:
    cur.execute(sql, params)
    return cur.fetchone()


def _delete_image_file(image_path: Optional[str]) -> bool:
    """ลบไฟล์รูปที่ image_path (path สัมพัทธ์ต่อ ALPL_IMAGE_DIR) ชี้อยู่

    คืน True เฉพาะตอนลบได้จริง และ **ไม่ throw ไม่ว่ากรณีไหน** เพราะทุกจุดที่
    เรียกใช้ถือว่า "ลบไฟล์ไม่ได้" ไม่ใช่เหตุให้ request ทั้งก้อนพัง (แถวใน DB
    ถูกลบ/อัปเดตไปแล้ว จะ rollback เพราะไฟล์ค้างไม่คุ้ม)

    กันไว้ 2 กรณีที่ห้ามแตะไฟล์เด็ดขาด:
      1. **ค่าที่เป็น URL เต็ม** — ข้อมูลตกค้างจากยุค MinIO ที่เก็บทั้ง URL ลง
         คอลัมน์นี้ (เจอจริงในแถว measurement_id=10:
         "http://172.20.10.4:8080/images/test.png") ไม่ใช่ path บนดิสก์เครื่องนี้
         ถ้าเอาไป join กับ ALPL_IMAGE_DIR จะได้ path มั่วๆ ที่ไม่ควรไปยุ่งด้วย
      2. **path ที่หลุดออกนอก ALPL_IMAGE_DIR** — ถ้าค่าใน DB มี ".." ปนมา
         (พลาดเองหรือถูกยัดมา) ต้องไม่ทำให้ลบไฟล์นอกโฟลเดอร์รูปได้
    """
    if not image_path:
        return False

    if "://" in image_path:
        log.warning(
            "ข้ามการลบไฟล์รูป: image_path เป็น URL ไม่ใช่ path บนดิสก์ (%s)", image_path
        )
        return False

    base = os.path.realpath(ALPL_IMAGE_DIR)
    target = os.path.realpath(os.path.join(base, image_path))
    if target != base and not target.startswith(base + os.sep):
        log.warning(
            "ข้ามการลบไฟล์รูป: path หลุดออกนอก ALPL_IMAGE_DIR (%s)", image_path
        )
        return False

    try:
        os.remove(target)
        return True
    except OSError:
        return False  # ไฟล์อาจถูกลบไปแล้ว/หาไม่เจอ — ไม่ใช่เรื่องผิดปกติ

# AGENT_HOST: เดิม hardcode เป็น "localhost" ตรงๆ (สมมติว่า Agent รันอยู่เครื่อง
# เดียวกับ Backend เสมอ) — ตอนนี้ Agent อาจย้ายไปรันบนเครื่องแยก (เช่น Raspberry
# Pi ที่ทำหน้าที่คุยกับ sensor/MCU โดยตรง) จึงต้องดึงจาก .env แทน ถ้าไม่ตั้งค่า
# ใน .env จะ fallback เป็น "localhost" เหมือนเดิมทุกประการ (เทสต์บน PC เครื่อง
# เดียวได้ปกติ ไม่กระทบ) พอมี Pi จริงแค่ตั้ง AGENT_HOST=<IP ของ Pi> ใน .env
# ไม่ต้องแก้โค้ดจุดนี้อีก
AGENT_HOST      = os.getenv("AGENT_HOST", "localhost")
AGENT_PORT      = int(os.getenv("AGENT_PORT", 9998))
AGENT_BASE_URL  = f"http://{AGENT_HOST}:{AGENT_PORT}"

# heartbeat: Agent ยิง POST /api/heartbeat มาทุก HEARTBEAT_INTERVAL วิ ระหว่างที่
# ยังรันอยู่ (ดู agent.py heartbeat_loop) — heartbeat_checker() ด้านล่างเช็คเป็น
# ระยะว่า session ที่ 'running' ยังได้ heartbeat ต่อเนื่องไหม ถ้าเงียบเกิน
# HEARTBEAT_TIMEOUT วิ (Agent process ตาย/แฮงค์กลาง session) จะ mark เป็น
# 'timeout' อัตโนมัติ (เอากลับมาใหม่ตามที่ตกลงกันไว้)
HEARTBEAT_INTERVAL = int(os.getenv("HEARTBEAT_INTERVAL", 5))
HEARTBEAT_TIMEOUT  = int(os.getenv("HEARTBEAT_TIMEOUT", 15))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [Server] %(message)s")
log = logging.getLogger(__name__)

# ── CORS ────────────────────────────────────────────────────────────────────
# ค่าเริ่มต้นเป็น "*" โดยตั้งใจ — เคยลองบีบให้เหลือเฉพาะ localhost แล้วพัง:
# เปิดหน้าเว็บผ่าน IP ของเครื่องในวง (เช่น http://192.168.1.50:8000) จะกลายเป็น
# cross-origin ทันที แล้ว SSE ถูกบล็อก ป้ายสถานะค้างที่ Offline ตลอดโดยไม่มี
# error อะไรให้เห็นนอกจาก console ของเบราว์เซอร์
#
# และต้องเข้าใจให้ตรงกันว่า CORS ไม่ใช่ระบบความปลอดภัยของ API — มันกันได้แค่
# JavaScript ในเบราว์เซอร์ ส่วน curl/Postman/สคริปต์ใดๆ ไม่สนใจ CORS เลย
# ช่องโหว่จริงของระบบนี้คือ "ไม่มี auth" ซึ่งต้องแก้ด้วยการทำ auth ไม่ใช่บีบ CORS
#
# ตั้ง CORS_ORIGINS ใน .env เป็นรายการคั่นด้วยคอมมาได้ถ้าอยากจำกัดจริงๆ เช่น
#   CORS_ORIGINS=http://localhost:8000,http://192.168.1.50:8000
CORS_ORIGINS = [
    o.strip() for o in os.getenv("CORS_ORIGINS", "*").split(",") if o.strip()
]

# ── SSE broadcast queue ──────────────────────────────────────────────────────
subscribers: List[asyncio.Queue] = []

# ── In-memory queue state สำหรับ session แบบ IPM/New/Rework ──────────────────
# เก็บ "คิว" ALPL + ตำแหน่งปัจจุบันของ session ที่เริ่มจาก Part Entry card
# (โหมด IPM/New/Rework) — เป็นตัวแปร memory ธรรมดา ไม่ใช่ column ใน DB เลย เพราะ
# schema ของ `sessions` ไม่มีที่เก็บลำดับ ALPL ทั้งคิว มีแค่ number_alpl ตัวเดียว
# (ที่เราใส่เป็น ALPL ตัวแรกในคิวไปแทน) — ถ้า server restart กลางที่ session
# กำลัง running อยู่ คิวนี้จะหาย (ยอมรับความเสี่ยงนี้ได้ตามที่คุยกันไว้)
#
# โครงสร้าง: { session_id: {"entry_mode": "IPM"|"New", "queue": [1011, 1002, ...],
#                            "position": 0} }
# หมายเหตุ: entry_mode มีแค่ 2 ค่า (ไม่มี "Rework") เพราะโหมด Rework ที่เลือก
# จากหน้าเว็บถูก map เป็น entry_mode='New' + note='Rework' ตั้งแต่ start_session
# แล้ว (ดู start_session) — ค่านี้ถูกเอาไปใส่คอลัมน์ measurements.measure_type ตรงๆ
session_queues: Dict[int, Dict[str, Any]] = {}


async def push_event(event_type: str, data: dict) -> None:
    """กระจาย (broadcast) เหตุการณ์ SSE หนึ่งรายการไปให้ทุก client ที่เปิดหน้า dashboard อยู่

    ทำไมต้องทำแบบนี้: backend คือ Single Source of Truth ของสถานะ session/measurement
    ดังนั้นเมื่อสถานะมีการเปลี่ยนแปลง (เริ่ม session, มี measurement ใหม่, timeout ฯลฯ)
    ทุกแท็บ dashboard ที่เปิดอยู่ต้องรู้ทันที แต่ละ subscriber มี asyncio.Queue ของตัวเอง
    (ดู sse_stream ด้านล่าง) เราแค่ส่ง payload เดียวกันลงไปในทุกคิว SSE ไหลทางเดียวจาก
    server ไป client เท่านั้น (ไม่ใช่ request/response แบบ 2 ทาง)
    """
    payload = json.dumps(data, default=str)
    for q in subscribers:
        await q.put({"event": event_type, "data": payload})
    log.info("SSE ▶ %s: %s", event_type, payload)


# ── DB helpers ───────────────────────────────────────────────────────────────
def get_db():
    """เปิด MySQL connection ใหม่สำหรับ 1 request

    ทำไมต้องเปิดใหม่ทุกครั้งแทนใช้ connection pool: นี่คือระบบที่ deploy บน PC เดียว
    มี concurrency ต่ำ ความเรียบง่ายของ "connect → ใช้งาน → close" จึงคุ้มกว่าความซับซ้อน
    ของการทำ pool ทุก endpoint ด้านล่างจะเปิด connection นี้ใน try/finally แล้วปิดเมื่อใช้เสร็จ

    หมายเหตุ (เพิ่มเข้ามาทีหลัง): endpoint ทุกตัวเรียก get_db() "ก่อน" เข้า try/finally
    ของตัวเอง ถ้า MySQL server ล่ม/ต่อไม่ติดเลย pymysql.connect() จะ raise
    OperationalError ซึ่งเป็น raw exception ที่ FastAPI ไม่รู้จัก — หลุดออกไปกลายเป็น
    500 ดิบไม่มี CORS header แนบมา (ปัญหาเดียวกับที่ _notify_agent_start เจอกับ Agent)
    จับไว้ตรงนี้ที่เดียวแล้ว raise เป็น HTTPException(503) แทน เพื่อให้ CORS header
    ยังติดมาด้วยเสมอ ไม่ต้องไปแก้ทุก endpoint
    """
    try:
        return pymysql.connect(**DB_CONFIG)
    except pymysql.MySQLError as exc:
        log.error("Database connection failed: %s", exc)
        raise HTTPException(503, f"เชื่อมต่อฐานข้อมูลไม่สำเร็จ: {exc}")


# ── Lifespan ─────────────────────────────────────────────────────────────────
async def _reload_session_queues() -> None:
    """โหลด session_queues กลับเข้า memory จากคอลัมน์ sessions.queue_state

    ทำไมต้องมี: session_queues เดิมอยู่ใน memory ของ backend ล้วนๆ ถ้า backend
    ถูก restart (reload ตอน dev, crash แล้ว auto-restart, deploy ใหม่) ระหว่างที่
    มี session แบบ IPM/New กำลัง running อยู่ คิวจะหายไปจาก memory ทันที —
    create_measurement หลังจากนั้นจะ fallback ไปใช้ req.number_alpl ที่ Agent
    ส่งมา ซึ่งเป็น ALPL ตัวแรกในคิวเสมอ (Agent ไม่เคยอัปเดตค่านี้เอง) ทำให้ทุก
    measurement ที่เหลือถูกบันทึกผิด ALPL ไปเรื่อยๆ แบบไม่มี error เตือนเลย

    จึงต้องรันตรงนี้ (ก่อน yield ให้แอปเริ่มรับ request) — ต้อง await ตรงๆ ไม่ใช่
    fire-and-forget แบบ _init_bucket_bg เพราะต้องมั่นใจว่า session_queues ถูก
    เติมกลับให้ครบก่อนที่ request แรกจาก Agent จะเข้ามาได้
    """
    try:
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT session_id, queue_state FROM sessions "
                    "WHERE state IN ('running', 'paused') AND queue_state IS NOT NULL"
                )
                rows = cur.fetchall()
        finally:
            db.close()

        for row in rows:
            try:
                session_queues[row["session_id"]] = json.loads(row["queue_state"])
                log.info("Restored queue_state for session %s from DB", row["session_id"])
            except Exception as exc:
                log.warning("Failed to parse queue_state for session %s: %s", row["session_id"], exc)
    except Exception as exc:
        # DB อาจยังไม่พร้อมตอน boot — อย่าทำให้แอปบูตไม่ขึ้นเพราะเรื่องนี้ แค่
        # log ไว้ (session ที่ running อยู่ตอน restart แบบนี้จะพลาดการกู้คืนคิว
        # แต่ยังใช้งานต่อได้ปกติถ้าไม่ใช่ queue-based หรือกด Stop แล้วเริ่มใหม่)
        log.warning("Reload session_queues failed: %s", exc)


async def heartbeat_checker() -> None:
    """ตรวจเป็นระยะว่า session ที่ 'running' ยังได้ heartbeat จาก Agent ต่อเนื่องไหม

    หมายเหตุ: ก่อนหน้านี้เคยถอดกลไกนี้ออกไปเพราะตอนนั้น Agent/Backend/Web รันอยู่
    บนเครื่องเดียวกันหมด คิดว่าไม่จำเป็น — ตอนนี้เอากลับมาใหม่ตามที่ตกลงกันไว้
    เป็น safety net เผื่อ Agent process ตาย/แฮงค์กลาง session (ไม่ใช่แค่ network
    หลุดข้ามเครื่องเหมือนเหตุผลเดิม) ถ้าเงียบเกิน HEARTBEAT_TIMEOUT วิ จะ mark
    session เป็น 'timeout' อัตโนมัติ แล้วแจ้ง web ผ่าน SSE (`session_timeout` —
    index.html มี handler นี้อยู่แล้ว แค่ก่อนหน้านี้ไม่มีใคร emit ให้)
    """
    while True:
        await asyncio.sleep(HEARTBEAT_INTERVAL)
        try:
            db = get_db()
        except HTTPException as exc:
            log.warning("heartbeat_checker: DB unreachable, skip this round: %s", exc.detail)
            continue

        timed_out: List[int] = []
        try:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT session_id FROM sessions "
                    "WHERE state = 'running' AND last_seen < NOW() - INTERVAL %s SECOND",
                    (HEARTBEAT_TIMEOUT,),
                )
                timed_out = [row["session_id"] for row in cur.fetchall()]
                for sid in timed_out:
                    cur.execute(
                        "UPDATE sessions SET state = 'timeout', ended_at = NOW() "
                        "WHERE session_id = %s",
                        (sid,),
                    )
        except Exception as exc:
            log.warning("heartbeat_checker: check failed: %s", exc)
            timed_out = []
        finally:
            db.close()

        for sid in timed_out:
            session_queues.pop(sid, None)
            measure_timeouts.pop(sid, None)  # คำถามค้างของ session ที่ตายไปแล้ว ไม่มีใครตอบได้อีก
            log.warning("Session %s: ไม่ได้ heartbeat เกิน %ss — mark เป็น 'timeout'", sid, HEARTBEAT_TIMEOUT)
            await push_event("session_timeout", {"session_id": sid})


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Hook ตอน FastAPI เริ่มทำงาน (startup) และตอนปิด (shutdown)

    ทำไม: ตรงนี้คือจุดที่ background task (heartbeat_checker) ถูกสั่งให้เริ่ม
    ทำงานตอนแอป boot แทนที่จะไปสั่งเริ่มภายใน request handler ส่วน
    _reload_session_queues() ต้อง await ให้เสร็จก่อน yield เพราะต้องกู้คืนคิว
    ให้ครบก่อนรับ request
    """
    asyncio.create_task(heartbeat_checker())     # fire-and-forget, never blocks
    asyncio.create_task(_deleted_purge_loop())   # ล้างถังขยะที่เกินอายุ
    await _reload_session_queues()               # ต้องเสร็จก่อนรับ request
    yield


async def _deleted_purge_loop():
    """ล้างถังขยะที่เกินอายุ — ทำทันทีตอนเริ่ม แล้ววนซ้ำวันละครั้ง

    ทำตอนเริ่มด้วยเพราะเครื่องหน้างานมักถูกปิดกลางคืน ถ้ารอครบ 24 ชั่วโมงอย่างเดียว
    อาจไม่มีวันได้ทำเลย
    """
    while True:
        try:
            _purge_old_deleted()
        except Exception as exc:
            log.warning("ล้างถังขยะไม่สำเร็จ: %s", exc)
        await asyncio.sleep(24 * 60 * 60)


app = FastAPI(title="TM-X Backend Server", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ══════════════════════════════════════════════════════════════════════════════
# SSE Stream
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/api/stream")
async def sse_stream(request: Request):
    """Endpoint SSE ที่ dashboard เชื่อมต่อเข้ามาเพื่อรับข้อมูล real-time

    ทำไม: แทนที่ frontend จะ poll backend ทุกวินาที มันเปิด connection ค้างไว้ทีเดียว
    ที่นี่ แล้วเรา push event ไปให้ตอนมันเกิดขึ้นจริง (ดู push_event) แต่ละ client
    จะมี queue ของตัวเองที่ลงทะเบียนใน `subscribers` เราจะ yield "ping" ทุก 25 วินาที
    ตอนไม่มีอะไรใหม่ แค่เพื่อ keep connection ไว้ไม่ให้ proxy/browser ตัดการเชื่อมต่อ
    ที่ idle อยู่ SSE เป็นทางเดียว (server → client เท่านั้น) — ฝั่ง frontend ยังใช้
    POST request ปกติในการส่งคำสั่งไปที่ backend
    """
    async def generator():
        queue: asyncio.Queue = asyncio.Queue()
        subscribers.append(queue)
        log.info("SSE client connected  (total=%d)", len(subscribers))
        try:
            while True:
                try:
                    event = await asyncio.wait_for(queue.get(), timeout=25)
                    yield event
                except asyncio.TimeoutError:
                    yield {"event": "ping", "data": ""}
        finally:
            if queue in subscribers:
                subscribers.remove(queue)
            log.info("SSE client disconnected (total=%d)", len(subscribers))

    return EventSourceResponse(generator())


# ══════════════════════════════════════════════════════════════════════════════
# Session endpoints
# ══════════════════════════════════════════════════════════════════════════════
class StopSessionRequest(BaseModel):
    session_id: int


@app.get("/api/session/state")
async def get_session_state():
    """คืนสถานะปัจจุบันของ session ล่าสุด

    ทำไม: ตอน dashboard โหลดครั้งแรก (หรือ refresh) มันต้องรู้ว่า "มี run การวัด
    กำลังทำงานอยู่ไหม" ก่อนที่ SSE connection จะเปิดเสียอีก นี่คือ snapshot
    แบบครั้งเดียวที่ใช้ sync ตอนเริ่ม หลังจากนั้น SSE event จะคอยอัปเดตให้ real-time
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            # queue_state แนบไปด้วย — frontend ใช้วาดแถบคิว ALPL ใน Live Telemetry
            # (ต้องได้คิวเต็มไม่ใช่แค่ ALPL ตัวแรก) และทำให้แถบนี้รอดการ refresh
            # หน้าเว็บกลาง session ด้วย เพราะอ่านคิวกลับจาก DB ได้ตรงๆ
            cur.execute(
                "SELECT session_id, number_alpl, state, target_count, measured_count, "
                "queue_state, last_seen, started_at, ended_at "
                "FROM sessions ORDER BY session_id DESC LIMIT 1"
            )
            row = cur.fetchone()
        return row or {"state": "idle"}
    finally:
        db.close()


# หมายเหตุ: เดิมมีฟังก์ชัน _insert_new_parts_from_payload() ที่ insert Part
# "ทุกตัวในคิว" ทีเดียวตอน start_session — เปลี่ยนพฤติกรรมแล้ว (ดู start_session
# และ create_measurement) เพราะถ้า user กด Stop กลางคัน ALPL ที่ยังไม่ทันวัดจะ
# ค้างเป็น Part "ผี" อยู่ใน DB ทั้งที่ไม่เคยมีการวัดจริงเกิดขึ้นเลย ตอนนี้จึง
# insert Part แค่ตัวแรกตอน start_session (จำเป็นเพราะ sessions.number_alpl มี
# FK ไป parts ต้องมี row อยู่ก่อนถึงจะ insert sessions ได้) ส่วนตัวที่เหลือใน
# คิวจะถูก insert ทีละตัว "ตอนได้ผลวัดจริงจาก Agent" ใน create_measurement เท่านั้น


def _get_template_name_for_ipm(cur, first_alpl: int) -> str:
    """Query หา template_name — ใช้เฉพาะกรณี IPM เพราะ JSON ของ IPM ไม่มี
    template_name ส่งมาด้วย ต่างจาก New ที่ frontend ส่ง template_name มาใน
    ก้อน JSON เลย ไม่ต้อง query

    schema: template_name ไม่ได้อยู่ใน `parts_specifications` ตรงๆ — ต้อง join
    ผ่าน part_number_id -> package_size_id -> template_id เป็นทอดๆ
    (parts_specifications.part_number_id -> part_number.package_size_id ->
    package_size.template_id -> template.template_name)

    ใช้ ALPL ตัวแรกในคิวเป็นตัวหา เพราะ IPM ต้องมี parts_specifications row
    อยู่แล้วทุกตัว (ลงทะเบียนไว้ก่อนหน้านี้) พร้อม part_number_id ที่ตั้งไว้แล้ว
    — ถ้าหาไม่เจอ แปลว่า frontend เช็คตกหรือมีคนลบ part/ยังไม่ได้ตั้ง part_number
    ให้ part นี้ ให้ raise error ชัดๆ ไปเลย ไม่เดา
    """
    # ใช้ LEFT JOIN แทน INNER JOIN เพื่อ "วินิจฉัย" ได้ว่าสายพังตรงข้อไหน
    # ของเดิมใช้ INNER JOIN แล้วได้ 0 แถว จึงบอกได้แค่ว่า "ยังไม่ได้ตั้ง part_number"
    # ซึ่งชี้ผิดจุดบ่อย เพราะจริงๆ อาจตั้ง Part Number ไว้แล้วแต่ Package Size
    # ของมันยังไม่ได้ผูก Template — ผู้ใช้ก็จะวนแก้ที่ Part Number ซ้ำๆ ไม่จบ
    cur.execute(
        "SELECT p.part_number_id, pn.part_number_name, pn.package_size_id, "
        "       ps.package_size, ps.template_id, t.template_name "
        "FROM parts_specifications p "
        "LEFT JOIN part_number pn  ON p.part_number_id  = pn.part_number_id "
        "LEFT JOIN package_size ps ON pn.package_size_id = ps.package_size_id "
        "LEFT JOIN template t      ON ps.template_id     = t.template_id "
        "WHERE p.number_alpl = %s",
        (first_alpl,),
    )
    row = cur.fetchone()

    head = f"เริ่มวัด ALPL {first_alpl} ไม่ได้ —"
    if row is None:
        raise HTTPException(404, f"{head} ยังไม่ได้ลงทะเบียน ALPL นี้ในตาราง Parts")
    if row["part_number_id"] is None:
        raise HTTPException(404, f"{head} ยังไม่ได้ตั้ง Part Number ให้ ALPL นี้ (แก้ที่หน้า Edit › Parts)")
    if row["part_number_name"] is None:
        raise HTTPException(
            404,
            f"{head} Part Number ที่ผูกไว้ถูกลบไปแล้ว (part_number_id = {row['part_number_id']}) "
            f"— เลือก Part Number ใหม่ให้ ALPL นี้ที่หน้า Edit › Parts",
        )
    if row["package_size_id"] is None:
        raise HTTPException(
            404,
            f"{head} Part Number \"{row['part_number_name']}\" ยังไม่ได้ผูก Package Size "
            f"(แก้ที่หน้า Edit › Lookup Tables › Part Number)",
        )
    if row["package_size"] is None:
        raise HTTPException(
            404,
            f"{head} Package Size ที่ Part Number \"{row['part_number_name']}\" ผูกไว้ถูกลบไปแล้ว "
            f"(package_size_id = {row['package_size_id']})",
        )
    if row["template_id"] is None:
        raise HTTPException(
            404,
            f"{head} Package Size \"{row['package_size']}\" ยังไม่ได้ตั้ง Template ของเครื่อง TM-X "
            f"(แก้ที่หน้า Edit › Lookup Tables › Package Size) — ไม่เกี่ยวกับ Part Number",
        )
    if row["template_name"] is None:
        raise HTTPException(
            404,
            f"{head} Template ที่ Package Size \"{row['package_size']}\" ผูกไว้ถูกลบไปแล้ว "
            f"(template_id = {row['template_id']})",
        )
    return row["template_name"]


async def _notify_agent_start(
    session_id: int,
    template_name: str,
    target_count: int,
    number_alpl: int,
) -> None:
    """ยิง POST ไปที่ Agent (`agent.py`) เพื่อบอกให้เริ่มวัด

    ใช้ร่วมกันทั้ง New และ IPM — สิ่งที่ Agent ต้องรู้เหมือนกันทุกกรณีคือ
    session_id, template_name (ให้โหลดเข้า TM-X), target_count, และ
    number_alpl ตัวแรกที่จะวัด (ตัวต่อไปในคิว Agent ไม่จำเป็นต้องรู้ เพราะ
    มันแค่ส่ง value_x/value_y มาเรื่อยๆ โดยไม่ต้องสนใจว่าเป็น ALPL ไหน —
    backend เป็นคนจับคู่กับ ALPL เองจากตำแหน่งในคิว)

    ครอบด้วย try/except เพราะถ้า Agent ไม่ได้รันอยู่ (หรือตอบช้าเกิน timeout)
    httpx จะ raise exception ที่ FastAPI ไม่รู้จัก (ConnectError/TimeoutException)
    — ถ้าปล่อยให้ exception นี้หลุดออกไปจากเอนด์พอยต์โดยไม่ catch จะกลายเป็น
    unhandled exception (500) ที่บางครั้งไม่มี CORS header แนบมาด้วย ทำให้
    browser เข้าใจผิดว่าเป็น CORS error ทั้งที่จริงๆคือ Agent ไม่ตอบ — log
    warning ไว้เฉยๆ แล้วให้ session ใน DB ยัง 'running' ต่อไปได้ (เหมือนที่
    stop_session ทำไว้อยู่แล้วตอน notify agent ตอน stop)
    """
    try:
        async with httpx.AsyncClient() as client:
            await client.post(
                f"{AGENT_BASE_URL}/command",
                json={
                    "action": "start",
                    "session_id": session_id,
                    "template_name": template_name,
                    "target_count": target_count,
                    "number_alpl": number_alpl,
                },
                timeout=10,
            )
    except Exception as exc:
        log.warning("Agent start notify failed: %s", exc)


@app.post("/api/session/start")
async def start_session(request: Request):
    """เริ่ม session การวัดใหม่ จาก Part Entry card (โหมด IPM, New หรือ Rework)

    **กรณี Measure_Type == "New"** (ลงทะเบียน Part ใหม่ + วัดในรอบเดียว):
      1. Insert Part เฉพาะ "ตัวแรก" ในคิวก่อน (ต้องทำก่อน insert session เพราะ
         sessions.number_alpl มี FOREIGN KEY ไป parts — ถ้า insert session
         ก่อนโดย ALPL ยังไม่มีอยู่จริง MySQL จะปฏิเสธทันที) — ALPL ตัวที่เหลือใน
         คิวจะ insert ทีละตัว "ตอนได้ผลวัดจริงจาก Agent" ใน create_measurement
         แทน ไม่ insert รวดเดียวทั้งคิว กันกรณีกด Stop กลางคันแล้วมี Part ที่
         ไม่เคยวัดจริงค้างอยู่ใน DB
      2. Insert sessions row (number_alpl = ALPL ตัวแรกในคิว) → ได้ session_id
      3. เก็บคิว ALPL ทั้งหมด + ตำแหน่งเริ่มต้น (0) + config ของ Part (ไว้ insert
         ตัวถัดๆไปแบบ lazy) ไว้ใน session_queues (memory)
      4. Notify Agent ให้เริ่มวัด (template_name มาจาก JSON ตรงๆ ไม่ query DB)

    **กรณี Measure_Type == "IPM"** (ALPL ลงทะเบียนไว้แล้ว):
      1. Query DB หา template_name จาก ALPL ตัวแรกในคิว (ไม่ insert parts เลย)
      2. Insert sessions row → ได้ session_id
      3. เก็บคิวไว้ใน session_queues เหมือนกัน
      4. Notify Agent

    **กรณี Measure_Type == "Rework"** (งานที่เคยลงทะเบียนผ่าน New แล้ว ไม่ผ่าน
    ถูกส่งไป Rework แล้วส่งกลับมาวัดใหม่):
      1. รับได้ทีละ 1 ALPL เท่านั้น (ต่างจาก New/IPM ที่รับเป็นคิวได้) — ALPL
         ต้องมี Part row อยู่แล้วจริง (ไม่งั้น 404 บอกให้ไปลงทะเบียนที่ New ก่อน)
      2. Update Part row เดิม (ไม่ insert ใหม่) ด้วยค่าที่ผู้ใช้กรอก/แก้ในฟอร์ม
         Rework — ฟอร์มนี้ auto-fill ข้อมูลเดิมของ ALPL มาให้แทบทุกช่องแล้ว
         ยกเว้น Receive Date ที่บังคับกรอกใหม่เสมอ (วันที่รับกลับมาจริง)
      3. Query template_name จาก DB หลัง update (เผื่อ package_size ถูกแก้)
      4. Insert sessions row → ได้ session_id, เก็บคิวไว้ (แค่ 1 ALPL) แล้ว
         notify Agent เหมือนโหมดอื่น

    ทั้ง 3 กรณี — Agent ไม่ต้องรู้ความต่างเลย ได้รับ payload หน้าตาเดียวกัน
    (action/session_id/template_name/target_count/number_alpl ตัวแรก) ส่วน
    การ map ALPL ตัวต่อๆไปในคิวเข้ากับ measurement ที่จะตามมา เป็นเรื่องที่
    backend จัดการเองทั้งหมดผ่าน session_queues (ดู create_measurement)

    หมายเหตุ (เพิ่มเข้ามาทีหลัง): Race condition ตอนกด Start ซ้ำเร็วๆ — ครอบ
    check+insert ด้วย MySQL GET_LOCK/RELEASE_LOCK กันสอง request แข่งกันผ่าน
    Button Guard พร้อมกันได้ (เดิมเช็คแล้ว insert คนละคำสั่ง ไม่มีอะไรล็อก
    ระหว่างนั้นเลย)
    """
    data = await request.json()
    log.info("📥 ได้รับ payload จาก /api/session/start:\n%s", json.dumps(data, ensure_ascii=False, indent=2))

    measure_type = data.get("Measure_Type")
    if measure_type not in ("New", "IPM", "Rework"):
        raise HTTPException(400, "Measure_Type ต้องเป็น 'New', 'IPM' หรือ 'Rework'")

    # parse number_alpl อย่างระมัดระวัง — ถ้า key หายไปหรือมีค่าที่แปลงเป็น
    # int ไม่ได้ จะได้ KeyError/ValueError ซึ่งเป็น raw exception (ไม่ใช่
    # HTTPException) ที่ FastAPI ไม่รู้จัก ทำให้ response กลายเป็น 500 แบบ
    # ไม่มี CORS header แนบไปด้วย (เกิดปัญหาเดียวกับที่เจอใน _notify_agent_start
    # ก่อนหน้านี้) จึงต้อง catch แล้ว raise เป็น HTTPException ให้ชัดเจน
    try:
        alpl_queue = [int(x) for x in data["number_alpl"]]
    except KeyError:
        raise HTTPException(400, "ต้องมี field 'number_alpl' ใน payload")
    except (ValueError, TypeError):
        raise HTTPException(400, "number_alpl ต้องเป็น array ของเลขจำนวนเต็มทั้งหมด")

    if not alpl_queue:
        raise HTTPException(400, "number_alpl ต้องมีอย่างน้อย 1 ค่า")
    first_alpl = alpl_queue[0]
    target_count = len(alpl_queue)

    db = get_db()
    try:
        # GET_LOCK ครอบทั้ง Button Guard + insert — ให้ทั้งสองเป็น atomic
        # section เดียวกันจริงๆ ในระดับ DB (ไม่ใช่แค่ระดับ Python) กันสอง
        # request "Start" ที่มาถึงพร้อมกันเป๊ะๆ ผ่าน check ทั้งคู่ก่อนจะมีใคร
        # insert ทัน — timeout 5 วิ พอสำหรับ critical section สั้นๆ นี้
        with db.cursor() as cur:
            cur.execute("SELECT GET_LOCK('tmx_start_session', 5) AS got")
            if not cur.fetchone()["got"]:
                raise HTTPException(503, "ระบบกำลังประมวลผลคำสั่ง Start อื่นอยู่ ลองใหม่อีกครั้ง")

        try:
            with db.cursor() as cur:
                # Button Guard — กันรัน 2 session ซ้อนกัน (เหมือนของเดิมก่อนหน้านี้)
                # นับ 'paused' เป็น session ที่ยังไม่จบด้วย ไม่งั้นตอนพักค้างไว้จะ
                # เริ่ม session ใหม่ซ้อนได้ แล้ว session เก่าค้างเป็น paused ตลอดกาล
                cur.execute("SELECT session_id FROM sessions WHERE state IN ('running', 'paused')")
                if cur.fetchone():
                    raise HTTPException(400, "A session is already running")

                if measure_type == "New":
                    # 1) Insert Part เฉพาะ "ตัวแรก" ในคิวก่อน (ต้องมาก่อน insert
                    # session เพราะ FK — sessions.number_alpl ต้องมี Part อยู่จริง
                    # ก่อนถึงจะ insert ได้) ส่วน ALPL ตัวที่เหลือในคิวจะถูก insert
                    # ทีละตัว "ตอนได้ผลวัดจริง" ใน create_measurement แทน ไม่ insert
                    # รวดเดียวทั้งคิวแบบเดิม — กันกรณีกด Stop กลางคันแล้วมี Part ที่
                    # ไม่เคยวัดจริงค้างอยู่ใน DB
                    #
                    # ครอบ try/except เพราะ MySQL อาจ throw error ได้หลายแบบตอน
                    # insert (ALPL ซ้ำ, ข้อมูลผิด type, ค่ายาวเกิน column ฯลฯ) —
                    # จับ pymysql.MySQLError (base class ของ error ทุกชนิดจาก MySQL)
                    # ไม่ใช่แค่ IntegrityError ตัวเดียว เพื่อกัน raw exception หลุด
                    # ออกไปทำให้ response 500 ไม่มี CORS header แนบมา
                    try:
                        _insert_part_row(cur, first_alpl, data)
                    except pymysql.MySQLError as exc:
                        raise HTTPException(409, f"Insert Part แรกในคิวไม่สำเร็จ: {exc}")
                    template_name = data.get("template_name")
                elif measure_type == "Rework":
                    # Rework: งานที่เคยลงทะเบียนผ่าน New แล้ว แต่ไม่ผ่าน ถูกส่งไป
                    # Rework แล้วส่งกลับมาวัดใหม่ — ALPL ต้องมี Part row อยู่แล้ว
                    # จริง (ห้ามใช้กับ ALPL ที่ไม่เคยลงทะเบียน ต้องไปสร้างที่ New
                    # ก่อน) จำกัดไว้ทีละ 1 ALPL เท่านั้น เพราะฟอร์ม Rework แสดง/
                    # แก้ config ของ Part เดิม 1 ตัวเป๊ะๆ (ไม่ใช่ config กลางที่ใช้
                    # ซ้ำกับหลาย ALPL แบบ New — ALPL แต่ละตัวที่ Rework กลับมามี
                    # ประวัติเดิมของตัวเองไม่เหมือนกัน จะ share config เดียวไม่ได้)
                    if len(alpl_queue) != 1:
                        raise HTTPException(400, "Rework รองรับทีละ 1 ALPL เท่านั้น")
                    cur.execute("SELECT 1 FROM parts_specifications WHERE number_alpl = %s", (first_alpl,))
                    if not cur.fetchone():
                        raise HTTPException(
                            404,
                            f"ALPL {first_alpl} ยังไม่เคยลงทะเบียน — ไปลงทะเบียนที่แท็บ New ก่อน",
                        )
                    try:
                        _update_part_row(cur, first_alpl, data)
                    except pymysql.MySQLError as exc:
                        raise HTTPException(409, f"Update Part สำหรับ Rework ไม่สำเร็จ: {exc}")
                    # ใช้ query แบบเดียวกับ IPM (ไม่เชื่อ template_name จาก payload)
                    # เพราะ package_size อาจถูกแก้ระหว่าง Rework — เอาค่าล่าสุดจาก
                    # DB หลัง update เสมอ กันส่ง template ผิดตัวไปให้ Agent
                    template_name = _get_template_name_for_ipm(cur, first_alpl)
                else:
                    # IPM: ไม่ insert parts เลย แค่ query หา template_name
                    template_name = _get_template_name_for_ipm(cur, first_alpl)

                # 2) Insert sessions row (ผ่าน FK ได้แน่นอนแล้ว ไม่ว่าจะ New หรือ IPM)
                cur.execute(
                    "INSERT INTO sessions (number_alpl, state, target_count, measured_count) "
                    "VALUES (%s, 'running', %s, 0)",
                    (first_alpl, target_count),
                )
                session_id = cur.lastrowid
        finally:
            with db.cursor() as cur:
                cur.execute("SELECT RELEASE_LOCK('tmx_start_session')")

        # 3) เก็บคิวไว้ใน memory ผูกกับ session_id นี้ (หลัง insert สำเร็จแล้ว
        # ค่อยผูก กัน insert fail แล้วมี state ค้างอยู่ใน session_queues)
        #
        # **การ map โหมดที่เลือกหน้าเว็บ → ค่าที่บันทึกลง measurements**
        # (ตามที่ตกลงกันไว้ — Rework ไม่ใช่ measure_type ของตัวเอง แต่ถือเป็น
        # การวัดแบบ New ที่มีหมายเหตุกำกับว่าเป็นงาน Rework):
        #   หน้าเว็บเลือก "Rework" → measure_type = 'New',  note = 'Rework'
        #   หน้าเว็บเลือก "New"    → measure_type = 'New',  note = NULL
        #   หน้าเว็บเลือก "IPM"    → measure_type = 'IPM',  note = NULL
        # entry_mode ตรงนี้คือค่าที่ create_measurement จะเอาไปใส่คอลัมน์
        # measure_type ตรงๆ (ดู create_measurement) จึงต้อง map ที่นี่ให้เรียบร้อย
        entry_mode = "New" if measure_type in ("New", "Rework") else "IPM"
        entry_note = "Rework" if measure_type == "Rework" else None

        queue_state = {
            "entry_mode": entry_mode,
            "queue": alpl_queue,
            "position": 0,
            "operator": data.get("Operator"),
            "note": entry_note,
            # เก็บ config เดิมของ New ไว้ (part_number/handler/vendor/
            # package_size/owner ฯลฯ) ให้ create_measurement เอาไป insert Part
            # ตัวถัดๆไปในคิวแบบ lazy ทีละตัวตอนวัดจริง — None ถ้าเป็น IPM/Rework
            # (IPM ใช้ Part ที่ลงทะเบียนไว้แล้วทั้งหมด, Rework จำกัดทีละ 1 ALPL ที่
            # update ไปเรียบร้อยแล้วข้างบน ไม่มี ALPL ตัวถัดไปในคิวให้ insert lazy)
            "new_part_config": data if measure_type == "New" else None,
        }
        session_queues[session_id] = queue_state
        measure_timeouts.pop(session_id, None)  # session ใหม่ต้องไม่มีคำถามค้างจากรอบก่อน

        # เขียนสำเนา queue_state ลง DB ด้วย (คอลัมน์ sessions.queue_state) — ถ้า
        # backend restart กลาง session นี้ จะโหลดกลับเข้า memory ได้ตอน boot
        # แทนที่จะ fallback ไปใช้ ALPL ตัวแรกผิดๆ ตลอดที่เหลือ (ดู create_measurement
        # และ lifespan())
        with db.cursor() as cur:
            cur.execute(
                "UPDATE sessions SET queue_state = %s WHERE session_id = %s",
                (json.dumps(queue_state), session_id),
            )

        # 4) Notify Agent ให้เริ่มวัด
        await _notify_agent_start(session_id, template_name, target_count, first_alpl)

        await push_event(
            "session_started",
            {
                "session_id": session_id,
                "number_alpl": first_alpl,
                "template_name": template_name,
                "target_count": target_count,
            },
        )
        return {"session_id": session_id, "template_name": template_name, "target_count": target_count}
    finally:
        db.close()


@app.post("/api/session/stop")
async def stop_session(req: StopSessionRequest):
    """หยุด session ที่กำลัง running จากปุ่ม Stop บน dashboard

    ทำไมเรื่องนี้สำคัญ: นี่คือ path "web-initiated stop" — มันอัปเดต DB
    (state='stopped', ended_at=NOW()) แล้วบอก Agent ให้หยุด ซึ่งต่างจากปุ่ม
    Stop ทางกายภาพที่ MCU (ในการ implement ปัจจุบันของ Agent) ที่แค่ flip
    flag ใน memory ฝั่ง Agent โดยไม่แตะ DB เลย — เป็นความไม่สมดุล (asymmetry)
    ที่รู้กันอยู่ระหว่าง stop ทั้ง 2 path นี้
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT state FROM sessions WHERE session_id = %s", (req.session_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Session not found")
            cur.execute(
                "UPDATE sessions SET state = 'stopped', ended_at = NOW() "
                "WHERE session_id = %s",
                (req.session_id,),
            )

        async with httpx.AsyncClient() as client:
            try:
                await client.post(
                    f"{AGENT_BASE_URL}/command", json={"action": "stop"}, timeout=10
                )
            except Exception as exc:
                log.warning("Agent stop notify failed: %s", exc)

        session_queues.pop(req.session_id, None)  # กดหยุดเองก่อนคิวหมด ก็เคลียร์ memory ทิ้งด้วย
        measure_timeouts.pop(req.session_id, None)  # กันคำถามค้างจาก session ที่จบไปแล้ว
        await push_event("session_stopped", {"session_id": req.session_id})
        return {"ok": True}
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Measure timeout — Pi ไม่ได้รับค่าการวัดกลับมา
# ══════════════════════════════════════════════════════════════════════════════
# ที่มา: send_command.py ส่ง T1 แล้ว "รอยืนยัน" ว่าค่าถูกบันทึกจริงไหม โดยดูว่า
# sessions.measured_count ขยับขึ้นหรือเปล่า (poll /api/session/state) ถ้าครบเวลา
# แล้วยังไม่ขยับ แปลว่าอย่างใดอย่างหนึ่ง:
#   - TM-X วัดไม่ติด (ส่ง -9999.999 มา ซึ่ง Recieve_tm-x.py ข้ามไม่บันทึก) —
#     จากข้อมูลจริงหน้างานเจอบ่อยมาก (8 ครั้งติดแค่ 1)
#   - FTP มาช้าเกินเวลาที่รอ หรือ Recieve_tm-x.py ไม่ได้รัน/ล่ม
# Pi แยกสองกรณีนี้ไม่ออก จึงโยนให้คนตัดสินใจแทน: แจ้งขึ้นหน้าเว็บว่าไม่ได้รับค่า
# แล้วถามว่าจะวัดต่อหรือหยุด
#
# ทำไมเก็บใน memory ไม่ลง DB: เป็นคำถามที่มีอายุแค่ไม่กี่วินาที ผูกกับ session ที่
# กำลัง running อยู่เท่านั้น ถ้า backend restart กลางคัน Pi จะ poll ไม่เจอแล้ว
# หลุด loop ไปเอง (ดู wait_for_decision ใน send_command.py) — เหมือน session_queues
# ที่ยอมรับข้อจำกัดเดียวกันอยู่แล้ว
measure_timeouts: dict[int, dict] = {}


class MeasureTimeoutRequest(BaseModel):
    session_id: int
    piece: int | None = None
    target: int | None = None


class MeasureTimeoutResolve(BaseModel):
    action: str  # "continue" | "stop"


@app.post("/api/measure-timeout")
async def report_measure_timeout(req: MeasureTimeoutRequest):
    """Pi แจ้งว่ารอค่าการวัดชิ้นนี้จนหมดเวลาแล้วยังไม่มาถึง

    บันทึกคำถามค้างไว้ใน memory แล้ว broadcast ให้หน้าเว็บเด้ง modal ถามผู้ใช้
    ตัว Pi จะ poll GET /api/measure-timeout/{session_id} รอคำตอบต่อเอง
    """
    measure_timeouts[req.session_id] = {
        "piece":    req.piece,
        "target":   req.target,
        "decision": None,
    }
    await push_event(
        "measure_timeout",
        {"session_id": req.session_id, "piece": req.piece, "target": req.target},
    )
    log.warning(
        "Measure timeout: session=%s piece=%s — รอผู้ใช้ตัดสินใจ", req.session_id, req.piece
    )
    return {"ok": True}


@app.get("/api/measure-timeout/{session_id}")
async def get_measure_timeout(session_id: int):
    """Pi poll ตรงนี้รอคำตอบจากผู้ใช้

    decision = None แปลว่ายังไม่มีใครตอบ · "continue" = วัดต่อ · "stop" = หยุด
    ถ้าไม่มี entry เลย (เช่น backend restart หรือ session จบไปแล้ว) คืน
    decision="stop" เพื่อให้ Pi หลุด loop ไม่ค้างรอตลอดกาล
    """
    pending = measure_timeouts.get(session_id)
    if pending is None:
        return {"decision": "stop", "reason": "ไม่พบคำถามค้างของ session นี้"}
    return {"decision": pending["decision"]}


@app.post("/api/measure-timeout/{session_id}/resolve")
async def resolve_measure_timeout(session_id: int, body: MeasureTimeoutResolve):
    """หน้าเว็บส่งคำตอบของผู้ใช้กลับมา

    หมายเหตุ: เคส "stop" ตรงนี้ทำแค่บอก Pi ให้เลิกรอ — ตัวการหยุด session จริง
    (อัปเดต DB, เคลียร์คิว, แจ้ง Agent, broadcast) ยังเป็นหน้าที่ของ
    POST /api/session/stop เหมือนเดิม ซึ่งหน้าเว็บจะเรียกต่อเองทันที
    ตั้งใจไม่รวมสองอย่างเข้าด้วยกัน เพื่อให้เส้นทางการหยุด session มีทางเดียว
    ตลอดทั้งระบบ (กดปุ่ม Stop กับตอบว่าหยุด เดินโค้ดชุดเดียวกันเป๊ะ)
    """
    if body.action not in ("continue", "stop"):
        raise HTTPException(400, "action ต้องเป็น 'continue' หรือ 'stop' เท่านั้น")

    pending = measure_timeouts.get(session_id)
    if pending is None:
        raise HTTPException(404, "ไม่พบคำถามค้างของ session นี้ (อาจหมดอายุไปแล้ว)")

    pending["decision"] = body.action
    log.info("Measure timeout: session=%s ผู้ใช้เลือก %s", session_id, body.action)
    return {"ok": True, "action": body.action}


async def _notify_agent(action: str) -> None:
    """แจ้ง Agent ด้วย action ง่ายๆ ที่ไม่มี payload อื่น (pause/resume/stop)

    ครอบ try/except เหมือน _notify_agent_start — ถ้า Agent ไม่ตอบ ไม่ควรทำให้
    คำสั่งฝั่งเว็บล้มเหลวตามไปด้วย (DB อัปเดตไปแล้ว) แค่ log ไว้พอ
    """
    try:
        async with httpx.AsyncClient() as client:
            await client.post(f"{AGENT_BASE_URL}/command", json={"action": action}, timeout=10)
    except Exception as exc:
        log.warning("Agent %s notify failed: %s", action, exc)


@app.post("/api/session/pause")
async def pause_session(req: StopSessionRequest):
    """พัก session ชั่วคราว (running → paused)

    ต่างจาก stop ตรงที่ session ยังมีชีวิตอยู่: คิว ALPL ใน session_queues ยัง
    อยู่ครบ ตำแหน่งปัจจุบันไม่ถูกรีเซ็ต กด Start อีกทีก็วัดต่อจากตัวเดิมได้เลย
    Agent จะหยุดรอที่ชิ้นถัดไปโดยไม่ trigger การวัดใหม่จนกว่าจะได้ resume
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT state FROM sessions WHERE session_id = %s", (req.session_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Session not found")
            if row["state"] != "running":
                raise HTTPException(400, f"พัก session ไม่ได้ — สถานะปัจจุบันคือ '{row['state']}'")
            cur.execute(
                "UPDATE sessions SET state = 'paused' WHERE session_id = %s", (req.session_id,)
            )

        await _notify_agent("pause")
        await push_event("session_paused", {"session_id": req.session_id})
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/session/resume")
async def resume_session(req: StopSessionRequest):
    """สั่งวัดต่อจากที่พักไว้ (paused → running)

    อัปเดต last_seen = NOW() ไปพร้อมกันด้วย — สำคัญมาก เพราะถ้าพักไว้นานกว่า
    HEARTBEAT_TIMEOUT แล้ว last_seen ยังค้างเป็นเวลาเดิม พอกลับมา running
    heartbeat_checker จะเห็นว่า "เงียบเกินกำหนด" แล้ว mark เป็น timeout ทันที
    ทั้งที่ Agent ยังทำงานปกติ
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT state FROM sessions WHERE session_id = %s", (req.session_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Session not found")
            if row["state"] != "paused":
                raise HTTPException(400, f"สั่งวัดต่อไม่ได้ — สถานะปัจจุบันคือ '{row['state']}'")
            cur.execute(
                "UPDATE sessions SET state = 'running', last_seen = NOW() WHERE session_id = %s",
                (req.session_id,),
            )

        await _notify_agent("resume")
        await push_event("session_resumed", {"session_id": req.session_id})
        return {"ok": True}
    finally:
        db.close()


class HeartbeatRequest(BaseModel):
    session_id: Optional[int] = None


@app.post("/api/heartbeat")
async def heartbeat(req: HeartbeatRequest):
    """รับ heartbeat จาก Agent (ดู agent.py heartbeat_loop — ยิงมาทุก
    HEARTBEAT_INTERVAL วิ ไม่ว่าจะมี session running อยู่หรือไม่)

    ถ้าไม่มี session_id (Agent ยัง idle ไม่มีงานอยู่) แค่ตอบ ok เฉยๆ ไม่ต้องแตะ DB
    ถ้ามี session_id จะอัปเดต sessions.last_seen = NOW() ให้ heartbeat_checker()
    เอาไปเทียบว่า session นี้ยังมี Agent ส่งสัญญาณชีพอยู่ไหม — เงื่อนไข
    `state IN ('running','paused')` กันไม่ให้ heartbeat ที่มาช้า/ค้างจาก session เก่าที่จบไป
    แล้วไปอัปเดต last_seen ของ session ผิดตัว
    """
    if req.session_id is None:
        return {"ok": True}
    db = get_db()
    try:
        with db.cursor() as cur:
            # รับ heartbeat ตอน 'paused' ด้วย — Agent ยังมีชีวิตอยู่ระหว่างพัก
            # ถ้าไม่อัปเดต last_seen ช่วงนี้ พอ resume กลับมาจะโดน mark timeout
            # ทันทีเพราะ last_seen ค้างเป็นเวลาก่อนพัก (heartbeat_checker ไม่แตะ
            # session ที่ paused อยู่แล้ว จึงไม่มีผลข้างเคียงอะไร)
            cur.execute(
                "UPDATE sessions SET last_seen = NOW() "
                "WHERE session_id = %s AND state IN ('running', 'paused')",
                (req.session_id,),
            )
    finally:
        db.close()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# Lookup endpoints (dropdown data สำหรับ index.html / edit.html)
# ══════════════════════════════════════════════════════════════════════════════
# Dropdown ทุกตัวนี้เป็นแบบ "ปิด" (closed) — frontend เลือกได้เฉพาะค่าที่มีอยู่
# จริงใน DB เท่านั้น ไม่มีช่องพิมพ์เพิ่มค่าใหม่ในฟอร์ม ถ้าต้องเพิ่ม
# owner/vendor/handler/operator ใหม่ ต้อง insert ตรงเข้า DB เอง
# (ตามที่คุยกันไว้ — ไม่ทำ "add new" inline ในฟอร์ม)
@app.get("/api/operators")
async def list_operators():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT operator_id, operator_name FROM operator ORDER BY operator_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/owners")
async def list_owners():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT owner_id, owner_name FROM owner ORDER BY owner_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/vendors")
async def list_vendors():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT vendor_id, vendor_name FROM vendor ORDER BY vendor_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/handlers")
async def list_handlers():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT handler_id, handler_name FROM handler ORDER BY handler_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/package-sizes")
async def list_package_sizes():
    """คืนรายการ package_size ทั้งหมด พร้อม template_name — ใช้เติม datalist
    ของช่อง Package Size ใน index.html/edit.html

    schema: nominal/tolerance ไม่ได้อยู่ที่ package_size แล้ว — ย้ายไปอยู่ที่
    part_number (1 part_number ผูก 1 ชุด nominal/tolerance ของตัวเอง เพราะ
    part_number เดียวกันอาจมี tolerance ต่างกันได้แม้ package_size เดียวกัน)
    ดู GET /api/part-numbers สำหรับรายละเอียดพวกนี้แทน
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT ps.package_size_id, ps.package_size, t.template_name "
                "FROM package_size ps "
                "LEFT JOIN template t ON ps.template_id = t.template_id "
                "ORDER BY ps.package_size"
            )
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/part-numbers")
async def list_part_numbers(package_size: str = Query(..., min_length=1)):
    """คืนรายการ part_number (จากตาราง catalog part_number ตรงๆ ไม่ใช่ derive
    จาก parts_specifications ที่เคยลงทะเบียนแล้ว) ที่ผูกกับ package_size ที่
    ระบุ — ใช้เป็น dropdown ของช่อง Part Number ที่ cascade จาก Package Size
    ในฟอร์ม Part Entry (New/Rework/IPM) เป็น dropdown "ปิด" เหมือน
    operator/vendor/handler/owner/package_size — เลือกได้เฉพาะ part_number ที่
    มีอยู่จริงในตาราง part_number เท่านั้น ไม่มีช่องพิมพ์เพิ่มค่าใหม่ (handler
    ของ part นั้นๆ ผูกมากับ part_number อยู่แล้ว ไม่ต้องให้ผู้ใช้เลือก Handler
    เองอีกที)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT pn.part_number_name FROM part_number pn "
                "JOIN package_size ps ON pn.package_size_id = ps.package_size_id "
                "WHERE ps.package_size = %s "
                "ORDER BY pn.part_number_name",
                (package_size,),
            )
            return [row["part_number_name"] for row in cur.fetchall()]
    finally:
        db.close()


@app.get("/api/templates")
async def list_templates():
    """คืนรายการ template ทั้งหมด — ใช้โดย Database Editor (edit.html) ตอน
    จัดการตาราง template (Add/Rename/Delete) และตอนสร้าง/แก้ package_size
    (เลือก template ที่จะผูกให้)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT template_id, template_name FROM template ORDER BY template_name")
            return cur.fetchall()
    finally:
        db.close()


@app.get("/api/part-numbers/all")
async def list_all_part_numbers():
    """คืนรายการ part_number ทั้งหมดพร้อมรายละเอียดครบ (package_size/handler/
    nominal/tolerance) — ใช้โดย Database Editor (edit.html) ต่างจาก
    GET /api/part-numbers (คืนแค่ชื่อ กรองด้วย package_size — ใช้เป็น cascading
    dropdown ของฟอร์ม Part Entry)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT pn.part_number_id, pn.part_number_name, ps.package_size, "
                "h.handler_name AS handler, pn.nominal_x, pn.nominal_y, "
                "pn.upper_tol, pn.lower_tol, pn.offset_tol "
                "FROM part_number pn "
                "JOIN package_size ps ON pn.package_size_id = ps.package_size_id "
                "JOIN handler h       ON pn.handler_id = h.handler_id "
                "ORDER BY pn.part_number_name"
            )
            return cur.fetchall()
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Lookup table management (Database Editor — Add/Rename/Delete)
# ══════════════════════════════════════════════════════════════════════════════
# operator/owner/vendor/handler/template เป็นตารางรูปแบบเดียวกันหมด (id + ชื่อ
# ตัวเดียว) เลยใช้ helper function ชุดเดียวกัน 3 ตัวนี้ร่วมกันได้ทั้งหมด แทนที่
# จะเขียน insert/rename/delete แยกกันซ้ำๆ ทีละตาราง — ยังคง endpoint แยกกัน
# ตามตารางเหมือนเดิม (ไม่ทำ dynamic routing) เพื่อให้ URL คาดเดาได้ตรงไปตรงมา
def _create_lookup(table: str, name_col: str, name: str) -> int:
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(f"INSERT INTO {table} ({name_col}) VALUES (%s)", (name,))
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่มไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return cur.lastrowid
    finally:
        db.close()


def _rename_lookup(table: str, id_col: str, name_col: str, id_value: int, name: str) -> None:
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(f"UPDATE {table} SET {name_col} = %s WHERE {id_col} = %s", (name, id_value))
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไขชื่อไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "ไม่พบข้อมูล")
    finally:
        db.close()


# ชื่อตารางจริงในหน้าจอ → ชื่อที่อ่านง่ายสำหรับ error message (ผู้ใช้ไม่ควรเห็น
# ชื่อตารางดิบๆ อย่าง "parts_specifications" ในข้อความแจ้งเตือน)
_TABLE_DISPLAY_NAME = {
    "parts_specifications": "Part",
    "measurements":         "Measurement",
    "part_number":          "Part Number",
    "package_size":         "Package Size",
}


def _delete_lookup(table: str, id_col: str, id_value: int, references: List[tuple],
                   kind: Optional[str] = None) -> None:
    """ลบ row ของตาราง lookup — ปฏิเสธถ้ายังมีตารางอื่นอ้างอิง id นี้อยู่จริง
    (`references` คือ list ของ (referencing_table, referencing_col)) เพื่อไม่ให้
    Part/Measurement ที่มีอยู่แล้วกลายเป็นข้อมูลกำพร้า (orphan FK) — แนะนำให้
    "เปลี่ยนชื่อ" (rename) แทนถ้ายังมีข้อมูลผูกอยู่ ไม่ใช่ลบทิ้งแล้วสร้างใหม่
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            for ref_table, ref_col in references:
                cur.execute(f"SELECT 1 FROM {ref_table} WHERE {ref_col} = %s LIMIT 1", (id_value,))
                if cur.fetchone():
                    display_name = _TABLE_DISPLAY_NAME.get(ref_table, ref_table)
                    raise HTTPException(
                        409,
                        f"ลบไม่ได้ — ยังมีข้อมูลใน {display_name} ที่อ้างอิงถึงอยู่ "
                        f"กรุณาเปลี่ยนชื่อแทนถ้าต้องการแก้ไข",
                    )
            # ── สำรองก่อนลบ ── ต้องอ่านตอนแถวยังอยู่ และหลังผ่านด่านเช็ค FK แล้ว
            # (ถ้าเช็คไม่ผ่านจะ raise 409 ไปก่อน ไม่มีอะไรถูกลบ ไม่ต้องสำรอง)
            row = _fetch_one(cur, f"SELECT * FROM {table} WHERE {id_col} = %s", (id_value,))
            if row is None:
                raise HTTPException(404, "ไม่พบข้อมูล")
            _archive_before_delete(
                kind=kind or table, table=table, pk={id_col: id_value}, row=row
            )

            cur.execute(f"DELETE FROM {table} WHERE {id_col} = %s", (id_value,))
            if cur.rowcount == 0:
                raise HTTPException(404, "ไม่พบข้อมูล")
    finally:
        db.close()


class LookupCreate(BaseModel):
    name: str


class LookupUpdate(BaseModel):
    name: str


@app.post("/api/operators", status_code=201)
async def create_operator(body: LookupCreate):
    return {"operator_id": _create_lookup("operator", "operator_name", body.name)}


@app.patch("/api/operators/{operator_id}")
async def rename_operator(operator_id: int, body: LookupUpdate):
    _rename_lookup("operator", "operator_id", "operator_name", operator_id, body.name)
    return {"ok": True}


@app.delete("/api/operators/{operator_id}")
async def delete_operator(operator_id: int):
    # operator ถูกอ้างอิงจาก measurements.operator_id เท่านั้น
    _delete_lookup("operator", "operator_id", operator_id, [("measurements", "operator_id")])
    return {"ok": True}


@app.post("/api/owners", status_code=201)
async def create_owner(body: LookupCreate):
    return {"owner_id": _create_lookup("owner", "owner_name", body.name)}


@app.patch("/api/owners/{owner_id}")
async def rename_owner(owner_id: int, body: LookupUpdate):
    _rename_lookup("owner", "owner_id", "owner_name", owner_id, body.name)
    return {"ok": True}


@app.delete("/api/owners/{owner_id}")
async def delete_owner(owner_id: int):
    # owner ถูกอ้างอิงจาก parts_specifications.owner_id เท่านั้น
    _delete_lookup("owner", "owner_id", owner_id, [("parts_specifications", "owner_id")])
    return {"ok": True}


@app.post("/api/vendors", status_code=201)
async def create_vendor(body: LookupCreate):
    return {"vendor_id": _create_lookup("vendor", "vendor_name", body.name)}


@app.patch("/api/vendors/{vendor_id}")
async def rename_vendor(vendor_id: int, body: LookupUpdate):
    _rename_lookup("vendor", "vendor_id", "vendor_name", vendor_id, body.name)
    return {"ok": True}


@app.delete("/api/vendors/{vendor_id}")
async def delete_vendor(vendor_id: int):
    # vendor ถูกอ้างอิงจาก parts_specifications.vendor_id เท่านั้น
    _delete_lookup("vendor", "vendor_id", vendor_id, [("parts_specifications", "vendor_id")])
    return {"ok": True}


@app.post("/api/handlers", status_code=201)
async def create_handler(body: LookupCreate):
    return {"handler_id": _create_lookup("handler", "handler_name", body.name)}


@app.patch("/api/handlers/{handler_id}")
async def rename_handler(handler_id: int, body: LookupUpdate):
    _rename_lookup("handler", "handler_id", "handler_name", handler_id, body.name)
    return {"ok": True}


@app.delete("/api/handlers/{handler_id}")
async def delete_handler(handler_id: int):
    # handler ถูกอ้างอิงจาก part_number.handler_id เท่านั้น (parts_specifications
    # ไม่มี handler_id ตรงๆ แล้ว — derive ผ่าน part_number)
    _delete_lookup("handler", "handler_id", handler_id, [("part_number", "handler_id")])
    return {"ok": True}


@app.post("/api/templates", status_code=201)
async def create_template(body: LookupCreate):
    return {"template_id": _create_lookup("template", "template_name", body.name)}


@app.patch("/api/templates/{template_id}")
async def rename_template(template_id: int, body: LookupUpdate):
    _rename_lookup("template", "template_id", "template_name", template_id, body.name)
    return {"ok": True}


@app.delete("/api/templates/{template_id}")
async def delete_template(template_id: int):
    # template ถูกอ้างอิงจาก package_size.template_id เท่านั้น
    _delete_lookup("template", "template_id", template_id, [("package_size", "template_id")])
    return {"ok": True}


class PackageSizeCreate(BaseModel):
    package_size:  str
    template_name: Optional[str] = None


class PackageSizeUpdate(BaseModel):
    package_size:  Optional[str] = None
    template_name: Optional[str] = None


@app.post("/api/package-sizes", status_code=201)
async def create_package_size(body: PackageSizeCreate):
    db = get_db()
    try:
        with db.cursor() as cur:
            template_id = _lookup_id(cur, "template", "template_id", "template_name", body.template_name)
            try:
                cur.execute(
                    "INSERT INTO package_size (package_size, template_id) VALUES (%s, %s)",
                    (body.package_size, template_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่ม Package Size ไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"package_size_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/package-sizes/{package_size_id}")
async def update_package_size(package_size_id: int, body: PackageSizeUpdate):
    db = get_db()
    try:
        with db.cursor() as cur:
            set_parts, values = [], []
            if body.package_size is not None:
                set_parts.append("package_size = %s")
                values.append(body.package_size)
            if body.template_name is not None:
                template_id = _lookup_id(cur, "template", "template_id", "template_name", body.template_name)
                set_parts.append("template_id = %s")
                values.append(template_id)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            try:
                cur.execute(
                    f"UPDATE package_size SET {', '.join(set_parts)} WHERE package_size_id = %s",
                    (*values, package_size_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไข Package Size ไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "Package Size not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/package-sizes/{package_size_id}")
async def delete_package_size(package_size_id: int):
    # package_size ถูกอ้างอิงจาก part_number.package_size_id เท่านั้น
    _delete_lookup("package_size", "package_size_id", package_size_id, [("part_number", "package_size_id")])
    return {"ok": True}


class PartNumberCreate(BaseModel):
    part_number_name: str
    package_size:     str
    handler:          str
    nominal_x:        float
    nominal_y:        float
    upper_tol:        float
    lower_tol:        float
    # มี default 0 เพื่อให้ payload เก่าที่ยังไม่ส่ง offset_tol มา ยัง POST ผ่านได้
    offset_tol:       float = 0


class PartNumberUpdate(BaseModel):
    part_number_name: Optional[str] = None
    package_size:     Optional[str] = None
    handler:          Optional[str] = None
    nominal_x:        Optional[float] = None
    nominal_y:        Optional[float] = None
    upper_tol:        Optional[float] = None
    lower_tol:        Optional[float] = None
    offset_tol:       Optional[float] = None


@app.post("/api/part-numbers", status_code=201)
async def create_part_number(body: PartNumberCreate):
    db = get_db()
    try:
        with db.cursor() as cur:
            package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", body.package_size)
            handler_id      = _lookup_id(cur, "handler",      "handler_id",      "handler_name",  body.handler)
            if package_size_id is None or handler_id is None:
                raise HTTPException(400, "ต้องระบุ package_size และ handler ที่มีอยู่จริงในระบบ")
            try:
                cur.execute(
                    "INSERT INTO part_number "
                    "(part_number_name, package_size_id, handler_id, nominal_x, nominal_y, "
                    " upper_tol, lower_tol, offset_tol) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    (
                        body.part_number_name, package_size_id, handler_id,
                        body.nominal_x, body.nominal_y, body.upper_tol, body.lower_tol,
                        body.offset_tol,
                    ),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"เพิ่ม Part Number ไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"part_number_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/part-numbers/{part_number_id}")
async def update_part_number(part_number_id: int, body: PartNumberUpdate):
    db = get_db()
    try:
        with db.cursor() as cur:
            set_parts, values = [], []
            if body.part_number_name is not None:
                set_parts.append("part_number_name = %s")
                values.append(body.part_number_name)
            if body.package_size is not None:
                package_size_id = _lookup_id(cur, "package_size", "package_size_id", "package_size", body.package_size)
                set_parts.append("package_size_id = %s")
                values.append(package_size_id)
            if body.handler is not None:
                handler_id = _lookup_id(cur, "handler", "handler_id", "handler_name", body.handler)
                set_parts.append("handler_id = %s")
                values.append(handler_id)
            for field_name in ("nominal_x", "nominal_y", "upper_tol", "lower_tol", "offset_tol"):
                value = getattr(body, field_name)
                if value is not None:
                    set_parts.append(f"{field_name} = %s")
                    values.append(value)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            try:
                cur.execute(
                    f"UPDATE part_number SET {', '.join(set_parts)} WHERE part_number_id = %s",
                    (*values, part_number_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"แก้ไข Part Number ไม่สำเร็จ (ชื่อใหม่นี้อาจมีอยู่แล้ว): {exc}")
            if cur.rowcount == 0:
                raise HTTPException(404, "Part Number not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/part-numbers/{part_number_id}")
async def delete_part_number(part_number_id: int):
    # part_number ถูกอ้างอิงจาก parts_specifications.part_number_id เท่านั้น
    _delete_lookup("part_number", "part_number_id", part_number_id, [("parts_specifications", "part_number_id")])
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# Parts endpoints
# ══════════════════════════════════════════════════════════════════════════════
# SELECT ที่ join parts_specifications กับทุกตาราง lookup ไว้ในที่เดียว — ใช้
# ร่วมกันทั้ง list_parts และ get_part เพื่อให้ response มีทั้งชื่อ (handler/
# vendor/owner/package_size) และรายละเอียดของ part_number (nominal/tolerance/
# template_name) ไม่ใช่แค่ id เปล่าๆ ที่ frontend เอาไปแสดงตรงๆ ไม่ได้ รวม
# recieve_date ด้วย — ใช้ prefill ฟอร์ม Rework (auto-fill ข้อมูลเดิมของ ALPL
# ที่กรอกกลับเข้ามา ยกเว้น recieve_date ที่ต้องเว้นว่างให้กรอกใหม่)
#
# schema: handler/package_size/nominal/tolerance/template_name ทั้งหมด derive
# มาจาก part_number_id (ไม่ได้เก็บตรงที่ parts_specifications เอง) — join ทอด
# parts_specifications -> part_number -> handler / package_size -> template
PARTS_SELECT = """
    SELECT p.part_id, p.number_alpl, pn.part_number_name AS part_number,
           p.description, p.po_number,
           p.recieve_date   AS recieve_date,
           h.handler_name   AS handler,
           v.vendor_name    AS vendor,
           o.owner_name     AS owner,
           ps.package_size  AS package_size,
           pn.nominal_x     AS nominal_x,
           pn.nominal_y     AS nominal_y,
           pn.upper_tol     AS upper_tol,
           pn.lower_tol     AS lower_tol,
           pn.offset_tol    AS offset_tol,
           t.template_name  AS template_name
    FROM parts_specifications p
    LEFT JOIN part_number pn  ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h       ON pn.handler_id = h.handler_id
    LEFT JOIN vendor v        ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o         ON p.owner_id = o.owner_id
    LEFT JOIN package_size ps ON pn.package_size_id = ps.package_size_id
    LEFT JOIN template t      ON ps.template_id = t.template_id
"""


def _lookup_id(cur, table: str, id_col: str, name_col: str, value: Optional[str]) -> Optional[int]:
    """แปลงชื่อ (เช่น vendor_name ที่ frontend ส่งมาจาก dropdown) เป็น id
    (เช่น vendor_id) จากตาราง lookup ที่เกี่ยวข้อง

    ทำไม: dropdown ฝั่ง frontend (Operator/Owner/Vendor/Handler/
    Package Size) เป็น dropdown "ปิด" — เลือกได้เฉพาะค่าที่มีอยู่แล้วใน DB
    เท่านั้น ไม่มีช่องพิมพ์ค่าใหม่ ดังนั้นค่าที่ส่งเข้ามาควรมีอยู่จริงเสมอ แต่ยัง
    defensive เช็คไว้กันกรณี frontend ค้างข้อมูลเก่า/ผิดพลาด — ถ้าหาไม่เจอ
    ให้ 400 ชัดเจนแทนที่จะปล่อยให้ FK constraint error กลายเป็น 500 ตอน insert
    """
    if value in (None, ""):
        return None
    cur.execute(f"SELECT {id_col} FROM {table} WHERE {name_col} = %s", (value,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(400, f"ไม่พบค่า '{value}' ใน {table} (เลือกจาก dropdown เท่านั้น)")
    return row[id_col]


def _block_if_session_running(cur, action: str) -> None:
    """เช็คว่ามี session ไหนกำลัง running อยู่ไหม — ถ้ามี ปฏิเสธการแก้ไข/ลบ Part
    หรือ Measurement ทันที (ทั้งคู่ ไม่ว่า ALPL ไหน) เพราะข้อมูลกำลังถูกวัดอยู่จริง

    ทำไมต้อง block กว้างขนาดนี้ (ไม่ใช่แค่ ALPL ที่กำลังวัดอยู่): ผู้ใช้เลือกไว้
    ชัดเจนว่าอยากให้ Edit/Delete กดไม่ได้เลยทั้ง Part และ Measurement ตราบใดที่
    ยังมี session running อยู่ — เพื่อความคาดเดาได้ง่าย ไม่ต้องตามว่า ALPL ไหน
    "ปลอดภัย" ไหม (ระบบรัน session พร้อมกันได้แค่ 1 อันเสมออยู่แล้ว — ดู
    Button Guard ใน start_session — เช็คแค่ "มี session running อยู่ไหม" จึง
    เทียบเท่ากับ "session ปัจจุบันกำลังวัดอยู่ไหม")
    """
    # นับ 'paused' ด้วย — session ที่พักไว้ยังวัดต่อได้อยู่ ข้อมูลจึงยังไม่นิ่ง
    cur.execute("SELECT 1 FROM sessions WHERE state IN ('running', 'paused') LIMIT 1")
    if cur.fetchone():
        raise HTTPException(
            409,
            f"ไม่สามารถ{action}ข้อมูลได้ขณะนี้ — กำลังมีการวัดอยู่ (session running/paused) "
            f"กรุณากด Stop ก่อนแล้วค่อยแก้ไข",
        )


class PartCreate(BaseModel):
    # schema: parts_specifications ไม่เก็บ handler/package_size/nominal/
    # tolerance/template_name ตรงๆ เลย — ทั้งหมด derive มาจาก part_number_id
    # ตัวเดียว (ดู init.sql: part_number ผูก package_size_id + handler_id +
    # nominal/tolerance ของตัวเองไว้แล้ว) เลือก part_number ก็ map ค่าพวกนี้
    # ให้อัตโนมัติหมด ไม่ต้องส่ง handler/package_size แยกมาที่ endpoint นี้อีก
    # (frontend ยังส่ง package_size มาด้วยเพื่อใช้ cascade dropdown Part Number
    # เฉยๆ — เป็นแค่ field ส่วนเกินที่ endpoint นี้เพิกเฉยไม่ได้ใช้)
    #
    # part_number เป็น Optional เพราะตอน IPM เจอ ALPL ที่ยังไม่เคยลงทะเบียน
    # (ดู POST /api/session/start) จะลงทะเบียน part ใหม่แบบขั้นต่ำผ่าน endpoint
    # นี้ — อาจมีแค่ number_alpl เท่านั้น ยังไม่รู้ part_number จริง (แต่ถ้าไม่รู้
    # part_number จะหา template_name ไม่เจอตอน start_session — ต้องตั้งให้
    # ครบก่อนถึงจะเริ่มวัดได้จริง)
    #
    # vendor/owner ยังเป็น FK ไป lookup table เหมือนเดิม — frontend ส่งมาเป็น
    # "ชื่อ" (string จาก dropdown) แล้ว backend resolve เป็น id เอง (ดู _lookup_id)
    #
    # recieve_date เป็น Optional — ถ้าไม่ส่งมา/ส่งค่าว่าง จะถูกบันทึกเป็น NULL
    # ตรงๆ (เว้นว่างแล้วว่างจริง ไม่เติมวันที่ปัจจุบันให้อัตโนมัติ — ดู
    # _insert_part_row)
    number_alpl:   int
    part_number:   Optional[str] = None
    description:   Optional[str] = None
    vendor:        Optional[str] = None
    po_number:     Optional[int] = None
    package_size:  Optional[str] = None  # ไม่ได้ใช้ resolve อะไรที่นี่ (เก็บไว้เผื่อ frontend ส่งมา)
    owner:         Optional[str] = None
    recieve_date:  Optional[str] = None


@app.get("/api/parts")
async def list_parts(
    limit:  int = Query(10, ge=1, le=1000),   # มีเพดาน กันยิง limit=999999 ดึงทั้งตาราง
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
):
    """คืน config ของ parts แบบ "แบ่งหน้า" (server-side pagination)

    ทำไมเปลี่ยนจากเดิมที่คืน array ของ parts ทั้งหมดทีเดียว มาเป็น object
    {items, total}: ตาราง parts มีแนวโน้มโตขึ้นเรื่อยๆ ตามการใช้งานจริง การ
    ดึงมาทั้งหมดทุกครั้งจะกินแบนด์วิดท์และหน่วยความจำ frontend โดยเปล่าประโยชน์
    จึงให้ดึงมาทีละหน้า (limit/offset) แทน — frontend ของ edit.html แสดงทีละ 10 แถว

    `total` มาจาก COUNT(*) ที่ใช้ WHERE ชุดเดียวกับ query หลัก (ไม่ใช่นับจาก
    items ของหน้าปัจจุบัน) เพื่อให้ frontend คำนวณจำนวนหน้า/ปิดปุ่ม Next ได้ถูก

    `search` (optional) — กรองด้วย number_alpl หรือ part_number แบบ LIKE
    เหตุผลที่ทำ search ฝั่ง server ไม่ใช่ฝั่ง client: เพื่อให้ค้นหาเจอ "ข้ามทุก
    หน้า" ไม่ใช่แค่ 10 แถวของหน้าที่โหลดมาแสดงอยู่ number_alpl เป็น INT จึงต้อง
    CAST เป็น CHAR ก่อนเทียบ LIKE เพื่อให้ค้นหาบางส่วนของตัวเลขได้ (เช่นพิมพ์
    "10" แล้วเจอทั้ง 1011, 1002 ที่ขึ้นต้นด้วย 10)

    หมายเหตุ: เพิ่ม ORDER BY number_alpl เพื่อให้ลำดับของหน้าคงที่ (stable) —
    ถ้าไม่กำหนด ORDER BY การไล่ LIMIT/OFFSET อาจได้ลำดับไม่แน่นอนข้ามหน้า
    """
    conditions, params = [], []
    if search:
        conditions.append("(CAST(p.number_alpl AS CHAR) LIKE %s OR pn.part_number_name LIKE %s)")
        like = f"%{search}%"
        params.extend([like, like])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) AS total FROM parts_specifications p "
                f"LEFT JOIN part_number pn ON p.part_number_id = pn.part_number_id {where}",
                params,
            )
            total = cur.fetchone()["total"]
            cur.execute(
                f"{PARTS_SELECT} {where} ORDER BY p.number_alpl LIMIT %s OFFSET %s",
                (*params, limit, offset),
            )
            items = cur.fetchall()
        return {"items": items, "total": total}
    finally:
        db.close()


@app.get("/api/parts/{part_id}")
async def get_part(part_id: int):
    """คืน config ของ part 1 ตัวตาม ALPL number (รวมชื่อ handler/vendor/
    owner/package_size ที่ join มาจาก lookup table แล้ว ไม่ใช่แค่ id)"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"{PARTS_SELECT} WHERE p.number_alpl = %s", (part_id,))
            row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Part not found")
        return row
    finally:
        db.close()


def _insert_part_row(cur, number_alpl: int, config: Dict[str, Any]) -> None:
    """Insert 1 row ลง table `parts_specifications` โดยใช้ number_alpl ที่ระบุ
    + field อื่นจาก `config` (dict ของ field part_number/vendor/owner ฯลฯ) —
    part_number/vendor/owner รับมาเป็น "ชื่อ" (ตรงกับค่าที่เลือกจาก dropdown
    ฝั่ง frontend) แล้ว resolve เป็น id ก่อน insert

    หมายเหตุ: handler/package_size ไม่ได้ resolve/insert ตรงนี้แล้ว เพราะ
    schema ใหม่ derive ค่าพวกนี้มาจาก part_number_id ทั้งหมด (part_number
    catalog ผูก package_size_id + handler_id ของตัวเองไว้แล้ว — ดู init.sql)

    ใช้ร่วมกันทั้งจาก endpoint POST /api/parts ปกติ, จาก flow ของ New queue ที่
    ALPL หลายตัวใช้ config เดียวกันซ้ำ, และจากการลงทะเบียน part แบบขั้นต่ำตอน
    IPM เจอ ALPL ที่ยังไม่เคยลงทะเบียน (ดู start_session / create_measurement)
    """
    part_number_id = _lookup_id(cur, "part_number", "part_number_id", "part_number_name", config.get("part_number"))
    vendor_id      = _lookup_id(cur, "vendor",      "vendor_id",      "vendor_name",      config.get("vendor"))
    owner_id       = _lookup_id(cur, "owner",       "owner_id",       "owner_name",       config.get("owner"))

    # recieve_date: ใส่คอลัมน์นี้ใน INSERT เสมอ แม้จะเป็นค่าว่าง (จะได้ NULL) —
    # ตั้งใจให้ "เว้นว่างแล้วว่างจริง" ไม่ใช่เติมวันที่ปัจจุบันให้อัตโนมัติ
    # (ถ้าไม่ใส่คอลัมน์นี้เลย DEFAULT CURRENT_TIMESTAMP ของ schema จะทำงานแทน
    # ซึ่งไม่ใช่พฤติกรรมที่ต้องการ)
    columns = [
        "number_alpl", "part_number_id", "description",
        "vendor_id", "po_number", "owner_id", "recieve_date",
    ]
    values: List[Any] = [
        number_alpl, part_number_id, config.get("description"),
        vendor_id, config.get("po_number"), owner_id,
        config.get("recieve_date") or None,
    ]

    placeholders = ", ".join(["%s"] * len(values))
    cur.execute(
        f"INSERT INTO parts_specifications ({', '.join(columns)}) VALUES ({placeholders})",
        tuple(values),
    )


def _update_part_row(cur, number_alpl: int, config: Dict[str, Any]) -> None:
    """Update 1 row ที่มีอยู่แล้วใน table `parts_specifications` (ตรงข้ามกับ
    _insert_part_row)

    ใช้เฉพาะกรณี Rework: ALPL นี้เคยผ่าน New มาแล้ว (มี Part row อยู่จริง) แต่
    งานไม่ผ่าน ถูกส่งไป Rework แล้วส่งกลับมาวัดใหม่ — ฟอร์ม Rework auto-fill
    ทุกช่องด้วยข้อมูลเดิมของ ALPL นี้ไว้ให้แล้ว (ดู GET /api/parts/{part_id})
    ยกเว้น recieve_date ที่ผู้ใช้ต้องกรอกวันที่รับกลับมาใหม่เอง — Save จึง
    เขียนทับ row เดิม (WHERE number_alpl = %s) แทนที่จะ insert row ใหม่ซ้อน
    ขึ้นมา ซึ่งจะชนกับ UNIQUE constraint ของ number_alpl ทันที
    """
    part_number_id = _lookup_id(cur, "part_number", "part_number_id", "part_number_name", config.get("part_number"))
    vendor_id      = _lookup_id(cur, "vendor",      "vendor_id",      "vendor_name",      config.get("vendor"))
    owner_id       = _lookup_id(cur, "owner",       "owner_id",       "owner_name",       config.get("owner"))
    cur.execute(
        "UPDATE parts_specifications SET part_number_id = %s, description = %s, "
        "vendor_id = %s, po_number = %s, owner_id = %s, "
        "recieve_date = %s "
        "WHERE number_alpl = %s",
        (
            part_number_id, config.get("description"),
            vendor_id, config.get("po_number"), owner_id,
            config.get("recieve_date") or None, number_alpl,
        ),
    )


@app.post("/api/parts", status_code=201)
async def create_part(part: PartCreate):
    """ลงทะเบียน ALPL part ใหม่: nominal X/Y, tolerance แยกแกน, และ template ของ TM-X ที่ใช้

    ทำไมต้องแยก endpoint นี้ออกจาก start_session: parts/templates ถูก config
    ไว้ล่วงหน้า (เป็นขั้นตอน setup) เพื่อให้ start_session แค่ lookup template
    จาก number_alpl ได้เลย ไม่ต้องให้ผู้ปฏิบัติงานพิมพ์เองทุกครั้งที่รัน
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _insert_part_row(cur, part.number_alpl, part.dict())
        return {"number_alpl": part.number_alpl}
    finally:
        db.close()


@app.patch("/api/parts/{part_id}")
async def update_part(part_id: int, data: Dict[str, Any] = Body(...)):
    """อัปเดต config ของ part แบบ partial (เฉพาะ field ที่ส่งมาใน body)

    ทำไมต้องมี whitelist (`allowed`): เพื่อไม่ให้ request body ไปเขียนทับ column
    ที่ไม่ควรแก้ผ่าน endpoint นี้ได้โดยไม่ตั้งใจ (หรือถูกใช้ในทางที่ไม่ดี)

    `number_alpl` แก้ไขได้ (จาก edit.html — Edit Part) แม้จะเป็น "business key"
    หลักที่ sessions/measurements อ้างอิงถึงก็ตาม — ถ้า ALPL ตัวนี้มีประวัติ
    session/measurement ผูกอยู่แล้ว MySQL จะปฏิเสธด้วย FK constraint error
    (เพราะ FOREIGN KEY ไม่มี ON UPDATE CASCADE) เราจับ error นั้นแล้วแปลงเป็น
    409 ที่อ่านง่ายแทนที่จะปล่อยให้เป็น 500 ดิบๆ
    """
    # field ที่แก้ตรงๆ ได้เลย ไม่ต้อง resolve ผ่าน lookup table
    direct_fields = {"number_alpl", "description", "po_number", "recieve_date"}
    # field ที่เป็น "ชื่อ" จาก dropdown — ต้อง resolve เป็น id ก่อน (key ที่รับจาก
    # request → (คอลัมน์จริงใน parts_specifications, ตาราง lookup, id column,
    # name column)) — part_number ย้ายมาอยู่ตรงนี้แล้ว (ไม่ใช่ direct_fields
    # อีกต่อไป) เพราะตอนนี้ต้อง resolve เป็น part_number_id ก่อน ไม่ใช่คอลัมน์
    # VARCHAR ตรงๆ — handler/package_size ไม่มีในนี้แล้ว เพราะ derive มาจาก
    # part_number_id ทั้งคู่ ไม่ได้เก็บที่ parts_specifications โดยตรง
    lookup_fields = {
        "part_number": ("part_number_id", "part_number", "part_number_id", "part_number_name"),
        "vendor":      ("vendor_id",      "vendor",      "vendor_id",      "vendor_name"),
        "owner":       ("owner_id",       "owner",       "owner_id",       "owner_name"),
    }
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "แก้ไข")
            set_parts, values = [], []
            for k, v in data.items():
                if k in direct_fields:
                    set_parts.append(f"{k} = %s")
                    values.append(v)
                elif k in lookup_fields:
                    col, table, id_col, name_col = lookup_fields[k]
                    set_parts.append(f"{col} = %s")
                    values.append(_lookup_id(cur, table, id_col, name_col, v))
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")
            set_clause = ", ".join(set_parts)
            try:
                cur.execute(
                    f"UPDATE parts_specifications SET {set_clause} WHERE number_alpl = %s",
                    (*values, part_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(
                    409,
                    f"บันทึกไม่สำเร็จ — ALPL ใหม่อาจซ้ำกับ part อื่น หรือ ALPL เดิมมี "
                    f"session/measurement ผูกอยู่แล้ว (เปลี่ยน ALPL ที่มีประวัติไม่ได้): {exc}",
                )
            if cur.rowcount == 0:
                raise HTTPException(404, "Part not found")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/parts/{part_id}")
async def delete_part(part_id: int):
    """ลบ Part 1 row ออกจากตาราง `parts_specifications`

    **กฎการลบ (ตามที่ตกลงกันไว้)**: ลบได้ต่อเมื่อ ALPL นี้ "ไม่มีข้อมูลการวัด
    (Measurement) เหลืออยู่เลย" เท่านั้น — ถ้ายังมี Measurement อยู่จะปฏิเสธ
    ทันทีด้วย 409 พร้อมข้อความบอกจำนวนรายการที่ติดอยู่ (frontend เอาไปโชว์เป็น
    popup เตือน) ไม่มีโหมด cascade ที่ลบ Measurement ตามไปด้วยอีกต่อไป —
    ประวัติการวัดถือเป็นข้อมูลจริงที่ห้ามหายไปโดยไม่ตั้งใจ ถ้าผู้ใช้ต้องการลบ
    จริงๆ ต้องไปลบ Measurement ทีละรายการเองที่ตาราง Measurements ก่อน

    ส่วนแถวใน `sessions` ที่อ้างถึง ALPL นี้อยู่ ถือเป็นข้อมูล bookkeeping ล้วนๆ
    (ไม่ใช่ผลการวัด) — พอไม่มี Measurement เหลือแล้ว session พวกนั้นก็ไม่มี
    ความหมายอะไรต่อ จึงลบทิ้งให้อัตโนมัติก่อนลบ Part เพื่อไม่ให้ FK block
    ทำทั้งหมดเป็น transaction เดียว (ปิด autocommit ชั่วคราว) กัน DB ค้างครึ่งๆ
    กลางๆ ถ้ามีขั้นไหนพังกลางทาง
    """
    db = get_db()
    try:
        db.autocommit(False)
        with db.cursor() as cur:
            _block_if_session_running(cur, "ลบ")

            # เช็คก่อนเลยว่ามี Measurement ของ ALPL นี้เหลืออยู่ไหม — ถ้ามี ปฏิเสธ
            # ทันที ไม่แตะอะไรใน DB เลย (ตรวจเองแทนที่จะรอ FK error เพื่อให้ได้
            # ข้อความบอกจำนวนที่ติดอยู่ชัดเจน ผู้ใช้จะได้รู้ว่าต้องไปจัดการอะไรต่อ)
            cur.execute("SELECT COUNT(*) AS n FROM measurements WHERE number_alpl = %s", (part_id,))
            measurement_count = cur.fetchone()["n"]
            if measurement_count:
                db.rollback()
                raise HTTPException(
                    409,
                    f"ลบไม่ได้ — ALPL {part_id} ยังมีข้อมูลการวัด (Measurement) อยู่ "
                    f"{measurement_count} รายการ กรุณาลบ Measurement ของ ALPL นี้ให้หมดก่อน "
                    f"แล้วค่อยลบ Part",
                )

            # ไม่มี Measurement เหลือแล้ว — เคลียร์ session ที่อ้างถึง ALPL นี้ทิ้งได้
            # อย่างปลอดภัย (ไม่มี measurement ตัวไหนชี้มาที่ session พวกนี้แล้ว)
            # ── สำรองก่อนลบ ── ต้องเก็บ sessions ที่จะถูกลบพ่วงไปด้วย ไม่งั้น
            # กู้ Part กลับมาได้แต่ประวัติ session หายไปแล้วกู้ไม่ได้อีก
            part_row = _fetch_one(
                cur, "SELECT * FROM parts_specifications WHERE number_alpl = %s", (part_id,)
            )
            if part_row is None:
                db.rollback()
                raise HTTPException(404, "Part not found")
            cur.execute("SELECT * FROM sessions WHERE number_alpl = %s", (part_id,))
            session_rows = cur.fetchall() or []
            _archive_before_delete(
                kind="part", table="parts_specifications",
                pk={"number_alpl": part_id}, row=part_row,
                related={"sessions": session_rows},
            )

            cur.execute("DELETE FROM sessions WHERE number_alpl = %s", (part_id,))

            try:
                cur.execute("DELETE FROM parts_specifications WHERE number_alpl = %s", (part_id,))
            except pymysql.MySQLError as exc:
                db.rollback()
                raise HTTPException(409, f"ลบ ALPL {part_id} ไม่สำเร็จ: {exc}")
            if cur.rowcount == 0:
                db.rollback()
                raise HTTPException(404, "Part not found")
        db.commit()
        return {"ok": True}
    finally:
        db.autocommit(True)
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Measurements endpoints
# ══════════════════════════════════════════════════════════════════════════════
# SELECT ที่ join measurements กับ operator ไว้ในที่เดียว — ใช้ร่วมกันทั้ง
# list_measurements และ export_csv เพื่อให้ response/CSV มีชื่อ operator (ไม่ใช่
# แค่ operator_id เปล่าๆ) หลังจากย้าย Operator จากคอลัมน์ VARCHAR ตรงๆ ไปเป็น
# FK ชี้ตาราง operator
# nominal/tolerance ที่ join มาด้วย ใช้ให้ frontend คำนวณเองได้ว่า value_x/value_y
# แต่ละตัว "ผ่านหรือไม่ผ่าน" แยกรายแกน (DB เก็บแค่ result รวมของทั้ง 2 แกน) —
# ตาราง Measurement ทั้งหน้า Home และ Edit ใช้ระบายสีเขียว/แดงรายค่า
# หมายเหตุ: CSV export ใช้ SELECT ชุดนี้ด้วย จึงมีคอลัมน์พวกนี้ติดไปในไฟล์ export
# ด้วย (ตั้งใจ — มีประโยชน์ตอนเอาไปวิเคราะห์ต่อใน Power BI/Excel)
MEASUREMENTS_SELECT = """
    SELECT m.*, op.operator_name AS operator_name,
           pn.nominal_x AS nominal_x, pn.nominal_y AS nominal_y,
           pn.upper_tol AS upper_tol, pn.lower_tol AS lower_tol
    FROM measurements m
    LEFT JOIN operator op ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn ON p.part_number_id = pn.part_number_id
"""


class MeasurementCreate(BaseModel):
    # session_id เป็น Optional แล้ว — None หมายถึง "manual add" จากหน้า
    # Database Editor (edit.html ปุ่ม + Add Measurement) ซึ่งไม่มี session ที่
    # Agent กำลัง running อยู่จริงให้อ้างอิงเลย ต่างจาก flow ปกติที่ Agent ส่ง
    # session_id ที่ได้จากตอนเริ่ม session มาด้วยเสมอ (ดู create_measurement)
    session_id:  Optional[int] = None
    number_alpl: int
    value_x:     float
    value_y:     float
    # ค่าที่ 3 จากไฟล์ .txt ของ TM-X (+0000.003) — ความเยื้องของชิ้นงาน
    # default 0 เพื่อให้สคริปต์เก่าที่ยังไม่ส่งฟิลด์นี้มา ยัง POST ผ่านได้เหมือนเดิม
    offset:      float = 0
    note:        Optional[str] = None
    # UUID ที่ Agent สร้างขึ้นต่อการวัด 1 ครั้ง (uuid4) — ส่งมาด้วยทุกครั้งที่มา
    # จาก agent.py (ไม่มีถ้าเป็น manual add จาก edit.html) ใช้กัน insert ซ้ำ
    # ตอน Agent retry POST นี้ (ดู create_measurement)
    client_uuid: Optional[str] = None


class ImageUpdate(BaseModel):
    # image_path เป็น Optional แล้ว — กรณี Agent จัดการรูปไม่สำเร็จ
    # จะ PATCH มาด้วย image_path=None,
    # upload_failed=True แทน เพื่อให้ backend รู้ว่า "พยายามแล้วแต่ไม่สำเร็จ"
    # ต่างจาก "ยังไม่เคยพยายามเลย" (NULL เฉยๆ ตอน insert)
    image_path:    Optional[str] = None
    upload_failed: bool = False


@app.get("/api/measurements")
async def list_measurements(
    number_alpl: Optional[int] = None,
    result:      Optional[str] = None,
    date_from:   Optional[str] = None,
    date_to:     Optional[str] = None,
    session_id:  Optional[int] = None,
    limit:  int = Query(100, le=1000),
    offset: int = Query(0, ge=0),
):
    """Query ประวัติ measurement พร้อม filter ที่เลือกได้ — ใช้ทั้งโดยตาราง dashboard
    (renderTable) และเป็นฐานของ /api/export/csv filter ทุกตัวเป็น optional
    และรวมกันด้วย AND

    เปลี่ยน response shape เป็น object {items, total} เหมือน /api/parts เพื่อรองรับ
    server-side pagination (เพิ่ม offset เข้ามาคู่กับ limit ที่มีอยู่เดิม) — measurements
    โตเร็วกว่า parts มาก จึงไม่ควรโหลดทั้งหมดมา slice ฝั่ง client

    `total` ใช้ COUNT(*) บน WHERE ชุดเดียวกับ query หลัก (filter เดียวกัน) เพื่อให้
    frontend รู้จำนวนทั้งหมดที่ตรงกับ filter ปัจจุบัน ไว้คำนวณหน้า/ปิดปุ่ม Next

    หมายเหตุ: /api/export/csv เป็น endpoint แยกที่ "ไม่" reuse ฟังก์ชันนี้ (มันสร้าง
    WHERE ของตัวเองและคืนไฟล์ CSV ไม่ใช่ JSON) — การเปลี่ยน shape ตรงนี้จึงไม่กระทบ export
    """
    conditions, params = [], []
    if number_alpl is not None:
        conditions.append("m.number_alpl = %s"); params.append(number_alpl)
    if result:
        conditions.append("m.result = %s"); params.append(result)
    if date_from:
        conditions.append("m.timestamp >= %s"); params.append(_day_start(date_from))
    if date_to:
        conditions.append("m.timestamp <= %s"); params.append(_day_end(date_to))
    if session_id is not None:
        conditions.append("m.session_id = %s"); params.append(session_id)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS total FROM measurements m {where}", params)
            total = cur.fetchone()["total"]
            cur.execute(
                f"{MEASUREMENTS_SELECT} {where} ORDER BY m.timestamp DESC LIMIT %s OFFSET %s",
                (*params, limit, offset),
            )
            items = cur.fetchall()
        return {"items": items, "total": total}
    finally:
        db.close()


@app.post("/api/measurements")
async def create_measurement(req: MeasurementCreate):
    """บันทึก measurement หนึ่งรายการที่ส่งมาจาก Agent และตัดสิน OK/NG

    Endpoint นี้ถูกเรียกครั้งละ 1 ชิ้นงานที่ TM-X วัดได้ ส่งต่อมาโดย Agent
    พร้อม session_id ที่ได้รับตอนเริ่ม session — Agent ส่ง `number_alpl` มาด้วย
    เหมือนเดิมเสมอ (ไม่ต้องแก้ agent.py) แต่ตอนนี้ backend จะตัดสินเองว่าจะใช้
    ALPL ไหนจริงๆ ตามประเภทของ session:

      - **Session แบบ queue-based (IPM/New — มี entry ใน session_queues)**:
        เพิกเฉยค่า `req.number_alpl` ที่ Agent ส่งมา แล้วใช้ ALPL ตามตำแหน่ง
        ปัจจุบันในคิวแทน (`session_queues[session_id]["queue"][position]`)
        เพราะ Agent ไม่รู้ (และไม่จำเป็นต้องรู้) ว่ากำลังวัดตัวไหนอยู่ในคิว
        มันรู้แค่ว่า "วัดเสร็จแล้ว ได้ value_x/value_y เท่านี้"
      - **Session แบบ manual (เดิม — ไม่มี entry ใน session_queues)**: ใช้
        `req.number_alpl` ตรงๆ ตามที่ Agent ส่งมา ไม่มีอะไรเปลี่ยนจากเดิม
    """
    is_manual = req.session_id is None
    qstate = None
    db = get_db()
    try:
        # กันการ insert ซ้ำถ้า Agent retry POST นี้ด้วย client_uuid เดิม (เช่น
        # ตอบกลับจาก request ครั้งก่อนหลุดหายระหว่างทาง ทั้งที่จริง backend
        # insert สำเร็จไปแล้ว) — เช็คก่อนทำอะไรอื่นเลย ถ้าเคยเห็น UUID นี้แล้ว
        # คืนผลเดิมไปตรงๆ ไม่ insert แถวใหม่ ไม่นับ measured_count ซ้ำ
        if req.client_uuid:
            with db.cursor() as cur:
                cur.execute(
                    "SELECT measurement_id, session_id, result FROM measurements "
                    "WHERE client_uuid = %s",
                    (req.client_uuid,),
                )
                dup = cur.fetchone()
            if dup:
                with db.cursor() as cur:
                    cur.execute(
                        "SELECT measured_count, target_count FROM sessions WHERE session_id = %s",
                        (dup["session_id"],),
                    )
                    s = cur.fetchone() or {}
                log.info(
                    "Duplicate measurement POST (client_uuid=%s) — คืนผลเดิม measurement_id=%d",
                    req.client_uuid, dup["measurement_id"],
                )
                return {
                    "measurement_id": dup["measurement_id"],
                    "result":  dup["result"],
                    "status":  "duplicate_ignored",
                    "measured": s.get("measured_count"),
                    "target":   s.get("target_count"),
                }

        with db.cursor() as cur:
            if is_manual:
                # ── Manual add จากหน้า Database Editor (edit.html) ──────────
                # ไม่มี session ของ Agent ที่ running อยู่จริงให้อ้างอิงเลย แต่
                # measurements.session_id เป็น NOT NULL + FK ไป sessions บังคับ
                # ต้องมี session อยู่จริงเสมอ จึงสร้าง session "จบในตัว" ขึ้นมา 1
                # แถวแทน (state='stopped', target=measured=1, ended_at=NOW())
                # ไม่ใช่ session ของ Agent เลย แค่เป็นที่ผูก FK ให้ record นี้เท่านั้น
                cur.execute(
                    "INSERT INTO sessions "
                    "(number_alpl, state, target_count, measured_count, ended_at) "
                    "VALUES (%s, 'stopped', 1, 1, NOW())",
                    (req.number_alpl,),
                )
                session_id = cur.lastrowid
                number_alpl = req.number_alpl
                measure_type = "Manual"
                operator_name = None
                note = req.note
            else:
                session_id = req.session_id
                # Session ต้องอยู่ในสถานะ running
                cur.execute(
                    "SELECT state, target_count, measured_count FROM sessions WHERE session_id = %s",
                    (session_id,),
                )
                session = cur.fetchone()
                if not session or session["state"] != "running":
                    raise HTTPException(400, "Session is not running")

                qstate = session_queues.get(session_id)  # None ถ้าเป็น manual session (เดิม)
                measure_type = None
                operator_name = None
                note = None

                if qstate is not None:
                    # ── Queue-based (IPM / New) ─────────────────────────────
                    queue = qstate["queue"]
                    pos = qstate["position"]
                    if pos >= len(queue):
                        raise HTTPException(400, "Measurement queue หมดแล้วสำหรับ session นี้")
                    number_alpl = queue[pos]
                    # entry_mode/note ถูก map ไว้แล้วตั้งแต่ start_session ตามโหมด
                    # ที่ผู้ใช้เลือกหน้าเว็บ: Rework → ('New', 'Rework'),
                    # New → ('New', None), IPM → ('IPM', None)
                    measure_type = qstate["entry_mode"]  # 'IPM' หรือ 'New'
                    operator_name = qstate.get("operator")
                    note = qstate.get("note")
                else:
                    # ── Manual session แบบเดิม (ผ่าน Agent แต่ไม่มีคิว) ─────────
                    number_alpl = req.number_alpl

                # ── New mode: insert Part ตัวนี้แบบ lazy ถ้ายังไม่เคยมีอยู่จริง ──
                # ตัวแรกในคิวถูก insert ไปแล้วตอน start_session (จำเป็นเพราะ FK
                # ของ sessions) ตัวที่เหลือยังไม่เคย insert เลย — insert "ตอนนี้"
                # ที่ได้ผลวัดจริงจาก Agent แล้วเท่านั้น เพื่อไม่ให้ ALPL ที่ยังไม่
                # ทันวัด (เช่น กด Stop กลางคัน) กลายเป็น Part ค้างอยู่ใน DB ทั้งที่
                # ไม่มีประวัติจริง
                if qstate is not None and qstate.get("new_part_config") is not None:
                    cur.execute("SELECT 1 FROM parts_specifications WHERE number_alpl = %s", (number_alpl,))
                    if not cur.fetchone():
                        try:
                            _insert_part_row(cur, number_alpl, qstate["new_part_config"])
                        except pymysql.MySQLError as exc:
                            raise HTTPException(409, f"Insert Part ALPL {number_alpl} ไม่สำเร็จ: {exc}")

            # หา nominal/tolerance ผ่าน part_number ที่ผูกกับ part นี้ — tolerance
            # ตัวเดียวใช้ร่วมกันทั้งแกน X/Y (upper_tol/lower_tol) เก็บอยู่ที่
            # part_number (ไม่ใช่ package_size อีกต่อไป — ดู init.sql)
            cur.execute(
                "SELECT pn.nominal_x, pn.nominal_y, pn.upper_tol, pn.lower_tol, pn.offset_tol "
                "FROM parts_specifications p "
                "JOIN part_number pn ON p.part_number_id = pn.part_number_id "
                "WHERE p.number_alpl = %s",
                (number_alpl,),
            )
            part = cur.fetchone()
            if not part:
                raise HTTPException(404, "Part not found (หรือยังไม่ได้ตั้ง part_number ให้ part นี้)")

            # เช็ค OK/NG — tolerance ตัวเดียวใช้ร่วมกันทั้งแกน X และ Y
            ok_x = _within_tolerance(req.value_x, part["nominal_x"], part["upper_tol"], part["lower_tol"])
            ok_y = _within_tolerance(req.value_y, part["nominal_y"], part["upper_tol"], part["lower_tol"])
            ok_offset = _offset_ok(req.offset, part.get("offset_tol"))
            # ตกข้อใดข้อหนึ่งใน 3 ข้อ = NG (ต้องผ่านครบทั้งหมดถึงจะ OK)
            result = "OK" if (ok_x and ok_y and ok_offset) else "NG"

            # resolve ชื่อ operator (จาก dropdown) เป็น operator_id ก่อน insert —
            # measurements.operator_id เป็น FK ไป operator table แล้ว (เดิมเป็น
            # คอลัมน์ VARCHAR ชื่อ "Oparetor" ที่เก็บชื่อ operator ตรงๆ)
            operator_id = _lookup_id(cur, "operator", "operator_id", "operator_name", operator_name)

            # Insert row ของ measurement (รวม measure_type/operator_id/note ถ้าเป็น
            # queue-based หรือ manual add — สำหรับ manual session (ผ่าน Agent) แบบเดิม
            # ทั้ง 3 ค่านี้จะเป็น NULL)
            try:
                cur.execute(
                    "INSERT INTO measurements "
                    "(session_id, number_alpl, value_x, value_y, `offset`, result, "
                    " measure_type, operator_id, note, client_uuid) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (session_id, number_alpl, req.value_x, req.value_y, req.offset, result,
                     measure_type, operator_id, note, req.client_uuid),
                )
            except pymysql.IntegrityError:
                # race เล็กๆ ที่ทฤษฎีมีได้: สอง request ที่มี client_uuid เดียวกัน
                # มาถึงพร้อมกันเป๊ะๆ ผ่านเช็ค dedup ด้านบนพร้อมกันทั้งคู่ (เช็คแล้ว
                # ยังไม่เจอ เพราะอีกฝั่งยัง insert ไม่เสร็จ) — unique index บน
                # client_uuid จะกันไม่ให้ insert ซ้ำจริงๆ ระดับ DB อยู่ดี แค่ต้อง
                # จับ error แล้วบอกให้รู้ว่าเป็นการซ้ำ ไม่ใช่ปล่อยเป็น 500 ดิบๆ
                raise HTTPException(409, "Measurement นี้ถูกบันทึกไปแล้ว (duplicate client_uuid)")
            measurement_id = cur.lastrowid

            if not is_manual:
                # เพิ่มตัวนับของ session — เฉพาะ session จริงของ Agent เท่านั้น
                # (manual session ที่สร้างเองข้างบน insert มาแบบ measured=target=1
                # อยู่แล้ว ไม่ต้องนับซ้ำ)
                cur.execute(
                    "UPDATE sessions SET measured_count = measured_count + 1 "
                    "WHERE session_id = %s",
                    (session_id,),
                )

                # อ่านค่าตัวนับล่าสุดอีกครั้ง เพื่อเช็คว่าครบ target แล้วหรือยัง
                cur.execute(
                    "SELECT measured_count, target_count FROM sessions WHERE session_id = %s",
                    (session_id,),
                )
                updated = cur.fetchone()
                measured = updated["measured_count"]
                target   = updated["target_count"]
            else:
                measured, target = 1, 1

        if not is_manual:
            # เพิ่มตำแหน่งในคิว (memory) แล้ว sync สำเนาลง DB ทันที (คอลัมน์
            # sessions.queue_state) — กัน backend restart กลาง session นี้แล้ว
            # ตำแหน่งคิวหาย ทำให้ measurement หลังจากนั้นถูกบันทึกผิด ALPL ไปเรื่อยๆ
            # แบบเงียบๆ (ดู lifespan() ที่โหลดค่านี้กลับตอน boot)
            if qstate is not None:
                qstate["position"] += 1
                with db.cursor() as cur:
                    cur.execute(
                        "UPDATE sessions SET queue_state = %s WHERE session_id = %s",
                        (json.dumps(qstate), session_id),
                    )

        # Auto-complete session เมื่อถึง target_count แล้ว — เฉพาะ session จริง
        # ของ Agent เท่านั้น (manual session จบในตัวเองไปแล้วตั้งแต่ insert)
        #
        # ⚠️ ห้ามใส่ `await` ระหว่าง UPDATE measured_count ข้างบน (ราวบรรทัด 2239)
        # กับ UPDATE state='stopped' ข้างล่าง — มีคนพึ่งพาช่วงนี้อยู่:
        #   send_command.py บน Pi poll GET /api/session/state เพื่อรอให้
        #   measured_count ขยับ พอครบ target มันจะหลุด loop เข้า finally แล้วอ่าน
        #   state อีกครั้ง ถ้าเจอ 'running' จะยิง POST /api/session/stop
        #   (ดูบล็อก "ปิด session ที่ค้าง running" ท้าย command_flow)
        # ตอนนี้ autocommit=True ทำให้ measured_count ถูก commit ทันที แต่ Pi ยัง
        # อ่านค่าคาบเกี่ยวไม่ได้ เพราะ handler นี้เป็น async def ที่เรียก pymysql
        # (sync ล้วน ไม่ยอมคืน control) → event loop สลับไปเสิร์ฟ
        # /api/session/state ระหว่างสอง UPDATE นี้ไม่ได้ Pi จึงเห็นทั้งคู่พร้อมกัน
        # เสมอ ไม่มีทางยิง stop ทับ session ที่กำลังจะปิดตัวเองอยู่พอดี
        #
        # การรับประกันนี้จะหายไปทันทีถ้า: ย้ายไป async DB driver, ยัด DB call ลง
        # threadpool, เปลี่ยน endpoint นี้เป็น `def` ธรรมดา (FastAPI จะโยนเข้า
        # threadpool ให้รันขนานได้), หรือรัน uvicorn หลาย worker
        # → ถ้าทำอย่างใดอย่างหนึ่ง ต้องรวมสอง UPDATE นี้เป็น transaction เดียว
        #   หรือให้ stop_session เช็ค state ก่อน UPDATE แทน
        status = "complete" if is_manual else "continue"
        if not is_manual and measured >= target:
            with db.cursor() as cur:
                cur.execute(
                    "UPDATE sessions SET state = 'stopped', ended_at = NOW() "
                    "WHERE session_id = %s",
                    (session_id,),
                )
            status = "complete"
            session_queues.pop(session_id, None)  # session จบแล้ว ลบคิวออกจาก memory
            measure_timeouts.pop(session_id, None)
            await push_event(
                "session_complete",
                {"session_id": session_id, "measured": measured, "target": target},
            )

        # ไม่ broadcast SSE เลยตอน manual add — เหตุผล: onNewMeasurement /
        # onSessionComplete ฝั่ง dashboard (index.html) ไม่ได้เช็คว่า session_id
        # ที่ได้รับตรงกับ session ที่กำลังแสดงอยู่ไหม เลยจะเขียนทับ measured_count/
        # telemetry ของ session จริงที่อาจกำลัง running อยู่พร้อมกันโดยไม่ตั้งใจ
        # (ดูรายละเอียดเพิ่มเติมในคำอธิบายที่คุยกันไว้) edit.html เองก็ไม่ได้พึ่ง
        # SSE อยู่แล้ว มัน refetch ตารางเองหลัง POST สำเร็จ
        if not is_manual:
            await push_event(
                "measurement",
                {
                    "measurement_id": measurement_id,
                    "session_id":     session_id,
                    "number_alpl":    number_alpl,
                    "value_x":        req.value_x,
                    "value_y":        req.value_y,
                    "result":         result,
                    # ผลแยกรายแกน + ช่วงที่รับได้ — ให้ Live Telemetry โชว์ได้ว่า
                    # "พังที่แกนไหน" ไม่ใช่รู้แค่ result รวม (DB เก็บแค่ result
                    # รวมอย่างเดียว ค่าพวกนี้จึงต้องส่งมาทาง event ตอนวัดเสร็จ
                    # ส่วนตอน refresh หน้าเว็บ frontend คำนวณเองจาก nominal/tol
                    # ที่ /api/measurements แนบมาให้ — ดู MEASUREMENTS_SELECT)
                    "offset":         req.offset,
                    "offset_tol":     part.get("offset_tol"),
                    "ok_x":           ok_x,
                    "ok_y":           ok_y,
                    "ok_offset":      ok_offset,
                    "nominal_x":      part["nominal_x"],
                    "nominal_y":      part["nominal_y"],
                    "upper_tol":      part["upper_tol"],
                    "lower_tol":      part["lower_tol"],
                    "measured":       measured,
                    "target":         target,
                },
            )
        return {
            "measurement_id": measurement_id,
            "result":  result,
            "status":  status,
            "measured": measured,
            "target":  target,
        }
    finally:
        db.close()


@app.patch("/api/measurements/{measurement_id}")
async def update_measurement(measurement_id: int, data: Dict[str, Any] = Body(...)):
    """แก้ไข measurement แบบ partial (เฉพาะ field ที่ส่งมาใน body) — ใช้โดยหน้า
    Database Editor (edit.html) ตอนกด Edit แล้ว Save

    ทำไมต้องมี endpoint นี้แยกจาก create_measurement: create_measurement เป็น flow
    ของ Agent (insert ค่าใหม่ที่ TM-X วัดได้ พร้อมตัดสิน OK/NG จาก tolerance) ส่วนการ
    "แก้" ค่าที่บันทึกไว้แล้วเป็นการ override ด้วยมือจากหน้า editor ซึ่งไม่มีมาก่อน

    ใช้ pattern เดียวกับ update_part(): whitelist เฉพาะ field ที่อนุญาตให้แก้ได้
    (number_alpl, operator) เพื่อกัน body ไปเขียนทับ column ที่ไม่ควรแก้ผ่าน
    endpoint นี้ (เช่น session_id, timestamp, image_path, value_x/value_y)

    `number_alpl` แก้ไขได้ (เพิ่มเข้ามาใหม่) — ใช้กรณี IPM พิมพ์/เลือก ALPL ผิดตัว
    (เลือกชิ้นที่มีอยู่จริงในระบบผิดตัว) วิธีแก้ที่ถูกคือ retarget measurement row
    นี้ไปที่ ALPL ที่ถูกต้อง ไม่ใช่ไปแก้ ALPL ที่ตัว Part (เพราะจะกลายเป็นเปลี่ยนชื่อ
    Part จริงที่มีประวัติของตัวเองอยู่แล้ว — ดู update_part สำหรับกรณีพิมพ์ ALPL
    ผิดตอน "New" ซึ่งเหมาะจะแก้ที่ตัว Part แทน)

    ไม่มี `result` ใน allowed อีกต่อไป (เดิมให้ผู้ใช้เลือก OK/NG เองตรงๆ) — เพราะ
    ตอนนี้ ALPL/value เปลี่ยนได้ ทำให้ tolerance ที่ใช้ตัดสิน OK/NG เปลี่ยนตามไปด้วย
    จึงคำนวณ result ใหม่เสมอหลัง update ทุกครั้ง (ไม่ว่าจะแก้ field ไหนก็ตาม) แทนที่
    จะรับค่าจาก frontend ตรงๆ กัน Result ค้างไม่ตรงกับ ALPL/ค่าที่วัดได้จริง
    """
    # แก้ได้เฉพาะ number_alpl กับ operator เท่านั้น (ตามที่ตกลงกันไว้) —
    # value_x/value_y เป็นผลวัดจริงจากเครื่อง ส่วน note/measure_type เป็นข้อมูล
    # ของ session ที่ระบบกำหนดตอนวัด ไม่ควรถูกแก้ย้อนหลังผ่าน endpoint นี้
    # operator ส่งมาเป็น "ชื่อ" จาก dropdown แล้ว resolve เป็น operator_id เอง
    # (เหมือน pattern ของ update_part) — measurements.operator_id เป็น NOT NULL
    # จึงไม่รับค่าว่าง ต้องเลือก operator ที่มีอยู่จริงเสมอ
    allowed = {"number_alpl"}
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "แก้ไข")

            set_parts, values = [], []
            for k, v in data.items():
                if k in allowed:
                    set_parts.append(f"{k} = %s")
                    values.append(v)
            if "operator" in data:
                operator_id = _lookup_id(cur, "operator", "operator_id", "operator_name", data["operator"])
                if operator_id is None:
                    raise HTTPException(400, "ต้องเลือก Operator (เว้นว่างไม่ได้)")
                set_parts.append("operator_id = %s")
                values.append(operator_id)
            if not set_parts:
                raise HTTPException(400, "No valid fields provided")

            set_clause = ", ".join(set_parts)
            try:
                cur.execute(
                    f"UPDATE measurements SET {set_clause} WHERE measurement_id = %s",
                    (*values, measurement_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(
                    409,
                    f"บันทึกไม่สำเร็จ — ALPL ใหม่อาจยังไม่ได้ลงทะเบียนใน Parts: {exc}",
                )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")

            # คำนวณ OK/NG ใหม่จากค่า value_x/value_y/number_alpl "ปัจจุบัน" ของ row
            # นี้เสมอ (หลัง update) — ครอบคลุมทั้งกรณีแก้ ALPL, แก้ value, หรือแก้แค่ note
            cur.execute(
                "SELECT m.value_x, m.value_y, m.`offset`, "
                "pn.nominal_x, pn.nominal_y, pn.upper_tol, pn.lower_tol, pn.offset_tol "
                "FROM measurements m "
                "JOIN parts_specifications p ON m.number_alpl = p.number_alpl "
                "JOIN part_number pn ON p.part_number_id = pn.part_number_id "
                "WHERE m.measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(
                    404,
                    "ไม่พบ part_number ของ ALPL นี้ — คำนวณ OK/NG ใหม่ไม่ได้ (ตรวจสอบว่า Part ตั้ง Part Number ไว้แล้ว)",
                )
            ok_x = _within_tolerance(row["value_x"], row["nominal_x"], row["upper_tol"], row["lower_tol"])
            ok_y = _within_tolerance(row["value_y"], row["nominal_y"], row["upper_tol"], row["lower_tol"])
            ok_offset = _offset_ok(row.get("offset"), row.get("offset_tol"))
            new_result = "OK" if (ok_x and ok_y and ok_offset) else "NG"
            cur.execute(
                "UPDATE measurements SET result = %s WHERE measurement_id = %s",
                (new_result, measurement_id),
            )
        return {"ok": True, "result": new_result}
    finally:
        db.close()


@app.patch("/api/measurements/{measurement_id}/image")
async def update_image(measurement_id: int, req: ImageUpdate):
    """แนบ path ของรูปภาพเข้ากับ measurement หลังจาก Agent จัดเก็บรูปเรียบร้อยแล้ว
    (เดิมคือหลังอัปโหลดขึ้น MinIO — architecture ใหม่จะเป็น path ของไฟล์ใน
    โฟลเดอร์บนเครื่อง PC แทน รอดีไซน์การจัดเก็บ finalize ก่อน)

    ทำไมต้องแยก call นี้ออกจาก create_measurement: Agent จัดเก็บรูป inspection
    *หลังจาก* ค่า measurement ถูกบันทึกไปแล้ว endpoint นี้แค่ patch path ของ
    รูปเข้ากับ measurement ทีหลัง การ broadcast 'image_updated' ทำให้ dashboard
    เปลี่ยนปุ่มรูปภาพในแถวนั้นได้โดยไม่ต้อง refresh ตารางทั้งหมด
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "UPDATE measurements SET image_path = %s, image_upload_failed = %s "
                "WHERE measurement_id = %s",
                (req.image_path, req.upload_failed, measurement_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")
        await push_event(
            "image_updated",
            {
                "measurement_id": measurement_id,
                "image_path": req.image_path,
                "upload_failed": req.upload_failed,
            },
        )
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/measurements/{measurement_id}/image-upload")
async def upload_measurement_image(measurement_id: int, file: UploadFile = File(...)):
    """รับไฟล์รูปจริง (multipart) จาก Agent แล้วบันทึกลงดิสก์ของเครื่อง PC ที่
    รัน backend นี้เอง — แทนที่ MinIO เดิมทั้งหมด (ดูหมายเหตุ ALPL_IMAGE_DIR
    ด้านบน) ต่างจาก update_image (PATCH /image) ตรงที่ endpoint นั้นรับแค่
    "path" ที่ Agent อ้างว่าเก็บไว้แล้ว (ใช้ได้ตอน Agent+Backend อยู่เครื่อง
    เดียวกัน) แต่ตอนนี้ Agent อยู่คนละเครื่อง (Pi) กับ backend (PC) จึงต้องรับ
    "เนื้อไฟล์จริง" มาด้วยเลย แล้ว backend เป็นคนตัดสินใจ path ปลายทางเอง

    path ปลายทาง (เปลี่ยนจากเดิมที่แยกตาม package_size มาเป็นแยกตามวันที่ ตามที่
    ตกลงกันไว้): ALPL_IMAGE_DIR/<DD-MM-YYYY พ.ศ.>/<number_alpl>_<DD-MM-YYYY พ.ศ.>.jpg
    เช่น "22-07-2569/203_22-07-2569.jpg" — ไฟล์ต้นทางจาก TM-X (ปกติเป็น .bmp)
    ถูกแปลงเป็น .jpg เสมอด้วย Pillow ก่อนเซฟ (ไม่เก็บไฟล์ต้นฉบับ .bmp ไว้เลย)
    เก็บเป็น "path สัมพัทธ์" (relative ต่อ ALPL_IMAGE_DIR) ลงคอลัมน์
    measurements.image_path — ไม่เก็บ absolute path เต็มๆ ลง DB เพื่อไม่ให้
    รั่วโครงสร้างไฟล์ระบบจริงออกไป และให้ /api/image-url ต่อ URL ได้ตรงๆ จาก
    ค่านี้ (ดู get_image_url ด้านล่าง กับ static mount /media/alpl ท้ายไฟล์)

    หมายเหตุ (อัปเดต): ชื่อไฟล์เปลี่ยนจาก "number_alpl + วันที่" เป็น
    "number_alpl + measurement_id + ตัวสุ่ม" (เช่น "203_5821_a3f9c8d4e5b6c7d8.jpg")
    — measurement_id ทำให้ชื่อไฟล์ไม่ซ้ำกันเองอยู่แล้วโดยธรรมชาติ (เลขรันของแต่ละ
    การวัด) ส่วนตัวสุ่ม (hex 16 ตัวอักษร จาก secrets.token_hex) กันคนในวง office
    network เดาชื่อไฟล์ ALPL ตัวอื่นถูกโดยไล่เลข measurement_id เอา (ตัว IP
    allowlist ของ office network เป็นชั้นป้องกันหลักที่กันคนนอกอยู่แล้ว ตัวสุ่มนี้
    เป็นชั้นเสริมเท่านั้น) โฟลเดอร์ปลายทางยังคงแยกตามวันที่วัดเหมือนเดิม (ดู
    dest_dir ด้านล่าง) แค่ชื่อไฟล์ข้างในไม่มีวันที่ปนแล้ว
    ผลข้างเคียง: ชื่อไฟล์ไม่ซ้ำกันเองอัตโนมัติอีกต่อไปเหมือนก่อนหน้านี้ (เดิมใช้
    number_alpl+วันที่ ล้วนๆ ทำให้เขียนทับไฟล์เดิมเองโดยไม่ต้องทำอะไรเพิ่ม) จึง
    ต้องลบไฟล์เก่าด้วยมือด้านล่าง (`old_image_path`) หลังเซฟไฟล์ใหม่สำเร็จ
    ไม่งั้นรูปเก่าจะค้างสะสมในโฟลเดอร์ทุกครั้งที่ ALPL เดียวกันถูกวัดซ้ำในวันเดียวกัน
    — ยังคงพฤติกรรมเดิมไว้ว่า "เก็บแค่รูปล่าสุดของ ALPL นั้นในแต่ละวัน"
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT number_alpl, image_path FROM measurements WHERE measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Measurement not found")

            old_image_path = row["image_path"]  # ไฟล์เก่า (ถ้ามี) ของ ALPL นี้ — ต้องลบทิ้งเองหลังเซฟไฟล์ใหม่สำเร็จ

            date_str = _thai_date_str()
            dest_dir = os.path.join(ALPL_IMAGE_DIR, date_str)
            os.makedirs(dest_dir, exist_ok=True)

            # ตัวสุ่มต่อท้ายชื่อไฟล์ — ใช้ secrets (CSPRNG) ไม่ใช่ random module
            # เพราะเป็นงานที่เกี่ยวกับความปลอดภัย (เดาไม่ได้จริงในทางปฏิบัติ)
            # 8 ไบต์ (16 ตัวอักษร hex) พอสำหรับ defense-in-depth ชั้นเสริม ไม่ต้อง
            # ยาวถึงระดับกัน brute-force จากอินเทอร์เน็ตแบบ token รีเซ็ตรหัสผ่าน
            token = secrets.token_hex(8)
            filename = f"{row['number_alpl']}_{measurement_id}_{token}.jpg"
            dest_path_abs = os.path.join(dest_dir, filename)
            image_path_rel = f"{date_str}/{filename}"  # เก็บลง DB แบบ forward-slash เสมอ (ใช้ต่อ URL ตรงๆ ได้)

            try:
                image_bytes = await file.read()
                img = Image.open(BytesIO(image_bytes))
                if img.mode != "RGB":
                    img = img.convert("RGB")  # JPEG ไม่รองรับ alpha/palette (RGBA/P/LA ฯลฯ)
                img.save(dest_path_abs, "JPEG", quality=90)
            except Exception as exc:
                raise HTTPException(500, f"บันทึก/แปลงไฟล์รูปเป็น .jpg ไม่สำเร็จ: {exc}")
            finally:
                await file.close()

            cur.execute(
                "UPDATE measurements SET image_path = %s, image_upload_failed = 0 "
                "WHERE measurement_id = %s",
                (image_path_rel, measurement_id),
            )

            # ลบไฟล์เก่าทิ้ง "หลัง" เซฟไฟล์ใหม่และอัปเดต DB สำเร็จแล้วเท่านั้น —
            # ชื่อไฟล์มีตัวสุ่มแล้วจึงไม่ทับกันเองอัตโนมัติแบบเดิมอีกต่อไป ถ้าไม่ลบ
            # เอง ไฟล์เก่าจะค้างสะสมในโฟลเดอร์ทุกครั้งที่วัด ALPL เดิมซ้ำวันเดียวกัน
            if old_image_path and old_image_path != image_path_rel:
                # ใช้ _delete_image_file เพื่อให้ได้การกัน 2 ชั้นเหมือนตอน DELETE
                # (URL ตกค้างจากยุค MinIO + path ที่หลุดออกนอกโฟลเดอร์รูป) —
                # เดิม os.remove ตรงๆ ถ้าเจอแถวเก่าที่เก็บ URL เต็มไว้จะได้ path
                # มั่วๆ ที่ไม่ควรไปแตะตั้งแต่แรก
                _delete_image_file(old_image_path)
        await push_event(
            "image_updated",
            {
                "measurement_id": measurement_id,
                "image_path": image_path_rel,
                "upload_failed": False,
            },
        )
        return {"ok": True, "image_path": image_path_rel}
    finally:
        db.close()


@app.delete("/api/measurements/{measurement_id}")
async def delete_measurement(measurement_id: int):
    """ลบ measurement 1 row (เช่น ลบค่าที่อ่านผิดพลาด/เป็นการทดสอบ) + ลบไฟล์รูปด้วย

    เดิมลบแค่แถวใน DB ทำให้ไฟล์รูปกลายเป็น "ไฟล์กำพร้า" ค้างสะสมใน
    ALPL_IMAGE_DIR ไปเรื่อยๆ โดยไม่มีอะไรอ้างถึงอีกเลย — และไล่เก็บทีหลังไม่ได้
    ด้วย เพราะข้อมูลที่จะใช้เทียบว่าไฟล์ไหนกำพร้า (แถวใน measurements) ถูกลบไป
    พร้อมกันแล้ว

    ลำดับสำคัญ: **อ่าน image_path ก่อน → ลบแถว → ค่อยลบไฟล์**
    ถ้าลบแถวไม่สำเร็จ (ไม่เจอ row / ติด session guard) ไฟล์ต้องยังอยู่ครบ
    ส่วนถ้าลบแถวสำเร็จแล้วลบไฟล์พลาด ก็ยอมให้ไฟล์ค้าง ดีกว่าทำให้ request พัง
    ทั้งที่ข้อมูลถูกลบไปเรียบร้อยแล้ว (autocommit=True แถวถูก commit ทันที)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            _block_if_session_running(cur, "ลบ")

            row = _fetch_one(
                cur, "SELECT * FROM measurements WHERE measurement_id = %s", (measurement_id,)
            )
            if not row:
                raise HTTPException(404, "Measurement not found")
            image_path = row["image_path"]

            # ── สำรองก่อนลบ ── ตัวนี้ "ย้าย" ไฟล์รูปเข้าถังขยะให้ด้วย ไม่ได้ลบ
            archived = _archive_before_delete(
                kind="measurement", table="measurements",
                pk={"measurement_id": measurement_id}, row=row,
                image_path=image_path,
            )

            cur.execute(
                "DELETE FROM measurements WHERE measurement_id = %s", (measurement_id,)
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "Measurement not found")

        # ลบไฟล์รูปเฉพาะตอนที่ "สำรองไม่สำเร็จ" เท่านั้น — ถ้าสำรองได้ ไฟล์ถูก
        # ย้ายเข้าถังขยะไปแล้ว (_delete_image_file จะหาไม่เจอและคืน False เอง
        # อยู่แล้ว แต่เขียนเงื่อนไขให้ชัดดีกว่าพึ่งผลข้างเคียง)
        if archived is None and _delete_image_file(image_path):
            log.info("ลบไฟล์รูปของ measurement %s แล้ว (%s)", measurement_id, image_path)
        return {"ok": True}
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Image URL endpoint (stub — รอดีไซน์การจัดเก็บรูปแบบ local folder)
# ══════════════════════════════════════════════════════════════════════════════
# เดิมตรงนี้มี 2 endpoint ที่ผูกกับ MinIO ทั้งคู่:
#   POST /api/upload-url          — ออก presigned PUT URL ให้ Agent อัปโหลดรูป
#   GET  /api/image-url/{id}      — ออก presigned GET URL ให้ dashboard ดูรูป
# architecture ใหม่เลิกใช้ MinIO แล้ว รูปจะเก็บเป็นไฟล์ในโฟลเดอร์บนเครื่อง PC
# แทน แต่ดีไซน์การจัดเก็บ (โครงสร้างโฟลเดอร์/ชื่อไฟล์/ใครเป็นคนย้ายไฟล์) ยังไม่
# fix — จึงตัด /api/upload-url ทิ้งไปเลย (Agent ตอนนี้ไม่อัปโหลดรูปแล้ว) ส่วน
# /api/image-url: ดีไซน์เสร็จแล้ว — image_path ใน DB เป็น path สัมพัทธ์ต่อ
# ALPL_IMAGE_DIR เสมอ (เช่น "22-07-2569/203_22-07-2569.jpg" — แยกโฟลเดอร์ตาม
# วันที่วัด ดู upload_measurement_image ด้านบน) จึงต่อ URL ตรงๆ ได้จาก static
# mount "/media/alpl" (ท้ายไฟล์) ไม่ต้อง
# ออก presigned URL แบบ MinIO เดิมอีกต่อไป (ไฟล์อยู่บนดิสก์เครื่องนี้ตรงๆ)
@app.get("/api/image-url/{measurement_id}")
async def get_image_url(measurement_id: int):
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT image_path, image_upload_failed FROM measurements WHERE measurement_id = %s",
                (measurement_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Measurement not found")
        if not row["image_path"]:
            detail = (
                "Agent อัปโหลดรูปไม่สำเร็จ (ลองครบ 3 ครั้งแล้ว)"
                if row["image_upload_failed"]
                else "ยังไม่มีรูปสำหรับ measurement นี้"
            )
            raise HTTPException(404, detail)
        return {"url": f"/media/alpl/{row['image_path']}"}
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Export (CSV) — เทมเพลตเลือกคอลัมน์ + filter + preview
# ══════════════════════════════════════════════════════════════════════════════
# SELECT ก้อนเดียวที่ join ครบทุกตารางที่ export อาจต้องใช้ — ดึงมาทั้งหมดเสมอ
# แล้วค่อยเลือก/จัดรูปแบบคอลัมน์ในฝั่ง Python ตามเทมเพลต (ไม่ประกอบ SQL แบบ
# dynamic จากค่าที่ผู้ใช้ส่งมา จึงไม่มีช่องให้ SQL injection เลย)
EXPORT_SELECT = """
    SELECT m.measurement_id, m.session_id, m.number_alpl, m.value_x, m.value_y,
           m.`offset` AS `offset`,
           m.result, m.note, m.measure_type, m.timestamp,
           op.operator_name,
           pn.part_number_name, pn.nominal_x, pn.nominal_y, pn.upper_tol, pn.lower_tol,
           pn.offset_tol,
           h.handler_name, ps.package_size, t.template_name,
           v.vendor_name, o.owner_name,
           p.po_number, p.description, p.recieve_date
    FROM measurements m
    LEFT JOIN operator op             ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p  ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn          ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h               ON pn.handler_id = h.handler_id
    LEFT JOIN package_size ps         ON pn.package_size_id = ps.package_size_id
    LEFT JOIN template t              ON ps.template_id = t.template_id
    LEFT JOIN vendor v                ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o                 ON p.owner_id = o.owner_id
"""


def _fmt_num(v, digits: int = 3):
    return "" if v is None else f"{float(v):.{digits}f}"


def _csv_header(key: str) -> str:
    """หัวคอลัมน์ในไฟล์ CSV — ใช้ csv_label ถ้ามี ไม่มีก็ใช้ label ตามปกติ

    มีไว้เพราะบางช่องชื่อที่เหมาะกับ "ชิปในหน้าแก้ผังรายงาน" กับ "หัวคอลัมน์ใน
    ไฟล์ CSV" ไม่ใช่ชื่อเดียวกัน เช่น timestamp: ในรายงานเลือกได้ว่าจะเอาแค่
    วันที่หรือเวลา จึงเรียกว่า "Date" แต่ CSV ออกทั้งวันและเวลาในช่องเดียวเสมอ
    เรียกว่า Date เฉยๆ จะสื่อผิด
    """
    c = EXPORT_COLUMNS[key]
    return c.get("csv_label") or c["label"]


def _fmt_timestamp(r, fmt: str = "datetime") -> str:
    """วันเวลาของการวัด — เลือกได้ว่าจะเอาอะไรบ้าง (ผู้ใช้ติ๊กในหน้าแก้ผังรายงาน)
    date      → 30/07/2569 แบบ DD/MM/YYYY
    time      → 14:23:11
    datetime  → ทั้งคู่
    """
    ts = r["timestamp"]
    if not ts:
        return ""
    if fmt == "time":
        return ts.strftime("%H:%M:%S")
    if fmt == "date":
        return ts.strftime("%d/%m/%Y")
    return ts.strftime("%d/%m/%Y %H:%M:%S")


# รูปแบบที่ให้เลือกได้ของช่องวันเวลา (frontend เอาไปทำ dropdown บนเซลล์)
_TIME_FORMATS = [
    {"key": "date",     "label": "Date"},
    {"key": "time",     "label": "Time"},
    {"key": "datetime", "label": "Date & Time"},
]


def _axis_state(r, axis: str) -> str:
    """ค่าที่วัดได้ของแกนนี้ อยู่ในสเปกไหม — คืน "OK"/"NG" ("" ถ้าเทียบไม่ได้)

    ต่างจาก Result ที่เป็นค่าดิบใน DB อยู่แล้ว — อันนี้ต้องคำนวณเทียบสเปกทีละแกน
    เพราะ Result เป็นผลรวมของทั้ง X และ Y (X ผ่านแต่ Y ไม่ผ่าน → Result = NG
    ทั้งที่ตัวเลข X เองอยู่ในสเปก) ถ้าเอา Result มาใช้ระบายสีช่อง X จะสีผิด
    """
    val = r[f"value_{axis}"]
    nom = r[f"nominal_{axis}"]
    if val is None or nom is None or r["upper_tol"] is None or r["lower_tol"] is None:
        return ""
    return "OK" if _within_tolerance(val, nom, r["upper_tol"], r["lower_tol"]) else "NG"


def _tolerance_spec(r) -> str:
    """สเปกขนาดชิ้นงานแบบย่อบรรทัดเดียว — ใช้ในรายงาน PDF/Excel เท่านั้น

    รูปแบบ: "<nominal>  +<upper>  -<lower>" เช่น "9.02  +0.03  -0.01"
    ถ้า nominal X กับ Y ไม่เท่ากัน (package ไม่ใช่ทรงจัตุรัส) จะคั่นด้วย "/"
    เป็น "3.28/7.43  +0.02  -0.01" — ที่ยุบเลขเดียวตอน X=Y เพราะรายงาน
    ที่ห้อง PM Kit ใช้อยู่เขียนแบบนั้น (ดูตัวอย่างที่ผู้ใช้ให้มา)

    ทศนิยม 2 ตำแหน่งเท่านั้น (ไม่ใช่ 3 ตำแหน่งแบบ _fmt_num ปกติ) เพราะสเปก
    บนแบบชิ้นงานเขียนแค่ 2 ตำแหน่ง — ส่วนค่าที่วัดได้จริง (value_x/value_y)
    ยังใช้ 3 ตำแหน่งเหมือนเดิม เพราะต้องเห็นความละเอียดของการวัด

    หมายเหตุ: หน้า Export CSV ยังใช้คอลัมน์แยก (nominal_x/nominal_y/
    upper_tol/lower_tol) เหมือนเดิม ไม่ได้ถูกยุบตาม — ตั้งใจให้ CSV แยกช่อง
    เพื่อเอาไปคำนวณต่อได้ ส่วนรายงานเน้นอ่านง่ายบนกระดาษ
    """
    d = 2
    nx, ny = r["nominal_x"], r["nominal_y"]
    if nx is None and ny is None:
        nom = ""
    elif nx is None or ny is None:
        nom = _fmt_num(nx if nx is not None else ny, d)
    elif abs(float(nx) - float(ny)) < _TOL_EPS:
        nom = _fmt_num(nx, d)
    else:
        nom = f"{_fmt_num(nx, d)}/{_fmt_num(ny, d)}"

    parts = [p for p in (nom,
                         f"+{_fmt_num(r['upper_tol'], d)}" if r["upper_tol"] is not None else "",
                         f"-{_fmt_num(r['lower_tol'], d)}" if r["lower_tol"] is not None else "") if p]
    return "  ".join(parts)


# catalog ของคอลัมน์ที่เลือกใส่เทมเพลตได้
#   key    → ชื่อที่เก็บใน export_template.columns_json (เป็น "สัญญา" กับ frontend)
#   label  → หัวคอลัมน์ในไฟล์ CSV
#   group  → ใช้จัดกลุ่มในหน้าเลือกคอลัมน์
#   get    → ฟังก์ชันดึง/จัดรูปแบบค่าจาก row ที่ EXPORT_SELECT คืนมา
# หมายเหตุ: nominal และ tolerance แยกเป็นคอลัมน์ละค่า (Nominal X, Nominal Y,
# Upper Tol, Lower Tol) ไม่รวมเป็นช่องเดียวแบบที่แสดงบนหน้าเว็บ ("4.030 / 4.030")
# เพราะไฟล์ export เอาไปคำนวณต่อใน Excel/Power BI ถ้ารวมเป็นสตริงจะต้องมาแยกเอง
EXPORT_COLUMNS: Dict[str, Dict[str, Any]] = {
    # row_number = ไม่ได้มาจาก DB — _render_report นับลำดับแถวให้ตอนคลี่ผัง
    # scope=report เพราะ CSV ไม่ต้องมีเลขลำดับ (Excel/โปรแกรมอื่นใส่เองได้)
    "item":          {"label": "Item",          "group": "ข้อมูลการวัด", "scope": "report",
                      "row_number": True, "get": lambda r: ""},
    "number_alpl":   {"label": "ALPL",          "group": "ข้อมูลการวัด",
                      "header": "Number ALPL", "get": lambda r: r["number_alpl"]},
    # Value X/Y เป็น scope=csv เพราะในรายงานมันไม่ยืนเดี่ยว — เป็นส่วนหนึ่งของ
    # บล็อก Tolerance เสมอ (ดู tolerance_spec ด้านล่าง)
    # values/state = ใช้ตั้ง "หน้าตาแยกตามค่า" ในรายงาน (เช่น เกินสเปกให้พื้นแดง)
    # state ของแกนต้องคำนวณเอง ไม่ใช้ Result เพราะ Result เป็นผลรวมของทั้ง X และ Y
    "value_x":       {"label": "Value X",       "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"], "state": lambda r: _axis_state(r, "x"),
                      "get": lambda r: _fmt_num(r["value_x"])},
    "value_y":       {"label": "Value Y",       "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"], "state": lambda r: _axis_state(r, "y"),
                      "get": lambda r: _fmt_num(r["value_y"])},
    # offset ไม่ได้เทียบกับช่วง nominal ± tol เหมือน X/Y แต่เทียบกับเพดาน
    # offset_tol ตัวเดียว จึงมี state เป็นของตัวเองไม่ใช้ _axis_state
    "offset":        {"label": "Offset",        "group": "ข้อมูลการวัด", "scope": "csv",
                      "values": ["OK", "NG"],
                      "state": lambda r: (
                          "" if r.get("offset") is None or r.get("offset_tol") is None
                          else ("OK" if _offset_ok(r.get("offset"), r.get("offset_tol")) else "NG")),
                      "get": lambda r: _fmt_num(r.get("offset"))},
    # ── บล็อก Tolerance (รายงานเท่านั้น) ────────────────────────────────
    # ลากครั้งเดียวได้ผังกว้าง 2 คอลัมน์ สูง 3 แถว:
    #   แถว 1  [        Tolerance        ]   ผสาน 2 คอลัมน์ — หัวตาราง
    #   แถว 2  [   ข้อมูล Tolerance      ]   ผสาน 2 คอลัมน์ — สเปกของกลุ่ม
    #   แถว 3   ข้อมูล Value X | ข้อมูล Value Y  แถวที่ทำซ้ำต่อ 1 การวัด
    # แถว 2 พิมพ์ครั้งเดียวต่อกลุ่ม เพราะรายงานถูกแบ่งกลุ่มด้วยสเปกนี้เสมอ
    #
    # label  = ชื่อชิปในแผงซ้าย (บอกว่าลากแล้วได้อะไรบ้าง)
    # header = ข้อความที่พิมพ์เป็นหัวตารางจริงในรายงาน — คนละอันกับ label
    #          เพราะหัวตารางบนกระดาษต้องสั้นว่า "Tolerance" เฉยๆ
    "tolerance_spec": {
        "label": "Tolerance + Value X/Y", "group": "ข้อมูลการวัด", "scope": "report",
        "get": _tolerance_spec,
        "block": {
            "cols": 2,
            "header": "Tolerance",
            "data": [{"key": "value_x", "label": "Value X"},
                     {"key": "value_y", "label": "Value Y"}],
        },
    },
    "result":        {"label": "Result",        "group": "ข้อมูลการวัด", "values": ["OK", "NG"],
                      "state": lambda r: r["result"] or "",
                      "get": lambda r: r["result"]},
    "note":          {"label": "Note",          "group": "ข้อมูลการวัด", "get": lambda r: r["note"] or ""},
    "operator":      {"label": "Operator",      "group": "ข้อมูลการวัด", "get": lambda r: r["operator_name"] or ""},
    "measure_type":  {"label": "Measure Type",  "group": "ข้อมูลการวัด", "get": lambda r: r["measure_type"] or ""},
    # header = ข้อความหัวตารางเริ่มต้นในรายงาน (ต่างจาก label ที่เป็นชื่อชิป/หัว CSV)
    # formats = รูปแบบที่ติ๊กเลือกได้บนเซลล์ — ค่าเริ่มต้นคือตัวแรกในลิสต์ (date)
    "timestamp":     {"label": "Date",          "csv_label": "Timestamp",
                      "group": "ข้อมูลการวัด",
                      "header": "Date", "formats": _TIME_FORMATS,
                      "get": lambda r: _fmt_timestamp(r, "datetime"),
                      "get_fmt": _fmt_timestamp},
    "session_id":    {"label": "Session",       "group": "ข้อมูลการวัด", "get": lambda r: r["session_id"]},

    "part_number":   {"label": "Part Number",   "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["part_number_name"] or ""},
    "handler":       {"label": "Handler",       "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["handler_name"] or ""},
    "package_size":  {"label": "Package Size",  "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["package_size"] or ""},
    "template_name": {"label": "Template",      "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["template_name"] or ""},
    # ── สเปกขนาด: CSV ใช้ 4 ช่องแยก / รายงาน PDF-Excel ใช้ช่องรวมช่องเดียว ──
    # scope บอกว่าคอลัมน์นี้โผล่ในหน้าไหน (ไม่ใส่ = โผล่ทั้งสองหน้า)
    "nominal_x":     {"label": "Nominal X",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["nominal_x"])},
    "nominal_y":     {"label": "Nominal Y",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["nominal_y"])},
    "upper_tol":     {"label": "Upper Tol",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["upper_tol"])},
    "lower_tol":     {"label": "Lower Tol",     "group": "ข้อมูลชิ้นงาน", "scope": "csv", "get": lambda r: _fmt_num(r["lower_tol"])},
    "vendor":        {"label": "Vendor",        "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["vendor_name"] or ""},
    "owner":         {"label": "Owner",         "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["owner_name"] or ""},
    "po_number":     {"label": "PO Number",     "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["po_number"] if r["po_number"] is not None else ""},
    "description":   {"label": "Description",   "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["description"] or ""},
    "recieve_date":  {"label": "Receive Date",  "group": "ข้อมูลชิ้นงาน", "get": lambda r: r["recieve_date"].strftime("%d/%m/%Y") if r["recieve_date"] else ""},
}


# kind ของเทมเพลตมี 3 ค่า: 'csv' | 'pdf' | 'excel' — แยกลิสต์กันคนละหน้า
# ('report' คือชื่อเก่าสมัยที่ PDF/Excel ใช้ลิสต์ร่วมกัน ยังอ่านของเดิมได้อยู่)
# ส่วน "ชุดคอลัมน์" มีแค่ 2 แบบ: csv ใช้คอลัมน์แยก / รายงานใช้บล็อก Tolerance
def _columns_scope(kind: str) -> str:
    return "csv" if kind == "csv" else "report"


@app.get("/api/export/columns")
async def list_export_columns(kind: str = "csv"):
    """คืน catalog คอลัมน์ที่ใส่ในเทมเพลตได้ (key/label/group)
    ใช้สร้างหน้าเลือกคอลัมน์ — frontend ไม่ต้อง hardcode รายชื่อเอง

    kind='csv'          → ได้ Nominal X / Nominal Y / Upper Tol / Lower Tol แยก 4 ช่อง
    kind='pdf'|'excel'  → ได้ช่องรวม "Tolerance" ช่องเดียวแทนทั้ง 4 ช่องนั้น
    คอลัมน์ที่ไม่ระบุ scope โผล่ทั้งสองแบบ
    """
    out = []
    for k, c in EXPORT_COLUMNS.items():
        if c.get("scope") not in (None, _columns_scope(kind)):
            continue
        item = {"key": k, "label": c["label"], "group": c["group"]}
        if c.get("block"):
            # แนบ values ของ "ลูก" ในบล็อกไปด้วย (value_x/value_y ก็ตั้งหน้าตา
            # แยกตามค่าได้เหมือน Result) — ลูกไม่ได้อยู่ใน catalog เป็นตัวของตัวเอง
            blk = dict(c["block"])
            blk["data"] = [
                {**d, **({"values": EXPORT_COLUMNS[d["key"]]["values"]}
                         if EXPORT_COLUMNS.get(d["key"], {}).get("values") else {})}
                for d in c["block"]["data"]
            ]
            item["block"] = blk             # ผังหลายเซลล์ที่ frontend ต้องกางออก
        if c.get("header"):
            item["header"] = c["header"]    # ข้อความหัวตารางเริ่มต้นในรายงาน
        if c.get("values"):
            item["values"] = c["values"]    # ค่าที่เป็นไปได้ → ตั้งหน้าตาแยกตามค่าได้
        if c.get("formats"):
            item["formats"] = c["formats"]  # รูปแบบที่ติ๊กเลือกได้บนเซลล์
        out.append(item)
    return out


# คอลัมน์รวมแบบเก่าที่เคยมีในเทมเพลตที่บันทึกไว้ก่อนหน้า → คอลัมน์ย่อยชุดใหม่
# (เทมเพลตเก่าที่ผู้ใช้เซฟไว้แล้วจะยังใช้งานได้ ไม่ใช่หายไปเงียบๆ)
_LEGACY_COLUMN_ALIASES = {
    "nominal_xy": ["nominal_x", "nominal_y"],
    "tolerance":  ["upper_tol", "lower_tol"],
}


def _parse_columns(raw) -> List[str]:
    """แปลงค่า columns_json จาก DB (str หรือ list) เป็น list ของ key ที่ใช้ได้จริง
    — แตก key แบบเก่าเป็นคอลัมน์ย่อย และตัด key ที่ไม่รู้จักทิ้ง
    """
    cols = raw
    if isinstance(cols, str):
        try:
            cols = json.loads(cols)
        except Exception:
            cols = []
    if not isinstance(cols, list):
        cols = []

    out: List[str] = []
    for c in cols:
        for key in _LEGACY_COLUMN_ALIASES.get(c, [c]):
            if key in EXPORT_COLUMNS and key not in out:
                out.append(key)
    return out


def _get_template(cur, export_template_id: int) -> Dict[str, Any]:
    cur.execute(
        "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
        "FROM export_template WHERE export_template_id = %s",
        (export_template_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบเทมเพลตนี้")
    return row


def _parse_layout(raw):
    """แปลง layout_json จาก DB (str หรือ dict) เป็น dict — None ถ้าไม่มี/พัง"""
    if raw is None:
        return None
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return None
    return raw if isinstance(raw, dict) else None


@app.get("/api/export/templates")
async def list_export_templates(kind: str = "csv"):
    """คืนเทมเพลตของชนิดที่ระบุ — csv / pdf / excel แยกลิสต์กันคนละชนิด

    PDF กับ Excel เคยใช้ kind='report' ร่วมกัน ทำให้เทมเพลตปนกันข้ามรูปแบบ
    ตอนนี้แยกแล้ว แต่ยังดึงของเก่าที่เป็น 'report' มาแสดงด้วย จะได้ไม่หายไปเฉยๆ
    """
    kinds = [kind, "report"] if kind in ("pdf", "excel") else [kind]
    ph = ", ".join(["%s"] * len(kinds))
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(
                "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                f"FROM export_template WHERE kind IN ({ph}) ORDER BY is_default DESC, name",
                kinds,
            )
            return [
                {
                    "export_template_id": r["export_template_id"],
                    "name": r["name"],
                    "kind": r["kind"],
                    "columns": _parse_columns(r["columns_json"]),
                    "layout": _parse_layout(r["layout_json"]),
                    "is_default": bool(r["is_default"]),
                }
                for r in cur.fetchall()
            ]
    finally:
        db.close()


class ExportTemplateBody(BaseModel):
    name: str
    # เทมเพลต CSV ใช้ columns ส่วนเทมเพลตรายงาน (PDF/Excel) ใช้ layout
    # อย่างใดอย่างหนึ่งต้องมี (ดู _validate_template_body)
    columns: Optional[List[str]] = None
    layout:  Optional[Dict[str, Any]] = None
    kind:    str = "csv"


def _validate_template_body(body: ExportTemplateBody) -> List[str]:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")
    cols = [c for c in (body.columns or []) if c in EXPORT_COLUMNS]
    if not cols:
        raise HTTPException(400, "ต้องเลือกอย่างน้อย 1 คอลัมน์")
    return cols


def _template_payload(body: ExportTemplateBody):
    """ตรวจ body แล้วคืน (kind, columns_json, layout_json) ที่พร้อมเขียนลง DB

    เก็บข้อมูลคนละคอลัมน์กันตามชนิด:
      csv          → columns_json (รายชื่อคอลัมน์ + ลำดับ)
      pdf / excel  → layout_json  (ผังตารางทั้งก้อนจากหน้า editor แบบสเปรดชีต)

    ⚠ ทุกอย่างที่ "ไม่ใช่ csv" ถือเป็นรายงานหมด — ห้ามเช็คแบบ `kind == "report"`
    เพราะตอนแยก PDF กับ Excel ออกจากกัน ค่า kind เปลี่ยนเป็น 'pdf'/'excel' แล้ว
    การเช็คชื่อตายตัวทำให้เทมเพลตรายงานตกไปเข้าเส้นทางของ CSV แล้วเด้ง
    "ต้องเลือกอย่างน้อย 1 คอลัมน์" ทั้งที่หน้าจอนั้นไม่มีให้เลือกคอลัมน์เลย
    """
    if not body.name.strip():
        raise HTTPException(400, "ต้องตั้งชื่อเทมเพลต")

    if body.kind != "csv":
        if not body.layout:
            raise HTTPException(400, "เทมเพลตรายงานต้องมีผังตาราง (layout)")
        kind = body.kind if body.kind in ("pdf", "excel", "report") else "pdf"
        return kind, None, json.dumps(body.layout, ensure_ascii=False)

    cols = _validate_template_body(body)
    return "csv", json.dumps(cols, ensure_ascii=False), None


@app.post("/api/export/templates", status_code=201)
async def create_export_template(body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                    "VALUES (%s, %s, %s, %s, 0)",
                    (body.name.strip(), kind, cols_json, layout_json),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"สร้างเทมเพลตไม่สำเร็จ (ชื่อนี้อาจมีอยู่แล้ว): {exc}")
            return {"export_template_id": cur.lastrowid}
    finally:
        db.close()


@app.patch("/api/export/templates/{export_template_id}")
async def update_export_template(export_template_id: int, body: ExportTemplateBody):
    kind, cols_json, layout_json = _template_payload(body)
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            # เทมเพลตตั้งต้นของระบบล็อกไว้ — ให้ Duplicate ไปแก้ตัวใหม่แทน
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นแก้ไขไม่ได้ — กด Duplicate แล้วแก้ตัวสำเนาแทน")
            try:
                cur.execute(
                    "UPDATE export_template SET name = %s, kind = %s, columns_json = %s, layout_json = %s "
                    "WHERE export_template_id = %s",
                    (body.name.strip(), kind, cols_json, layout_json, export_template_id),
                )
            except pymysql.MySQLError as exc:
                raise HTTPException(409, f"บันทึกไม่สำเร็จ (ชื่อนี้อาจซ้ำกับเทมเพลตอื่น): {exc}")
        return {"ok": True}
    finally:
        db.close()


@app.delete("/api/export/templates/{export_template_id}")
async def delete_export_template(export_template_id: int):
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            if row["is_default"]:
                raise HTTPException(403, "เทมเพลตค่าเริ่มต้นลบไม่ได้")
            cur.execute(
                "DELETE FROM export_template WHERE export_template_id = %s", (export_template_id,)
            )
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/export/templates/{export_template_id}/duplicate", status_code=201)
async def duplicate_export_template(export_template_id: int):
    """ทำสำเนาเทมเพลต (ใช้ได้กับทุกตัวรวมทั้งตัวค่าเริ่มต้น) — ตั้งชื่อใหม่ให้
    อัตโนมัติแบบ "<ชื่อเดิม> (สำเนา)" และเติมเลขต่อท้ายถ้าชื่อนั้นถูกใช้ไปแล้ว
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            row = _get_template(cur, export_template_id)
            cur.execute("SELECT name FROM export_template")
            taken = {r["name"] for r in cur.fetchall()}
            base = f"{row['name']} (สำเนา)"
            name, n = base, 2
            while name in taken:
                name, n = f"{base} {n}", n + 1
            # สำเนาต้องเก็บทั้ง 2 ฟิลด์ตามชนิดของต้นฉบับ (csv ใช้ columns_json,
            # report ใช้ layout_json) — คัดลอกดิบๆ ไปเลยไม่ต้องแปลง
            cur.execute(
                "INSERT INTO export_template (name, kind, columns_json, layout_json, is_default) "
                "VALUES (%s, %s, %s, %s, 0)",
                (
                    name,
                    row.get("kind") or "csv",
                    json.dumps(_parse_columns(row["columns_json"]), ensure_ascii=False)
                        if row.get("columns_json") else None,
                    json.dumps(_parse_layout(row["layout_json"]), ensure_ascii=False)
                        if row.get("layout_json") else None,
                ),
            )
            return {"export_template_id": cur.lastrowid, "name": name}
    finally:
        db.close()


# subquery หา measurement_id ของ "การวัดครั้งล่าสุดของแต่ละ ALPL"
# คิดจากข้อมูลทั้งตารางก่อนเสมอ (ไม่ขึ้นกับ filter อื่น) แล้วค่อยเอาไป AND กับ
# filter ที่เหลือ — ให้ความหมายตรงกับ ISLatest ที่ใช้ใน Power BI คือ
# "แถวนี้เป็นค่าล่าสุดของ ALPL นั้นไหม" เป็นคุณสมบัติของแถว ไม่ใช่ผลของการกรอง
# (เช่นกรอง Result=NG + ล่าสุด = ALPL ที่ "สถานะปัจจุบันยังไม่ผ่าน" ไม่ใช่
#  "เคยมี NG ครั้งล่าสุดในบรรดา NG ทั้งหมด" ซึ่งคนละความหมายกันคนละเรื่อง)
def _latest_only_sql(inner_where: str) -> str:
    """เงื่อนไข "เอาเฉพาะการวัดล่าสุดของแต่ละ ALPL"

    ⚠ จุดที่เคยพลาด: ของเดิม subquery เป็น `FROM measurements` เปล่าๆ ไม่มี WHERE
    จึงหา "ล่าสุด" จากทั้งตารางก่อน แล้วค่อยเอาผลไปกรองด้วยเงื่อนไขของผู้ใช้ทีหลัง
    ผลคือถ้า ALPL ตัวไหนถูกวัดซ้ำ "นอกช่วง" ที่ผู้ใช้กรอง แถวล่าสุดของมันจะตก
    ตัวกรอง แล้ว ALPL ตัวนั้นหายไปจากรายงานทั้งตัว ทั้งที่ในช่วงที่เลือกมีการวัดจริง
    เช่น กรอง 1–15 ก.ค. แต่ ALPL 400 ถูกวัดซ้ำวันที่ 20 ก.ค. → 400 หายไปเฉยๆ

    ตอนนี้ subquery ใช้ FROM/JOIN และ WHERE ชุดเดียวกับ query หลัก จึงหมายถึง
    "ล่าสุดภายในชุดข้อมูลที่กรองแล้ว" ตามที่ผู้ใช้คาดหวังจริงๆ
    (พารามิเตอร์ของ inner_where ต้องถูกส่งซ้ำอีกชุด — ดู _export_filters)
    """
    return f"""m.measurement_id IN (
    SELECT measurement_id FROM (
        SELECT m.measurement_id,
               ROW_NUMBER() OVER (
                   PARTITION BY m.number_alpl
                   ORDER BY m.timestamp DESC, m.measurement_id DESC
               ) AS rn
        {_EXPORT_FROM}
        {inner_where}
    ) AS latest WHERE latest.rn = 1
)"""


# รูปแบบที่ยอมรับต่อ 1 ท่อน: "400" หรือ "400-407" (เว้นวรรครอบๆ ได้)
_RANGE_PART_RE = re.compile(r"^\s*(\d+)\s*(?:-\s*(\d+)\s*)?$")


def _parse_int_ranges(spec: str, field_label: str):
    """แปลงข้อความแบบ "400-407,500,600" เป็น (ช่วง, ตัวเดี่ยว)

    คืนค่าเป็น ([(400, 407)], [500, 600]) — แยกช่วงกับตัวเดี่ยวออกจากกัน
    **ตั้งใจไม่แตกช่วงออกเป็นรายตัว** เพราะถ้าผู้ใช้พิมพ์ "1-999999" การแตกจะได้
    ค่าเกือบล้านตัวไปต่อเป็น SQL ยาว ~6.5 MB ซึ่งชน max_allowed_packet ของ MySQL
    หรือช้าจนใช้งานไม่ได้ — เก็บเป็นช่วงแล้วแปลงเป็น BETWEEN ตอนสร้าง SQL แทน
    ต้นทุนเท่ากับไม่กรองเลย (จำนวนแถวขึ้นกับข้อมูลจริง ไม่ใช่ความกว้างของช่วง)

    เคสที่รองรับ: เว้นวรรค, พิมพ์กลับด้าน (407-400), ค่าซ้ำ, ท่อนว่างจากคอมมาเกิน
    """
    ranges, singles = [], []
    for part in str(spec).split(","):
        if not part.strip():
            continue  # เช่น "400,,500" หรือคอมมาท้ายสุด — ข้ามไปเฉยๆ
        m = _RANGE_PART_RE.match(part)
        if not m:
            raise HTTPException(
                400,
                f'รูปแบบ {field_label} ไม่ถูกต้องตรง "{part.strip()}" — '
                f'ใช้ได้เช่น 400 หรือ 400-407 หรือ 400-407,500,600',
            )
        lo = int(m.group(1))
        if m.group(2) is None:
            singles.append(lo)
            continue
        hi = int(m.group(2))
        if lo > hi:
            lo, hi = hi, lo          # พิมพ์กลับด้าน (407-400) ก็ยังใช้ได้
        if lo == hi:
            singles.append(lo)       # "400-400" ก็คือตัวเดียว
        else:
            ranges.append((lo, hi))
    return sorted(set(ranges)), sorted(set(singles))


_DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _day_end(v):
    """ครอบปลายช่วงให้เต็มวัน — "2026-07-30" → "2026-07-30 23:59:59"

    ถ้าไม่ทำ MySQL จะตีค่าเป็น 00:00:00 แล้ว `timestamp <= ...` จะตัดการวัด
    ของวันสุดท้ายที่ผู้ใช้เลือกทิ้งทั้งวัน หน้าเว็บเติมเวลาให้อยู่แล้ว แต่ย้าย
    มาบังคับที่ backend ด้วย เพราะเป็นกติกาที่หน้าใหม่ (หรือคนที่ยิง API ตรงๆ)
    ไม่มีทางรู้ แล้วจะเจอบั๊กเดิมซ้ำ
    """
    return f"{v} 23:59:59" if isinstance(v, str) and _DATE_ONLY_RE.match(v.strip()) else v


def _day_start(v):
    return f"{v} 00:00:00" if isinstance(v, str) and _DATE_ONLY_RE.match(v.strip()) else v


def _export_filters(f: Dict[str, Any]):
    """สร้าง WHERE + params ของหน้า Export — ใช้ร่วมกันทั้ง preview และ csv
    เพื่อให้สิ่งที่เห็นในตัวอย่างตรงกับไฟล์ที่ดาวน์โหลดจริงเสมอ
    """
    conditions, params = [], []

    def eq(col, key):
        """เงื่อนไขเท่ากับ — รองรับทั้งค่าเดียวและหลายค่า (multi-select)

        หลายค่าจะกลายเป็น IN (...) ซึ่งหมายถึง "เอาอันไหนก็ได้ในกลุ่มนี้" (OR กัน
        ภายในช่องเดียวกัน) ส่วนต่างช่องกันยังเป็น AND เหมือนเดิม เช่น
        เลือก Vendor = A,B และ Result = NG → (vendor A หรือ B) และ ต้องเป็น NG
        """
        v = f.get(key)
        if v is None or v == "":
            return
        if isinstance(v, (list, tuple, set)):
            vals = [x for x in v if x not in (None, "")]
            if not vals:
                return
            conditions.append(f"{col} IN ({','.join(['%s'] * len(vals))})")
            params.extend(vals)
        else:
            conditions.append(f"{col} = %s"); params.append(v)

    # ALPL รับได้ทั้งตัวเดียว, หลายตัวคั่นคอมมา และช่วง — ประกอบเป็นเงื่อนไข
    # เดียวที่ OR กันเอง เช่น (BETWEEN 400 AND 407 OR BETWEEN 500 AND 507 OR IN (600))
    alpl_spec = f.get("number_alpl")
    if alpl_spec not in (None, ""):
        alpl_ranges, alpl_singles = _parse_int_ranges(alpl_spec, "ALPL")
        or_parts = []
        for lo, hi in alpl_ranges:
            or_parts.append("m.number_alpl BETWEEN %s AND %s")
            params.extend([lo, hi])
        if alpl_singles:
            or_parts.append(f"m.number_alpl IN ({','.join(['%s'] * len(alpl_singles))})")
            params.extend(alpl_singles)
        if or_parts:
            conditions.append("(" + " OR ".join(or_parts) + ")")

    eq("m.result",          "result")
    eq("m.session_id",      "session_id")
    eq("op.operator_name",  "operator")
    eq("m.measure_type",    "measure_type")
    eq("v.vendor_name",     "vendor")
    eq("o.owner_name",      "owner")
    eq("pn.part_number_name", "part_number")
    eq("h.handler_name",    "handler")
    eq("ps.package_size",   "package_size")
    eq("p.po_number",       "po_number")

    # ช่วงวันที่ของ "วันที่วัด" (measurements.timestamp)
    if f.get("date_from"):
        conditions.append("m.timestamp >= %s"); params.append(_day_start(f["date_from"]))
    if f.get("date_to"):
        conditions.append("m.timestamp <= %s"); params.append(_day_end(f["date_to"]))
    # ช่วงวันที่ของ "วันที่รับชิ้นงาน" (parts_specifications.recieve_date)
    if f.get("recv_from"):
        conditions.append("p.recieve_date >= %s"); params.append(_day_start(f["recv_from"]))
    if f.get("recv_to"):
        conditions.append("p.recieve_date <= %s"); params.append(_day_end(f["recv_to"]))

    # Description ค้นแบบมีคำนี้อยู่ข้างใน (ไม่ต้องพิมพ์ตรงเป๊ะ)
    if f.get("description"):
        conditions.append("p.description LIKE %s"); params.append(f"%{f['description']}%")

    # ต้องต่อท้ายสุดเสมอ — เงื่อนไขนี้ห่อ WHERE ของ "ทุกข้อข้างบน" ไว้ข้างใน
    # ลำดับพารามิเตอร์จึงเป็น [ของข้อข้างบน..., ของข้อข้างบนซ้ำอีกรอบ...]
    # (ถ้าย้ายไปไว้กลางๆ ลำดับ %s จะเพี้ยนแล้วได้ข้อมูลผิดแบบเงียบๆ)
    if f.get("latest_only"):
        inner_where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        inner_params = list(params)
        conditions.append(_latest_only_sql(inner_where))
        params.extend(inner_params)

    return (("WHERE " + " AND ".join(conditions)) if conditions else ""), params


# FROM + JOIN ชุดเดียวกับ EXPORT_SELECT — ใช้ตอนนับ COUNT(*) เพราะ WHERE อ้างถึง
# คอลัมน์ของตาราง join ด้วย (vendor/owner/handler ฯลฯ) จะนับจาก measurements
# เปล่าๆ ไม่ได้
_EXPORT_FROM = """
    FROM measurements m
    LEFT JOIN operator op             ON m.operator_id = op.operator_id
    LEFT JOIN parts_specifications p  ON m.number_alpl = p.number_alpl
    LEFT JOIN part_number pn          ON p.part_number_id = pn.part_number_id
    LEFT JOIN handler h               ON pn.handler_id = h.handler_id
    LEFT JOIN package_size ps         ON pn.package_size_id = ps.package_size_id
    LEFT JOIN template t              ON ps.template_id = t.template_id
    LEFT JOIN vendor v                ON p.vendor_id = v.vendor_id
    LEFT JOIN owner o                 ON p.owner_id = o.owner_id
"""


def _fetch_export_rows(cols: List[str], where: str, params: list, limit: Optional[int] = None):
    """ดึงข้อมูลแล้วแปลงเป็น list ของ list ตามลำดับคอลัมน์ใน `cols`"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp DESC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            rows = cur.fetchall()
    finally:
        db.close()
    data = [[EXPORT_COLUMNS[c]["get"](r) for c in cols] for r in rows]
    return data, total


def _count_export_rows(where: str, params: list) -> int:
    """นับจำนวนแถวอย่างเดียว — ใช้เช็คเพดานก่อนดึงข้อมูลจริงมาสร้างไฟล์"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            return cur.fetchone()["n"]
    finally:
        db.close()


def _fetch_export_raw(where: str, params: list, limit: Optional[int] = None):
    """เหมือน _fetch_export_rows แต่คืน row ดิบ (dict) ไม่ได้แปลงเป็นคอลัมน์
    ใช้กับรายงาน PDF/Excel ที่ต้องหยิบค่าทีละช่องตามผังตาราง ไม่ใช่เรียงเป็นแถว
    เรียง ASC ตามเวลา เพราะรายงานบนกระดาษอ่านจากเก่าไปใหม่ (ต่างจาก CSV ที่
    เรียง DESC ให้เห็นตัวล่าสุดก่อน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n {_EXPORT_FROM} {where}", params)
            total = cur.fetchone()["n"]
            sql = f"{EXPORT_SELECT} {where} ORDER BY m.timestamp ASC, m.measurement_id ASC"
            if limit:
                sql += f" LIMIT {int(limit)}"
            cur.execute(sql, params)
            return cur.fetchall(), total
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════
# ตัวเรนเดอร์รายงาน — ใช้ร่วมกันทั้ง Preview / Excel / PDF
# ══════════════════════════════════════════════════════════════════════════
# รับผังจาก report-template.html (layout_json) + ข้อมูลจริง แล้วคลี่ออกเป็น
# ตาราง 2 มิติที่พร้อมวาด ไม่ว่าจะวาดด้วย HTML (preview/print) หรือ openpyxl
#
# โครงของผัง 1 ชุด:
#   แถว 0 .. dataRow-1   = ส่วนหัว   — พิมพ์ซ้ำทุกกลุ่ม (มีเซลล์ spec ของกลุ่ม)
#   แถว dataRow          = แถวข้อมูล — ทำซ้ำ 1 แถวต่อ 1 การวัด
#   แถว dataRow+1 .. จบ  = ส่วนท้าย  — พิมพ์ครั้งเดียวตอนจบรายงาน (ช่องเซ็นชื่อ)
#
# รายงานถูกแบ่งกลุ่มด้วย tolerance_spec เสมอ (ดู GROUP_BY ใน frontend) ชิ้นงาน
# คนละสเปกจึงไม่ปนกันในตารางเดียว
REPORT_GROUP_BY = "tolerance_spec"


def _cell_out(cell: Dict[str, Any], text: str) -> Dict[str, Any]:
    """แปลงเซลล์ในผังเป็นเซลล์ผลลัพธ์ — เก็บเฉพาะที่ตัววาดต้องใช้"""
    out = {"v": text, "s": cell.get("s") or {}}
    span = cell.get("span")
    if span:
        out["span"] = {"r": int(span.get("r", 1)), "c": int(span.get("c", 1))}
    if cell.get("hidden"):
        out["hidden"] = True
    # ทำเครื่องหมายไว้ว่าเซลล์ไหนมาจากข้อมูลจริง — ตัววาดใช้ตัดสินใจตีเส้นขอบ
    if cell.get("f") or cell.get("spec"):
        out["data"] = True
    if cell.get("hdr"):
        out["head"] = True
    return out


def _render_report(layout: Dict[str, Any], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """คลี่ผัง + ข้อมูล ออกเป็นตารางพร้อมวาด

    คืน {"nCols": int, "rows": [[cell, ...], ...]} โดย cell = {v, s, span?, hidden?}
    """
    grid = layout.get("grid") or []
    n_rows = int(layout.get("nRows") or len(grid))
    n_cols = int(layout.get("nCols") or (len(grid[0]) if grid else 0))
    data_row = int(layout.get("dataRow", 2))
    if not grid or n_cols == 0:
        return {"nCols": 0, "rows": []}
    data_row = max(0, min(data_row, n_rows - 1))

    def tpl_row(r: int) -> List[Dict[str, Any]]:
        row = grid[r] if r < len(grid) else []
        return [row[c] if c < len(row) else {} for c in range(n_cols)]

    def clamp(cells: List[Dict[str, Any]], r: int, last: int) -> List[Dict[str, Any]]:
        """ตัด rowspan ไม่ให้ทะลุออกนอกบล็อกที่กำลังพิมพ์อยู่

        ผู้ใช้ผสานเซลล์คร่อมแถวข้อมูลได้ในหน้าแก้ไข แต่ตอนคลี่ออกมา ส่วนหัวของ
        กลุ่มถัดไปจะมาต่อท้าย ถ้าปล่อย rowspan เดิมไว้มันจะกินทับแถวของกลุ่มถัดไป
        แล้วช่วงผสานซ้อนกัน → Excel เปิดไฟล์ไม่ขึ้น บอกว่าไฟล์เสียหาย
        """
        for cell in cells:
            sp = cell.get("span")
            if sp and sp["r"] > 1:
                sp["r"] = max(1, min(sp["r"], last - r + 1))
        return cells

    def render_static(r: int, spec_text: str) -> List[Dict[str, Any]]:
        """แถวส่วนหัว/ส่วนท้าย — แทนเฉพาะเซลล์ spec ด้วยค่าของกลุ่มปัจจุบัน
        เซลล์ f (ข้อมูลรายชิ้น) ที่หลุดมาอยู่นอกแถวข้อมูลถือว่าไม่มีความหมาย
        ปล่อยว่างไว้ ดีกว่าเอาค่าของชิ้นแรกมาใส่แบบมั่วๆ
        """
        out = []
        for cell in tpl_row(r):
            if cell.get("spec"):
                out.append(_cell_out(cell, spec_text))
            elif cell.get("f"):
                out.append(_cell_out(cell, ""))
            else:
                out.append(_cell_out(cell, cell.get("v") or ""))
        return out

    def render_data(row: Dict[str, Any], item_no: int) -> List[Dict[str, Any]]:
        """แถวข้อมูล 1 แถวต่อ 1 การวัด — rowspan ถูกบีบเป็น 1 เพราะแถวนี้ถูก
        ทำซ้ำหลายรอบ ถ้าปล่อย rowspan เดิมไว้จะกินทับแถวของการวัดชิ้นถัดไป
        """
        out = []
        for cell in tpl_row(data_row):
            key = cell.get("f")
            col = EXPORT_COLUMNS.get(key) if key else None

            # ช่องลำดับแถว (Item) — ค่าไม่ได้อยู่ใน DB ต้องนับตอนคลี่ผัง
            if col and col.get("row_number"):
                text = str(item_no)
            # รูปแบบที่ผู้ใช้ติ๊กไว้บนเซลล์ (ตอนนี้มีแค่ช่องวันเวลา: date/time/ทั้งคู่)
            elif col and col.get("get_fmt") and cell.get("fmt"):
                text = col["get_fmt"](row, cell["fmt"])
            elif col:
                text = col["get"](row)
            else:
                text = cell.get("v") or ""

            c = _cell_out(cell, str(text))
            if "span" in c:
                c["span"]["r"] = 1

            # หน้าตาแยกตามค่า — เช่น Result OK พื้นเขียว / NG พื้นแดง
            # "สถานะ" ของเซลล์มาจากฟังก์ชัน state ของคอลัมน์ ไม่ใช่ข้อความที่แสดง
            # เพราะบางคอลัมน์สถานะกับค่าที่แสดงเป็นคนละเรื่อง เช่น Value X แสดง
            # "4.089" แต่สถานะคือ NG (เกินสเปก) — ถ้าจับคู่ด้วยข้อความจะไม่เจอเลย
            variants = cell.get("variants") or {}
            if variants:
                state = col["state"](row) if col and col.get("state") else str(text)
                if state in variants:
                    c["s"] = variants[state]
            out.append(c)
        return out

    # จัดกลุ่มตามสเปก โดยรักษาลำดับที่เจอครั้งแรกไว้ (ไม่เรียงใหม่)
    #
    # แบ่งกลุ่มเฉพาะเมื่อผังมีช่อง Tolerance อยู่จริงเท่านั้น — ถ้าผู้ใช้ทำ
    # เทมเพลตที่ไม่ได้ลาก Tolerance ลงไป การแบ่งกลุ่มจะทำให้หัวตารางถูกพิมพ์ซ้ำ
    # หลายรอบโดยไม่มีอะไรบนกระดาษบอกว่าทำไมถึงแยกก้อน ผู้ใช้จะงงว่าเป็นบั๊ก
    has_spec = any(
        (cell or {}).get("spec") or (cell or {}).get("hdr") == REPORT_GROUP_BY
        for row in grid for cell in row
    )
    groups: Dict[str, List[Dict[str, Any]]] = {}
    if has_spec:
        for r in rows:
            groups.setdefault(EXPORT_COLUMNS[REPORT_GROUP_BY]["get"](r), []).append(r)
    elif rows:
        groups[""] = list(rows)

    # แถวส่วนหัวแบ่งเป็น 2 พวก:
    #   - แถวที่มีช่องข้อมูลผูกอยู่ (ชื่อคอลัมน์ / สเปกของกลุ่ม) → ซ้ำทุกกลุ่ม
    #     เพราะแต่ละกลุ่มเป็นตารางของมันเอง ต้องมีหัวตารางกำกับ
    #   - แถวที่เป็นข้อความที่ผู้ใช้พิมพ์เองล้วนๆ (เช่น "ตรวจสอบการวัดวันที่ ...")
    #     → พิมพ์ครั้งเดียวตอนกลุ่มแรก เพราะเป็นหัวเรื่องของทั้งรายงาน ไม่ใช่ของ
    #     ตารางใดตารางหนึ่ง ถ้าซ้ำทุกกลุ่มจะกลายเป็นชื่อเรื่องโผล่กลางหน้า
    # เช็คที่เซลล์ต้นของการผสาน (ข้าม hidden) เพราะหัวเรื่องมักผสานยาวทั้งแถว
    def _repeats_per_group(r: int) -> bool:
        return any(
            (c.get("hdr") or c.get("spec") or c.get("f"))
            for c in tpl_row(r) if not c.get("hidden")
        )

    repeat_head = {r: _repeats_per_group(r) for r in range(0, data_row)}

    out_rows: List[List[Dict[str, Any]]] = []
    # ลำดับแถว (Item) นับต่อเนื่องทั้งรายงาน ไม่รีเซ็ตตอนขึ้นกลุ่มใหม่ —
    # ตรงกับรายงานกระดาษที่ห้อง PM Kit ใช้ ซึ่งเลขลำดับคือ "ชิ้นที่เท่าไรของใบนี้"
    item_no = 0
    for gi, (spec_text, members) in enumerate(groups.items()):
        for r in range(0, data_row):
            if gi and not repeat_head[r]:
                continue
            out_rows.append(clamp(render_static(r, spec_text), r, data_row - 1))
        for m in members:
            item_no += 1
            out_rows.append(render_data(m, item_no))
    # ส่วนท้ายพิมพ์ครั้งเดียวตอนจบ (ปกติเป็นช่องเซ็นชื่อ/ผู้ตรวจ)
    for r in range(data_row + 1, n_rows):
        out_rows.append(clamp(render_static(r, ""), r, n_rows - 1))

    # ── ตัดคอลัมน์/แถวว่างที่ห้อยท้ายทิ้ง ──────────────────────────────
    # ผังในหน้าแก้ไขมี 10 คอลัมน์ 5 แถวเสมอ แต่ผู้ใช้ใช้จริงไม่กี่ช่อง ถ้าปล่อย
    # ช่องว่างติดไปด้วย เวลาพิมพ์ PDF ตารางจะถูกจัดกึ่งกลางโดยนับช่องว่างพวกนั้น
    # เข้าไปด้วย ทำให้ส่วนที่มีเนื้อหาดูเบี้ยวไปทางซ้าย และมีเส้นจางๆ ห้อยอยู่
    # ทางขวาของกระดาษ
    def filled(cell: Dict[str, Any]) -> bool:
        return bool((cell.get("v") or "").strip() or cell.get("data") or cell.get("head"))

    last_col = 0
    for row in out_rows:
        for c, cell in enumerate(row):
            if not cell.get("hidden") and filled(cell):
                last_col = max(last_col, c + (cell.get("span", {}) or {}).get("c", 1))
    last_col = last_col or n_cols

    for row in out_rows:
        del row[last_col:]
        for c, cell in enumerate(row):
            sp = cell.get("span")
            if sp and c + sp["c"] > last_col:
                sp["c"] = max(1, last_col - c)

    while out_rows and not any(filled(x) for x in out_rows[-1]):
        out_rows.pop()

    return {"nCols": last_col, "rows": out_rows, "groups": len(groups)}


def _load_report_layout(cur, export_template_id: int) -> Dict[str, Any]:
    tpl = _get_template(cur, export_template_id)
    layout = _parse_layout(tpl.get("layout_json"))
    if not layout or not layout.get("grid"):
        raise HTTPException(400, f'เทมเพลต "{tpl["name"]}" ยังไม่มีผังตาราง — เปิดแก้ไขแล้วบันทึกใหม่อีกครั้ง')
    return {"name": tpl["name"], "kind": tpl.get("kind") or "pdf", "layout": layout}


# พารามิเตอร์ filter ที่ /api/export/preview กับ /api/export/csv รับเหมือนกันทุกตัว
# (ประกาศเป็น Pydantic model แล้วใช้ Depends เพื่อไม่ต้องเขียนซ้ำ 2 ที่ และกัน
#  ลืมเพิ่มข้างใดข้างหนึ่งจนตัวอย่างกับไฟล์จริงไม่ตรงกัน)
def export_filters_dep(
    # ALPL เป็น str ไม่ใช่ int — รับได้ทั้ง "400", "400,500,600" และ
    # "400-407,500-507" (ดู _parse_int_ranges)
    number_alpl:  Optional[str] = None,
    date_from:    Optional[str] = None,
    date_to:      Optional[str] = None,
    recv_from:    Optional[str] = None,
    recv_to:      Optional[str] = None,
    session_id:   Optional[int] = None,
    po_number:    Optional[int] = None,
    description:  Optional[str] = None,
    # ── ช่องแบบเลือกได้หลายค่า (multi-select) ────────────────────────────
    # ส่งมาเป็น query param ซ้ำๆ เช่น ?vendor=A&vendor=B → กลายเป็น IN ('A','B')
    # ต้องประกาศเป็น dependency แบบฟังก์ชัน ไม่ใช่ Pydantic model + Depends()
    # เพราะ field ชนิด List ใน model ไม่ผูกกับ query param (ทดสอบแล้ว)
    result:       Optional[List[str]] = Query(None),
    operator:     Optional[List[str]] = Query(None),
    measure_type: Optional[List[str]] = Query(None),
    vendor:       Optional[List[str]] = Query(None),
    owner:        Optional[List[str]] = Query(None),
    part_number:  Optional[List[str]] = Query(None),
    handler:      Optional[List[str]] = Query(None),
    package_size: Optional[List[str]] = Query(None),
    # ค่าเริ่มต้นเป็น True — หน้า Export ต้องการ "สถานะล่าสุดของแต่ละ ALPL"
    # เป็นหลัก ไม่ใช่ประวัติทุกครั้งที่เคยวัด (ติ๊กออกได้ถ้าอยากได้ทั้งหมด)
    latest_only:  bool = True,
) -> Dict[str, Any]:
    """พารามิเตอร์ filter ที่ /api/export/preview กับ /api/export/csv รับเหมือนกัน
    ประกาศไว้ที่เดียวแล้วใช้ร่วมกันผ่าน Depends — กันลืมเพิ่มข้างใดข้างหนึ่ง
    จนตัวอย่างบนหน้าจอกับไฟล์ที่ดาวน์โหลดจริงไม่ตรงกัน
    """
    return {
        "number_alpl": number_alpl, "date_from": date_from, "date_to": date_to,
        "recv_from": recv_from, "recv_to": recv_to, "session_id": session_id,
        "po_number": po_number, "description": description,
        "result": result, "operator": operator, "measure_type": measure_type,
        "vendor": vendor, "owner": owner, "part_number": part_number,
        "handler": handler, "package_size": package_size,
        "latest_only": latest_only,
    }


@app.get("/api/export/preview")
async def export_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    limit:   int = Query(5, ge=1, le=50),
):
    """คืนหัวคอลัมน์ + ข้อมูลตัวอย่างไม่กี่แถว + จำนวนแถวทั้งหมดที่ตรงกับ filter
    ให้หน้าเว็บโชว์ก่อนกดดาวน์โหลดจริง (จะได้รู้ว่ากรองถูกไหม ไฟล์ใหญ่แค่ไหน)
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            tpl = _get_template(cur, export_template_id)
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, total = _fetch_export_rows(cols, where, params, limit=limit)
    return {
        "template_name": tpl["name"],
        "columns": [_csv_header(c) for c in cols],
        "column_keys": cols,
        "rows": data,
        "total": total,
    }


@app.get("/api/export/csv")
async def export_csv(
    export_template_id: Optional[int] = None,
    filename: Optional[str] = None,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """Export ประวัติ measurement (พร้อม filter) เป็นไฟล์ CSV ให้ดาวน์โหลด

    export_template_id: เลือกว่าจะเอาคอลัมน์ไหนและเรียงลำดับยังไง — ถ้าไม่ส่งมา
    จะใช้เทมเพลตที่ is_default = 1 ให้อัตโนมัติ (พฤติกรรมเดิมของ endpoint นี้
    คือ dump ทุกคอลัมน์ ซึ่งตรงกับเทมเพลตค่าเริ่มต้นพอดี ของเดิมที่เรียกมาโดย
    ไม่ส่งพารามิเตอร์นี้จึงยังใช้งานได้เหมือนเดิม)

    ใช้ utf-8-sig encoding เพื่อให้เปิดใน Excel ได้ถูกต้องแม้มีตัวอักษรไทยอยู่ใน
    ไฟล์ (BOM ช่วยไม่ให้ตัวอักษรเพี้ยน) และใช้ filter ชุดเดียวกับ /api/export/preview
    เพื่อให้ไฟล์ที่ได้ตรงกับตัวอย่างที่เห็นบนหน้าจอเสมอ
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            if export_template_id is not None:
                tpl = _get_template(cur, export_template_id)
            else:
                cur.execute(
                    "SELECT export_template_id, name, kind, columns_json, layout_json, is_default "
                    "FROM export_template WHERE kind = 'csv' AND is_default = 1 LIMIT 1"
                )
                tpl = cur.fetchone()
                if not tpl:
                    raise HTTPException(404, "ยังไม่มีเทมเพลตค่าเริ่มต้นในระบบ")
    finally:
        db.close()

    cols = _parse_columns(tpl["columns_json"])
    where, params = _export_filters(filters)
    data, _ = _fetch_export_rows(cols, where, params)

    df  = pd.DataFrame(data, columns=[_csv_header(c) for c in cols])
    buf = StringIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)

    # ชื่อไฟล์ใส่วันที่ให้ด้วย — ดาวน์โหลดหลายรอบจะได้ไม่ทับกันใน Downloads
    # ชื่อไฟล์ที่ผู้ใช้กรอกในหน้า Export มาก่อน — ถ้าไม่ส่งมาค่อยตั้งชื่อตามเวลาให้
    # กรองตัวอักษรที่ใช้ในชื่อไฟล์ Windows ไม่ได้ออกอีกชั้น (ฝั่งหน้าเว็บกรองแล้ว
    # แต่ endpoint นี้เรียกตรงจาก URL ได้ จึงต้องกันเองด้วย ไม่เชื่อ input จากข้างนอก)
    safe_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "", (filename or "")).strip(". ")
    fname = (f"{safe_name}.csv" if safe_name
             else f"measurements_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
    return StreamingResponse(
        iter(["﻿" + buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── รายงาน PDF/Excel — ใช้ผังจาก report-template.html ────────────────────────
# preview กับไฟล์จริงเรียก _render_report ตัวเดียวกัน สิ่งที่เห็นบนจอจึงตรงกับ
# ไฟล์ที่ได้เสมอ (ถ้าแยกโค้ดกัน 2 ชุดจะเพี้ยนกันเงียบๆ ตอนแก้ข้างเดียว)

# preview ไม่ดึงทั้งหมด — ผังใหญ่ๆ ที่มีหมื่นแถวจะทำให้หน้าเว็บค้าง
REPORT_PREVIEW_LIMIT = 300

# เพดานของ "ไฟล์จริง" — ทั้ง Excel และ PDF สร้างตารางทั้งก้อนไว้ในหน่วยความจำ
# ก่อนส่งออก ถ้าไม่จำกัดแล้วเจอข้อมูลหลักหมื่นแถว จะกินแรมหนักและ request ค้าง
# จนหมดเวลาแบบเงียบๆ — ปฏิเสธไปเลยพร้อมบอกให้กรองข้อมูลก่อน ดีกว่าปล่อยให้ค้าง
REPORT_MAX_ROWS = int(os.getenv("REPORT_MAX_ROWS", 20000))


def _guard_report_size(total: int) -> None:
    if total > REPORT_MAX_ROWS:
        raise HTTPException(
            413,
            f"ข้อมูลที่เลือกมี {total:,} แถว เกินเพดาน {REPORT_MAX_ROWS:,} แถวของรายงาน "
            f"— กรองให้แคบลงก่อน (เช่น จำกัดช่วงวันที่) หรือใช้ Export CSV แทน "
            f"ถ้าต้องการข้อมูลดิบทั้งหมด",
        )


@app.get("/api/export/report-preview")
async def export_report_preview(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
    full: int = 0,
):
    """คืนผังที่คลี่แล้วพร้อมข้อมูลจริง ให้หน้าเว็บวาดเป็นตาราง preview

    full=0 → ตัดที่ REPORT_PREVIEW_LIMIT แถว (ดูบนจอเฉยๆ ไม่ต้องครบ)
    full=1 → เอาครบทุกแถว ใช้ตอนสั่งพิมพ์เป็น PDF จริง — ไฟล์ที่ได้ต้องไม่ขาด
    """
    db = get_db()
    try:
        with db.cursor() as cur:
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    if full:
        _guard_report_size(total_check := _count_export_rows(where, params))
    rows, total = _fetch_export_raw(where, params, limit=None if full else REPORT_PREVIEW_LIMIT)
    out = _render_report(tpl["layout"], rows)
    out.update({
        "template_name": tpl["name"],
        "total": total,
        "shown": len(rows),
        "truncated": total > len(rows),
    })
    return out


@app.get("/api/export/xlsx")
async def export_xlsx(
    export_template_id: int,
    filters: Dict[str, Any] = Depends(export_filters_dep),
):
    """สร้างไฟล์ .xlsx จากผังรายงาน — คงฟอนต์/สี/การผสานเซลล์ตามที่จัดไว้"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            500,
            "ยังไม่ได้ติดตั้ง openpyxl — รัน `pip install openpyxl` ในเครื่องที่รัน Backend ก่อน",
        )

    db = get_db()
    try:
        with db.cursor() as cur:
            tpl = _load_report_layout(cur, export_template_id)
    finally:
        db.close()

    where, params = _export_filters(filters)
    rows, total = _fetch_export_raw(where, params)
    _guard_report_size(total)
    rendered = _render_report(tpl["layout"], rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    thick = Side(style="medium", color="000000")
    box = Border(left=thick, right=thick, top=thick, bottom=thick)

    def rgb(v: Optional[str]) -> Optional[str]:
        """#RRGGBB → RRGGBB (openpyxl ไม่รับเครื่องหมาย #)"""
        if not v or not isinstance(v, str):
            return None
        v = v.strip().lstrip("#")
        return v.upper() if re.fullmatch(r"[0-9A-Fa-f]{6}", v) else None

    # ผังเก็บค่าการจัดวางเป็นคำของ CSS แต่ openpyxl ใช้คำคนละชุด ถ้าส่งคำที่มัน
    # ไม่รู้จักไปจะโยน ValueError ทันที (เคสจริงที่เจอ: valign='middle' ของ CSS
    # ซึ่ง openpyxl เรียกว่า 'center') — แปลงและกรองให้เหลือเฉพาะค่าที่ยอมรับได้
    _H = {"left": "left", "center": "center", "right": "right", "justify": "justify"}
    _V = {"top": "top", "middle": "center", "center": "center", "bottom": "bottom"}

    def align_of(s: Dict[str, Any]) -> "Alignment":
        try:
            indent = max(0, min(250, int(s.get("indent") or 0)))
        except (TypeError, ValueError):
            indent = 0
        return Alignment(
            horizontal=_H.get(str(s.get("align") or "").lower(), "center"),
            vertical=_V.get(str(s.get("valign") or "").lower(), "center"),
            indent=indent,
            wrap_text=False,
        )

    def size_of(s: Dict[str, Any]) -> float:
        try:
            v = float(s.get("size") or 11)
        except (TypeError, ValueError):
            return 11.0
        return v if 1 <= v <= 409 else 11.0

    for r, row in enumerate(rendered["rows"], start=1):
        for c, cell in enumerate(row, start=1):
            if cell.get("hidden"):
                continue
            s = cell.get("s") or {}
            x = ws.cell(row=r, column=c, value=cell.get("v") or "")
            x.font = Font(
                name=str(s.get("font") or "Tahoma"),
                size=size_of(s),
                bold=bool(s.get("bold")),
                italic=bool(s.get("italic")),
                underline="single" if s.get("underline") else None,
                color=rgb(s.get("color")) or "000000",
            )
            fill = rgb(s.get("fill"))
            if fill:
                x.fill = PatternFill("solid", fgColor=fill)
            x.alignment = align_of(s)
            # เส้นขอบหนารอบเซลล์ที่มีเนื้อหา — ตรงกับที่ตกลงไว้ว่าในหน้าแก้ไข
            # ไม่ต้องโชว์เส้น แต่ตอน export ต้องมี
            if cell.get("v") or cell.get("data") or cell.get("head"):
                x.border = box

            span = cell.get("span")
            if span and (span["r"] > 1 or span["c"] > 1):
                r2 = r + max(1, span["r"]) - 1
                c2 = min(c + max(1, span["c"]) - 1, rendered["nCols"])
                # ข้ามถ้าซ้อนกับช่วงที่ผสานไปแล้ว — openpyxl ไม่ห้ามตอนสร้าง แต่
                # ไฟล์ที่ได้จะเปิดไม่ขึ้นใน Excel (บอกว่าไฟล์เสียหาย) ยอมให้ผสาน
                # ไม่ครบดีกว่าได้ไฟล์ที่เปิดไม่ได้เลย
                if not any(
                    m.min_row <= r2 and r <= m.max_row and m.min_col <= c2 and c <= m.max_col
                    for m in ws.merged_cells.ranges
                ):
                    ws.merge_cells(start_row=r, start_column=c, end_row=r2, end_column=c2)

    for c in range(1, (rendered["nCols"] or 1) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        log.exception("สร้างไฟล์ xlsx ไม่สำเร็จ")
        raise HTTPException(500, f"สร้างไฟล์ Excel ไม่สำเร็จ: {type(exc).__name__}: {exc}")
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", tpl["name"]).strip("_") or "report"
    fname = f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# ถังขยะ — ดูรายการ / กู้คืน / ลบอัตโนมัติเมื่อเกินอายุ
# ══════════════════════════════════════════════════════════════════════════════
# ⚠ ไฟล์ในถังขยะ "ห้าม" mount เป็น static เด็ดขาด — ต่างจาก image_ALPL ที่เปิด
#   ให้เข้าถึงผ่าน /media/alpl ถังขยะเข้าถึงได้เฉพาะผ่าน endpoint ด้านล่างเท่านั้น

# ป้ายชื่อภาษาไทยของแต่ละ kind — ใช้แสดงในตารางหน้าเว็บ
_DELETED_KIND_LABEL = {
    "measurement":  "ผลการวัด",
    "part":         "Part (ALPL)",
    "operator":     "Operator",
    "owner":        "Owner",
    "vendor":       "Vendor",
    "handler":      "Handler",
    "template":     "Template",
    "package_size": "Package Size",
    "part_number":  "Part Number",
}


def _deleted_files():
    """ไล่หาไฟล์ .json ทุกอันในถังขยะ — คืน (โฟลเดอร์วัน, ชื่อไฟล์, path เต็ม)"""
    if not os.path.isdir(DELETED_DIR):
        return []
    out = []
    for day in sorted(os.listdir(DELETED_DIR), reverse=True):
        day_dir = os.path.join(DELETED_DIR, day)
        if not os.path.isdir(day_dir):
            continue
        for name in sorted(os.listdir(day_dir), reverse=True):
            if name.endswith(".json"):
                out.append((day, name, os.path.join(day_dir, name)))
    return out


def _deleted_summary(payload: Dict[str, Any]) -> str:
    """ข้อความสรุปสั้นๆ ให้คนอ่านออกว่าแถวที่ลบไปคืออะไร โดยไม่ต้องเปิดไฟล์ดู"""
    kind, row = payload.get("kind"), payload.get("row", {})
    if kind == "measurement":
        vx, vy = row.get("value_x"), row.get("value_y")
        return (f"ALPL {row.get('number_alpl')} · {vx} / {vy} · {row.get('result')}"
                f" · {str(row.get('timestamp') or '')[:19]}")
    if kind == "part":
        n = len(payload.get("related", {}).get("sessions", []))
        return f"ALPL {row.get('number_alpl')}" + (f" (+ {n} sessions)" if n else "")
    # lookup ทั้งหมดมีคอลัมน์ชื่อลงท้ายด้วย _name
    for k, v in row.items():
        if k.endswith("_name") or k == "package_size":
            return str(v)
    return ", ".join(f"{k}={v}" for k, v in list(row.items())[:3])


def _purge_old_deleted() -> int:
    """ลบโฟลเดอร์ในถังขยะที่เก่ากว่า DELETED_RETENTION_DAYS วัน — คืนจำนวนที่ลบ

    อ่านวันจาก "ชื่อโฟลเดอร์" (DD-MM-YYYY พ.ศ.) ไม่ใช่ mtime ของไฟล์ เพราะ mtime
    เปลี่ยนได้ง่ายเวลาก๊อป/ย้ายไฟล์ ส่วนชื่อโฟลเดอร์คือวันที่ลบจริงเสมอ
    """
    if not os.path.isdir(DELETED_DIR):
        return 0
    cutoff = datetime.now().date() - timedelta(days=DELETED_RETENTION_DAYS)
    removed = 0
    for day in os.listdir(DELETED_DIR):
        day_dir = os.path.join(DELETED_DIR, day)
        if not os.path.isdir(day_dir):
            continue
        try:
            d, m, y = day.split("-")
            day_date = date(int(y) - 543, int(m), int(d))   # แปลง พ.ศ. → ค.ศ.
        except (ValueError, TypeError):
            log.warning("ข้ามโฟลเดอร์ในถังขยะที่ชื่อไม่ใช่รูปแบบวันที่: %s", day)
            continue
        if day_date < cutoff:
            try:
                shutil.rmtree(day_dir)
                removed += 1
                log.info("ลบถังขยะของวันที่ %s (เกิน %s วัน)", day, DELETED_RETENTION_DAYS)
            except OSError as exc:
                log.warning("ลบโฟลเดอร์ถังขยะ %s ไม่สำเร็จ: %s", day, exc)
    return removed


@app.get("/api/deleted")
async def list_deleted():
    """รายการทุกอย่างที่อยู่ในถังขยะ เรียงจากลบล่าสุดก่อน"""
    items = []
    for day, name, path in _deleted_files():
        try:
            with open(path, encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as exc:
            log.warning("อ่านไฟล์ถังขยะ %s ไม่สำเร็จ: %s", name, exc)
            continue
        kind = payload.get("kind", "?")
        items.append({
            "id":         f"{day}/{name}",          # ใช้เป็นตัวอ้างตอนกู้คืน
            "deleted_at": payload.get("deleted_at"),
            "kind":       kind,
            "kind_label": _DELETED_KIND_LABEL.get(kind, kind),
            "summary":    _deleted_summary(payload),
            "has_image":  bool(payload.get("image_file")),
        })
    return {"items": items, "total": len(items), "retention_days": DELETED_RETENTION_DAYS}


def _deleted_path(item_id: str) -> str:
    """แปลง id ("<วัน>/<ไฟล์>.json") เป็น path จริง พร้อมกันหลุดออกนอกถังขยะ"""
    base = os.path.realpath(DELETED_DIR)
    target = os.path.realpath(os.path.join(base, item_id))
    if not target.startswith(base + os.sep) or not target.endswith(".json"):
        raise HTTPException(404, "ไม่พบรายการนี้ในถังขยะ")
    if not os.path.isfile(target):
        raise HTTPException(404, "ไม่พบรายการนี้ในถังขยะ")
    return target


@app.post("/api/deleted/restore")
async def restore_deleted(body: Dict[str, str] = Body(...)):
    """กู้คืน 1 รายการจากถังขยะ — insert แถวกลับ + ย้ายไฟล์รูปกลับที่เดิม

    ทำเป็น transaction เดียว ถ้าขั้นไหนพังจะ rollback ทั้งหมด แล้วไฟล์ JSON
    ยังอยู่ในถังขยะเหมือนเดิม (ลองใหม่ได้)

    ⚠ ลำดับการกู้สำคัญ: ถ้าแถวนี้อ้าง FK ไปหาของที่ถูกลบไปแล้วเหมือนกัน
      (เช่นกู้ measurement ที่อ้าง operator ที่โดนลบไป) MySQL จะปฏิเสธ
      ต้องกู้ตัวที่ถูกอ้างก่อน — ข้อความ error จะบอกให้ผู้ใช้รู้
    """
    item_id = (body or {}).get("id", "")
    path = _deleted_path(item_id)
    with open(path, encoding="utf-8") as f:
        payload = json.load(f)

    table   = payload["table"]
    row     = payload["row"]
    related = payload.get("related") or {}

    db = get_db()
    try:
        db.autocommit(False)
        with db.cursor() as cur:
            _block_if_session_running(cur, "กู้คืน")

            def insert(tbl: str, data: Dict[str, Any]):
                cols = ", ".join(f"`{c}`" for c in data)
                marks = ", ".join(["%s"] * len(data))
                cur.execute(f"INSERT INTO `{tbl}` ({cols}) VALUES ({marks})", list(data.values()))

            try:
                insert(table, row)
                # sessions ต้องเข้าหลัง parts_specifications เสมอ (FK ชี้ไปหา)
                for tbl, rows in related.items():
                    for r in rows:
                        insert(tbl, r)
            except pymysql.err.IntegrityError as exc:
                db.rollback()
                code = exc.args[0] if exc.args else 0
                if code == 1062:
                    raise HTTPException(409, "กู้คืนไม่ได้ — มีข้อมูล id นี้อยู่ในระบบแล้ว")
                if code == 1452:
                    raise HTTPException(
                        409,
                        "กู้คืนไม่ได้ — ข้อมูลนี้อ้างอิงถึงรายการอื่นที่ถูกลบไปแล้ว "
                        "กรุณากู้คืนรายการที่ถูกอ้างถึงก่อน (เช่น Operator / Part Number)",
                    )
                raise HTTPException(409, f"กู้คืนไม่สำเร็จ: {exc}")

        # ── ย้ายไฟล์รูปกลับที่เดิม (หลัง DB สำเร็จแล้วเท่านั้น) ──────────────
        image_file  = payload.get("image_file")
        image_path  = row.get("image_path")
        if image_file and image_path:
            src = os.path.join(os.path.dirname(path), image_file)
            dst = os.path.join(ALPL_IMAGE_DIR, image_path)
            try:
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.move(src, dst)
            except OSError as exc:
                # DB กู้แล้วแต่รูปย้ายไม่สำเร็จ — ไม่ rollback เพราะข้อมูลสำคัญกว่ารูป
                log.warning("กู้ข้อมูลสำเร็จแต่ย้ายรูปกลับไม่ได้ (%s): %s", image_file, exc)

        db.commit()
        os.remove(path)   # ออกจากถังขยะแล้ว
        log.info("กู้คืนจากถังขยะ: %s", item_id)
        return {"ok": True, "kind": payload.get("kind"), "table": table}
    finally:
        db.autocommit(True)
        db.close()


@app.post("/api/deleted/remove")
async def remove_deleted(body: Dict[str, str] = Body(...)):
    """ลบ 1 รายการออกจากถังขยะ **ถาวร** — กู้คืนไม่ได้อีก

    ลบทั้งไฟล์ .json และไฟล์รูปที่เก็บคู่กัน ไม่แตะฐานข้อมูลเลย (แถวนั้นถูกลบ
    ไปตั้งแต่ตอนกดลบครั้งแรกแล้ว ตรงนี้แค่ทิ้งตัวสำรอง)
    """
    path = _deleted_path((body or {}).get("id", ""))
    try:
        with open(path, encoding="utf-8") as f:
            image_file = json.load(f).get("image_file")
    except Exception:
        image_file = None

    if image_file:
        # ยึด "โฟลเดอร์ของไฟล์ json" เป็นฐานเสมอ + เอาเฉพาะชื่อไฟล์ กัน image_file
        # ที่มี path ปนมาพาไปลบไฟล์นอกถังขยะ
        img = os.path.join(os.path.dirname(path), os.path.basename(image_file))
        try:
            os.remove(img)
        except OSError:
            pass

    os.remove(path)
    log.info("ลบถาวรจากถังขยะ: %s", (body or {}).get("id"))
    return {"ok": True}


@app.delete("/api/deleted")
async def purge_deleted_now():
    """สั่งลบของในถังขยะที่เกินอายุทันที (ปกติทำอัตโนมัติตอน backend เริ่ม)"""
    return {"ok": True, "removed_days": _purge_old_deleted()}


# ══════════════════════════════════════════════════════════════════════════════
# Static image files (รูป ALPL ที่ upload_measurement_image เซฟไว้)
# ══════════════════════════════════════════════════════════════════════════════
# ต้อง mount ก่อน static mount ที่ "/" ด้านล่างเสมอ (ตัวนั้นเป็น catch-all จับ
# ทุก path ที่เหลือ ถ้า mount ทีหลังจะไม่มีทางไปถึง route นี้เลย) — สร้างโฟลเดอร์
# ไว้ก่อนด้วยเผื่อยังไม่เคยมีรูปมาเลยสักใบ (StaticFiles ต้องการให้ directory
# มีอยู่จริงตอน mount ไม่งั้น import พังทันที)
os.makedirs(ALPL_IMAGE_DIR, exist_ok=True)
app.mount("/media/alpl", StaticFiles(directory=ALPL_IMAGE_DIR), name="alpl-images")


# ══════════════════════════════════════════════════════════════════════════════
# Static dashboard files (index.html / edit.html)
# ══════════════════════════════════════════════════════════════════════════════
# ต้องอยู่ล่างสุดของไฟล์เสมอ — mount ที่ "/" ทำหน้าที่เป็น catch-all ให้ทุก
# path ที่ไม่ตรงกับ route ไหนเลยด้านบน ถ้า register ไว้ก่อน (เช่นบนสุดของไฟล์)
# มันจะดักจับ request ของ /api/... ไปหมดก่อนถึง route จริง ทำให้ API พังทันที
#
# โครงสร้างจริงของโปรเจกต์เป็นแบบนี้ (คนละโฟลเดอร์กับ main.py):
#   TM-X_Project/
#     Backend-server/main.py   (ไฟล์นี้)
#     Backend-pc_station/agent_real.py
#     Frontend/index.html, edit.html, ...
# ดังนั้นต้องถอยขึ้นไป 1 ชั้นจาก main.py แล้วเข้าโฟลเดอร์ Frontend แทนที่จะใช้
# โฟลเดอร์เดียวกับไฟล์นี้ตรงๆ (ที่พังก่อนหน้านี้เพราะ index.html ไม่ได้อยู่ใน
# Backend-server/ ด้วย)
#
# html=True ทำให้เข้า "/" แล้วได้ index.html อัตโนมัติ และเข้า "/edit.html"
# ได้ตรงๆ — เหตุผลที่ทำแบบนี้แทนรัน web server แยก: จะได้มีแค่ process เดียว
# (uvicorn) ให้ autostart/ผูก host=127.0.0.1 ตัวเดียวจบ ไม่ต้องเปิดอีก process
# มาเสิร์ฟไฟล์ static ต่างหาก
_frontend_dir = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Frontend")
)
app.mount(
    "/",
    StaticFiles(directory=_frontend_dir, html=True),
    name="static",
)
